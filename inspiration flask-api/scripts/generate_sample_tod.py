"""
Generate a sample TOD Results Excel to preview the output format.
Uses mock data — no API calls, no evidence folder needed.

Usage:
    python generate_sample_tod.py
"""

import sys
import os
import time

# Navigate up from scripts/ to flask-api/
_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _project_root)
sys.path.insert(0, os.path.join(_project_root, "engines"))

from engines.TOD_Engine import (
    RCMRow, ControlSchema, TOESampleResult, TODControlResult, TOEControlResult,
    aggregate_tod_result, _normalize_attr_value, is_sample_fail,
)


def main():
    # ── Mock RCM rows ──
    rcm_rows = [
        RCMRow(
            process="Procure to Pay",
            subprocess="Invoice Processing",
            control_objective="Ensure invoices are properly approved before payment",
            risk_id="R-P2P-001",
            risk_title="Unauthorized payments",
            risk_description="Payments may be processed without proper authorization, leading to financial loss",
            risk_level="High",
            control_id="C-P2P-001",
            control_description="Three-way match (PO, GRN, Invoice) is performed and approved by AP Manager before payment release",
            control_owner="AP Manager",

            nature_of_control="Preventive",
            control_type="Manual",
            control_frequency="Per occurrence",
            application_system="SAP S/4HANA",
            count_of_samples="1",
        ),
        RCMRow(
            process="Procure to Pay",
            subprocess="Vendor Management",
            control_objective="Ensure new vendors are properly vetted before onboarding",
            risk_id="R-P2P-002",
            risk_title="Fictitious vendor fraud",
            risk_description="Fictitious vendors may be created to siphon company funds",
            risk_level="High",
            control_id="C-P2P-002",
            control_description="New vendor requests require dual approval (Procurement Lead + Finance Controller) with background check documentation",
            control_owner="Procurement Lead",

            nature_of_control="Preventive",
            control_type="Manual",
            control_frequency="Per occurrence",
            application_system="SAP S/4HANA",
            count_of_samples="1",
        ),
        RCMRow(
            process="Revenue",
            subprocess="Billing",
            control_objective="Ensure revenue is recognized in the correct period",
            risk_id="R-REV-001",
            risk_title="Revenue misstatement",
            risk_description="Revenue may be recorded in the wrong period leading to misstated financial statements",
            risk_level="High",
            control_id="C-REV-001",
            control_description="System-enforced billing cutoff: sales orders cannot be invoiced after period close. Monthly reconciliation of deferred revenue by Revenue Accountant",
            control_owner="Revenue Accountant",

            nature_of_control="Detective",
            control_type="IT Dependent Manual",
            control_frequency="Monthly",
            application_system="Oracle EBS",
            count_of_samples="1",
        ),
        RCMRow(
            process="Financial Close",
            subprocess="Journal Entries",
            control_objective="Ensure manual journal entries are properly authorized",
            risk_id="R-FC-001",
            risk_title="Unauthorized journal entries",
            risk_description="Manual journal entries may be posted without proper review, leading to errors or fraud",
            risk_level="Medium",
            control_id="C-FC-001",
            control_description="All manual journal entries above $10,000 require secondary approval from Financial Controller before posting",
            control_owner="Financial Controller",

            nature_of_control="Preventive",
            control_type="IT Dependent Manual",
            control_frequency="Per occurrence",
            application_system="SAP S/4HANA",
            count_of_samples="1",
        ),
    ]

    rcm_lookup = {r.control_id: r for r in rcm_rows}

    # ── Mock schemas (what the LLM would generate) ──
    schemas = {
        "C-P2P-001": ControlSchema(
            control_id="C-P2P-001",
            worksteps=[
                "1. Obtained a sample invoice and traced to corresponding Purchase Order (PO) and Goods Receipt Note (GRN).",
                "2. Verified that the three-way match (PO amount, GRN quantity, Invoice amount) was performed prior to payment.",
                "3. Confirmed AP Manager approval signature and date on the payment voucher.",
                "4. Verified payment was released only after all three documents were matched and approved.",
            ],
            attributes=[
                {"id": "A", "name": "PO Exists", "description": "A valid Purchase Order exists for the invoice and matches the vendor and line items"},
                {"id": "B", "name": "GRN Matched", "description": "Goods Receipt Note confirms physical receipt of goods/services matching the PO"},
                {"id": "C", "name": "Three-Way Match", "description": "Invoice amount, PO amount, and GRN quantity are reconciled within tolerance"},
                {"id": "D", "name": "AP Manager Approval", "description": "AP Manager has reviewed and approved the payment voucher with signature and date"},
            ],
            sample_columns=[
                {"key": "invoice_number", "header": "Invoice Number"},
                {"key": "vendor_name", "header": "Vendor Name"},
                {"key": "invoice_amount", "header": "Invoice Amount"},
                {"key": "po_number", "header": "PO Number"},
            ],
        ),
        "C-P2P-002": ControlSchema(
            control_id="C-P2P-002",
            worksteps=[
                "1. Selected a newly onboarded vendor and obtained the vendor creation request form.",
                "2. Verified dual approval signatures (Procurement Lead + Finance Controller) on the request.",
                "3. Confirmed background check / D&B report was obtained and reviewed prior to approval.",
                "4. Verified vendor master data in SAP matches the approved request form.",
            ],
            attributes=[
                {"id": "A", "name": "Request Form Complete", "description": "Vendor creation request form is fully completed with all required fields"},
                {"id": "B", "name": "Procurement Lead Approval", "description": "Procurement Lead has signed and dated the vendor request"},
                {"id": "C", "name": "Finance Controller Approval", "description": "Finance Controller has signed and dated the vendor request"},
                {"id": "D", "name": "Background Check", "description": "Background check or D&B credit report was obtained and reviewed before approval"},
                {"id": "E", "name": "Master Data Accuracy", "description": "Vendor master data in SAP matches the information on the approved request form"},
            ],
            sample_columns=[
                {"key": "vendor_id", "header": "Vendor ID"},
                {"key": "vendor_name", "header": "Vendor Name"},
                {"key": "creation_date", "header": "Creation Date"},
            ],
        ),
        "C-REV-001": ControlSchema(
            control_id="C-REV-001",
            worksteps=[
                "1. Obtained the month-end billing cutoff report from Oracle EBS.",
                "2. Verified that no sales orders were invoiced after the period close date.",
                "3. Reviewed the deferred revenue reconciliation prepared by Revenue Accountant.",
                "4. Confirmed reconciliation was reviewed and approved by Revenue Manager.",
            ],
            attributes=[
                {"id": "A", "name": "System Cutoff Enforced", "description": "Oracle EBS system configuration prevents invoicing after period close"},
                {"id": "B", "name": "Cutoff Report Reviewed", "description": "Billing cutoff report was generated and reviewed for exceptions"},
                {"id": "C", "name": "Deferred Revenue Reconciled", "description": "Monthly deferred revenue reconciliation was performed with no unexplained differences"},
                {"id": "D", "name": "Manager Review", "description": "Revenue Manager reviewed and approved the reconciliation with sign-off"},
            ],
            sample_columns=[
                {"key": "period", "header": "Period"},
                {"key": "cutoff_date", "header": "Cutoff Date"},
                {"key": "deferred_balance", "header": "Deferred Revenue Balance"},
            ],
        ),
        "C-FC-001": ControlSchema(
            control_id="C-FC-001",
            worksteps=[
                "1. Selected a manual journal entry above $10,000 from the JE listing.",
                "2. Verified the journal entry was prepared with supporting documentation.",
                "3. Confirmed Financial Controller approval was obtained before posting.",
                "4. Verified the journal entry was posted to the correct accounts and period.",
            ],
            attributes=[
                {"id": "A", "name": "Supporting Documentation", "description": "Journal entry has adequate supporting documentation attached"},
                {"id": "B", "name": "Controller Approval", "description": "Financial Controller approved the journal entry before posting"},
                {"id": "C", "name": "Correct Posting", "description": "Journal entry was posted to the correct GL accounts and reporting period"},
            ],
            sample_columns=[
                {"key": "je_number", "header": "JE Number"},
                {"key": "je_amount", "header": "Amount"},
                {"key": "posting_date", "header": "Posting Date"},
                {"key": "preparer", "header": "Prepared By"},
            ],
        ),
    }

    # ── Mock sample results (what the LLM evaluation would return) ──

    # C-P2P-001: PASS
    sr1 = TOESampleResult(
        control_id="C-P2P-001", sample_id="C-P2P-001",
        result="PASS", operated_effectively="Yes",
        control_performed="Yes", timely_execution="Yes",
        accurate_execution="Yes", authorized_performer="Yes",
        evidence_sufficient="Yes",
        remarks="Three-way match was properly performed. PO #4500012345 matched to GRN and invoice. AP Manager J. Smith approved on 2025-01-15, two days before payment date.",
        deviation_details="None",
        raw_evidence="Invoice #INV-2025-0042 from Acme Corp for $18,500...",
        sample_details={
            "Invoice Number": "INV-2025-0042",
            "Vendor Name": "Acme Corp Ltd",
            "Invoice Amount": "$18,500.00",
            "PO Number": "PO-4500012345",
        },
        attribute_results={"A": "Yes", "B": "Yes", "C": "Yes", "D": "Yes"},
        attribute_reasoning={
            "A": "PO #4500012345 exists in SAP, matching vendor Acme Corp and line items for office supplies.",
            "B": "GRN #5000067890 confirms receipt of 500 units on 2025-01-10, matching PO quantities.",
            "C": "Invoice amount $18,500 matches PO amount $18,500 and GRN confirmed full delivery. Within tolerance.",
            "D": "AP Manager J. Smith signed payment voucher on 2025-01-15 with date stamp. Approval is before payment release date of 2025-01-17.",
        },
    )

    # C-P2P-002: FAIL (missing background check)
    sr2 = TOESampleResult(
        control_id="C-P2P-002", sample_id="C-P2P-002",
        result="FAIL", operated_effectively="No",
        control_performed="Yes", timely_execution="Yes",
        accurate_execution="No", authorized_performer="Yes",
        evidence_sufficient="No",
        remarks="Vendor creation form was approved by both parties, but no background check or D&B report was obtained prior to vendor onboarding.",
        deviation_details="Background check / D&B credit report was not obtained before vendor approval. Control design requires background verification but no evidence of this step was found.",
        raw_evidence="Vendor creation request for TechSupply Inc...",
        sample_details={
            "Vendor ID": "V-100456",
            "Vendor Name": "TechSupply Inc",
            "Creation Date": "2025-02-03",
        },
        attribute_results={"A": "Yes", "B": "Yes", "C": "Yes", "D": "No", "E": "Yes"},
        attribute_reasoning={
            "A": "Vendor creation request form is fully completed with company name, address, tax ID, and bank details.",
            "B": "Procurement Lead M. Johnson signed the request on 2025-01-28.",
            "C": "Finance Controller R. Chen signed the request on 2025-01-30.",
            "D": "No background check or D&B credit report found in the vendor file. The control requires this verification before approval, but it was not performed.",
            "E": "Vendor master data in SAP (name, address, bank details) matches the approved request form.",
        },
    )

    # C-REV-001: PASS
    sr3 = TOESampleResult(
        control_id="C-REV-001", sample_id="C-REV-001",
        result="PASS", operated_effectively="Yes",
        control_performed="Yes", timely_execution="Yes",
        accurate_execution="Yes", authorized_performer="Yes",
        evidence_sufficient="Yes",
        remarks="System cutoff was enforced. Deferred revenue reconciliation for January 2025 was completed by Revenue Accountant on Feb-3 and approved by Revenue Manager on Feb-5. No exceptions noted.",
        deviation_details="None",
        raw_evidence="January 2025 billing cutoff report and deferred revenue reconciliation...",
        sample_details={
            "Period": "January 2025",
            "Cutoff Date": "2025-01-31",
            "Deferred Revenue Balance": "$1,245,000",
        },
        attribute_results={"A": "Yes", "B": "Yes", "C": "Yes", "D": "Yes"},
        attribute_reasoning={
            "A": "Oracle EBS period status shows January 2025 was closed on 2025-01-31 at 23:59. No invoices were created after this timestamp.",
            "B": "Billing cutoff report was generated on 2025-02-01 and reviewed. Zero exceptions identified.",
            "C": "Deferred revenue reconciliation shows opening balance $1,180,000, additions $320,000, releases $255,000, closing $1,245,000. All differences explained.",
            "D": "Revenue Manager K. Patel signed the reconciliation on 2025-02-05 with 'Reviewed and Approved' notation.",
        },
    )

    # C-FC-001: PASS
    sr4 = TOESampleResult(
        control_id="C-FC-001", sample_id="C-FC-001",
        result="PASS", operated_effectively="Yes",
        control_performed="Yes", timely_execution="Yes",
        accurate_execution="Yes", authorized_performer="Yes",
        evidence_sufficient="Yes",
        remarks="Manual journal entry JE-2025-0198 for $45,000 was properly supported with invoice documentation and approved by Financial Controller before posting.",
        deviation_details="None",
        raw_evidence="Journal entry JE-2025-0198 for intercompany allocation...",
        sample_details={
            "JE Number": "JE-2025-0198",
            "Amount": "$45,000.00",
            "Posting Date": "2025-01-31",
            "Prepared By": "Senior Accountant L. Wang",
        },
        attribute_results={"A": "Yes", "B": "Yes", "C": "Yes"},
        attribute_reasoning={
            "A": "JE is supported by intercompany allocation schedule and underlying invoice from subsidiary.",
            "B": "Financial Controller D. Martinez approved the JE on 2025-01-30, one day before posting date.",
            "C": "JE debits account 6500 (Intercompany Expense) and credits 2100 (Intercompany Payable) in period 01-2025. Correct accounts and period.",
        },
    )

    sample_results = {
        "C-P2P-001": sr1,
        "C-P2P-002": sr2,
        "C-REV-001": sr3,
        "C-FC-001": sr4,
    }

    # ── Aggregate TOD results ──
    tod_results = []
    for cid in ["C-P2P-001", "C-P2P-002", "C-REV-001", "C-FC-001"]:
        rcm = rcm_lookup[cid]
        sr = sample_results[cid]
        tod_r = aggregate_tod_result(cid, rcm, sr)
        tod_r.schema = schemas[cid]
        tod_results.append(tod_r)

    # ── Convert to TOE format and export (same format as real engine) ──
    from engines.TOD_Engine import TOEControlResult
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    toe_results = []
    for r in tod_results:
        sr = r.sample_result
        is_fail = r.result == "FAIL"
        toe_r = TOEControlResult(
            control_id=r.control_id,
            risk_id=r.risk_id,
            risk_title=r.risk_title,
            control_type=r.control_type,
            nature_of_control=r.nature_of_control,
            control_frequency=r.control_frequency,
            total_samples=1,
            passed_samples=0 if is_fail else 1,
            failed_samples=1 if is_fail else 0,
            deviation_rate=1.0 if is_fail else 0.0,
            operating_effectiveness="Not Effective" if is_fail else "Effective",
            deficiency_type=r.deficiency_type,
            overall_remarks=r.overall_remarks,
            sample_results=[sr] if sr else [],
            evaluation_timestamp=r.evaluation_timestamp,
            schema=r.schema,
        )
        toe_results.append(toe_r)

    # Now build the workpaper manually (replicating the TOE export format)
    # We can't call tester.export_toe_workpaper() without a real tester instance,
    # so we replicate the key parts here.

    wb = Workbook()

    hfont = Font(bold=True, color="FFFFFF", size=10)
    bfont = Font(bold=True, size=10)
    nfont = Font(size=10)
    sfont = Font(size=9)
    bdr = Border(left=Side("thin"), right=Side("thin"),
                 top=Side("thin"), bottom=Side("thin"))
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    fill_label = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_attr = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_red_banner = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_green_banner = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fill_amber_banner = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

    fill_rcm = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    fill_result = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_agg = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    exc_fill = PatternFill(start_color="FFEB9C", fill_type="solid")
    sd_fill = PatternFill(start_color="F4B084", fill_type="solid")
    mw_fill = PatternFill(start_color="FF4444", fill_type="solid")
    fill_remediation = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

    # ============================================================
    #  SHEET 1: Control Summary
    # ============================================================
    ws_sum = wb.active
    ws_sum.title = "Control Summary"

    sum_headers = [
        ("Process", fill_rcm), ("Sub Process", fill_rcm), ("Control Objective", fill_rcm),
        ("Risk ID", fill_rcm), ("Risk Title", fill_rcm), ("Risk Description", fill_rcm),
        ("Risk Level", fill_rcm), ("Control ID", fill_rcm), ("Control Description", fill_rcm),
        ("Control Owner", fill_rcm), ("Nature of Control", fill_rcm),
        ("Control Type", fill_rcm), ("Control Frequency", fill_rcm), ("Application/System", fill_rcm),
        ("Detailed Testing Steps", fill_rcm), ("RCM Sample Size", fill_rcm),
        ("Samples Tested", fill_agg), ("Passed / Total", fill_agg),
        ("Failed / Total", fill_agg), ("Deviation Rate", fill_agg),
        ("Validator Overrides", fill_agg), ("Deficiency Type", fill_result),
        ("Failed Sample Details", fill_result), ("Overall Remarks", fill_result),
        ("Remedial Actions", fill_remediation),
    ]
    for c, (h, f) in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = f
        cell.border = bdr
        cell.alignment = center

    for ri, r in enumerate(toe_results, 2):
        rcm = rcm_lookup.get(r.control_id)
        fd = []
        for sr in r.sample_results:
            if is_sample_fail(sr):
                d = sr.deviation_details if sr.deviation_details and sr.deviation_details.lower() not in ("none", "n/a") else sr.remarks
                fd.append(f"{sr.sample_id}: {d}")
        testing_steps = ""
        if r.schema and r.schema.worksteps:
            testing_steps = "\n".join(r.schema.worksteps)

        # Simple remediation
        if r.deficiency_type == "None":
            remedial = "No remediation required -- control design is adequate."
        else:
            remedial = f"Investigate design gap. Redesign control to address: {fd[0] if fd else 'identified deficiency'}."

        vals = [
            rcm.process if rcm else "", rcm.subprocess if rcm else "",
            rcm.control_objective if rcm else "", r.risk_id, r.risk_title,
            rcm.risk_description if rcm else "", rcm.risk_level if rcm else "",
            r.control_id, rcm.control_description if rcm else "",
            rcm.control_owner if rcm else "",
            r.nature_of_control, r.control_type, r.control_frequency,
            rcm.application_system if rcm else "", testing_steps,
            rcm.count_of_samples if rcm else "",
            r.total_samples, f"{r.passed_samples}/{r.total_samples}",
            f"{r.failed_samples}/{r.total_samples}", f"{r.deviation_rate:.1%}",
            "", r.deficiency_type, "; ".join(fd) if fd else "None", r.overall_remarks,
            remedial,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=ri, column=c, value=v)
            cell.alignment = wrap
            cell.border = bdr
        dc = ws_sum.cell(row=ri, column=22)
        if r.deficiency_type == "None":
            dc.fill = fill_pass
        elif "Control" in r.deficiency_type:
            dc.fill = exc_fill
        elif "Significant" in r.deficiency_type:
            dc.fill = sd_fill
        else:
            dc.fill = mw_fill

    widths_sum = [18, 18, 35, 10, 25, 40, 10, 12, 40, 16, 14, 12, 14, 16, 45, 14, 14, 14, 14, 12, 14, 22, 60, 60, 55]
    for i, w in enumerate(widths_sum, 1):
        if i <= len(widths_sum):
            ws_sum.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    #  PER-CONTROL WORKPAPER SHEETS
    # ============================================================
    for r in toe_results:
        rcm = rcm_lookup.get(r.control_id)
        if not rcm:
            continue
        schema = r.schema

        sheet_name = r.control_id[:31]
        ws = wb.create_sheet(title=sheet_name)

        def label_val(row, label, value, merge_to=8):
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = bfont
            lc.fill = fill_label
            lc.border = bdr
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = nfont
            vc.alignment = wrap
            vc.border = bdr
            if merge_to > 2:
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=merge_to)

        row = 1
        ws.cell(row=row, column=1, value="Test of Design (TOD) Workpaper").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 2

        label_val(row, "Process:", rcm.process); row += 1
        label_val(row, "Sub Process:", rcm.subprocess); row += 1
        label_val(row, "Control No:", rcm.control_id); row += 1
        label_val(row, "Control Description:", rcm.control_description); row += 1
        label_val(row, "Control Type:", f"{rcm.control_type} | {rcm.nature_of_control}"); row += 1
        label_val(row, "Control Frequency:", rcm.control_frequency); row += 1
        label_val(row, "Control Owner:", rcm.control_owner); row += 1
        label_val(row, "RCM Sample Size:", rcm.count_of_samples); row += 1
        label_val(row, "Sample Size:", "1 (TOD -- design walkthrough)"); row += 1

        # Worksteps
        if schema and schema.worksteps:
            ws_text = "\n".join(schema.worksteps)
            label_val(row, "Worksteps Performed:", ws_text)
            ws.row_dimensions[row].height = max(30, 15 * len(schema.worksteps))
            row += 1

        row += 1

        # Attributes legend
        if schema and schema.attributes:
            c1 = ws.cell(row=row, column=1, value="Attribute")
            c1.font = bfont; c1.fill = fill_attr; c1.border = bdr; c1.alignment = center
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=2).fill = fill_attr; ws.cell(row=row, column=2).border = bdr
            c3 = ws.cell(row=row, column=3, value="Attribute to be tested")
            c3.font = bfont; c3.fill = fill_attr; c3.border = bdr
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
            for cc in range(3, 11):
                ws.cell(row=row, column=cc).fill = fill_attr
                ws.cell(row=row, column=cc).border = bdr
            row += 1
            for attr in schema.attributes:
                ws.cell(row=row, column=1, value=attr["id"]).font = bfont
                ws.cell(row=row, column=1).alignment = center
                ws.cell(row=row, column=1).border = bdr
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.cell(row=row, column=2).border = bdr
                desc_text = f"{attr['name']}: {attr['description']}" if attr.get('description') else attr['name']
                ws.cell(row=row, column=3, value=desc_text).font = nfont
                ws.cell(row=row, column=3).alignment = wrap
                ws.cell(row=row, column=3).border = bdr
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
                for cc in range(3, 11):
                    ws.cell(row=row, column=cc).border = bdr
                row += 1
            row += 1

        # Testing table
        row += 1
        table_start_row = row

        sr_no_col = 2
        sample_no_col = 3
        detail_start_col = 4

        sample_col_keys = []
        if schema and schema.sample_columns:
            for sc in schema.sample_columns:
                sample_col_keys.append(sc["header"])
        else:
            sample_col_keys = ["Source Document"]
        n_detail_cols = len(sample_col_keys)
        detail_end_col = detail_start_col + n_detail_cols - 1

        attr_start_col = detail_end_col + 1
        attr_ids = []
        attr_names = []
        if schema and schema.attributes:
            for attr in schema.attributes:
                attr_ids.append(attr["id"])
                attr_names.append(attr["name"])
        n_attr_cols = len(attr_ids)
        attr_end_col = attr_start_col + n_attr_cols - 1 if n_attr_cols > 0 else attr_start_col - 1

        analysis_col = attr_end_col + 1
        remarks_col = analysis_col + 1
        wpref_col = remarks_col + 1
        last_col = wpref_col

        h1 = row
        h2 = row + 1

        c = ws.cell(row=h1, column=sr_no_col, value="Sr. No.")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        ws.merge_cells(start_row=h1, start_column=sr_no_col, end_row=h2, end_column=sr_no_col)

        tsd_start = sample_no_col
        tsd_end = detail_end_col
        c = ws.cell(row=h1, column=tsd_start, value="Testing Sample Details")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        if tsd_end > tsd_start:
            ws.merge_cells(start_row=h1, start_column=tsd_start, end_row=h1, end_column=tsd_end)

        c = ws.cell(row=h2, column=sample_no_col, value="Sample No.")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        for idx_sc, sc_header in enumerate(sample_col_keys):
            col_num = detail_start_col + idx_sc
            c = ws.cell(row=h2, column=col_num, value=sc_header)
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center

        attr_hfont = Font(bold=True, color="000000", size=10)
        attr_sfont = Font(bold=True, color="000000", size=9)
        for idx_a, (aid, aname) in enumerate(zip(attr_ids, attr_names)):
            col_num = attr_start_col + idx_a
            c1 = ws.cell(row=h1, column=col_num, value=f"Attribute {aid}")
            c1.font = attr_hfont; c1.fill = fill_attr; c1.border = bdr; c1.alignment = center
            c2 = ws.cell(row=h2, column=col_num, value=aname)
            c2.font = attr_sfont; c2.fill = fill_attr; c2.border = bdr; c2.alignment = center

        c = ws.cell(row=h1, column=analysis_col, value="Attribute Analysis")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        ws.merge_cells(start_row=h1, start_column=analysis_col, end_row=h2, end_column=analysis_col)

        c = ws.cell(row=h1, column=remarks_col, value="Remarks")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        ws.merge_cells(start_row=h1, start_column=remarks_col, end_row=h2, end_column=remarks_col)

        c = ws.cell(row=h1, column=wpref_col, value="Workpaper Reference")
        c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
        ws.merge_cells(start_row=h1, start_column=wpref_col, end_row=h2, end_column=wpref_col)

        for col_num in range(2, last_col + 1):
            for hr in (h1, h2):
                ws.cell(row=hr, column=col_num).border = bdr

        row = h2 + 1

        # Sample data row (just 1 for TOD)
        for idx, sr in enumerate(r.sample_results, 1):
            ws.cell(row=row, column=sr_no_col, value=idx).border = bdr
            ws.cell(row=row, column=sr_no_col).alignment = center

            ws.cell(row=row, column=sample_no_col, value=sr.sample_id).border = bdr
            ws.cell(row=row, column=sample_no_col).alignment = center
            ws.cell(row=row, column=sample_no_col).font = nfont

            if schema and schema.sample_columns:
                for idx_sc, sc in enumerate(schema.sample_columns):
                    col_num = detail_start_col + idx_sc
                    val = ""
                    if sr.sample_details and isinstance(sr.sample_details, dict):
                        val = sr.sample_details.get(sc["header"], "")
                    cell = ws.cell(row=row, column=col_num, value=val)
                    cell.border = bdr; cell.alignment = wrap; cell.font = nfont

            for idx_a, aid in enumerate(attr_ids):
                col_num = attr_start_col + idx_a
                val = ""
                if sr.attribute_results and isinstance(sr.attribute_results, dict):
                    val = sr.attribute_results.get(aid, "")
                cell = ws.cell(row=row, column=col_num, value=val)
                cell.border = bdr; cell.alignment = center; cell.font = nfont
                norm_val = _normalize_attr_value(str(val))
                if norm_val == "Yes":
                    cell.fill = fill_pass
                elif norm_val == "No":
                    cell.fill = fill_fail

                if sr.attribute_reasoning and isinstance(sr.attribute_reasoning, dict):
                    reason_text = sr.attribute_reasoning.get(aid, "")
                    if reason_text and str(reason_text).strip().lower() not in ("", "n/a", "none"):
                        cell.comment = Comment(
                            f"LLM Assessment:\n{reason_text}",
                            "TOD Evaluator", width=350, height=120
                        )

            # Attribute Analysis column
            analysis_parts = []
            if sr.attribute_reasoning and isinstance(sr.attribute_reasoning, dict):
                attr_name_map = {a["id"]: a["name"] for a in schema.attributes} if schema and schema.attributes else {}
                for aid in attr_ids:
                    aval = sr.attribute_results.get(aid, "?") if sr.attribute_results else "?"
                    areason = sr.attribute_reasoning.get(aid, "")
                    aname = attr_name_map.get(aid, f"Attr {aid}")
                    if areason:
                        analysis_parts.append(f"{aid} ({aname}): {aval} -- {areason}")
                    else:
                        analysis_parts.append(f"{aid} ({aname}): {aval}")
            ac = ws.cell(row=row, column=analysis_col, value="\n".join(analysis_parts))
            ac.border = bdr; ac.alignment = wrap; ac.font = nfont

            # Remarks
            is_fail_sample = is_sample_fail(sr)
            rmk = ""
            if is_fail_sample:
                rmk = f"[FAIL] {sr.deviation_details}" if sr.deviation_details and sr.deviation_details.lower() not in ("none", "n/a") else f"[FAIL] {sr.remarks}"
            else:
                rmk = sr.remarks if sr.remarks else ""
            cell = ws.cell(row=row, column=remarks_col, value=rmk)
            cell.border = bdr; cell.alignment = wrap; cell.font = nfont
            if is_fail_sample:
                cell.font = Font(size=10, color="C00000")
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            wp_ref = f"{r.control_id}.T.{sr.sample_id}"
            ws.cell(row=row, column=wpref_col, value=wp_ref).border = bdr
            ws.cell(row=row, column=wpref_col).alignment = center
            ws.cell(row=row, column=wpref_col).font = nfont

            row += 1

        # Summary totals
        ws.cell(row=row, column=sr_no_col, value="TOTAL").font = bfont
        ws.cell(row=row, column=sr_no_col).alignment = center
        ws.cell(row=row, column=sr_no_col).border = bdr
        ws.cell(row=row, column=sr_no_col).fill = fill_label
        total_text = (f"Samples: {r.total_samples} | "
                      f"Passed: {r.passed_samples} | "
                      f"Failed: {r.failed_samples} | "
                      f"Deviation Rate: {r.deviation_rate:.1%}")
        tc = ws.cell(row=row, column=sample_no_col, value=total_text)
        tc.font = bfont; tc.alignment = wrap; tc.border = bdr; tc.fill = fill_label
        ws.merge_cells(start_row=row, start_column=sample_no_col,
                       end_row=row, end_column=last_col)
        for cc in range(sample_no_col, last_col + 1):
            ws.cell(row=row, column=cc).border = bdr
            ws.cell(row=row, column=cc).fill = fill_label
        row += 1

        data_end_row = row - 1

        # "Testing:" label
        testing_cell = ws.cell(row=table_start_row, column=1, value="Testing:")
        testing_cell.font = bfont
        testing_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        testing_cell.border = bdr
        if data_end_row > table_start_row:
            ws.merge_cells(start_row=table_start_row, start_column=1,
                           end_row=data_end_row, end_column=1)
        for rr in range(table_start_row, data_end_row + 1):
            ws.cell(row=rr, column=1).border = bdr

        row += 1

        # Conclusion banner
        conc_text = r.operating_effectiveness
        if r.deficiency_type != "None":
            conc_text += f" -- {r.deficiency_type}"
        conc_fill = fill_green_banner if r.operating_effectiveness == "Effective" else \
                    fill_amber_banner if r.operating_effectiveness == "Effective with Exceptions" else \
                    fill_red_banner

        ws.cell(row=row, column=1, value="Conclusion:").font = bfont
        ws.cell(row=row, column=1).border = bdr
        conc_cell = ws.cell(row=row, column=2, value=conc_text)
        conc_cell.font = Font(bold=True, color="FFFFFF", size=11)
        conc_cell.fill = conc_fill
        conc_cell.border = bdr
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 2

        # Sign-off
        ws.cell(row=row, column=1, value="Prepared by").font = bfont
        ws.cell(row=row, column=1).border = bdr
        ws.cell(row=row, column=2, value="").border = bdr
        ws.cell(row=row, column=4, value="Reviewed by").font = bfont
        ws.cell(row=row, column=4).border = bdr
        ws.cell(row=row, column=5, value="").border = bdr
        row += 1
        ws.cell(row=row, column=1, value="Date").font = bfont
        ws.cell(row=row, column=1).border = bdr
        ws.cell(row=row, column=2, value=time.strftime("%d-%b-%Y")).border = bdr
        ws.cell(row=row, column=4, value="Date").font = bfont
        ws.cell(row=row, column=4).border = bdr
        ws.cell(row=row, column=5, value="").border = bdr

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        for c_idx in range(detail_start_col, detail_end_col + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 22
        for c_idx in range(attr_start_col, attr_end_col + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 20
        ws.column_dimensions[get_column_letter(analysis_col)].width = 55
        ws.column_dimensions[get_column_letter(remarks_col)].width = 55
        ws.column_dimensions[get_column_letter(wpref_col)].width = 18

    # Save
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Sample_TOD_Results.xlsx")
    wb.save(output_path)

    print(f"\nSample TOD workpaper saved to: {output_path}")
    print(f"\nContents:")
    print(f"  Sheet 1: Control Summary (4 controls)")
    print(f"  Sheet 2: C-P2P-001 -- PASS (three-way match)")
    print(f"  Sheet 3: C-P2P-002 -- FAIL (missing background check)")
    print(f"  Sheet 4: C-REV-001 -- PASS (billing cutoff)")
    print(f"  Sheet 5: C-FC-001  -- PASS (journal entry approval)")
    print(f"\nThis is the same format TOE produces -- summary + per-control workpaper sheets")
    print(f"with attributes, sample details, LLM reasoning, and conclusion banners.")


if __name__ == "__main__":
    main()
