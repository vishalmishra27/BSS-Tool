"""Evidence Validation Module — compare uploaded evidence against Required Documents list.

Runs AFTER the user provides the evidence folder and BEFORE the actual TOD/TOE test.
Extracts evidence files once and uses them for:
  1. LLM-based matching against the Required Documents list (this module)
  2. The TOD/TOE evaluation engine (same loaded evidence, no re-parsing)

Uses an LLM (not embeddings) to match evidence files against required documents.
This allows one evidence file to cover multiple required documents and understands
semantic context far better than cosine similarity.

Returns a per-control match report showing which required documents are found
(matched) and which are missing.
"""

from __future__ import annotations

import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np

logger = logging.getLogger("agent.tools.evidence_validator")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class EvidenceFileInfo:
    """Metadata + embedding for one evidence file."""
    file_name: str
    file_path: str
    control_id: str
    content_summary: str  # first ~500 chars of extracted text
    embedding: Optional[np.ndarray] = None


@dataclass
class DocumentMatch:
    """Match result for one required document."""
    document_name: str
    document_description: str
    matched: bool
    matched_file: Optional[str] = None
    similarity_score: float = 0.0


@dataclass
class ControlValidationResult:
    """Validation result for one control."""
    control_id: str
    total_required: int
    total_matched: int
    total_missing: int
    documents: List[DocumentMatch] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Embedding helpers
# ---------------------------------------------------------------------------

def _get_embeddings(texts: List[str], api_key: str, endpoint: str,
                    api_version: str, model: str) -> np.ndarray:
    """Batch-embed a list of texts using Azure OpenAI."""
    from openai import AzureOpenAI

    client = AzureOpenAI(
        api_key=api_key,
        azure_endpoint=endpoint,
        api_version=api_version,
    )

    # Azure embedding API has a batch limit; chunk if needed
    BATCH_SIZE = 16
    all_embeddings = []

    for i in range(0, len(texts), BATCH_SIZE):
        batch = texts[i:i + BATCH_SIZE]
        # Truncate each text to avoid token limits
        batch = [t[:8000] if len(t) > 8000 else t for t in batch]
        try:
            resp = client.embeddings.create(model=model, input=batch)
            for d in resp.data:
                all_embeddings.append(d.embedding)
        except Exception as exc:
            logger.error("Embedding batch %d failed: %s", i // BATCH_SIZE, exc)
            # Append zero vectors as fallback
            for _ in batch:
                all_embeddings.append([0.0] * 1536)

    return np.array(all_embeddings, dtype=np.float32)


def _cosine_similarity(a: np.ndarray, b: np.ndarray) -> float:
    """Cosine similarity between two vectors."""
    norm = np.linalg.norm(a) * np.linalg.norm(b)
    if norm == 0:
        return 0.0
    return float(np.dot(a, b) / norm)


# ---------------------------------------------------------------------------
# LLM-based document matching
# ---------------------------------------------------------------------------

_LLM_MATCH_SYSTEM = (
    "You are an expert audit evidence reviewer. You are given:\n"
    "1. A list of REQUIRED DOCUMENTS that an auditor expects for a specific control.\n"
    "2. A list of EVIDENCE FILES actually uploaded, with their content previews.\n\n"
    "Your task: for each required document, determine whether ANY of the evidence files\n"
    "cover or satisfy that requirement. One evidence file CAN cover MULTIPLE required\n"
    "documents (e.g. a single policy PDF may satisfy both a 'policy document' and an\n"
    "'approval workflow document' requirement).\n\n"
    "RULES:\n"
    "- Match based on CONTENT, not just filename.\n"
    "- A required document is MATCHED if any evidence file contains information that\n"
    "  substantively addresses what the required document describes.\n"
    "- Be generous but not reckless — if the content clearly relates to the requirement,\n"
    "  mark it as matched. If there is no relevant content at all, mark it missing.\n"
    "- Return ONLY valid JSON, no markdown fences.\n\n"
    "Return JSON:\n"
    "{\n"
    '  "matches": [\n'
    "    {\n"
    '      "required_document": "<name of required doc>",\n'
    '      "matched": true/false,\n'
    '      "matched_file": "<filename that covers it, or null>",\n'
    '      "confidence": "<high/medium/low>",\n'
    '      "reason": "<brief reason for match or why missing>"\n'
    "    }\n"
    "  ]\n"
    "}"
)


def _llm_match_for_control(
    control_id: str,
    evidence_files: List[Dict],
    required_docs: List[Dict],
    api_key: str,
    endpoint: str,
    api_version: str,
    llm_model: str,
) -> List[DocumentMatch]:
    """Use LLM to match evidence files against required documents for one control.

    One evidence file can cover multiple required documents.
    Returns a DocumentMatch for each required document.
    """
    from openai import AzureOpenAI

    if not required_docs:
        return []

    # If no evidence files at all, everything is missing
    if not evidence_files:
        return [
            DocumentMatch(
                document_name=d.get("name", ""),
                document_description=d.get("description", ""),
                matched=False, matched_file=None, similarity_score=0.0,
            )
            for d in required_docs
        ]

    # Build the user prompt with evidence previews (truncated for token budget)
    _PREVIEW_MAX = 1500
    ev_parts = []
    for i, ef in enumerate(evidence_files, 1):
        preview = ef.get("content_preview", ef.get("file_name", ""))
        if len(preview) > _PREVIEW_MAX:
            preview = preview[:_PREVIEW_MAX] + "..."
        ev_parts.append(f"  {i}. **{ef['file_name']}**\n     Content: {preview}")

    doc_parts = []
    for i, d in enumerate(required_docs, 1):
        doc_parts.append(
            f"  {i}. **{d.get('name', f'Document {i}')}**\n"
            f"     Description: {d.get('description', 'N/A')}"
        )

    user_prompt = (
        f"CONTROL: {control_id}\n\n"
        f"EVIDENCE FILES ({len(evidence_files)}):\n"
        + "\n".join(ev_parts) + "\n\n"
        f"REQUIRED DOCUMENTS ({len(required_docs)}):\n"
        + "\n".join(doc_parts) + "\n\n"
        "Match each required document to the evidence files. "
        "One evidence file can cover multiple required documents."
    )

    try:
        client = AzureOpenAI(
            api_key=api_key,
            azure_endpoint=endpoint,
            api_version=api_version,
        )
        resp = client.chat.completions.create(
            model=llm_model,
            messages=[
                {"role": "system", "content": _LLM_MATCH_SYSTEM},
                {"role": "user", "content": user_prompt},
            ],
            max_completion_tokens=4096,
            response_format={"type": "json_object"},
        )
        text = resp.choices[0].message.content or "{}"
        result = json.loads(text)
        raw_matches = result.get("matches", [])

        # Build DocumentMatch objects from LLM response
        # Map by required document name for lookup
        doc_name_map = {d.get("name", ""): d for d in required_docs}
        matched_results: List[DocumentMatch] = []
        seen_docs = set()

        for m in raw_matches:
            doc_name = m.get("required_document", "")
            is_matched = m.get("matched", False)
            matched_file = m.get("matched_file") if is_matched else None
            confidence = m.get("confidence", "low")
            confidence_score = {"high": 0.95, "medium": 0.75, "low": 0.5}.get(confidence, 0.5)

            # Find the matching required doc
            doc_info = doc_name_map.get(doc_name, {})
            matched_results.append(DocumentMatch(
                document_name=doc_name,
                document_description=doc_info.get("description", ""),
                matched=is_matched,
                matched_file=matched_file,
                similarity_score=confidence_score if is_matched else 0.0,
            ))
            seen_docs.add(doc_name)

        # Any required docs the LLM didn't mention → mark as missing
        for d in required_docs:
            dname = d.get("name", "")
            if dname not in seen_docs:
                matched_results.append(DocumentMatch(
                    document_name=dname,
                    document_description=d.get("description", ""),
                    matched=False, matched_file=None, similarity_score=0.0,
                ))

        logger.info(
            "LLM match for %s: %d/%d required docs covered",
            control_id,
            sum(1 for m in matched_results if m.matched),
            len(required_docs),
        )
        return matched_results

    except Exception as exc:
        logger.error("LLM evidence matching failed for %s: %s", control_id, exc)
        # Fallback: mark all as missing
        return [
            DocumentMatch(
                document_name=d.get("name", ""),
                document_description=d.get("description", ""),
                matched=False, matched_file=None, similarity_score=0.0,
            )
            for d in required_docs
        ]


# ---------------------------------------------------------------------------
# Evidence file scanning
# ---------------------------------------------------------------------------

def _scan_evidence_files(
    evidence_folder: str,
    control_ids: List[str],
    extract_cache: Optional[Dict[str, tuple]] = None,
    per_sample: bool = False,
) -> Tuple[Dict[str, Any], Dict[str, tuple]]:
    """Scan evidence folder and return file info per control (or per sample).

    Also builds an extraction cache (file_path -> (content, doc_type, ok)) that
    can be passed to the TOD/TOE engine so files are extracted only ONCE.

    Parameters
    ----------
    extract_cache : dict, optional
        If provided, reuse cached extractions instead of re-extracting.
    per_sample : bool
        If True (for TOE), returns files grouped by sample subfolder:
        ``{control_id: {sample_name: [file_info, ...]}}``
        If False (for TOD), returns files flat per control:
        ``{control_id: [file_info, ...]}``

    Returns
    -------
    (file_info_dict, extract_cache)
        file_info_dict: see per_sample above
        extract_cache:  dict[abs_file_path -> (content_text, doc_type, ok)]
    """
    ev_path = Path(evidence_folder)
    result: Dict[str, List[Dict]] = {}
    if extract_cache is None:
        extract_cache = {}

    # Build case-insensitive mapping
    folder_map: Dict[str, Path] = {}
    for d in ev_path.iterdir():
        if d.is_dir() and not d.name.startswith("_"):
            folder_map[d.name.strip().upper()] = d

    cid_upper_map = {cid.strip().upper(): cid for cid in control_ids}

    # Max chars for the embedding preview — Azure text-embedding models
    # support ~8191 tokens (~6000 chars of dense text).  We keep a generous
    # budget so the semantic comparison sees enough of the document content
    # to match against the required documents list.  The FULL (un-truncated)
    # text is stored separately in extract_cache for the TOD/TOE engine.
    _EMBED_PREVIEW_MAX = 6000

    def _normalize_cache_key(p: Path) -> str:
        """Normalize path for cache key consistency across Phase 2 and Phase 3."""
        return os.path.normcase(os.path.normpath(str(p.resolve())))

    def _extract_one(f: Path) -> Tuple[str, str, str, bool]:
        """Extract one file. Returns (abs_key, content, doc_type, ok)."""
        abs_key = _normalize_cache_key(f)

        # Already cached
        if abs_key in extract_cache:
            content, doc_type, ok = extract_cache[abs_key]
            return abs_key, content, doc_type, ok

        ext = f.suffix.lower()
        text_exts = {".txt", ".md", ".csv", ".log", ".json"}

        if ext in text_exts:
            try:
                full_text = f.read_text(encoding="utf-8", errors="replace")
                return abs_key, full_text, "Text File", True
            except Exception:
                return abs_key, "", "Text File", False
        else:
            try:
                import sys
                engines_dir = os.path.join(os.path.dirname(__file__), "..", "..", "engines")
                engines_dir = os.path.abspath(engines_dir)
                if engines_dir not in sys.path:
                    sys.path.insert(0, engines_dir)
                from Document_Intelligence import extract_text as _di_extract
                content, doc_type, ok = _di_extract(str(f))
                return abs_key, content, doc_type, ok
            except Exception as exc:
                logger.warning("Extraction failed for %s: %s", f.name, exc)
                return abs_key, "", "Unknown", False

    # ── Phase 1: Collect all files that need extraction ──
    all_files_to_extract: List[Tuple[str, Path]] = []  # (control_id, file_path)

    for upper_cid, cid in cid_upper_map.items():
        folder = folder_map.get(upper_cid)
        if not folder:
            continue

        if per_sample:
            sample_dirs = sorted(d for d in folder.iterdir() if d.is_dir())
            if sample_dirs:
                for sample_dir in sample_dirs:
                    for sf in sorted(sample_dir.iterdir()):
                        if sf.is_file() and sf.name.lower() not in ("readme.txt", "readme.md"):
                            all_files_to_extract.append((cid, sf))
            else:
                for f in sorted(folder.iterdir()):
                    if f.is_file() and f.name.lower() not in ("readme.txt", "readme.md"):
                        all_files_to_extract.append((cid, f))
        else:
            for f in sorted(folder.iterdir()):
                if not f.is_file():
                    for sf in sorted(f.iterdir()):
                        if sf.is_file() and sf.name.lower() not in ("readme.txt", "readme.md"):
                            all_files_to_extract.append((cid, sf))
                    continue
                if f.name.lower() in ("readme.txt", "readme.md"):
                    continue
                all_files_to_extract.append((cid, f))

    # ── Phase 2: Parallel extraction of all files ──
    # Filter to only files not already in cache
    files_needing_extraction = [
        (cid, f) for cid, f in all_files_to_extract
        if _normalize_cache_key(f) not in extract_cache
    ]
    logger.info(
        "Evidence scan: %d total files, %d need extraction, %d already cached",
        len(all_files_to_extract), len(files_needing_extraction),
        len(all_files_to_extract) - len(files_needing_extraction),
    )

    if files_needing_extraction:
        import os as _os
        parse_workers = int(_os.environ.get("EVIDENCE_PARSE_WORKERS", "8"))
        from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _as_completed

        with _TPE(max_workers=parse_workers) as pool:
            future_map = {pool.submit(_extract_one, f): (cid, f) for cid, f in files_needing_extraction}
            done = 0
            for future in _as_completed(future_map):
                cid_f, f_path = future_map[future]
                done += 1
                try:
                    abs_key, content, doc_type, ok = future.result()
                    extract_cache[abs_key] = (content, doc_type, ok)
                except Exception as exc:
                    abs_key = _normalize_cache_key(f_path)
                    extract_cache[abs_key] = ("", "Unknown", False)
                    logger.warning("Extraction failed for %s: %s", f_path.name, exc)
                if done % 10 == 0 or done == len(files_needing_extraction):
                    logger.info("Extracted %d/%d files", done, len(files_needing_extraction))

    # ── Phase 3: Build result dict from cached extractions ──
    for cid_file, f in all_files_to_extract:
        abs_key = _normalize_cache_key(f)
        content, doc_type, ok = extract_cache.get(abs_key, ("", "Unknown", False))
        if ok and content and content.strip():
            preview = content.strip()[:_EMBED_PREVIEW_MAX]
        else:
            preview = f.stem.replace("_", " ").replace("-", " ")

        file_info = {
            "file_name": f.name,
            "file_path": str(f),
            "content_preview": preview,
        }

        if per_sample:
            # Group by sample subfolder
            sample_name = f.parent.name if f.parent != folder_map.get(cid_file.strip().upper()) else f"sample_{len(result.get(cid_file, {})) + 1}"
            result.setdefault(cid_file, {})
            result[cid_file].setdefault(sample_name, []).append(file_info)
        else:
            result.setdefault(cid_file, []).append(file_info)

    return result, extract_cache


def _get_file_preview(file_path: Path, max_chars: int = 2000) -> str:
    """Extract actual content from evidence file for embedding.

    Uses Document Intelligence for PDFs, DOCX, images, etc. — the same
    extraction pipeline that the TOD/TOE engine uses.  Falls back to
    filename-based preview only when extraction fails.
    """
    ext = file_path.suffix.lower()
    text_exts = {".txt", ".md", ".csv", ".log", ".json"}

    if ext in text_exts:
        try:
            return file_path.read_text(encoding="utf-8", errors="replace")[:max_chars]
        except Exception:
            return file_path.stem

    # For all other formats, use Document Intelligence to extract real content
    try:
        import sys
        engines_dir = os.path.join(os.path.dirname(__file__), "..", "..", "engines")
        engines_dir = os.path.abspath(engines_dir)
        if engines_dir not in sys.path:
            sys.path.insert(0, engines_dir)

        from Document_Intelligence import extract_text as _di_extract
        content, doc_type, ok = _di_extract(str(file_path))
        if ok and content and content.strip():
            # Include filename for additional context
            preview = f"{file_path.stem}: {content.strip()}"
            return preview[:max_chars]
        else:
            logger.warning(
                "Document Intelligence extraction empty for %s, falling back to filename",
                file_path.name,
            )
    except Exception as exc:
        logger.warning(
            "Document Intelligence unavailable for %s: %s — falling back to filename",
            file_path.name, exc,
        )

    # Fallback: filename-based preview (only when DI fails)
    stem = file_path.stem.replace("_", " ").replace("-", " ")
    ext_label = {
        ".pdf": "PDF document",
        ".docx": "Word document",
        ".doc": "Word document",
        ".xlsx": "Excel spreadsheet",
        ".xls": "Excel spreadsheet",
        ".pptx": "PowerPoint presentation",
        ".png": "image/screenshot",
        ".jpg": "image/screenshot",
        ".jpeg": "image/screenshot",
        ".tiff": "image/screenshot",
        ".eml": "email message",
        ".msg": "email message",
    }.get(ext, f"{ext} file")

    return f"{stem} ({ext_label})"


# ---------------------------------------------------------------------------
# Main validation function
# ---------------------------------------------------------------------------

def validate_evidence_against_documents(
    evidence_folder: str,
    document_list_rows: List[Dict],
    control_ids: List[str],
    api_key: str,
    endpoint: str,
    api_version: str,
    embedding_model: str = "",
    similarity_threshold: float = 0.45,
    per_sample: bool = False,
    llm_model: str = "",
) -> Dict[str, Any]:
    """Compare evidence files against the Required Documents list using LLM matching.

    For each control (or control+sample for TOE), sends the evidence file contents
    and required documents list to an LLM which determines which requirements are
    covered. One evidence file can cover multiple required documents.

    Parameters
    ----------
    evidence_folder : str
        Path to the evidence folder.
    document_list_rows : list
        Output rows from generate_document_list_excel():
        [{"control_id", "required_documents": [{"name", "description"}, ...]}]
    control_ids : list
        All control IDs to validate.
    api_key, endpoint, api_version : str
        Azure OpenAI credentials.
    embedding_model : str
        Kept for backward compatibility (unused with LLM matching).
    similarity_threshold : float
        Kept for backward compatibility (unused with LLM matching).
    per_sample : bool
        If True (for TOE), validates each sample subfolder independently.
        If False (for TOD), validates at the control level (all files pooled).
    llm_model : str
        Azure OpenAI deployment name for the chat model. If empty, falls back
        to the engine config default.

    Returns
    -------
    dict with keys:
        controls: list of validation results (as dicts)
        total_required, total_matched, total_missing
        evidence_file_map: raw scan results (for reuse)
        extract_cache: dict[abs_file_path -> (content, doc_type, ok)]
    """

    # Resolve LLM model
    if not llm_model:
        try:
            from engines.config import AZURE_OPENAI_DEPLOYMENT
            llm_model = AZURE_OPENAI_DEPLOYMENT
        except Exception:
            llm_model = "gpt-5.2-chat"

    # Set parse phase for log separation
    try:
        from Document_Intelligence import set_parse_phase
        set_parse_phase("TOE" if per_sample else "TOD")
    except ImportError:
        pass

    # 1. Scan evidence files and build extraction cache
    logger.info("Scanning and extracting evidence files in %s (per_sample=%s)...",
                evidence_folder, per_sample)
    evidence_files, extract_cache = _scan_evidence_files(
        evidence_folder, control_ids, per_sample=per_sample,
    )

    if per_sample:
        total_evidence_files = sum(
            sum(len(files) for files in samples.values())
            for samples in evidence_files.values()
            if isinstance(samples, dict)
        )
    else:
        total_evidence_files = sum(len(files) for files in evidence_files.values())
    logger.info(
        "Found %d evidence files across %d controls (%d cached extractions)",
        total_evidence_files, len(evidence_files), len(extract_cache),
    )

    # Build document list lookup
    doc_lookup: Dict[str, List[Dict]] = {}
    for row in document_list_rows:
        cid = row.get("control_id", "")
        docs = row.get("required_documents", [])
        if cid and docs:
            doc_lookup[cid] = docs

    total_required_docs = sum(len(docs) for docs in doc_lookup.values())
    if not evidence_files and not doc_lookup:
        logger.warning("No evidence files or documents to compare")
        return {
            "controls": [], "total_required": 0, "total_matched": 0,
            "total_missing": 0, "evidence_file_map": evidence_files,
            "extract_cache": extract_cache,
        }

    logger.info(
        "LLM evidence matching: %d evidence files vs %d required documents across %d controls",
        total_evidence_files, total_required_docs, len(doc_lookup),
    )

    # ── Helper: convert DocumentMatch list to result dict list ──
    def _matches_to_dicts(matches: List[DocumentMatch]) -> List[Dict]:
        return [
            {
                "document_name": m.document_name,
                "document_description": m.document_description,
                "matched": m.matched,
                "matched_file": m.matched_file,
                "similarity_score": m.similarity_score,
            }
            for m in matches
        ]

    # 2. Run LLM matching per control (parallelised)
    results: List[Dict] = []
    total_required = 0
    total_matched = 0

    if per_sample:
        # ── TOE: validate EACH sample independently (parallel across all samples) ──
        # Collect all (control, sample) tasks first, then run them all in parallel
        _sample_tasks = []          # (cid, sample_name, sample_files, docs)
        _no_evidence_controls = []  # controls with zero evidence

        for cid in sorted(set(list(doc_lookup.keys()) + list(evidence_files.keys()))):
            docs = doc_lookup.get(cid, [])
            if not docs:
                continue
            samples = evidence_files.get(cid, {})
            if not isinstance(samples, dict) or not samples:
                _no_evidence_controls.append((cid, docs))
                continue
            for sample_name in sorted(samples.keys()):
                _sample_tasks.append((cid, sample_name, samples[sample_name], docs))

        # Handle controls with no evidence (no LLM call needed)
        for cid, docs in _no_evidence_controls:
            no_ev_matches = [
                DocumentMatch(document_name=d.get("name", ""), document_description=d.get("description", ""),
                              matched=False, matched_file=None, similarity_score=0.0)
                for d in docs
            ]
            total_required += len(docs)
            results.append({
                "control_id": cid,
                "samples": [{
                    "sample_name": "(no evidence)",
                    "total_required": len(docs),
                    "total_matched": 0,
                    "total_missing": len(docs),
                    "documents": _matches_to_dicts(no_ev_matches),
                }],
                "total_required": len(docs),
                "total_matched": 0,
                "total_missing": len(docs),
            })

        # Parallel LLM matching across ALL samples of ALL controls
        if _sample_tasks:
            logger.info("Running LLM matching for %d sample tasks in parallel...", len(_sample_tasks))
            _sample_results_map: Dict[str, List] = {}  # cid -> list of sample result dicts

            def _validate_one_sample(task):
                cid, sample_name, sample_files, docs = task
                matches = _llm_match_for_control(
                    control_id=f"{cid}/{sample_name}",
                    evidence_files=sample_files,
                    required_docs=docs,
                    api_key=api_key,
                    endpoint=endpoint,
                    api_version=api_version,
                    llm_model=llm_model,
                )
                matched_count = sum(1 for m in matches if m.matched)
                return cid, sample_name, matches, matched_count

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {
                    executor.submit(_validate_one_sample, task): task
                    for task in _sample_tasks
                }
                done_count = 0
                for future in as_completed(futures):
                    done_count += 1
                    try:
                        cid, sample_name, matches, matched_count = future.result()
                        missing_count = len(matches) - matched_count
                        _sample_results_map.setdefault(cid, []).append({
                            "sample_name": sample_name,
                            "total_required": len(matches),
                            "total_matched": matched_count,
                            "total_missing": missing_count,
                            "documents": _matches_to_dicts(matches),
                        })
                        if done_count % 5 == 0 or done_count == len(_sample_tasks):
                            logger.info("LLM matching: %d/%d samples done", done_count, len(_sample_tasks))
                    except Exception as exc:
                        task = futures[future]
                        logger.error("LLM validation failed for %s/%s: %s", task[0], task[1], exc)

            # Aggregate per-control results
            for cid in sorted(_sample_results_map.keys()):
                sample_results = sorted(_sample_results_map[cid], key=lambda s: s["sample_name"])
                control_total_req = sum(s["total_required"] for s in sample_results)
                control_total_match = sum(s["total_matched"] for s in sample_results)
                total_required += control_total_req
                total_matched += control_total_match
                results.append({
                    "control_id": cid,
                    "samples": sample_results,
                    "total_required": control_total_req,
                    "total_matched": control_total_match,
                    "total_missing": control_total_req - control_total_match,
                })

    else:
        # ── TOD: validate at control level (all files pooled, parallel LLM calls) ──
        all_cids = sorted(set(list(doc_lookup.keys()) + list(evidence_files.keys())))
        cids_with_docs = [cid for cid in all_cids if doc_lookup.get(cid)]

        def _validate_one_control(cid: str) -> Optional[Dict]:
            docs = doc_lookup.get(cid, [])
            if not docs:
                return None
            ev_files = evidence_files.get(cid, [])
            matches = _llm_match_for_control(
                control_id=cid,
                evidence_files=ev_files,
                required_docs=docs,
                api_key=api_key,
                endpoint=endpoint,
                api_version=api_version,
                llm_model=llm_model,
            )
            matched_count = sum(1 for m in matches if m.matched)
            return {
                "control_id": cid,
                "total_required": len(matches),
                "total_matched": matched_count,
                "total_missing": len(matches) - matched_count,
                "documents": _matches_to_dicts(matches),
            }

        # Parallel LLM calls (one per control)
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(_validate_one_control, cid): cid
                for cid in cids_with_docs
            }
            for future in as_completed(futures):
                cid = futures[future]
                try:
                    result = future.result()
                    if result:
                        results.append(result)
                        total_required += result["total_required"]
                        total_matched += result["total_matched"]
                except Exception as exc:
                    logger.error("LLM validation failed for %s: %s", cid, exc)

        # Sort results by control_id for consistent output
        results.sort(key=lambda r: r["control_id"])

    total_missing = total_required - total_matched

    logger.info(
        "Evidence validation complete: %d/%d documents matched across %d controls%s",
        total_matched, total_required, len(results),
        " (per-sample)" if per_sample else "",
    )

    return {
        "controls": results,
        "total_required": total_required,
        "total_matched": total_matched,
        "total_missing": total_missing,
        "evidence_file_map": evidence_files,
        "extract_cache": extract_cache,
    }


# ---------------------------------------------------------------------------
# Excel export for the validation report
# ---------------------------------------------------------------------------

def export_validation_excel(
    validation_controls: List[Dict],
    output_dir: str,
    phase: str = "TOD",
    state: Any = None,
) -> Optional[str]:
    """Write the evidence validation report as a styled Excel and return the path.

    Parameters
    ----------
    validation_controls : list
        The ``controls`` list from ``validate_evidence_against_documents()``.
    output_dir : str
        Directory to write the file into.
    phase : str
        ``"TOD"`` or ``"TOE"``.
    state : AgentState, optional
        If provided, artifact is tracked and uploaded to blob.

    Returns
    -------
    str or None — file path (or blob path) on success, None on failure.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from datetime import datetime

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evidence Validation"

        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        wrap_top = Alignment(vertical="top", wrap_text=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")

        # Detect if results are per-sample (TOE) or per-control (TOD)
        has_samples = any("samples" in ctrl for ctrl in validation_controls)

        if has_samples:
            headers = [
                "Control ID", "Sample", "Required Document",
                "Description", "Status", "Matched File", "Similarity",
            ]
        else:
            headers = [
                "Control ID", "Required Document",
                "Description", "Status", "Matched File", "Similarity",
            ]

        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.column_dimensions["A"].width = 16
        if has_samples:
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 50
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 30
            ws.column_dimensions["G"].width = 12
        else:
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 12

        row = 2
        for ctrl in validation_controls:
            cid = ctrl.get("control_id", "")

            if has_samples:
                # TOE: one row per document per sample
                for sample in ctrl.get("samples", []):
                    sample_name = sample.get("sample_name", "")
                    for doc in sample.get("documents", []):
                        matched = doc.get("matched", False)
                        status = "FOUND" if matched else "MISSING"

                        ws.cell(row=row, column=1, value=cid).border = thin_border
                        ws.cell(row=row, column=2, value=sample_name).border = thin_border
                        ws.cell(row=row, column=3, value=doc.get("document_name", "")).border = thin_border
                        ws.cell(row=row, column=3).alignment = wrap_top
                        ws.cell(row=row, column=4, value=doc.get("document_description", "")).border = thin_border
                        ws.cell(row=row, column=4).alignment = wrap_top

                        status_cell = ws.cell(row=row, column=5, value=status)
                        status_cell.border = thin_border
                        status_cell.alignment = Alignment(horizontal="center", vertical="top")
                        if matched:
                            status_cell.fill = green_fill
                            status_cell.font = green_font
                        else:
                            status_cell.fill = red_fill
                            status_cell.font = red_font

                        ws.cell(row=row, column=6, value=doc.get("matched_file") or "—").border = thin_border
                        ws.cell(row=row, column=7, value=doc.get("similarity_score", 0)).border = thin_border
                        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="top")

                        row += 1
            else:
                # TOD: one row per document per control
                for doc in ctrl.get("documents", []):
                    matched = doc.get("matched", False)
                    status = "FOUND" if matched else "MISSING"

                    ws.cell(row=row, column=1, value=cid).border = thin_border
                    ws.cell(row=row, column=2, value=doc.get("document_name", "")).border = thin_border
                    ws.cell(row=row, column=2).alignment = wrap_top
                    ws.cell(row=row, column=3, value=doc.get("document_description", "")).border = thin_border
                    ws.cell(row=row, column=3).alignment = wrap_top

                    status_cell = ws.cell(row=row, column=4, value=status)
                    status_cell.border = thin_border
                    status_cell.alignment = Alignment(horizontal="center", vertical="top")
                    if matched:
                        status_cell.fill = green_fill
                        status_cell.font = green_font
                    else:
                        status_cell.fill = red_fill
                        status_cell.font = red_font

                    ws.cell(row=row, column=5, value=doc.get("matched_file") or "—").border = thin_border
                    ws.cell(row=row, column=6, value=doc.get("similarity_score", 0)).border = thin_border
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="top")

                    row += 1

        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Evidence_Validation_{phase.upper()}_{ts}.xlsx"
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        logger.info("Evidence validation Excel saved: %s (%d rows)", output_path, row - 2)

        # Upload to blob
        blob_path = output_path
        if state is not None:
            try:
                from server.blob_store import get_blob_store
                store = get_blob_store()
                if store.available:
                    session_key = "default"
                    if getattr(state, "output_dir", None):
                        session_key = os.path.basename(state.output_dir)
                    bp = f"artifacts/{session_key}/{filename}"
                    result = store.upload_file(output_path, bp)
                    if result:
                        blob_path = result
                        state.artifacts.append(blob_path)
            except Exception as exc:
                logger.warning("Blob upload failed for validation Excel: %s", exc)

        return blob_path

    except Exception as exc:
        logger.error("Failed to export validation Excel: %s", exc)
        return None
