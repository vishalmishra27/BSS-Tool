"""Export pending attribute schemas as an editable Excel file.

Shared helper used by preview_tod_attributes and preview_toe_attributes
to produce a downloadable Excel that the user can edit and re-upload via
upload_modified_attributes.

Format:
  Column A: Control ID
  Column B: Attribute #
  Column C: Attribute Name
  Column D: Attribute Description
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Dict, Optional

logger = logging.getLogger("agent.tools.export_attributes_excel")


def export_attributes_excel(
    schemas: Dict[str, Any],
    output_dir: str,
    phase: str = "TOD",
    state: Any = None,
) -> Optional[str]:
    """Write an editable attributes Excel and return the file path (or blob path).

    Returns None if export fails.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attributes"

        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        wrap_top = Alignment(vertical="top", wrap_text=True)

        headers = ["Control ID", "Attribute #", "Attribute Name", "Attribute Description"]
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 60

        row = 2
        for cid in sorted(schemas.keys()):
            schema_data = schemas[cid]
            attrs = (
                schema_data.attributes
                if hasattr(schema_data, "attributes")
                else schema_data.get("attributes", [])
            )
            for i, attr in enumerate(attrs, 1):
                ws.cell(row=row, column=1, value=cid).border = thin_border
                ws.cell(row=row, column=2, value=i).border = thin_border
                ws.cell(row=row, column=3, value=attr.get("name", "")).border = thin_border
                ws.cell(row=row, column=3).alignment = wrap_top
                ws.cell(row=row, column=4, value=attr.get("description", "")).border = thin_border
                ws.cell(row=row, column=4).alignment = wrap_top
                row += 1

        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Editable_Attributes_{phase.upper()}_{ts}.xlsx"
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        logger.info("Editable attributes Excel saved: %s (%d rows)", output_path, row - 2)

        # Upload to blob if available
        blob_path = output_path
        if state is not None:
            try:
                from server.blob_store import get_blob_store
                store = get_blob_store()
                if store.available:
                    session_key = "default"
                    if getattr(state, "output_dir", None):
                        session_key = os.path.basename(state.output_dir)
                    bp = f"artifacts/{session_key}/{filename}"
                    result = store.upload_file(output_path, bp)
                    if result:
                        blob_path = result
                        state.artifacts.append(blob_path)
                        logger.info("Editable attributes Excel uploaded to blob: %s", blob_path)
            except Exception as exc:
                logger.warning("Blob upload failed for attributes Excel: %s", exc)

        return blob_path

    except Exception as exc:
        logger.error("Failed to export editable attributes Excel: %s", exc)
        return None
