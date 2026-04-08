"""Generate a list of required documents for each control based on approved attributes.

Auto-runs as part of TOD/TOE execution after the user approves attributes.
NOT a standalone agent tool — called internally by run_test_of_design and
run_test_of_effectiveness.

Reads control attributes + control description + control ID, uses AI to
produce a document checklist per control, and writes a styled Excel workbook.
"""

from __future__ import annotations

import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger("agent.tools.document_list")


# ---------------------------------------------------------------------------
# Blob upload helper
# ---------------------------------------------------------------------------

def _upload_artifact(local_path: str, state: Any = None) -> str:
    """Upload a generated artifact to blob storage. Returns blob path, or local path as fallback."""
    try:
        from server.blob_store import get_blob_store
        store = get_blob_store()
        if not store.available:
            logger.warning("Blob Storage not available — artifact stays local: %s", local_path)
            return local_path
        filename = os.path.basename(local_path)
        session_key = "default"
        if state and getattr(state, "output_dir", None):
            session_key = os.path.basename(state.output_dir)
        blob_path = f"artifacts/{session_key}/{filename}"
        result = store.upload_file(local_path, blob_path)
        if not result:
            logger.warning("Blob upload failed for %s — using local path", local_path)
            return local_path
        return result
    except Exception as exc:
        logger.warning("Blob upload error for %s: %s — using local path", local_path, exc)
        return local_path


# ---------------------------------------------------------------------------
# Core LLM call — one control at a time (parallelised by caller)
# ---------------------------------------------------------------------------

def _generate_documents_for_control(
    llm_client,
    control_id: str,
    control_description: str,
    attributes: List[Dict[str, str]],
) -> Dict[str, Any]:
    """Ask the LLM to produce a list of required documents for one control.

    Returns ``{"documents": [{"number": int, "name": str, "description": str}, ...]}``
    """
    attr_text = "\n".join(
        f"  {i+1}. {a.get('name', 'Unnamed')} — {a.get('description', '')}"
        for i, a in enumerate(attributes)
    )

    prompt = f"""You are an audit documentation expert.

Given the following control and its testing attributes, list ALL the documents
that an auditor would need to collect in order to evaluate a sample for this control.

Control ID: {control_id}
Control Description: {control_description}

Testing Attributes:
{attr_text}

For each required document, provide:
1. A short document name (e.g. "Purchase Order", "Three-Way Match Report")
2. A one-sentence description of what the document should contain and why it is needed

Return your answer as a JSON array. Each element must have exactly these keys:
  "name"  — short document name
  "description" — one-sentence explanation

Return ONLY the JSON array, no extra text. Example:
[
  {{"name": "Purchase Order", "description": "The original PO showing authorized amount, vendor, and date to verify approval limits."}},
  {{"name": "Goods Receipt Note", "description": "Confirms physical receipt of goods matching the PO quantity and specifications."}}
]"""

    try:
        raw = llm_client.complete(
            system="You are an audit documentation expert. Return only valid JSON.",
            user=prompt,
            max_tokens=2048,
        )
        # Strip markdown fences if present
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[-1]
        if cleaned.endswith("```"):
            cleaned = cleaned.rsplit("```", 1)[0]
        cleaned = cleaned.strip()

        docs = json.loads(cleaned)
        if not isinstance(docs, list):
            docs = []
        numbered = []
        for i, d in enumerate(docs, 1):
            numbered.append({
                "number": i,
                "name": d.get("name", f"Document {i}"),
                "description": d.get("description", ""),
            })
        return {"documents": numbered}
    except Exception as exc:
        logger.error("Document list generation failed for %s: %s", control_id, exc)
        return {"documents": [], "error": str(exc)}


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_excel(
    results: Dict[str, Dict],
    schemas: Dict[str, Any],
    desc_lookup: Dict[str, str],
    output_path: str,
) -> None:
    """Write the Required Documents Excel workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Required Documents"

    # Styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_top = Alignment(vertical="top", wrap_text=True)

    headers = [
        "Control ID",
        "Control Description",
        "Control Attributes",
        "Required Documents",
        "Document Descriptions",
    ]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 55

    row = 2
    for cid in sorted(results.keys()):
        docs = results[cid].get("documents", [])
        control_desc = desc_lookup.get(cid, "")

        schema_data = schemas.get(cid, {})
        attrs = (
            schema_data.attributes
            if hasattr(schema_data, "attributes")
            else schema_data.get("attributes", [])
        )

        attr_text = "\n".join(
            f"{i+1}. {a.get('name', 'Unnamed')} — {a.get('description', '')}"
            for i, a in enumerate(attrs)
        )
        doc_names = "\n".join(f"{d['number']}. {d['name']}" for d in docs)
        doc_descs = "\n".join(f"{d['number']}. {d['description']}" for d in docs)

        for col, val in enumerate(
            [cid, control_desc, attr_text, doc_names, doc_descs], 1
        ):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = wrap_top
            cell.border = thin_border

        row += 1

    wb.save(output_path)
    logger.info("Document list Excel saved: %s (%d rows)", output_path, row - 2)


# ---------------------------------------------------------------------------
# Public API — called by TOD / TOE tools
# ---------------------------------------------------------------------------

def generate_document_list_excel(
    schemas: Dict[str, Any],
    rcm_df: pd.DataFrame,
    output_dir: str,
    phase: str = "TOD",
    state: Any = None,
) -> Dict[str, Any]:
    """Generate the Required Documents Excel and return metadata.

    Parameters
    ----------
    schemas : dict
        ``{control_id: {"attributes": [...], ...}}`` or ControlSchema objects.
    rcm_df : pd.DataFrame
        The loaded RCM DataFrame (needs ``Control Id`` and ``Control Description``).
    output_dir : str
        Directory to write the Excel file into.
    phase : str
        ``"TOD"`` or ``"TOE"`` — used in the filename.
    state : AgentState, optional
        If provided, artifacts list and blob upload are handled.

    Returns
    -------
    dict with keys:
        controls_processed, total_documents, output_excel, output_blob_path, rows
    """
    from ..config import get_config
    from ..llm import LLMClient

    config = get_config()
    llm = LLMClient(
        api_key=config.openai_api_key,
        model=config.openai_model,
        azure_endpoint=config.azure_openai_endpoint,
        azure_api_version=config.azure_openai_api_version,
    )

    # Build description lookup
    desc_lookup: Dict[str, str] = {}
    if "Control Id" in rcm_df.columns and "Control Description" in rcm_df.columns:
        for _, row in rcm_df.iterrows():
            cid = str(row["Control Id"]).strip()
            desc_lookup[cid] = str(row.get("Control Description", "")).strip()

    # Generate in parallel
    results: Dict[str, Dict] = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_cid = {}
        for cid, schema_data in schemas.items():
            attrs = (
                schema_data.attributes
                if hasattr(schema_data, "attributes")
                else schema_data.get("attributes", [])
            )
            future = executor.submit(
                _generate_documents_for_control, llm, cid,
                desc_lookup.get(cid, ""), attrs,
            )
            future_to_cid[future] = cid

        for future in as_completed(future_to_cid):
            cid = future_to_cid[future]
            try:
                results[cid] = future.result()
                logger.info(
                    "Document list for %s: %d documents",
                    cid, len(results[cid].get("documents", [])),
                )
            except Exception as exc:
                logger.error("Document list failed for %s: %s", cid, exc)
                results[cid] = {"documents": [], "error": str(exc)}

    # Write Excel
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Required_Documents_{phase.upper()}_{ts}.xlsx"
    output_path = os.path.join(output_dir, filename)
    _build_excel(results, schemas, desc_lookup, output_path)

    # Upload to blob
    blob_path = output_path
    if state is not None:
        blob_path = _upload_artifact(output_path, state)
        state.artifacts.append(blob_path or output_path)

    total_docs = sum(len(r.get("documents", [])) for r in results.values())

    # Build JSON-friendly rows for the tool result
    output_rows = []
    for cid in sorted(results.keys()):
        docs = results[cid].get("documents", [])
        schema_data = schemas.get(cid, {})
        attrs = (
            schema_data.attributes
            if hasattr(schema_data, "attributes")
            else schema_data.get("attributes", [])
        )
        output_rows.append({
            "control_id": cid,
            "control_description": desc_lookup.get(cid, ""),
            "attributes": [
                {"name": a.get("name", ""), "description": a.get("description", "")}
                for a in attrs
            ],
            "required_documents": [
                {"number": d["number"], "name": d["name"], "description": d["description"]}
                for d in docs
            ],
        })

    return {
        "controls_processed": len(results),
        "total_documents": total_docs,
        "output_excel": blob_path or output_path,
        "output_blob_path": blob_path if blob_path != output_path else None,
        "rows": output_rows,
    }
