"""Wrapper tool to run the Control Scoping Engine from sibling folder 'sox_package 2'.

Supports Azure Blob Storage paths for trial balance files, SOP paths,
and override Excel files. Generated artifacts are uploaded to blob.
"""

from __future__ import annotations

import concurrent.futures
import importlib.util
import logging
import math
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from unittest.mock import patch

from .base import Tool
from ..config import get_config
from ..core.progress import update_progress, clear_progress, finish_pipeline, make_counting_as_completed
from ..types import AgentState, ToolCategory, ToolParameter, ToolResult

logger = logging.getLogger("agent.tools.sox_scoping_engine")


def _resolve_blob_file(path: str) -> str:
    """If path is a blob path, download to local cache and return local path."""
    from server.blob_store import get_blob_store, BlobStore
    if BlobStore.is_blob_path(path):
        store = get_blob_store()
        if not store.available:
            raise RuntimeError("Azure Blob Storage is not available")
        local = store.ensure_local(path)
        if not local:
            raise RuntimeError(f"Failed to download blob file: {path}")
        logger.info("Downloaded blob file %s → %s", path, local)
        return local
    return path


def _resolve_blob_directory(path: str) -> str:
    """If path is a blob prefix, download all files to local cache."""
    from server.blob_store import get_blob_store, BlobStore
    if BlobStore.is_blob_path(path):
        store = get_blob_store()
        if not store.available:
            raise RuntimeError("Azure Blob Storage is not available")
        local = store.ensure_local_directory(path)
        if not local:
            raise RuntimeError(f"Failed to download blob directory: {path}")
        logger.info("Downloaded blob directory %s → %s", path, local)
        return local
    return path


def _upload_artifact(local_path: str) -> Optional[str]:
    """Upload a generated artifact to blob storage.

    Derives the session key from the parent directory name (e.g. sox_agent_20240101_120000).
    Returns the blob path on success, or the local path as fallback.
    """
    try:
        from server.blob_store import get_blob_store
        store = get_blob_store()
        if not store.available:
            logger.warning("Blob Storage not available — artifact stays local: %s", local_path)
            return local_path
        filename = os.path.basename(local_path)
        # Use parent dir name as session key (e.g. sox_agent_20240101_120000)
        session_key = os.path.basename(os.path.dirname(local_path)) or "default"
        blob_path = f"artifacts/{session_key}/{filename}"
        result = store.upload_file(local_path, blob_path)
        if not result:
            logger.warning("Blob upload failed for %s — using local path", local_path)
            return local_path
        return result
    except Exception as exc:
        logger.warning("Blob upload error for %s: %s — using local path", local_path, exc)
        return local_path


def _load_external_engine_module():
    """Load sox_scoping_engine.py from sibling workspace folder 'sox_package 2'."""
    workspace_root = Path(__file__).resolve().parents[3]
    engine_path = workspace_root / "sox_package 2" / "sox_scoping_engine.py"

    if not engine_path.exists():
        raise FileNotFoundError(
            f"Scoping engine file not found at expected location: {engine_path}. "
            "Expected sibling folder: 'sox_package 2'."
        )

    module_name = "external_sox_scoping_engine"
    spec = importlib.util.spec_from_file_location(module_name, str(engine_path))
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Failed to create module spec for {engine_path}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ---------------------------------------------------------------------------
# Helpers — defensively extract data from engine objects whose internals
# we do not control (loaded dynamically from 'sox_package 2').
# ---------------------------------------------------------------------------

def _acct_name(obj) -> str:
    for attr in ("raw_name", "name", "account_name", "description", "label", "account"):
        v = getattr(obj, attr, None)
        if v and str(v).strip():
            return str(v).strip()
    return str(obj)


def _acct_balance(obj) -> Optional[float]:
    for attr in ("balance", "amount", "value", "ending_balance", "net_balance"):
        v = getattr(obj, attr, None)
        if v is not None:
            try:
                return round(float(v), 2)
            except (TypeError, ValueError):
                pass
    return None


def _extract_account_list(engine, max_items: int = 200) -> List[Dict]:
    """Account names, balances and categories after ingest+categorize."""
    raw = getattr(engine, "accounts", []) or []
    result = []
    for acc in raw[:max_items]:
        entry: Dict[str, Any] = {"name": _acct_name(acc)}
        bal = _acct_balance(acc)
        if bal is not None:
            entry["balance"] = bal
        for attr in ("category", "group_level", "fs_line_item", "account_type", "type", "classification"):
            v = getattr(acc, attr, None)
            if v and str(v).strip():
                entry["category"] = str(v).strip()
                break
        # Always include group_level and financial_statement as dedicated fields
        gl = getattr(acc, "group_level", "") or ""
        if gl:
            entry["group_level"] = gl
        fs = getattr(acc, "financial_statement", "") or ""
        if fs:
            entry["financial_statement"] = fs
        result.append(entry)
    return result


def _extract_benchmark_values(engine) -> Dict[str, float]:
    """
    Financial totals per category so the user can see what each benchmark
    base actually is (e.g. Revenue: $123.4M) before choosing.

    Returns title-case keys matching the benchmark enum:
    Revenue, Assets, EBITDA, PBT, Net Income, Net Interest.

    If compute_threshold() has already been called, uses engine.consolidated
    for accuracy.  Otherwise aggregates from engine.accounts using group_level.
    """
    # Primary: engine.consolidated is populated after compute_threshold()
    consolidated = getattr(engine, "consolidated", None)
    if consolidated and isinstance(consolidated, dict):
        revenue = abs(consolidated.get("REVENUE", 0))
        assets_direct = abs(consolidated.get("ASSETS", 0))
        liab = abs(consolidated.get("LIABILITIES", 0))
        equity = abs(consolidated.get("EQUITY", 0))
        total_assets = assets_direct if assets_direct > 0 else (liab + equity)
        net_income = consolidated.get("net_income", 0)
        ebitda = consolidated.get("ebitda", 0)
        pbt = consolidated.get("pbt", 0)
        interest = abs(consolidated.get("INTEREST", 0))
        out: Dict[str, float] = {}
        if revenue:    out["Revenue"]      = round(revenue, 2)
        if total_assets:   out["Assets"]       = round(total_assets, 2)
        if net_income:     out["Net Income"]   = round(abs(net_income), 2)
        if ebitda and ebitda != net_income:
            out["EBITDA"] = round(abs(ebitda), 2)
            out["Adjusted EBITDA"] = out["EBITDA"]
        if pbt and pbt != net_income:
            out["PBT"] = round(abs(pbt), 2)
        if interest:       out["Net Interest"] = round(interest, 2)
        if out:
            return out

    # Fallback: aggregate from accounts using group_level (before compute_threshold)
    accounts = getattr(engine, "fs_accounts", []) or getattr(engine, "accounts", []) or []
    raw: Dict[str, float] = {}
    for acc in accounts:
        # Account objects use group_level; FSAccount objects use group
        g = getattr(acc, "group_level", "") or getattr(acc, "group", "") or ""
        g_upper = g.upper().strip()
        bal = _acct_balance(acc)
        if not g_upper or bal is None:
            continue
        if any(x in g_upper for x in ["REVENUE", "SALES", "OTHER INCOME", "INCOME FROM OPERATIONS"]):
            raw["Revenue"] = raw.get("Revenue", 0.0) + bal
        elif any(x in g_upper for x in ["ASSETS", "FIXED ASSETS", "INTANGIBLE ASSETS",
                                         "INVESTMENTS", "PPE", "PROPERTY", "PLANT"]):
            raw["Assets"] = raw.get("Assets", 0.0) + abs(bal)
        elif any(x in g_upper for x in ["LIABILIT", "PAYABLE", "BORROWING", "PROVISION", "DEBT"]):
            raw["_liab"] = raw.get("_liab", 0.0) + abs(bal)
        elif any(x in g_upper for x in ["EQUITY", "SHARE CAPITAL", "RETAINED", "RESERVE"]):
            raw["_equity"] = raw.get("_equity", 0.0) + abs(bal)

    # Assets fallback via L+E if not computed directly
    if not raw.get("Assets"):
        raw["Assets"] = raw.pop("_liab", 0.0) + raw.pop("_equity", 0.0)
    else:
        raw.pop("_liab", None)
        raw.pop("_equity", None)

    return {k: round(v, 2) for k, v in raw.items() if v != 0}


def _extract_group_level_distribution(engine) -> List[Dict]:
    """
    Aggregate ALL accounts by group_level for the Trial Balance Distribution chart.
    Returns a list of {group, balance, count} sorted by balance descending.
    Uses all engine.accounts regardless of max_items truncation.
    """
    accounts = getattr(engine, "accounts", []) or []
    agg: Dict[str, Dict] = {}
    for acc in accounts:
        grp = (getattr(acc, "group_level", "") or "").strip().upper() or "OTHER"
        bal = _acct_balance(acc)
        if bal is None:
            bal = 0.0
        if grp not in agg:
            agg[grp] = {"group": grp, "balance": 0.0, "count": 0}
        agg[grp]["balance"] += abs(bal)
        agg[grp]["count"] += 1
    result = sorted(agg.values(), key=lambda x: x["balance"], reverse=True)
    for item in result:
        item["balance"] = round(item["balance"], 2)
    return result


def _extract_quantitative_results(engine) -> Dict:
    """Quantitative in/out counts + per-account detail after run_quantitative().

    After run_quantitative(), results live on engine.accounts (not fs_accounts,
    which is only populated after run_scoping).  The key attribute is
    ``quant_significant`` (bool).
    """
    accounts = getattr(engine, "accounts", []) or []
    threshold = None
    for attr in ("threshold", "materiality_threshold", "quant_threshold"):
        v = getattr(engine, attr, None)
        if v is not None:
            try:
                threshold = round(float(v), 2)
            except (TypeError, ValueError):
                pass
            break

    in_scope: List[Dict] = []
    out_scope: List[Dict] = []
    unclassified: List[Dict] = []

    for acc in accounts:
        name = _acct_name(acc)
        bal = _acct_balance(acc)
        entry: Dict[str, Any] = {"name": name}
        if bal is not None:
            entry["balance"] = bal

        # Primary: quant_significant (bool) set by engine.run_quantitative()
        decision = getattr(acc, "quant_significant", None)
        if decision is None:
            # Fallback: probe other possible attribute names
            for attr in ("quantitative_scope", "quant_result", "quant_scope",
                         "quantitative_result", "above_threshold"):
                v = getattr(acc, attr, None)
                if v is not None:
                    decision = v
                    break

        if decision is None:
            unclassified.append(entry)
        elif decision is True or str(decision).lower() in ("in-scope", "in_scope", "in scope", "true", "yes", "1"):
            in_scope.append(entry)
        else:
            out_scope.append(entry)

    return {
        "materiality_threshold": threshold,
        "quantitatively_in_scope": in_scope,
        "quantitatively_out_of_scope": out_scope,
        "in_count": len(in_scope),
        "out_count": len(out_scope),
        "unclassified_count": len(unclassified),
    }


def _extract_qualitative_results(engine) -> Dict:
    """Qualitative risk assessment results after run_qualitative().

    After run_qualitative(), each account in engine.accounts has:
      - qual_score (float)
      - qual_risk ("High", "Moderate", "Low")
      - quant_significant (bool) — from prior quantitative phase

    "Qualitatively added" = accounts that are NOT quant_significant but have
    High qual_risk (these will be scoped IN by Rule 4 in run_scoping).
    "Qualitatively removed" = not applicable in this engine (qual only adds risk).
    """
    accounts = getattr(engine, "accounts", []) or []
    threshold = getattr(engine, "threshold", None)

    high: List[Dict] = []
    moderate: List[Dict] = []
    low: List[Dict] = []
    additions: List[Dict] = []

    for acc in accounts:
        name = _acct_name(acc)
        bal = _acct_balance(acc)
        risk = getattr(acc, "qual_risk", "") or ""
        score = getattr(acc, "qual_score", 0.0)
        quant_sig = getattr(acc, "quant_significant", False)

        entry: Dict[str, Any] = {"name": name, "qual_risk": risk, "qual_score": round(score, 2)}
        if bal is not None:
            entry["balance"] = bal

        if risk == "High":
            high.append(entry)
            if not quant_sig:
                additions.append({"name": name, "reason": f"High qualitative risk (score={score:.2f})"})
        elif risk == "Moderate":
            moderate.append(entry)
        elif risk == "Low":
            low.append(entry)

    return {
        "qualitatively_added": additions,
        "additions_count": len(additions),
        "removals_count": 0,
        "high_risk_count": len(high),
        "moderate_risk_count": len(moderate),
        "low_risk_count": len(low),
        "high_risk_accounts": high,
        "moderate_risk_accounts": moderate[:20],
        "low_risk_accounts": low[:10],
    }


def _extract_inscope_accounts(engine) -> List[Dict]:
    """Final in-scope account list after run_scoping()."""
    fs_accounts = getattr(engine, "fs_accounts", []) or []
    result = []
    for fa in fs_accounts:
        scope = getattr(fa, "scope", None)
        if scope and str(scope).lower() in ("in-scope", "in_scope", "in scope"):
            entry: Dict[str, Any] = {"name": _acct_name(fa)}
            bal = _acct_balance(fa)
            if bal is not None:
                entry["balance"] = bal
            for attr in ("category", "fs_line_item", "account_type", "group", "group_level", "type", "classification"):
                v = getattr(fa, attr, None)
                if v and str(v).strip():
                    entry["category"] = str(v).strip()
                    break
            # Include scoping rule so the user understands why it's in-scope
            rule = getattr(fa, "rule", 0)
            if rule:
                entry["rule"] = rule
                entry["rule_description"] = _SCOPING_RULE_LABELS.get(rule, f"Rule {rule}")
            result.append(entry)
    return result


_SCOPING_RULE_LABELS = {
    1: "Quantitatively significant + High qualitative risk",
    2: "Quantitatively significant + Moderate qualitative risk",
    3: "Quantitatively significant + Low qualitative risk",
    4: "High qualitative risk (not quant significant, nonzero balance)",
    5: "High qualitative risk but zero balance — Out-of-scope",
    6: "Moderate qualitative risk, not quant significant — Out-of-scope",
    7: "Low qualitative risk, not quant significant — Out-of-scope",
    8: "LLM validation override — recommended to scope in",
}


# ---------------------------------------------------------------------------
# Override helpers — apply user-uploaded Excel to engine state
# ---------------------------------------------------------------------------

def _read_override_excel(path: str) -> List[Dict[str, Any]]:
    """Read an Excel file into a list of row dicts. Returns [] on failure."""
    try:
        import pandas as pd  # type: ignore
    except ImportError:
        logger.warning("pandas not available — cannot read override Excel")
        return []
    try:
        df = pd.read_excel(path, sheet_name=0)
        # Normalise column names: strip, lower, underscores
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        df = df.where(df.notna(), other=None)
        return df.to_dict(orient="records")
    except Exception as exc:
        logger.error("Failed to read override Excel '%s': %s", path, exc)
        return []


def _apply_accounts_override(engine, rows: List[Dict[str, Any]]) -> tuple[int, int]:
    """Apply user overrides from an accounts Excel (Phase 0 output).

    Matches rows to engine.accounts by account name (case-insensitive).
    Updatable columns: balance, financial_statement, group_level, fs_account
    (mapped from Excel column names exported by _export_accounts_to_excel).

    Returns (updated_count, unmatched_count).
    """
    if not rows:
        return 0, 0
    accounts = getattr(engine, "accounts", []) or []
    # Build lookup by normalised name — collect lists to handle duplicates
    acct_map: Dict[str, List[Any]] = {}
    for acc in accounts:
        key = _acct_name(acc).strip().lower()
        acct_map.setdefault(key, []).append(acc)

    # Column name mapping: Excel header (lowered) → Account attribute
    col_map = {
        "account_name": "raw_name",
        "balance": "balance",
        "financial_statement": "financial_statement",
        "group_level": "group_level",
        "fs_line_item": "fs_account",
        "confidence": "confidence",
    }

    updated = 0
    unmatched = 0
    for row in rows:
        # Try to match by account name
        name_val = row.get("account_name") or row.get("name") or row.get("account") or ""
        if not name_val:
            continue
        key = str(name_val).strip().lower()
        matching_accts = acct_map.get(key)
        if not matching_accts:
            unmatched += 1
            continue
        for acc in matching_accts:
            changed = False
            for excel_col, attr in col_map.items():
                val = row.get(excel_col)
                if val is not None and str(val).strip():
                    old = getattr(acc, attr, None)
                    if attr == "balance":
                        try:
                            val = float(val)
                        except (TypeError, ValueError):
                            continue
                    else:
                        val = str(val).strip()
                    if val != old:
                        setattr(acc, attr, val)
                        changed = True
            if changed:
                updated += 1

    logger.info("Accounts override applied: %d updated, %d unmatched out of %d rows", updated, unmatched, len(rows))
    return updated, unmatched


def _apply_quantitative_override(engine, rows: List[Dict[str, Any]]) -> tuple[int, int]:
    """Apply user overrides from a quantitative Excel (Phase 1 output).

    Matches rows to engine.accounts by name and updates quant_significant
    based on a 'decision' column (In-scope / Out-of-scope).

    Returns (updated_count, unmatched_count).
    """
    if not rows:
        return 0, 0
    accounts = getattr(engine, "accounts", []) or []
    acct_map: Dict[str, List[Any]] = {}
    for acc in accounts:
        key = _acct_name(acc).strip().lower()
        acct_map.setdefault(key, []).append(acc)

    updated = 0
    unmatched = 0
    for row in rows:
        name_val = row.get("name") or row.get("account_name") or row.get("account") or ""
        if not name_val:
            continue
        key = str(name_val).strip().lower()
        matching_accts = acct_map.get(key)
        if not matching_accts:
            unmatched += 1
            continue
        decision = str(row.get("decision") or "").strip().lower()
        if decision in ("in-scope", "in_scope", "in scope"):
            new_val = True
        elif decision in ("out-of-scope", "out_of_scope", "out of scope"):
            new_val = False
        else:
            continue
        for acc in matching_accts:
            if getattr(acc, "quant_significant", None) != new_val:
                acc.quant_significant = new_val
                updated += 1
            # Also update balance if provided
            bal = row.get("balance")
            if bal is not None:
                try:
                    acc.balance = float(bal)
                except (TypeError, ValueError):
                    pass

    logger.info("Quantitative override applied: %d updated, %d unmatched out of %d rows", updated, unmatched, len(rows))
    return updated, unmatched


def _apply_qualitative_override(engine, rows: List[Dict[str, Any]]) -> tuple[int, int]:
    """Apply user overrides from a qualitative Excel (Phase 2 output).

    Matches rows to engine.accounts by name and updates qual_score,
    qual_risk, and individual qual_factors from the Excel columns.

    Returns (updated_count, unmatched_count).
    """
    if not rows:
        return 0, 0
    accounts = getattr(engine, "accounts", []) or []
    acct_map: Dict[str, List[Any]] = {}
    for acc in accounts:
        # Match on fs_account (used in qualitative export) or raw_name
        fs_name = (getattr(acc, "fs_account", "") or "").strip().lower()
        raw_name = _acct_name(acc).strip().lower()
        if fs_name:
            acct_map.setdefault(fs_name, []).append(acc)
        acct_map.setdefault(raw_name, []).append(acc)

    factor_keys = [
        "contingent_liability", "nature_of_account", "size_composition",
        "transaction_volume", "exposure_to_losses", "accounting_complexity",
        "susceptibility_misstatement", "related_party", "historical_experience",
    ]
    # Map display labels (lowered+underscored) back to factor keys
    label_to_key = {
        "contingent_liability": "contingent_liability",
        "nature_of_account": "nature_of_account",
        "size_&_composition": "size_composition",
        "size_composition": "size_composition",
        "transaction_volume": "transaction_volume",
        "exposure_to_losses": "exposure_to_losses",
        "accounting_complexity": "accounting_complexity",
        "susceptibility/fraud": "susceptibility_misstatement",
        "susceptibility_misstatement": "susceptibility_misstatement",
        "related_party": "related_party",
        "historical_experience": "historical_experience",
    }

    updated = 0
    unmatched = 0
    for row in rows:
        name_val = row.get("account") or row.get("name") or row.get("account_name") or ""
        if not name_val:
            continue
        key = str(name_val).strip().lower()
        matching_accts = acct_map.get(key)
        if not matching_accts:
            unmatched += 1
            continue

        for acc in matching_accts:
            changed = False

            # Update risk level if provided
            risk = row.get("risk") or row.get("qual_risk")
            if risk and str(risk).strip() in ("High", "Moderate", "Low"):
                if getattr(acc, "qual_risk", "") != str(risk).strip():
                    acc.qual_risk = str(risk).strip()
                    changed = True

            # Update score if provided
            score = row.get("score") or row.get("qual_score")
            if score is not None:
                try:
                    new_score = round(float(score), 2)
                    if getattr(acc, "qual_score", 0.0) != new_score:
                        acc.qual_score = new_score
                        changed = True
                except (TypeError, ValueError):
                    pass

            # Update individual factors if provided (Yes/No columns)
            factors = getattr(acc, "qual_factors", {}) or {}
            factors_changed = False
            for col_name, factor_key in label_to_key.items():
                val = row.get(col_name)
                if val is not None:
                    bool_val = str(val).strip().lower() in ("yes", "true", "1")
                    if factors.get(factor_key) != bool_val:
                        factors[factor_key] = bool_val
                        factors_changed = True
            if factors_changed:
                acc.qual_factors = factors
                changed = True

            # Update balance if provided
            bal = row.get("balance")
            if bal is not None:
                try:
                    acc.balance = float(bal)
                except (TypeError, ValueError):
                    pass

            if changed:
                updated += 1

    logger.info("Qualitative override applied: %d updated, %d unmatched out of %d rows", updated, unmatched, len(rows))
    return updated, unmatched


def _apply_scoping_override(engine, rows: List[Dict[str, Any]]) -> tuple[int, int]:
    """Apply user overrides from a scoping Excel (Phase 3 output).

    Matches rows to engine.fs_accounts by name and updates scope and rule.

    Returns (updated_count, unmatched_count).
    """
    if not rows:
        return 0, 0
    fs_accounts = getattr(engine, "fs_accounts", []) or []
    fa_map: Dict[str, List[Any]] = {}
    for fa in fs_accounts:
        key = _acct_name(fa).strip().lower()
        fa_map.setdefault(key, []).append(fa)

    updated = 0
    unmatched = 0
    for row in rows:
        name_val = row.get("name") or row.get("account_name") or row.get("account") or ""
        if not name_val:
            continue
        key = str(name_val).strip().lower()
        matching_fas = fa_map.get(key)
        if not matching_fas:
            unmatched += 1
            continue

        # Update scope decision
        scope = row.get("scope") or row.get("decision") or ""
        scope_str = str(scope).strip()
        if scope_str.lower() in ("in-scope", "in_scope", "in scope"):
            scope_str = "In-scope"
        elif scope_str.lower() in ("out-of-scope", "out_of_scope", "out of scope"):
            scope_str = "Out-of-scope"
        else:
            scope_str = ""

        # Update rule if provided
        rule_val = row.get("rule")
        rule_int = None
        if rule_val is not None:
            rule_str_raw = str(rule_val).strip().upper().replace("R", "")
            try:
                rule_int = int(rule_str_raw)
            except (TypeError, ValueError):
                pass

        for fa in matching_fas:
            changed = False

            if scope_str and getattr(fa, "scope", "") != scope_str:
                fa.scope = scope_str
                changed = True

            if rule_int is not None and getattr(fa, "rule", 0) != rule_int:
                fa.rule = rule_int
                changed = True

            # Update qual_risk if provided
            risk = row.get("qual_risk") or row.get("risk")
            if risk and str(risk).strip() in ("High", "Moderate", "Low"):
                if getattr(fa, "qual_risk", "") != str(risk).strip():
                    fa.qual_risk = str(risk).strip()
                    changed = True

            # Update quant_significant if provided
            qs = row.get("quant_significant")
            if qs is not None:
                qs_bool = str(qs).strip().lower() in ("yes", "true", "1")
                if getattr(fa, "quant_significant", False) != qs_bool:
                    fa.quant_significant = qs_bool
                    changed = True

            if changed:
                updated += 1

    logger.info("Scoping override applied: %d updated, %d unmatched out of %d rows", updated, unmatched, len(rows))
    return updated, unmatched


def _export_qualitative_to_excel(engine, output_dir: str) -> Optional[str]:
    """Export qualitative analysis results to a two-sheet Excel workbook.

    Sheet 1 — 'Qualitative Factor Assessment': per-account Yes/No for each of
    the 9 PCAOB factors, plus weighted score and risk level.
    Sheet 2 — 'Weighted Average Summary': factor weights, flag counts/rates,
    weighted contributions, risk distribution, and overall average score.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not available — cannot export qualitative results")
        return None

    accounts = getattr(engine, "accounts", []) or []
    if not accounts:
        return None

    factor_keys = getattr(engine, "FACTOR_KEYS", [
        "contingent_liability", "nature_of_account", "size_composition",
        "transaction_volume", "exposure_to_losses", "accounting_complexity",
        "susceptibility_misstatement", "related_party", "historical_experience",
    ])
    factor_labels = [
        "Contingent Liability", "Nature of Account", "Size & Composition",
        "Transaction Volume", "Exposure to Losses", "Accounting Complexity",
        "Susceptibility/Fraud", "Related Party", "Historical Experience",
    ]
    config = getattr(engine, "config", None)
    default_weights = {
        "contingent_liability": 5, "nature_of_account": 10, "size_composition": 20,
        "transaction_volume": 20, "exposure_to_losses": 5, "accounting_complexity": 20,
        "susceptibility_misstatement": 10, "related_party": 5, "historical_experience": 5,
    }
    qual_weights = getattr(config, "qual_weights", default_weights) if config else default_weights
    high_threshold = getattr(config, "high_threshold", 0.70) if config else 0.70
    moderate_threshold = getattr(config, "moderate_threshold", 0.35) if config else 0.35

    wb = openpyxl.Workbook()

    # -- Styling helpers --
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yes_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    no_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    mod_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    low_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # ====================================================================
    # SHEET 1 — Qualitative Factor Assessment
    # ====================================================================
    ws1 = wb.active
    ws1.title = "Qualitative Factor Assessment"
    headers = ["ID", "Account", "FS", "Group", "Balance"] + factor_labels + ["Score", "Risk"]
    ws1.append(headers)
    style_header(ws1)

    factor_col_start = 6  # column F (1-based)
    score_col = factor_col_start + len(factor_keys)
    risk_col = score_col + 1

    for acc in accounts:
        risk = getattr(acc, "qual_risk", "") or ""
        if not risk:
            continue
        factors = getattr(acc, "qual_factors", {}) or {}
        row = [
            getattr(acc, "id", ""),
            getattr(acc, "fs_account", "") or _acct_name(acc),
            getattr(acc, "financial_statement", ""),
            getattr(acc, "group_level", ""),
            _acct_balance(acc),
        ]
        for k in factor_keys:
            row.append("Yes" if factors.get(k, False) else "No")
        row.append(round(getattr(acc, "qual_score", 0.0), 2))
        row.append(risk)
        ws1.append(row)

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin_border
        # Balance column ($)
        bal_cell = row[4]
        if isinstance(bal_cell.value, (int, float)):
            bal_cell.number_format = '$#,##0.00'
        # Yes/No colouring
        for idx in range(factor_col_start - 1, factor_col_start - 1 + len(factor_keys)):
            cell = row[idx]
            if cell.value == "Yes":
                cell.fill = yes_fill
            elif cell.value == "No":
                cell.fill = no_fill
            cell.alignment = Alignment(horizontal="center")
        # Risk colouring
        risk_cell = row[risk_col - 1]
        if risk_cell.value == "High":
            risk_cell.fill = high_fill
        elif risk_cell.value == "Moderate":
            risk_cell.fill = mod_fill
        elif risk_cell.value == "Low":
            risk_cell.fill = low_fill
        # Score formatting
        score_cell = row[score_col - 1]
        if isinstance(score_cell.value, (int, float)):
            score_cell.number_format = '0.00'

    auto_width(ws1)

    # ====================================================================
    # SHEET 2 — Weighted Average Summary
    # ====================================================================
    ws2 = wb.create_sheet("Weighted Average Summary")

    # --- Section A: Factor breakdown ---
    ws2.append(["#", "Qualitative Factor", "Weight (/100)", "Accounts Flagged",
                "Flag Rate (%)", "Weighted Contribution"])
    style_header(ws2)

    assessed = [a for a in accounts if (getattr(a, "qual_risk", "") or "")]
    total_accounts = len(assessed)

    total_contribution = 0.0
    for i, (key, label) in enumerate(zip(factor_keys, factor_labels), 1):
        yes_count = sum(
            1 for a in assessed
            if (getattr(a, "qual_factors", {}) or {}).get(key, False)
        )
        weight = qual_weights.get(key, 0)
        flag_rate = (yes_count / total_accounts * 100) if total_accounts else 0
        contribution = (yes_count / total_accounts) * (weight / 100) if total_accounts else 0
        total_contribution += contribution
        ws2.append([
            i, label, weight,
            f"{yes_count} / {total_accounts}",
            round(flag_rate, 1),
            round(contribution, 3),
        ])

    # Total row
    ws2.append(["", "TOTAL", 100, "", "", round(total_contribution, 3)])
    total_row_idx = ws2.max_row
    for cell in ws2[total_row_idx]:
        cell.font = Font(bold=True)
        cell.border = thin_border

    for row in ws2.iter_rows(min_row=2, max_row=total_row_idx - 1):
        for cell in row:
            cell.border = thin_border

    ws2.append([])

    # --- Section B: Risk Distribution ---
    ws2.append(["Risk Level", "Count", "Accounts"])
    dist_header_row = ws2.max_row
    for cell in ws2[dist_header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    high_accts = [a for a in assessed if getattr(a, "qual_risk", "") == "High"]
    mod_accts = [a for a in assessed if getattr(a, "qual_risk", "") == "Moderate"]
    low_accts = [a for a in assessed if getattr(a, "qual_risk", "") == "Low"]

    for risk_label, acct_list, fill in [("High", high_accts, high_fill),
                                         ("Moderate", mod_accts, mod_fill),
                                         ("Low", low_accts, low_fill)]:
        names = ", ".join(
            getattr(a, "fs_account", "") or _acct_name(a) for a in acct_list
        )
        ws2.append([risk_label, len(acct_list), names])
        r = ws2.max_row
        ws2.cell(row=r, column=1).fill = fill
        for cell in ws2[r]:
            cell.border = thin_border
        ws2.cell(row=r, column=3).alignment = Alignment(wrap_text=True)

    ws2.append([])

    # --- Section C: Overall average ---
    avg_score = (
        sum(getattr(a, "qual_score", 0.0) for a in assessed) / total_accounts
    ) if total_accounts else 0
    if avg_score >= high_threshold:
        avg_risk = "High"
    elif avg_score >= moderate_threshold:
        avg_risk = "Moderate"
    else:
        avg_risk = "Low"

    ws2.append(["Overall Weighted Average Score (all accounts):", round(avg_score, 2), avg_risk])
    summary_row = ws2.max_row
    ws2.cell(row=summary_row, column=1).font = Font(bold=True)
    ws2.cell(row=summary_row, column=2).number_format = '0.00'
    ws2.cell(row=summary_row, column=3).font = Font(bold=True)

    auto_width(ws2)
    # Widen the Accounts column for readability
    ws2.column_dimensions["C"].width = max(ws2.column_dimensions["C"].width, 60)

    output_path = os.path.join(output_dir, "qualitative_analysis_results.xlsx")
    wb.save(output_path)
    logger.info("Exported qualitative analysis (%d accounts) to %s", total_accounts, output_path)
    blob_path = _upload_artifact(output_path)
    return blob_path or output_path


def _export_scoping_to_excel(engine, output_dir: str) -> Optional[str]:
    """Export scoping results to a two-sheet Excel workbook.

    Sheet 1 — 'Scoping Results': every FS-level account with quant/qual inputs,
    applied rule, and final in-scope / out-of-scope decision.
    Sheet 2 — 'Scoping Logic Map': the 7-rule decision matrix reference table
    followed by a per-account breakdown showing which rule matched and why.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not available — cannot export scoping results")
        return None

    fs_accounts = getattr(engine, "fs_accounts", []) or []
    if not fs_accounts:
        return None

    wb = openpyxl.Workbook()

    # -- Styling helpers --
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    in_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    out_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row_num=1):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # ====================================================================
    # SHEET 1 — Scoping Results (mirrors the old _export_phase_to_excel output)
    # ====================================================================
    ws1 = wb.active
    ws1.title = "Scoping Results"
    ws1.append(["SI#", "Name", "Financial Statement", "Group", "Balance",
                "Quant Significant", "Qual Score", "Qual Risk",
                "Rule", "Rule Description", "Scope"])
    style_header_row(ws1)

    sorted_accounts = sorted(
        fs_accounts,
        key=lambda x: (0 if getattr(x, "scope", "") == "In-scope" else 1, -abs(getattr(x, "balance", 0))),
    )
    for i, fa in enumerate(sorted_accounts, 1):
        scope = getattr(fa, "scope", "")
        rule = getattr(fa, "rule", 0)
        ws1.append([
            i,
            _acct_name(fa),
            getattr(fa, "fs", "") or getattr(fa, "financial_statement", ""),
            getattr(fa, "group", "") or getattr(fa, "group_level", ""),
            _acct_balance(fa),
            "Yes" if getattr(fa, "quant_significant", False) else "No",
            round(getattr(fa, "qual_score", 0.0), 2),
            getattr(fa, "qual_risk", ""),
            f"R{rule}",
            _SCOPING_RULE_LABELS.get(rule, ""),
            scope,
        ])

    # Format data rows
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin_border
        # Balance ($)
        if isinstance(row[4].value, (int, float)):
            row[4].number_format = '$#,##0.00'
        # Score
        if isinstance(row[6].value, (int, float)):
            row[6].number_format = '0.00'
        # Row colouring by decision
        decision = row[10].value or ""
        fill = in_fill if decision == "In-scope" else out_fill
        for cell in row:
            cell.fill = fill

    # Coverage summary
    in_scope_list = [fa for fa in fs_accounts if getattr(fa, "scope", "") == "In-scope"]
    out_scope_list = [fa for fa in fs_accounts if getattr(fa, "scope", "") != "In-scope"]
    total_bal = sum(abs(getattr(fa, "balance", 0)) for fa in fs_accounts)
    in_bal = sum(abs(getattr(fa, "balance", 0)) for fa in in_scope_list)
    coverage = (in_bal / total_bal * 100) if total_bal else 0

    ws1.append([])
    ws1.append(["", "COVERAGE SUMMARY"])
    ws1.append(["", f"In-scope: {len(in_scope_list)} accounts", "", "", in_bal])
    ws1.append(["", f"Out-of-scope: {len(out_scope_list)} accounts", "", "", total_bal - in_bal])
    ws1.append(["", f"Balance Coverage: {coverage:.1f}%", "", "", total_bal])
    for r in range(ws1.max_row - 3, ws1.max_row + 1):
        for cell in ws1[r]:
            cell.font = Font(bold=True)
        if isinstance(ws1.cell(row=r, column=5).value, (int, float)):
            ws1.cell(row=r, column=5).number_format = '$#,##0.00'

    auto_width(ws1)

    # ====================================================================
    # SHEET 2 — Scoping Logic Map
    # ====================================================================
    ws2 = wb.create_sheet("Scoping Logic Map")

    # --- Section A: Rule Reference Table ---
    ws2.append(["Rule", "Quant Significant", "Qual Risk", "Balance Condition", "Scoping Decision"])
    style_header_row(ws2)

    rule_matrix = [
        ("R1", "Yes", "High",     "Any",      "In-scope"),
        ("R2", "Yes", "Moderate", "Any",      "In-scope"),
        ("R3", "Yes", "Low",      "Any",      "In-scope"),
        ("R4", "No",  "High",     "Non-zero", "In-scope"),
        ("R5", "No",  "High",     "Zero",     "Out-of-scope"),
        ("R6", "No",  "Moderate", "Any",      "Out-of-scope"),
        ("R7", "No",  "Low",      "Any",      "Out-of-scope"),
        ("R8", "No",  "Any",      "Any",      "In-scope (LLM Override)"),
    ]
    for rule_row in rule_matrix:
        ws2.append(list(rule_row))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        decision = row[4].value or ""
        if "In-scope" in decision:
            for cell in row:
                cell.fill = in_fill
        else:
            for cell in row:
                cell.fill = out_fill

    # Spacer
    ws2.append([])
    ws2.append([])

    # --- Section B: Per-account logic breakdown ---
    breakdown_header_row = ws2.max_row + 1
    ws2.append(["SI#", "FS Account", "Balance", "Quant Significant",
                "Qual Score", "Qual Risk", "Matched Rule", "Rule Logic", "Decision"])
    for cell in ws2[breakdown_header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for i, fa in enumerate(sorted_accounts, 1):
        scope = getattr(fa, "scope", "")
        rule = getattr(fa, "rule", 0)
        quant_sig = getattr(fa, "quant_significant", False)
        qual_risk = getattr(fa, "qual_risk", "")
        qual_score = round(getattr(fa, "qual_score", 0.0), 2)

        ws2.append([
            i,
            _acct_name(fa),
            _acct_balance(fa),
            "Yes" if quant_sig else "No",
            qual_score,
            qual_risk,
            f"R{rule}",
            _SCOPING_RULE_LABELS.get(rule, ""),
            scope,
        ])

    for row in ws2.iter_rows(min_row=breakdown_header_row + 1, max_row=ws2.max_row):
        for cell in row:
            cell.border = thin_border
        # Balance ($)
        if isinstance(row[2].value, (int, float)):
            row[2].number_format = '$#,##0.00'
        # Score
        if isinstance(row[4].value, (int, float)):
            row[4].number_format = '0.00'
        # Row colouring
        decision = row[8].value or ""
        fill = in_fill if decision == "In-scope" else out_fill
        for cell in row:
            cell.fill = fill

    auto_width(ws2)
    # Widen rule logic column
    ws2.column_dimensions["H"].width = max(ws2.column_dimensions["H"].width, 55)

    output_path = os.path.join(output_dir, "scoping_results.xlsx")
    wb.save(output_path)
    logger.info("Exported scoping results (%d accounts) to %s", len(fs_accounts), output_path)
    blob_path = _upload_artifact(output_path)
    return blob_path or output_path


def _export_phase_to_excel(data: List[Dict], output_dir: str, filename: str, sheet_title: str) -> Optional[str]:
    """Export a list of dicts to a styled Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not available — cannot export phase results")
        return None

    if not data:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = list(data[0].keys())
    # Pretty-print header names
    display_headers = [h.replace("_", " ").title() for h in headers]
    ws.append(display_headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx in range(1, len(display_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in data:
        row_values = []
        for h in headers:
            val = row.get(h)
            # Ensure complex objects are converted to strings for Excel cells
            if isinstance(val, (list, dict, tuple)):
                val = str(val)
            row_values.append(val)
        ws.append(row_values)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)
    logger.info("Exported %d rows to %s", len(data), output_path)
    blob_path = _upload_artifact(output_path)
    return blob_path or output_path


def _export_accounts_to_excel(engine, output_dir: str) -> Optional[str]:
    """Export all extracted accounts to an Excel file after ingestion."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        logger.warning("openpyxl not available — cannot export accounts to Excel")
        return None

    accounts = getattr(engine, "accounts", []) or []
    if not accounts:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Accounts"

    # Header row
    headers = ["#", "Account Name", "Balance", "Financial Statement", "Group Level",
               "FS Line Item", "Confidence"]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill  # type: ignore
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for i, acc in enumerate(accounts, 1):
        name = _acct_name(acc)
        bal = _acct_balance(acc)
        fs = getattr(acc, "financial_statement", "") or ""
        group = getattr(acc, "group_level", "") or ""
        fs_item = getattr(acc, "fs_account", "") or ""
        confidence = getattr(acc, "confidence", "") or ""
        ws.append([i, name, bal, fs, group, fs_item, confidence])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output_path = os.path.join(output_dir, "extracted_accounts.xlsx")
    wb.save(output_path)
    logger.info("Exported %d accounts to %s", len(accounts), output_path)
    blob_path = _upload_artifact(output_path)
    return blob_path or output_path


def _read_rcm_from_excel(output_path: str) -> Dict:
    """
    Read the generated RCM workbook and return its sheets as lists of records.
    Tries to import pandas; if not available, returns an empty dict.
    """
    try:
        import pandas as pd  # type: ignore
    except ImportError:
        return {"error": "pandas not available — open the workbook directly"}

    try:
        xl = pd.ExcelFile(output_path)
        sheets: Dict[str, Any] = {}
        for sheet in xl.sheet_names:
            # Sheet "10. RCM" has a two-row header: Row 1 is merged group
            # headers ("RCM", "Financial Assertions", etc.) and Row 2 has
            # the actual column names.  Use header=1 so pandas reads Row 2
            # as the header instead of the merged group row.
            header_row = 1 if "rcm" in sheet.lower() else 0
            df = xl.parse(sheet, header=header_row)
            # Replace NaN with None so JSON serialisation works
            df = df.where(df.notna(), other=None)
            sheets[sheet] = df.to_dict(orient="records")
        return {"sheets": list(xl.sheet_names), "data": sheets}
    except Exception as exc:
        return {"error": f"Could not read workbook: {exc}"}


class SoxScopingEngineTool(Tool):
    @property
    def name(self) -> str:
        return "run_sox_scoping_engine"

    @property
    def description(self) -> str:
        return (
            "Run the Control Scoping Engine (from folder 'sox_package 2') using dynamic input files. "
            "Accepts a trial balance file path and optional SOP/policy paths provided at runtime. "
            "No hardcoded input files are used."
        )

    @property
    def category(self) -> ToolCategory:
        return ToolCategory.ANALYSIS

    @property
    def parameters(self) -> List[ToolParameter]:
        return [
            ToolParameter(
                "trial_balance_path",
                "string",
                "Absolute path to trial balance file (.xlsx/.xls/.csv/.tsv/.pdf/.json)",
            ),
            ToolParameter(
                "sop_paths",
                "array",
                "Optional list of SOP/policy file paths OR folder paths for downstream process/control extraction. "
                "If a folder path is given, all supported files (.docx, .pdf, .txt, .xlsx) inside it are used.",
                required=False,
                items={"type": "string"},
            ),
            ToolParameter(
                "benchmark",
                "string",
                "Materiality benchmark (ask user after accounts are fetched from trial balance)",
                required=False,
                enum=[
                    "EPS",
                    "Revenue",
                    "Assets",
                    "EBITDA",
                    "Adjusted EBITDA",
                    "Net Interest",
                    "PBT",
                    "Net Income",
                ],
            ),
            ToolParameter(
                "materiality_pct",
                "number",
                "Materiality percentage, e.g. 1.5 (ask user after accounts are fetched)",
                required=False,
            ),
            ToolParameter(
                "run_downstream",
                "boolean",
                "Whether to run downstream mapping and control/risk extraction phases",
                required=False,
                default=True,
            ),
            ToolParameter(
                "output_filename",
                "string",
                "Output Excel filename (saved inside the current session output directory)",
                required=False,
                default="sox_scoping_output.xlsx",
            ),
            ToolParameter(
                "max_workers",
                "integer",
                "Concurrency for engine LLM calls",
                required=False,
                default=5,
            ),
            ToolParameter(
                "top_k_chunks",
                "integer",
                "RAG retrieval chunk count per process in extraction phase",
                required=False,
                default=15,
            ),
            ToolParameter(
                "override_excel_path",
                "string",
                "Optional path to a user-modified Excel file to override the results of the "
                "previous phase. The file must follow the same format as the exported Excel "
                "from that phase. Accepted after any phase: accounts (Phase 0), quantitative "
                "(Phase 1), qualitative (Phase 2), or scoping (Phase 3).",
                required=False,
            ),
        ]

    def execute(self, args: Dict[str, Any], state: AgentState) -> ToolResult:  # noqa: C901
        cfg = get_config()
        # Clear stale override info from previous calls
        state.last_override = None

        trial_balance_path = str(args["trial_balance_path"]).strip().strip("'\"")
        # Resolve blob path if needed
        try:
            trial_balance_path = _resolve_blob_file(trial_balance_path)
        except RuntimeError as exc:
            # Blob download failed — fall back to cached path if engine exists
            if state.scoping_engine is not None and state.scoping_trial_balance_path:
                logger.warning(
                    "Blob resolution failed for '%s' (%s) — using cached path '%s'",
                    trial_balance_path, exc, state.scoping_trial_balance_path,
                )
                trial_balance_path = state.scoping_trial_balance_path
            else:
                return ToolResult(success=False, data={}, error=str(exc))
        trial_balance_path = os.path.abspath(os.path.expanduser(trial_balance_path))

        if not os.path.exists(trial_balance_path):
            # If the engine is already cached, fall back to the cached path
            # rather than failing — the LLM likely mangled/hallucinated the path.
            if state.scoping_engine is not None and state.scoping_trial_balance_path:
                cached = state.scoping_trial_balance_path
                if os.path.exists(cached):
                    logger.warning(
                        "Trial balance path '%s' not found — falling back to cached '%s'",
                        trial_balance_path, cached,
                    )
                    trial_balance_path = cached
                else:
                    return ToolResult(
                        success=False, data={},
                        error=f"Trial balance file not found: {trial_balance_path}",
                    )
            else:
                return ToolResult(
                    success=False, data={},
                    error=f"Trial balance file not found: {trial_balance_path}",
                )

        sop_paths = args.get("sop_paths") or []
        valid_sops: List[str] = []
        invalid_sops: List[str] = []
        _SOP_EXTENSIONS = {".docx", ".pdf", ".txt", ".xlsx", ".xls", ".doc", ".pptx", ".csv"}
        for p in sop_paths:
            # Resolve blob paths for SOP files/directories
            resolved = _resolve_blob_file(str(p).strip().strip("'\""))
            if resolved != str(p).strip().strip("'\""):
                # Was a blob file, use the resolved local path
                full = resolved
            else:
                # Try as blob directory
                resolved_dir = _resolve_blob_directory(str(p).strip().strip("'\""))
                full = os.path.abspath(os.path.expanduser(resolved_dir))
            if os.path.isdir(full):
                # Expand directory into individual supported files
                logger.info("Expanding SOP directory: %s", full)
                found_any = False
                for entry in os.listdir(full):
                    entry_path = os.path.join(full, entry)
                    if os.path.isfile(entry_path):
                        ext = os.path.splitext(entry)[1].lower()
                        if ext in _SOP_EXTENSIONS:
                            valid_sops.append(entry_path)
                            found_any = True
                        else:
                            logger.debug("Skipping non-SOP file: %s", entry_path)
                if not found_any:
                    invalid_sops.append(f"{full} (directory contains no supported SOP files)")
            elif os.path.isfile(full):
                valid_sops.append(full)
            else:
                invalid_sops.append(full)

        benchmark = str(args.get("benchmark", "")).strip()
        materiality_pct_arg = args.get("materiality_pct", None)

        # --- Override Excel path (optional user-uploaded modified results) ---
        override_excel_raw = args.get("override_excel_path")
        override_excel_path: Optional[str] = None
        if override_excel_raw:
            # Resolve blob path if needed
            resolved_override = _resolve_blob_file(str(override_excel_raw).strip().strip("'\""))
            override_excel_path = os.path.abspath(
                os.path.expanduser(resolved_override)
            )
            if not os.path.isfile(override_excel_path):
                return ToolResult(
                    success=False,
                    data={},
                    error=f"Override Excel file not found: {override_excel_path}",
                )

        # ----------------------------------------------------------------
        # Guard: block re-entry while waiting for user input.
        #
        # After any intermediate phase result (accounts_fetched,
        # quantitative_done, qualitative_done, scoped_done) is returned,
        # scoping_awaiting_input is set to True.  If the agent calls the
        # tool again WITHOUT providing the required next input, reject the
        # call immediately so it cannot loop and produce repeated output.
        #
        # Reset conditions:
        #   - accounts_fetched waiting → user must supply benchmark
        #     OR override_excel_path (user uploading modified accounts)
        #   - quantitative_done waiting → user already confirmed (same args = advance)
        #     OR override_excel_path (user uploading modified quant results)
        #   - qualitative_done waiting  → same
        #     OR override_excel_path (user uploading modified qual results)
        #   - scoped_done waiting       → user must supply sop_paths
        #     OR override_excel_path (user uploading modified scoping results)
        # ----------------------------------------------------------------
        if state.scoping_awaiting_input:
            phase = state.scoping_phase
            has_new_input = False
            if override_excel_path:
                has_new_input = True          # user uploaded override Excel → advance
            elif phase == "ingested" and benchmark:
                has_new_input = True          # benchmark provided → advance to quant
            elif phase == "scoped_done" and valid_sops:
                has_new_input = True          # SOP paths provided → advance to downstream
            elif phase == "none":
                has_new_input = True          # fresh start
            elif phase == "complete":
                # Already done — only allow re-entry if user provided NEW SOPs
                # (different from the cached set) to re-run downstream.
                if valid_sops:
                    sorted_new = sorted(valid_sops)
                    sorted_cached = sorted(state.scoping_sop_paths or [])
                    has_new_input = sorted_new != sorted_cached
                else:
                    has_new_input = False
            # NOTE: quantitative_done and qualitative_done do NOT accept
            # benchmark as "new input" — the benchmark is already cached from
            # the previous call.  The LLM must STOP and show the upload
            # question.  On the NEXT user message, scoping_awaiting_input is
            # cleared by process_message() in loop.py, so the guard won't
            # trigger and the tool proceeds normally.

            if not has_new_input:
                logger.warning(
                    "Blocked re-entry: scoping_awaiting_input=True phase=%s — agent called without required input",
                    phase,
                )
                return ToolResult(
                    success=False,
                    data={
                        "status": phase,
                        "blocked": True,
                        "user_upload_prompt": (
                            "You can download the results Excel, make any modifications, "
                            "and upload the modified file. Would you like to upload a "
                            "modified Excel, or continue with the current results?"
                        ),
                    },
                    error=(
                        "WAITING_FOR_USER: The scoping engine already returned its results for the "
                        f"'{phase}' phase and is waiting for the user to respond. "
                        "Do NOT call run_sox_scoping_engine again. "
                        "You MUST show the previous results to the user and ask: "
                        "'Would you like to upload a modified Excel, or continue with the current results?' "
                        "Include the user_upload_prompt text in your response. STOP and wait for their reply."
                    ),
                )
            # User provided the required input — clear the flag and proceed
            state.scoping_awaiting_input = False
            logger.info("scoping_awaiting_input cleared — user provided new input (phase=%s)", phase)

        allowed_benchmarks = {
            "EPS", "Revenue", "Assets", "EBITDA",
            "Adjusted EBITDA", "Net Interest", "PBT", "Net Income",
        }
        benchmark_options = sorted(allowed_benchmarks)

        run_downstream = bool(args.get("run_downstream", True))
        max_workers = int(args.get("max_workers", 5))
        top_k_chunks = int(args.get("top_k_chunks", 15))

        output_filename = str(args.get("output_filename", "sox_scoping_output.xlsx")).strip()
        if not output_filename:
            output_filename = "sox_scoping_output.xlsx"
        if not output_filename.lower().endswith(".xlsx"):
            output_filename += ".xlsx"

        if state.output_dir:
            output_dir = state.output_dir
        else:
            tb_dir = os.path.dirname(trial_balance_path)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.join(tb_dir, f"sox_agent_{ts}")
            os.makedirs(output_dir, exist_ok=True)
            state.output_dir = output_dir

        output_path = os.path.join(output_dir, output_filename)

        try:
            engine_module = _load_external_engine_module()
            if hasattr(engine_module, "MAX_WORKERS"):
                engine_module.MAX_WORKERS = max_workers
            Config = engine_module.Config
            SOXScopingEngine = engine_module.SOXScopingEngine

            # ----------------------------------------------------------------
            # Phase 0 — Ingest + Categorize
            # Only re-runs if trial balance path changed or no cached engine.
            # Also runs categorize() here so benchmark reference values are
            # available for the user before they pick a benchmark.
            # ----------------------------------------------------------------
            # Determine if we need a fresh ingest.  Use os.path.samefile()
            # when possible to handle Windows case-insensitive paths and
            # short/long name differences (e.g. TEJWAN~1 vs Tejwant).
            # Also fall back to the cached path when the LLM hallucinates
            # a non-existent path but the cached engine is still valid.
            if state.scoping_engine is not None and state.scoping_trial_balance_path:
                _paths_match = False
                try:
                    if os.path.exists(trial_balance_path) and os.path.exists(state.scoping_trial_balance_path):
                        _paths_match = os.path.samefile(trial_balance_path, state.scoping_trial_balance_path)
                    else:
                        _paths_match = (
                            os.path.normcase(os.path.normpath(trial_balance_path))
                            == os.path.normcase(os.path.normpath(state.scoping_trial_balance_path))
                        )
                except (OSError, ValueError):
                    _paths_match = trial_balance_path == state.scoping_trial_balance_path

                if not _paths_match and not os.path.exists(trial_balance_path):
                    # LLM hallucinated a non-existent path — fall back to cached
                    logger.warning(
                        "Trial balance path '%s' does not exist — using cached path '%s'",
                        trial_balance_path, state.scoping_trial_balance_path,
                    )
                    trial_balance_path = state.scoping_trial_balance_path
                    _paths_match = True

                need_ingest = not _paths_match
            else:
                need_ingest = True

            if need_ingest:
                logger.info("Ingesting trial balance: %s", trial_balance_path)
                state.scoping_progress = {
                    "phase": "starting", "message": "Starting scoping engine...",
                    "current": 0, "total": 1, "pct": 0.0,
                }
                engine_cfg = Config(
                    api_key=cfg.openai_api_key,
                    azure_endpoint=cfg.azure_openai_endpoint,
                    azure_api_version=cfg.azure_openai_api_version,
                    model=cfg.openai_model,
                    embedding_model=cfg.azure_openai_embedding_deployment,
                    vision_model=cfg.openai_model,
                    trial_balance_path=trial_balance_path,
                    sop_paths=[],
                    output_path=output_path,
                    benchmark="",
                    materiality_pct=0.0,
                    top_k_chunks=top_k_chunks,
                )
                engine = SOXScopingEngine(engine_cfg)
                with patch("builtins.input", return_value="n"):
                    update_progress(state, "ingest", 0, 1, "Reading and parsing trial balance file…")
                    engine.ingest(trial_balance_path)
                    num_parsed = len(engine.accounts)
                    update_progress(state, "ingest", 1, 1, f"Trial balance parsed — {num_parsed} accounts found")

                    # Categorize with progress — wrap as_completed to count batches
                    num_accounts = len(engine.accounts)
                    cat_batch_size = 40
                    num_cat_batches = math.ceil(num_accounts / cat_batch_size) if num_accounts > 0 else 1
                    update_progress(state, "categorize", 0, num_cat_batches, f"Categorising {num_accounts} accounts (0/{num_cat_batches} batches)")

                    _orig_ac = getattr(engine_module, "as_completed", concurrent.futures.as_completed)
                    engine_module.as_completed = make_counting_as_completed(
                        state, "categorize", num_cat_batches, _orig_ac,
                        "Categorising accounts (batch {current}/{total})",
                    )
                    try:
                        engine.categorize()
                    finally:
                        engine_module.as_completed = _orig_ac
                state.scoping_engine = engine
                state.scoping_trial_balance_path = trial_balance_path
                state.scoping_phase = "ingested"
                state.scoping_benchmark = None
                state.scoping_materiality_pct = None
                state.scoping_sop_paths = None
                logger.info("Engine cached after ingest+categorize: %d accounts", len(engine.accounts))

                accounts_list = _extract_account_list(engine, max_items=50)
                benchmark_values = _extract_benchmark_values(engine)
                group_distribution = _extract_group_level_distribution(engine)

                # Export all accounts to Excel for user download
                accounts_excel_path = _export_accounts_to_excel(engine, output_dir)
                artifacts = [accounts_excel_path] if accounts_excel_path else []
                if accounts_excel_path:
                    state.artifacts.append(accounts_excel_path)

                state.scoping_awaiting_input = True
                clear_progress(state)
                return ToolResult(
                    success=True,
                    data={
                        "status": "accounts_fetched",
                        "accounts_ingested": len(engine.accounts),
                        "accounts": accounts_list,
                        "group_level_distribution": group_distribution,
                        "accounts_excel": accounts_excel_path,
                        "benchmark_reference_values": benchmark_values,
                        "benchmark_options": benchmark_options,
                        "override_applied": getattr(state, "last_override", None),
                        "user_upload_prompt": (
                            "You can download the exported accounts Excel, make any modifications "
                            "(adjust balances, reclassify accounts, fix names), and upload the modified "
                            "file. Would you like to upload a modified Excel, or continue with the "
                            "current results?"
                        ),
                        "phase_steps": [
                            "Parsed and ingested the trial balance file (extracted all accounts with balances)",
                            "Normalised account names and cleaned data",
                            "Categorised each account into financial statement lines (BS/IS, group level, FS line item) using LLM",
                            "Computed benchmark reference values (Revenue, Assets, etc.) for materiality calculation",
                            "Exported all extracted accounts to Excel for download",
                        ],
                        "next_action": (
                            "Present ONLY the phase_steps above as a single numbered list (do NOT repeat steps from any previous phase). "
                            "Show the benchmark reference values ONCE (do not duplicate them). "
                            "Mention that all accounts have been exported to the Excel file. "
                            "Note: The upload question will be appended automatically — do NOT include it yourself. "
                            "End your response with the upload question. Do NOT ask for benchmark or"
                            "materiality_pct yet. Those will be asked AFTER the user decides on the upload. "
                            "Do NOT ask for SOPs yet. "
                            "STOP — do NOT call run_sox_scoping_engine again until the user replies. "
                            "If the user uploads a modified Excel: call run_sox_scoping_engine with "
                            "override_excel_path set to the uploaded file path (without benchmark — "
                            "the engine will re-show accounts and THEN you ask for benchmark). "
                            "If the user says 'continue' or declines to upload: THEN ask for "
                            "(1) materiality benchmark (show benchmark_options as a numbered list), "
                            "(2) materiality percentage (e.g. 1.5). Wait for the user's answer."
                        ),
                    },
                    artifacts=artifacts,
                    summary=f"Accounts fetched: {len(engine.accounts)} — exported to {accounts_excel_path or 'N/A'} — awaiting benchmark and materiality_pct",
                )

            # Reuse cached engine
            engine = state.scoping_engine
            logger.info("Reusing cached engine — skipping ingest (%d accounts)", len(engine.accounts))

            # ----------------------------------------------------------------
            # Apply user-uploaded override Excel (if provided)
            # The override is applied to the engine state for the CURRENT phase,
            # so subsequent phases use the user's modifications.
            # ----------------------------------------------------------------
            if override_excel_path:
                override_rows = _read_override_excel(override_excel_path)
                if not override_rows:
                    return ToolResult(
                        success=False,
                        data={"override_excel_path": override_excel_path},
                        error="Could not read any data from the override Excel file. "
                              "Ensure it is a valid .xlsx file with data in the first sheet.",
                    )

                phase = state.scoping_phase
                override_count = 0
                override_unmatched = 0
                override_phase_label = ""

                if phase == "ingested":
                    override_count, override_unmatched = _apply_accounts_override(engine, override_rows)
                    override_phase_label = "accounts (Phase 0)"
                    # Re-export updated accounts
                    accounts_excel_path = _export_accounts_to_excel(engine, output_dir)
                    if accounts_excel_path:
                        state.artifacts.append(accounts_excel_path)

                elif phase == "quantitative_done":
                    override_count, override_unmatched = _apply_quantitative_override(engine, override_rows)
                    override_phase_label = "quantitative (Phase 1)"
                    # Update cached quantitative results
                    state.scoping_quantitative_results = _extract_quantitative_results(engine)

                elif phase == "qualitative_done":
                    override_count, override_unmatched = _apply_qualitative_override(engine, override_rows)
                    override_phase_label = "qualitative (Phase 2)"

                elif phase == "scoped_done":
                    override_count, override_unmatched = _apply_scoping_override(engine, override_rows)
                    override_phase_label = "scoping (Phase 3)"

                else:
                    return ToolResult(
                        success=False,
                        data={"phase": phase, "override_excel_path": override_excel_path},
                        error=f"Cannot apply override at phase '{phase}'. "
                              "Overrides are accepted after phases: ingested, quantitative_done, "
                              "qualitative_done, scoped_done.",
                    )

                logger.info(
                    "Override applied: %d items updated, %d unmatched from '%s' at phase '%s'",
                    override_count, override_unmatched, override_excel_path, phase,
                )

                # Store override metadata so the agent can report what happened
                warning = None
                if override_unmatched > 0:
                    warning = (
                        f"{override_unmatched} row(s) in the uploaded Excel did not match "
                        "any existing account names and were skipped. Check for typos or "
                        "extra whitespace in the account name column."
                    )

                state.last_override = {
                    "path": override_excel_path,
                    "phase": override_phase_label,
                    "rows_read": len(override_rows),
                    "items_updated": override_count,
                    "rows_unmatched": override_unmatched,
                    "warning": warning,
                }

                # ----------------------------------------------------------------
                # IMPORTANT: After applying an override, immediately return the
                # CURRENT phase's updated results so the user can review them and
                # decide whether to upload again or continue. Do NOT fall through
                # to the next phase.
                # ----------------------------------------------------------------
                state.scoping_awaiting_input = True
                clear_progress(state)

                if phase == "ingested":
                    # Re-surface updated accounts (same as the normal accounts_fetched return)
                    accounts_list = _extract_account_list(engine, max_items=50)
                    benchmark_values = _extract_benchmark_values(engine)
                    return ToolResult(
                        success=True,
                        data={
                            "status": "accounts_fetched",
                            "accounts_ingested": len(engine.accounts),
                            "accounts": accounts_list,
                            "benchmark_reference_values": benchmark_values,
                            "benchmark_options": sorted(allowed_benchmarks),
                            "override_applied": state.last_override,
                            "user_upload_prompt": (
                                "Your modifications have been applied. You can upload another "
                                "modified Excel if you'd like to make further changes, or continue "
                                "with the current results."
                            ),
                            "next_action": (
                                "Tell the user their override was applied (show items_updated count). "
                                "If override_applied.warning is present, show it to the user. "
                                "Show the updated accounts and benchmark reference values. "
                                "Note: The upload question will be appended automatically — do NOT include it yourself. "
                                "End your response with the upload question. Do NOT ask for benchmark yet."
                                "STOP — wait for the user's reply. "
                                "If user uploads again: call with override_excel_path. "
                                "If user says 'continue': THEN ask for benchmark and materiality_pct."
                            ),
                        },
                        artifacts=[accounts_excel_path] if accounts_excel_path else [],
                        summary=f"Override applied: {override_count} accounts updated — re-showing results",
                    )

                elif phase == "quantitative_done":
                    # Re-surface updated quantitative results
                    quant = state.scoping_quantitative_results or _extract_quantitative_results(engine)
                    quant_excel = _export_phase_to_excel(
                        quant.get("quantitatively_in_scope", []) + quant.get("quantitatively_out_of_scope", []),
                        output_dir, "quantitative_analysis_results.xlsx", "Quantitative Analysis",
                    )
                    return ToolResult(
                        success=True,
                        data={
                            "status": "quantitative_done",
                            "benchmark": state.scoping_benchmark,
                            "materiality_pct": state.scoping_materiality_pct,
                            "quantitative_analysis": quant,
                            "results_excel": quant_excel,
                            "override_applied": state.last_override,
                            "user_upload_prompt": (
                                "Your modifications have been applied. You can upload another "
                                "modified Excel if you'd like to make further changes, or continue "
                                "to qualitative analysis."
                            ),
                            "next_action": (
                                "Tell the user their override was applied (show items_updated count). "
                                "If override_applied.warning is present, show it to the user. "
                                "Show the updated quantitative analysis summary. "
                                "Note: The upload question will be appended automatically — do NOT include it yourself. "
                                "End your response with the upload question. Do NOT ask about qualitative yet."
                                "STOP — wait for the user's reply. "
                                "If user uploads again: call with override_excel_path. "
                                "If user says 'continue': proceed to qualitative."
                            ),
                        },
                        artifacts=[quant_excel] if quant_excel else [],
                        summary=f"Override applied: {override_count} accounts updated — re-showing quantitative results",
                    )

                elif phase == "qualitative_done":
                    # Re-surface updated qualitative results
                    qual = _extract_qualitative_results(engine)
                    qual_excel = _export_qualitative_to_excel(engine, output_dir)
                    return ToolResult(
                        success=True,
                        data={
                            "status": "qualitative_done",
                            "benchmark": state.scoping_benchmark,
                            "materiality_pct": state.scoping_materiality_pct,
                            "qualitative_analysis": qual,
                            "results_excel": qual_excel,
                            "override_applied": state.last_override,
                            "user_upload_prompt": (
                                "Your modifications have been applied. You can upload another "
                                "modified Excel if you'd like to make further changes, or continue "
                                "to final scoping."
                            ),
                            "next_action": (
                                "Tell the user their override was applied (show items_updated count). "
                                "If override_applied.warning is present, show it to the user. "
                                "Show the updated qualitative analysis summary. "
                                "Note: The upload question will be appended automatically — do NOT include it yourself. "
                                "End your response with the upload question. Do NOT ask about scoping yet."
                                "STOP — wait for the user's reply. "
                                "If user uploads again: call with override_excel_path. "
                                "If user says 'continue': proceed to final scoping."
                            ),
                        },
                        artifacts=[qual_excel] if qual_excel else [],
                        summary=f"Override applied: {override_count} accounts updated — re-showing qualitative results",
                    )

                elif phase == "scoped_done":
                    # Re-surface updated scoping results
                    inscope = _extract_inscope_accounts(engine)
                    fs_accounts = getattr(engine, "fs_accounts", []) or []
                    out_count = sum(
                        1 for fa in fs_accounts
                        if str(getattr(fa, "scope", "")).lower() not in ("in-scope", "in_scope", "in scope")
                    )
                    scoping_excel = _export_scoping_to_excel(engine, output_dir)
                    return ToolResult(
                        success=True,
                        data={
                            "status": "scoped_done",
                            "benchmark": state.scoping_benchmark,
                            "materiality_pct": state.scoping_materiality_pct,
                            "in_scope_accounts": [
                                {"name": a.get("name"), "category": a.get("category"),
                                 "balance": a.get("balance"), "rule": a.get("rule"),
                                 "rule_description": a.get("rule_description")}
                                for a in inscope
                            ],
                            "in_scope_count": len(inscope),
                            "out_of_scope_count": out_count,
                            "results_excel": scoping_excel,
                            "override_applied": state.last_override,
                            "user_upload_prompt": (
                                "Your modifications have been applied. You can upload another "
                                "modified Excel if you'd like to make further changes, or continue "
                                "with these results."
                            ),
                            "next_action": (
                                "Tell the user their override was applied (show items_updated count). "
                                "If override_applied.warning is present, show it to the user. "
                                "Show the updated in-scope accounts list. "
                                "Note: The upload question will be appended automatically — do NOT include it yourself. "
                                "End your response with the upload question. Do NOT ask for SOPs yet."
                                "STOP — wait for the user's reply. "
                                "If user uploads again: call with override_excel_path. "
                                "If user says 'continue': THEN ask for SOP file paths."
                            ),
                        },
                        artifacts=[scoping_excel] if scoping_excel else [],
                        summary=f"Override applied: {override_count} accounts updated — re-showing scoping results",
                    )

            # For Phase 2+ calls the LLM may omit benchmark/materiality_pct
            # (e.g. after context compaction). Fall back to cached values so
            # the engine can proceed without requiring the LLM to repeat them.
            if state.scoping_phase in ("quantitative_done", "qualitative_done", "scoped_done", "complete"):
                if not benchmark and state.scoping_benchmark:
                    benchmark = state.scoping_benchmark
                    logger.info("Using cached benchmark: %s", benchmark)
                if materiality_pct_arg is None and state.scoping_materiality_pct is not None:
                    materiality_pct_arg = state.scoping_materiality_pct
                    logger.info("Using cached materiality_pct: %s", materiality_pct_arg)

            # If benchmark still not provided, re-surface accounts for user
            if not benchmark:
                accounts_list = _extract_account_list(engine, max_items=50)
                benchmark_values = _extract_benchmark_values(engine)
                group_distribution = _extract_group_level_distribution(engine)
                state.scoping_awaiting_input = True
                return ToolResult(
                    success=True,
                    data={
                        "status": "accounts_fetched",
                        "accounts_ingested": len(engine.accounts),
                        "accounts": accounts_list,
                        "group_level_distribution": group_distribution,
                        "benchmark_reference_values": benchmark_values,
                        "benchmark_options": benchmark_options,
                        "user_upload_prompt": (
                            "You can download the exported accounts Excel, make any modifications "
                            "(adjust balances, reclassify accounts, fix names), and upload the modified "
                            "file. Would you like to upload a modified Excel, or continue with the "
                            "current results?"
                        ),
                        "next_action": (
                            "Show accounts and benchmark reference values. "
                            "Note: The upload question will be appended automatically — do NOT include it yourself. "
                            "End your response with the upload question. Do NOT ask for benchmark or"
                            "materiality_pct yet. Wait for the user's reply about upload. "
                            "If they decline, THEN ask for benchmark and materiality_pct. "
                            "Do NOT ask for SOPs yet. "
                            "STOP — do NOT call run_sox_scoping_engine again until the user replies."
                        ),
                    },
                    summary=f"{len(engine.accounts)} accounts cached — awaiting benchmark and materiality inputs",
                )

            # Validate inputs
            if benchmark not in allowed_benchmarks:
                return ToolResult(
                    success=False,
                    data={
                        "status": "invalid_input",
                        "benchmark": benchmark,
                        "benchmark_options": benchmark_options,
                    },
                    error=f"Invalid benchmark '{benchmark}'. Choose one of: {', '.join(sorted(allowed_benchmarks))}.",
                )

            if materiality_pct_arg is None:
                return ToolResult(
                    success=False,
                    data={"status": "invalid_input", "benchmark": benchmark},
                    error="materiality_pct is required. Please provide a numeric value (example: 1.5).",
                )

            try:
                materiality_pct = float(materiality_pct_arg)
            except (TypeError, ValueError):
                return ToolResult(
                    success=False,
                    data={"status": "invalid_input"},
                    error=f"Invalid materiality_pct '{materiality_pct_arg}'. Must be a numeric value.",
                )

            if materiality_pct <= 0:
                return ToolResult(
                    success=False,
                    data={"status": "invalid_input", "materiality_pct": materiality_pct},
                    error="materiality_pct must be greater than 0.",
                )

            # ----------------------------------------------------------------
            # Phase 1 — Quantitative analysis
            # Runs when benchmark/pct changed or phase is still "ingested".
            # Returns results and pauses — user must confirm to continue.
            # ----------------------------------------------------------------
            need_quantitative = (
                state.scoping_phase == "ingested"
                or state.scoping_benchmark != benchmark
                or state.scoping_materiality_pct != materiality_pct
            )

            if need_quantitative:
                logger.info("Running quantitative analysis | benchmark=%s | pct=%.2f", benchmark, materiality_pct)
                with patch("builtins.input", return_value="n"):
                    update_progress(state, "quantitative", 0, 3, "Setting materiality benchmark...", sub_step="Setting materiality benchmark")
                    engine.set_materiality(benchmark=benchmark, pct=materiality_pct)
                    update_progress(state, "quantitative", 1, 3, "Computing materiality threshold...", sub_step="Computing materiality threshold")
                    engine.compute_threshold()
                    update_progress(state, "quantitative", 2, 3, "Classifying accounts by significance...", sub_step="Classifying accounts")
                    engine.run_quantitative()
                    update_progress(state, "quantitative", 3, 3, "Quantitative analysis complete")
                state.scoping_benchmark = benchmark
                state.scoping_materiality_pct = materiality_pct
                state.scoping_phase = "quantitative_done"
                state.scoping_sop_paths = None

                quant = _extract_quantitative_results(engine)
                # Cache quantitative results so qualitative phase can carry them forward
                # without re-extracting (run_qualitative may modify engine.accounts)
                state.scoping_quantitative_results = quant

                # Export quantitative results to Excel — include all accounts so file is always generated
                for item in quant.get("quantitatively_in_scope", []):
                    item["decision"] = "In-scope"
                for item in quant.get("quantitatively_out_of_scope", []):
                    item["decision"] = "Out-of-scope"
                _in_names = {a["name"] for a in quant.get("quantitatively_in_scope", [])}
                _out_names = {a["name"] for a in quant.get("quantitatively_out_of_scope", [])}
                _unclassified_rows = [
                    {"name": _acct_name(acc), "balance": _acct_balance(acc), "decision": "Unclassified"}
                    for acc in (getattr(engine, "accounts", []) or [])
                    if _acct_name(acc) not in _in_names and _acct_name(acc) not in _out_names
                ]
                quant_export_data = (
                    quant.get("quantitatively_in_scope", [])
                    + quant.get("quantitatively_out_of_scope", [])
                    + _unclassified_rows
                )
                quant_excel = _export_phase_to_excel(
                    quant_export_data, output_dir,
                    "quantitative_analysis_results.xlsx", "Quantitative Analysis"
                )
                artifacts = [quant_excel] if quant_excel else []
                if quant_excel:
                    state.artifacts.append(quant_excel)

                benchmark_ref_values = _extract_benchmark_values(engine)

                state.scoping_awaiting_input = True
                clear_progress(state)
                return ToolResult(
                    success=True,
                    data={
                        "status": "quantitative_done",
                        "benchmark": benchmark,
                        "materiality_pct": materiality_pct,
                        "benchmark_reference_values": benchmark_ref_values,
                        "accounts_ingested": len(getattr(engine, "accounts", []) or []),
                        "quantitative_analysis": quant,
                        "results_excel": quant_excel,
                        "phase_steps": [
                            "Set materiality benchmark and percentage",
                            "Compute materiality threshold from benchmark base value",
                            "Classify each account as quantitatively significant (balance >= threshold) or not",
                        ],
                        "override_applied": getattr(state, "last_override", None),
                        "user_upload_prompt": (
                            "You can download the quantitative results Excel, make any modifications "
                            "(change which accounts are in-scope or out-of-scope, adjust balances), and "
                            "upload the modified file. Would you like to upload a modified Excel, or "
                            "continue to qualitative analysis?"
                        ),
                        "next_action": (
                            "Present ONLY the phase_steps above as a single numbered list (do NOT repeat steps from any previous phase). "
                            "If override_applied is present, mention that the user's modifications were applied "
                            "(show items_updated count). If override_applied.warning is present, show it to the user. "
                            "Show the quantitative analysis summary (in-scope count, out-of-scope count, threshold). "
                            "Mention the downloadable Excel file with full results. "
                            "Note: The upload question will be appended automatically — do NOT include it yourself. "
                            "End your response with the upload question. Do NOT ask about qualitative analysis yet."
                            "Wait for the user's reply. "
                            "If the user uploads an override Excel: call run_sox_scoping_engine with "
                            "override_excel_path set to the uploaded file path. "
                            "If the user says 'continue' or declines: THEN proceed to qualitative analysis "
                            "by calling run_sox_scoping_engine with the same args. "
                            "STOP — do NOT call run_sox_scoping_engine again until the user replies."
                        ),
                    },
                    artifacts=artifacts,
                    summary=(
                        f"Quantitative analysis done — "
                        f"{quant.get('in_count', '?')} in-scope, "
                        f"{quant.get('out_count', '?')} out-of-scope "
                        f"(threshold: {quant.get('materiality_threshold', '?')})"
                    ),
                )

            # ----------------------------------------------------------------
            # Phase 2 — Qualitative analysis
            # Runs when phase is "quantitative_done".
            # Returns results and pauses — user must confirm to continue.
            # ----------------------------------------------------------------
            if state.scoping_phase == "quantitative_done":
                logger.info("Running qualitative analysis")
                num_accounts = len(engine.accounts)
                qual_batch_size = 30
                num_qual_batches = math.ceil(num_accounts / qual_batch_size) if num_accounts > 0 else 1
                update_progress(state, "qualitative", 0, num_qual_batches, f"Assessing risk factors for {num_accounts} accounts (0/{num_qual_batches} batches)")

                _orig_ac = getattr(engine_module, "as_completed", concurrent.futures.as_completed)
                engine_module.as_completed = make_counting_as_completed(
                    state, "qualitative", num_qual_batches, _orig_ac,
                    "Assessing risk factors (batch {current}/{total})",
                )
                try:
                    with patch("builtins.input", return_value="n"):
                        engine.run_qualitative()
                finally:
                    engine_module.as_completed = _orig_ac
                state.scoping_phase = "qualitative_done"

                qual = _extract_qualitative_results(engine)

                # Export full qualitative results (factor Yes/No + weighted summary)
                qual_excel = _export_qualitative_to_excel(engine, output_dir)
                artifacts = [qual_excel] if qual_excel else []
                if qual_excel:
                    state.artifacts.append(qual_excel)

                # Carry-forward materiality + quant data so the full picture is
                # available even on page-reload from this phase's cached result.
                # Use cached quantitative results (from quantitative_done phase) to avoid
                # re-extracting after run_qualitative() which may have modified engine.accounts.
                _bm_ref = _extract_benchmark_values(engine)
                _quant_cf = state.scoping_quantitative_results or _extract_quantitative_results(engine)

                state.scoping_awaiting_input = True
                clear_progress(state)
                return ToolResult(
                    success=True,
                    data={
                        "status": "qualitative_done",
                        # carry-forward from earlier phases
                        "benchmark": state.scoping_benchmark or benchmark,
                        "materiality_pct": state.scoping_materiality_pct or materiality_pct,
                        "benchmark_reference_values": _bm_ref,
                        "accounts_ingested": len(getattr(engine, "accounts", []) or []),
                        "quantitative_analysis": _quant_cf,
                        # qualitative results
                        "qualitative_analysis": qual,
                        "results_excel": qual_excel,
                        "phase_steps": [
                            "Assessed qualitative risk factors for each account (fraud risk, complexity, estimation uncertainty, volume of transactions, regulatory sensitivity)",
                            "Scored each account on a 0-1 qualitative risk scale using LLM analysis",
                            "Classified accounts as High / Moderate / Low qualitative risk",
                            "Identified accounts that should be added to scope despite being quantitatively insignificant (High qual risk with nonzero balance)",
                        ],
                        "override_applied": getattr(state, "last_override", None),
                        "user_upload_prompt": (
                            "You can download the qualitative results Excel, make any modifications "
                            "(change risk scores, risk levels, or factor assessments), and upload the "
                            "modified file. Would you like to upload a modified Excel, or continue to "
                            "final scoping?"
                        ),
                        "next_action": (
                            "Present ONLY the phase_steps above as a single numbered list (do NOT repeat steps from any previous phase). "
                            "If override_applied is present, mention that the user's modifications were applied "
                            "(show items_updated count). If override_applied.warning is present, show it to the user. "
                            "Show the qualitative analysis summary (high/moderate/low risk counts, additions, removals). "
                            "Mention the downloadable Excel file with full results. "
                            "Note: The upload question will be appended automatically — do NOT include it yourself. "
                            "End your response with the upload question. Do NOT ask about proceeding to scoping yet."
                            "Wait for the user's reply. "
                            "If the user uploads an override Excel: call run_sox_scoping_engine with "
                            "override_excel_path set to the uploaded file path. "
                            "If the user says 'continue' or declines: THEN proceed to final scoping "
                            "by calling run_sox_scoping_engine with the same args. "
                            "STOP after sending this message — do NOT call run_sox_scoping_engine again. "
                            "Wait for the user to reply."
                        ),
                    },
                    artifacts=artifacts,
                    summary=(
                        f"Qualitative analysis done — "
                        f"{qual.get('additions_count', 0)} accounts added, "
                        f"{qual.get('removals_count', 0)} removed"
                    ),
                )

            # ----------------------------------------------------------------
            # Phase 3 — Final scoping (in-scope / out-of-scope determination)
            # Runs when phase is "qualitative_done".
            # Returns in-scope account list and pauses.
            # ----------------------------------------------------------------
            if state.scoping_phase == "qualitative_done":
                logger.info("Running final scoping")
                update_progress(state, "scoping", 0, 1, "Applying scoping rules...")
                with patch("builtins.input", return_value="n"):
                    engine.run_scoping()
                update_progress(state, "scoping", 1, 1, "Scoping rules applied")
                state.scoping_phase = "scoped_done"

                inscope = _extract_inscope_accounts(engine)
                fs_accounts = getattr(engine, "fs_accounts", []) or []
                out_count = sum(
                    1 for fa in fs_accounts
                    if str(getattr(fa, "scope", "")).lower() not in ("in-scope", "in_scope", "in scope")
                )

                # Count LLM override accounts (Rule 8)
                llm_overrides = [a for a in inscope if a.get("rule") == 8]

                # Export scoping results + logic map to Excel
                scoping_excel = _export_scoping_to_excel(engine, output_dir)
                artifacts = [scoping_excel] if scoping_excel else []
                if scoping_excel:
                    state.artifacts.append(scoping_excel)

                state.scoping_awaiting_input = True
                clear_progress(state)

                # Pre-format the in-scope accounts as a markdown table so the
                # agent can display it directly without reconstructing from JSON.
                # This avoids empty/broken tables when the data gets large.
                _inscope_data = [
                    {
                        "name": a.get("name"),
                        "category": a.get("category"),
                        "balance": a.get("balance"),
                        "rule": a.get("rule"),
                        "rule_description": a.get("rule_description"),
                    }
                    for a in inscope
                ]
                _md_rows = []
                for a in _inscope_data:
                    bal = a.get("balance")
                    bal_str = f"{bal:,.2f}" if isinstance(bal, (int, float)) else str(bal or "")
                    _md_rows.append(
                        f"| {a.get('name', '')} | {a.get('category', '')} | {bal_str} "
                        f"| {a.get('rule', '')} | {a.get('rule_description', '')} |"
                    )
                _md_table = (
                    "| Account Name | Category | Balance | Rule | Rule Description |\n"
                    "|---|---|---|---|---|\n"
                    + "\n".join(_md_rows)
                ) if _md_rows else "(no in-scope accounts)"

                return ToolResult(
                    success=True,
                    data={
                        "status": "scoped_done",
                        "benchmark": state.scoping_benchmark or benchmark,
                        "materiality_pct": state.scoping_materiality_pct or materiality_pct,
                        "accounts_ingested": len(getattr(engine, "accounts", []) or []),
                        "in_scope_accounts_table": _md_table,
                        "in_scope_accounts": _inscope_data,
                        "in_scope_count": len(inscope),
                        "out_of_scope_count": out_count,
                        "llm_override_count": len(llm_overrides),
                        "llm_overrides": [{"name": a.get("name")} for a in llm_overrides],
                        "results_excel": scoping_excel,
                        "phase_steps": [
                            "Aggregated detailed accounts into FS-level accounts (roll-up)",
                            "Applied 7-rule scoping matrix (quantitative significance + qualitative risk)",
                            "Ran LLM validation — reviewed all scoping decisions for gaps and overrides (Rule 8)",
                            "Calculated total balance coverage of in-scope accounts",
                        ],
                        "override_applied": getattr(state, "last_override", None),
                        "user_upload_prompt": (
                            "You can download the scoping results Excel, make any modifications "
                            "(change which accounts are in-scope or out-of-scope, adjust scoping rules), "
                            "and upload the modified file. Would you like to upload a modified Excel, "
                            "or continue with these results?"
                        ),
                        "next_action": (
                            "Present ONLY the phase_steps above as a single numbered list (do NOT repeat steps from any previous phase). "
                            "If override_applied is present, mention that the user's modifications were applied "
                            "(show items_updated count). If override_applied.warning is present, show it to the user. "
                            "Display the in_scope_accounts_table markdown EXACTLY as provided — do NOT rebuild it from the JSON array. "
                            "If there are LLM override accounts (Rule 8), explain them clearly. "
                            "Mention the downloadable Excel file with all scoping results. "
                            "Note: The upload question will be appended automatically — do NOT include it yourself. "
                            "End your response with the upload question. Do NOT ask for SOPs yet."
                            "Wait for the user's reply. "
                            "If the user uploads an override Excel: call run_sox_scoping_engine with "
                            "override_excel_path set to the uploaded file path. "
                            "If the user says 'continue' or declines: THEN ask the user to provide "
                            "SOP/policy file paths or a folder path — SOPs are REQUIRED for process "
                            "mapping and RCM generation. Accept either individual file paths OR a folder "
                            "path containing SOPs (.docx, .pdf, .txt, .xlsx). "
                            "STOP after sending this message — do NOT call run_sox_scoping_engine again "
                            "until the user explicitly replies. "
                            "Do NOT call the tool just because the user says 'yes' or 'proceed' — "
                            "you must have actual SOP file paths before calling again."
                        ),
                    },
                    artifacts=artifacts,
                    summary=f"Scoping done — {len(inscope)} in-scope, {out_count} out-of-scope, {len(llm_overrides)} LLM overrides",
                )

            # ----------------------------------------------------------------
            # Phase 4 — Downstream mapping + export
            # Runs when phase is "scoped_done" or when SOPs changed on re-run.
            # This is the final phase — exports the RCM workbook.
            # ----------------------------------------------------------------
            if state.scoping_phase in ("scoped_done", "complete"):
                # Guard: if phase is already "complete" and SOPs haven't changed,
                # return cached results instead of re-running the entire export
                # (which includes expensive COSO mapping LLM calls).
                if state.scoping_phase == "complete":
                    sorted_new_sops = sorted(valid_sops) if valid_sops else []
                    sorted_cached_sops = sorted(state.scoping_sop_paths or [])
                    if sorted_new_sops == sorted_cached_sops or not valid_sops:
                        logger.info("Phase already complete and SOPs unchanged — returning cached result")
                        # Re-set the flag so the agent cannot loop-call this tool
                        state.scoping_awaiting_input = True
                        # Try to read the existing exported workbook
                        rcm_workbook = _read_rcm_from_excel(output_path) if os.path.exists(output_path) else None
                        output_blob = None
                        if os.path.exists(output_path):
                            output_blob = _upload_artifact(output_path)
                        fs_accounts = getattr(engine, "fs_accounts", []) or []
                        in_scope = sum(1 for fa in fs_accounts if str(getattr(fa, "scope", "")).lower() in ("in-scope", "in_scope", "in scope"))
                        out_scope = len(fs_accounts) - in_scope
                        process_map = getattr(engine, "process_map", None) or {}
                        process_count = len(sorted(set(p for ps in process_map.values() for p in ps))) if process_map else 0
                        return ToolResult(
                            success=False,
                            data={
                                "status": "complete",
                                "blocked": True,
                                "output_excel": output_blob or output_path,
                                "already_complete": True,
                                "accounts_ingested": len(getattr(engine, "accounts", []) or []),
                                "fs_accounts_total": len(fs_accounts),
                                "in_scope_accounts": in_scope,
                                "out_of_scope_accounts": out_scope,
                                "process_count": process_count,
                                "rcm_workbook": rcm_workbook,
                            },
                            error=(
                                "ALREADY_COMPLETE: The RCM workbook was already exported. "
                                "Do NOT call run_sox_scoping_engine again. "
                                "Display the RCM content and mention the downloadable Excel file."
                            ),
                            artifacts=[output_blob or output_path] if (output_blob or output_path) else [],
                            summary=(
                                f"Control scoping already complete: {in_scope} in-scope accounts, "
                                f"{process_count} processes mapped, output={output_blob or output_path}"
                            ),
                        )

                # Guard: SOPs are required to proceed. If none provided, stop and ask.
                # This prevents the agent from looping when the user says "yes/proceed"
                # without supplying actual SOP file paths.
                if state.scoping_phase == "scoped_done" and not valid_sops:
                    fs_accounts_count = len(getattr(engine, "fs_accounts", []) or [])
                    in_scope_count = sum(
                        1 for fa in (getattr(engine, "fs_accounts", []) or [])
                        if str(getattr(fa, "scope", "")).lower() in ("in-scope", "in_scope", "in scope")
                    )
                    # Block re-entry so the LLM cannot loop calling this tool
                    # without SOPs in the same turn.  The flag is cleared when
                    # the next user message arrives (providing SOP paths).
                    state.scoping_awaiting_input = True
                    return ToolResult(
                        success=True,
                        data={
                            "status": "awaiting_sops",
                            "in_scope_count": in_scope_count,
                            "fs_accounts_total": fs_accounts_count,
                            "next_action": (
                                "Scoping is complete. You must now ask the user to provide SOP file paths "
                                "before calling run_sox_scoping_engine again. "
                                "Do NOT call this tool again until you have actual file paths from the user. "
                                "Ask: 'Please provide the file paths or folder path for your SOP/policy documents "
                                "(.docx, .pdf, .txt, .xlsx) so I can proceed with process mapping and RCM generation.'"
                            ),
                        },
                        summary=f"Awaiting SOP paths — {in_scope_count} accounts in scope, scoping complete",
                    )

                sorted_new_sops = sorted(valid_sops)
                sorted_cached_sops = sorted(state.scoping_sop_paths or [])
                need_downstream = run_downstream and (
                    state.scoping_phase == "scoped_done"
                    or sorted_new_sops != sorted_cached_sops
                )

                if need_downstream:
                    logger.info("Running downstream phases | sops=%d", len(valid_sops))

                    _orig_ac = getattr(engine_module, "as_completed", concurrent.futures.as_completed)

                    with patch("builtins.input", return_value="n"):
                        # --- map_to_processes (batched, batch_size=40) ---
                        fs_accounts = getattr(engine, "fs_accounts", []) or []
                        in_scope_list = [
                            fa for fa in fs_accounts
                            if str(getattr(fa, "scope", "")).lower() in ("in-scope", "in_scope", "in scope")
                        ]
                        map_batches = math.ceil(len(in_scope_list) / 40) if in_scope_list else 1
                        update_progress(state, "map_to_processes", 0, map_batches, "Mapping accounts to business processes...")
                        engine_module.as_completed = make_counting_as_completed(
                            state, "map_to_processes", map_batches, _orig_ac,
                            "Mapping to processes (batch {current}/{total})",
                        )
                        try:
                            engine.map_to_processes()
                        finally:
                            engine_module.as_completed = _orig_ac

                        # --- ingest_sops ---
                        if valid_sops:
                            num_sops = len(valid_sops)
                            update_progress(state, "ingest_sops", 0, num_sops, f"Parsing SOP files (0/{num_sops})")
                            engine_module.as_completed = make_counting_as_completed(
                                state, "ingest_sops", num_sops, _orig_ac,
                                "Parsing SOP files ({current}/{total})",
                            )
                            try:
                                engine.ingest_sops(valid_sops)
                            finally:
                                engine_module.as_completed = _orig_ac

                        # --- validate process-SOP coverage (instant) ---
                        if hasattr(engine, "_validate_process_sop_coverage"):
                            update_progress(state, "validate_coverage", 0, 1, "Validating process-SOP coverage...")
                            engine._validate_process_sop_coverage()
                            update_progress(state, "validate_coverage", 1, 1, "Coverage validation complete")
                        else:
                            logger.warning("Engine missing _validate_process_sop_coverage — skipping")

                        # --- extract_and_map (1 future per unique process) ---
                        process_map = getattr(engine, "process_map", None) or {}
                        unique_procs = sorted(set(p for ps in process_map.values() for p in ps)) if process_map else []
                        num_procs = len(unique_procs) if unique_procs else 1
                        update_progress(state, "extract_and_map", 0, num_procs, f"Extracting controls (0/{num_procs} processes)")
                        engine_module.as_completed = make_counting_as_completed(
                            state, "extract_and_map", num_procs, _orig_ac,
                            "Extracting controls (process {current}/{total})",
                        )
                        try:
                            engine.extract_and_map()
                        finally:
                            engine_module.as_completed = _orig_ac

                        # --- completeness review (instant) ---
                        if hasattr(engine, "_completeness_review"):
                            update_progress(state, "completeness_review", 0, 1, "Running completeness review...")
                            engine._completeness_review()
                            update_progress(state, "completeness_review", 1, 1, "Completeness review done")
                        else:
                            logger.warning("Engine missing _completeness_review — skipping")

                    state.scoping_sop_paths = list(valid_sops)

                    update_progress(state, "export", 0, 1, "Exporting RCM workbook...")
                    # Ensure output directory exists (may have been cleaned by
                    # blob cache cleanup or temp file reaper during long runs).
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    engine.export(output_path)
                    update_progress(state, "export", 1, 1, "RCM workbook exported")
                else:
                    logger.info("Reusing downstream results — SOPs unchanged, skipping export")
                state.scoping_phase = "complete"
                # Prevent the agent from re-calling this tool in the same turn
                # (which would re-run the expensive COSO mapping export).
                state.scoping_awaiting_input = True

                fs_accounts = getattr(engine, "fs_accounts", []) or []
                in_scope = sum(1 for fa in fs_accounts if str(getattr(fa, "scope", "")).lower() in ("in-scope", "in_scope", "in scope"))
                out_scope = sum(1 for fa in fs_accounts if str(getattr(fa, "scope", "")).lower() not in ("in-scope", "in_scope", "in scope"))
                process_map = getattr(engine, "process_map", None) or {}
                process_count = len(sorted(set(p for ps in process_map.values() for p in ps))) if process_map else 0
                coverage_stats = getattr(engine, "process_sop_coverage", {}) if run_downstream else {}

                # Read the exported workbook so the agent can display the RCM to the user
                rcm_workbook = _read_rcm_from_excel(output_path)

                # Upload final RCM workbook to blob
                output_blob = _upload_artifact(output_path)

                pipeline_timing = finish_pipeline(state)
                return ToolResult(
                    success=True,
                    data={
                        "status": "success",
                        "output_excel": output_blob or output_path,
                        "sop_paths_used": valid_sops,
                        "sop_paths_invalid": invalid_sops,
                        "benchmark": benchmark,
                        "materiality_pct": materiality_pct,
                        "run_downstream": run_downstream,
                        "accounts_ingested": len(getattr(engine, "accounts", []) or []),
                        "fs_accounts_total": len(fs_accounts),
                        "in_scope_accounts": in_scope,
                        "out_of_scope_accounts": out_scope,
                        "process_count": process_count,
                        "sop_coverage": coverage_stats,
                        "rcm_workbook": rcm_workbook,
                        "pipeline_total_elapsed": pipeline_timing.get("pipeline_total_elapsed"),
                        "pipeline_total_elapsed_seconds": pipeline_timing.get("pipeline_total_elapsed_seconds"),
                        "next_action": (
                            "Display the RCM content from rcm_workbook.data — show each sheet as a labelled table. "
                            "Mention the downloadable Excel file. "
                            "Do NOT call run_sox_scoping_engine again."
                        ),
                    },
                    artifacts=[output_blob or output_path],
                    summary=(
                        f"Control scoping complete: {in_scope} in-scope accounts, "
                        f"{process_count} processes mapped, output={output_blob or output_path}"
                    ),
                )

            # Unexpected state — reset so next call starts clean
            bad_phase = state.scoping_phase
            finish_pipeline(state)
            logger.warning("Unexpected scoping_phase '%s' — resetting cache", bad_phase)
            state.scoping_engine = None
            state.scoping_phase = "none"
            return ToolResult(
                success=False,
                data={
                    "status": "reset",
                    "bad_phase": bad_phase,
                    "cache_cleared": True,
                    "retry_note": "Cache has been reset. Retrying will restart from Phase 0 (re-ingest).",
                },
                error=(
                    f"Unexpected engine state (phase='{bad_phase}') — cache reset. "
                    "Call run_sox_scoping_engine again with trial_balance_path to restart."
                ),
            )

        except Exception as exc:
            finish_pipeline(state)  # Clear progress + pipeline timer on error
            failed_phase = state.scoping_phase  # capture before clearing
            logger.exception("Control scoping engine execution failed at phase '%s'", failed_phase)

            phase_labels = {
                "none":              "ingesting and categorising the trial balance",
                "ingested":          "running quantitative analysis",
                "quantitative_done": "running qualitative analysis",
                "qualitative_done":  "determining in-scope accounts",
                "scoped_done":       "running process mapping and SOP extraction",
                "complete":          "re-exporting the RCM workbook",
            }
            phase_desc = phase_labels.get(failed_phase, f"phase '{failed_phase}'")

            # Only clear the full engine cache for early-phase failures.
            # For downstream/SOP failures (scoped_done, complete), preserve the
            # engine so that a retry with corrected SOP paths does NOT re-ingest
            # the trial balance from scratch.
            if failed_phase in ("scoped_done", "complete"):
                # Roll back to scoped_done so downstream re-runs on retry
                state.scoping_phase = "scoped_done"
                state.scoping_sop_paths = None  # force SOP re-processing
                cache_cleared = False
                retry_note = (
                    "The engine encountered an error during process mapping / SOP extraction, "
                    "but the trial balance and scoping results are preserved. "
                    "Retrying will re-run only the downstream phases (process mapping + SOP extraction). "
                    "You can provide corrected SOP file paths or folder paths."
                )
            else:
                state.scoping_engine = None
                state.scoping_trial_balance_path = None
                state.scoping_phase = "none"
                cache_cleared = True
                retry_note = (
                    "Engine cache has been cleared. "
                    "Retrying will restart from Phase 0 (re-ingest the trial balance)."
                )

            return ToolResult(
                success=False,
                data={
                    "trial_balance_path": trial_balance_path,
                    "sop_paths_provided": sop_paths,
                    "sop_paths_valid": valid_sops,
                    "sop_paths_invalid": invalid_sops,
                    "failed_phase": failed_phase,
                    "cache_cleared": cache_cleared,
                    "retry_note": retry_note,
                },
                error=f"Control scoping engine failed while {phase_desc}: {exc}",
            )
