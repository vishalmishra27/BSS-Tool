"""
RCM Semantic Deduplication Engine - LLM VERSION (Azure OpenAI)
===============================================================
Uses Azure OpenAI API for intelligent duplicate detection
- Understands domain context and nuance
- Application/System uniqueness rule
- SubProcess filtering
- Many-to-Many relationship awareness (1 risk → many controls, many risks → 1 control)
- Per-process RCM file analysis
- Structured reasoning for each comparison

Author: Rishi
Date: February 2026
"""

import pandas as pd
import numpy as np
import json
import time
import os
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import AzureOpenAI

# =====================================================
# CONFIGURATION
# =====================================================

# Azure OpenAI Configuration — imported from central config
from engines.config import (
    AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY,
    AZURE_OPENAI_API_VERSION, AZURE_OPENAI_DEPLOYMENT as AZURE_OPENAI_DEPLOYMENT_NAME,
    OPENAI_API_KEY, OPENAI_MODEL,
)

# Input: Folder containing per-process RCM Excel files
# OR a single Excel file with a "Process" column
RCM_INPUT = "/Users/rishi/Downloads/Sample_Data/output.xlsx"

# Set to True if RCM_INPUT is a folder with per-process files
# Set to False if RCM_INPUT is a single file with "Process" column
INPUT_IS_FOLDER = False

# Output settings
OUTPUT_FOLDER = "/Users/rishi/Downloads"
OUTPUT_EXCEL_NAME = "RCM_LLM_Duplicates_Review"
OUTPUT_JSON_NAME = "rcm_llm_duplicates"

# LLM settings
MAX_RETRIES = 3
RETRY_DELAY = 2  # seconds

# Parallel processing settings
MAX_WORKERS = 5  # Number of parallel threads for LLM calls

# Batch size: how many pairs to evaluate in one API call
# Higher = fewer API calls but longer responses
BATCH_SIZE = 5

# Columns used for comparison
COMPARE_COLUMNS = [
    "Risk Id",
    "Risk Title",
    "Risk Description",
    "Control Id",
    "Control Description",
    "Control Objective",
    "Control Owner",
    "Control Rating",
    "Nature of Control",
    "Control Type",
    "Control Frequency",
    "Application/System",
    "Process",
    "SubProcess",
    "risk_level",
]

# =====================================================
# END CONFIGURATION
# =====================================================


def create_llm_client():
    return AzureOpenAI(
        api_key=OPENAI_API_KEY,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_version=AZURE_OPENAI_API_VERSION,
    )


client = create_llm_client()


# =====================================================
# SYSTEM PROMPT FOR LLM
# =====================================================
SYSTEM_PROMPT = """You are an expert compliance auditor specializing in Risk Control Matrices (RCM).
Your task is to analyze pairs of RCM entries and determine if they are semantic duplicates.

DEFINITION OF SEMANTIC DUPLICATE:
Two RCM entries are duplicates if they describe the SAME underlying risk AND the SAME control mechanism, 
just expressed with different words, vocabulary, or sentence structure. 

The test: "If I merged these two entries into one, would any meaningful information be lost?"
- If NO = DUPLICATE (one is redundant)
- If YES = UNIQUE (both contain distinct information)

DUPLICATE DETECTION RULES:

1. SEMANTIC SIMILARITY ASSESSMENT:
   - Focus on MEANING, not exact wording
   - Different vocabulary describing the same concept = DUPLICATE
   - Paraphrased descriptions of the same risk/control = DUPLICATE
   - Industry synonyms and equivalent terms should be treated as the same
   - Ask: "Are these two entries trying to say the same thing?"

2. APPLICATION/SYSTEM RULE:
   - DIFFERENT Application/System = AUTOMATICALLY UNIQUE (not duplicates)
   - Even identical risk/control text on different systems = UNIQUE
   - Same or equivalent system naming + same meaning = could be DUPLICATE

3. MANY-TO-MANY RELATIONSHIP AWARENESS:
   
   ONE RISK → MULTIPLE CONTROLS (NOT duplicates):
   - A single risk SHOULD have multiple different controls addressing it
   - Example: Risk "Unauthorized Access" may have:
     * Control A: Approval workflow (preventive)
     * Control B: Quarterly access review (detective)
     * Control C: System-enforced SoD (automated)
   - These are THREE UNIQUE entries, not duplicates
   
   MULTIPLE RISKS → ONE CONTROL (NOT duplicates):
   - A single control can mitigate multiple different risks
   - Example: Control "Access Review" may address:
     * Risk A: Unauthorized access
     * Risk B: SoD violations  
     * Risk C: Terminated employee access
   - These are THREE UNIQUE entries, not duplicates

4. WHAT MAKES ENTRIES UNIQUE (not duplicates):
   - Different risk concepts (even if related to same process area)
   - Different control mechanisms (approval vs review vs reconciliation vs system check)
   - Different control timing (preventive vs detective vs corrective)
   - Different Application/System
   - Different stages of a process (setup vs execution vs review)
   - Complementary controls that work together (not redundant)

5. WHAT MAKES ENTRIES DUPLICATES:
   - Same risk concept expressed with different vocabulary
   - Same control mechanism described with different wording
   - Paraphrased versions of each other
   - One entry is essentially a restatement of the other
   - No new information would be gained by keeping both

6. CONFIDENCE LEVELS:
   - HIGH: Clearly same meaning, just different wording - definitely redundant
   - MEDIUM: Very similar meaning, minor differences that likely don't matter
   - LOW: Related concepts, but may have subtle distinctions worth human review

DECISION FRAMEWORK:
1. First check: Different Application/System? → UNIQUE (stop here)
2. Then ask: Is the RISK concept the same or different?
3. Then ask: Is the CONTROL mechanism the same or different?
4. If BOTH risk AND control are semantically the same → DUPLICATE
5. If EITHER risk OR control is meaningfully different → UNIQUE

Respond ONLY in valid JSON format with no additional text."""


def format_entry_for_llm(row, row_num, available_cols):
    """Format a single RCM entry for LLM comparison"""
    entry = f"[Row {row_num}]\n"
    for col in available_cols:
        if col in row.index and pd.notna(row[col]):
            entry += f"  {col}: {row[col]}\n"
    return entry


def build_comparison_prompt(pairs, df, available_cols):
    """Build prompt for comparing multiple pairs"""
    prompt = "Analyze the following RCM entry pairs and determine if each pair contains semantic duplicates.\n\n"
    prompt += "For EACH pair, respond with a JSON object.\n\n"
    
    for idx, (i, j) in enumerate(pairs, 1):
        row_i = df.iloc[i]
        row_j = df.iloc[j]
        
        prompt += f"=== PAIR {idx} ===\n"
        prompt += f"ENTRY A:\n{format_entry_for_llm(row_i, i, available_cols)}\n"
        prompt += f"ENTRY B:\n{format_entry_for_llm(row_j, j, available_cols)}\n"
    
    prompt += f"""
Respond with a JSON array containing exactly {len(pairs)} objects, one per pair, in order:
[
  {{
    "pair_number": 1,
    "is_duplicate": true/false,
    "confidence": "HIGH"/"MEDIUM"/"LOW",
    "similarity_assessment": "brief 1-line similarity summary",
    "reasoning": "2-3 sentence explanation of why duplicate or not",
    "application_match": true/false,
    "risk_match": true/false,
    "control_mechanism_match": true/false,
    "why_controls_differ": "if control_mechanism_match is false, explain how controls differ (type, timing, activity, threshold, etc.)",
    "recommendation": "REMOVE_A"/"REMOVE_B"/"KEEP_BOTH"/"MANUAL_REVIEW"
  }}
]

REMEMBER:
- Different Application/System = AUTOMATICALLY not a duplicate
- Same Risk but Different Control Mechanism = NOT a duplicate (proper SOX design)
- Different Risks but Same Control = NOT a duplicate (1 control mitigates many risks)
- Only flag as duplicate if SAME app + SAME risk + SAME control mechanism (just different wording)

Respond with ONLY the JSON array, no other text."""
    
    return prompt


def call_llm(prompt, retries=MAX_RETRIES):
    """Call Azure OpenAI API with retry logic"""
    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": prompt}
                ],
                max_completion_tokens=16384,
            )
            
            response_text = response.choices[0].message.content.strip()
            
            # Clean response if wrapped in markdown
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            results = json.loads(response_text)
            return results
            
        except json.JSONDecodeError as e:
            print(f"    JSON parse error (attempt {attempt + 1}): {e}")
            if attempt < retries - 1:
                time.sleep(RETRY_DELAY)
            else:
                print(f"    Returning raw response for manual parsing")
                return None
                
        except Exception as e:
            print(f"    API error (attempt {attempt + 1}): {e}")
            if attempt < retries - 1:
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                return None


def pre_filter_pairs(df):
    """
    Generate candidate pairs for LLM comparison.
    Skips pairs that are obviously unique based on:
    1. Different Application/System (automatic unique)
    2. Different SubProcess (different area = different controls)
    3. Different Nature of Control (Preventive vs Detective = different controls)
    4. Different Control Type (Manual vs Automated = different controls)
    """
    all_pairs = []
    skipped_app = 0
    skipped_subprocess = 0
    skipped_nature = 0
    skipped_type = 0
    
    for i in range(len(df)):
        for j in range(i + 1, len(df)):
            # Rule 1: Different Application/System = unique
            app_i = str(df.iloc[i].get("Application/System", "")).strip().lower()
            app_j = str(df.iloc[j].get("Application/System", "")).strip().lower()
            
            if app_i and app_j and app_i != app_j:
                skipped_app += 1
                continue
            
            # Rule 2: Different SubProcess = different area, different controls
            subprocess_i = str(df.iloc[i].get("SubProcess", "")).strip().lower()
            subprocess_j = str(df.iloc[j].get("SubProcess", "")).strip().lower()
            
            if subprocess_i and subprocess_j and subprocess_i != subprocess_j:
                skipped_subprocess += 1
                continue
            
            # Rule 3: Different Nature of Control = different controls for same risk
            # (Preventive vs Detective vs Corrective)
            nature_i = str(df.iloc[i].get("Nature of Control", "")).strip().lower()
            nature_j = str(df.iloc[j].get("Nature of Control", "")).strip().lower()
            
            if nature_i and nature_j and nature_i != nature_j:
                skipped_nature += 1
                continue
            
            # Rule 4: Different Control Type = different controls
            # (Manual vs Automated vs IT-Dependent Manual)
            type_i = str(df.iloc[i].get("Control Type", "")).strip().lower()
            type_j = str(df.iloc[j].get("Control Type", "")).strip().lower()
            
            if type_i and type_j and type_i != type_j:
                skipped_type += 1
                continue
            
            # Passed all pre-filters - send to LLM for detailed comparison
            all_pairs.append((i, j))
    
    return all_pairs, skipped_app, skipped_subprocess, skipped_nature, skipped_type


def process_batch(batch_num, batch_pairs, df, available_cols):
    """Process a single batch of pairs - used for parallel execution"""
    prompt = build_comparison_prompt(batch_pairs, df, available_cols)
    results = call_llm(prompt)
    return batch_num, batch_pairs, results


def process_single_rcm(df, process_name):
    """Process a single process RCM and find duplicates using LLM"""
    
    print(f"\n{'=' * 70}")
    print(f"PROCESSING: {process_name}")
    print(f"{'=' * 70}")
    print(f"Records: {len(df)}")
    
    available_cols = [c for c in COMPARE_COLUMNS if c in df.columns]
    
    # Phase 1: Pre-filter pairs
    print(f"\n>> Phase 1: Pre-filtering pairs...")
    candidate_pairs, skipped_app, skipped_subprocess, skipped_nature, skipped_type = pre_filter_pairs(df)
    total_possible = len(df) * (len(df) - 1) // 2
    total_skipped = skipped_app + skipped_subprocess + skipped_nature + skipped_type
    
    print(f"   Total possible pairs: {total_possible}")
    print(f"   Skipped (different Application/System): {skipped_app}")
    print(f"   Skipped (different SubProcess): {skipped_subprocess}")
    print(f"   Skipped (different Nature - Preventive/Detective): {skipped_nature}")
    print(f"   Skipped (different Type - Manual/Automated): {skipped_type}")
    print(f"   Total skipped (obviously unique): {total_skipped}")
    print(f"   Candidate pairs for LLM review: {len(candidate_pairs)}")
    
    if not candidate_pairs:
        print(f"   No candidate pairs found - all entries are unique!")
        return []
    
    # Phase 2: LLM comparison in batches (PARALLEL)
    total_batches = (len(candidate_pairs) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"\n>> Phase 2: LLM comparison ({len(candidate_pairs)} pairs, {total_batches} batches, {MAX_WORKERS} parallel workers)...")
    
    # Prepare all batches
    batches = []
    for batch_num in range(total_batches):
        start = batch_num * BATCH_SIZE
        end = min(start + BATCH_SIZE, len(candidate_pairs))
        batch_pairs = candidate_pairs[start:end]
        batches.append((batch_num, batch_pairs))
    
    # Process batches in parallel
    all_results = []
    batch_results = {}
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Submit all batch jobs
        future_to_batch = {
            executor.submit(process_batch, batch_num, batch_pairs, df, available_cols): batch_num
            for batch_num, batch_pairs in batches
        }
        
        # Collect results as they complete
        for future in as_completed(future_to_batch):
            batch_num = future_to_batch[future]
            try:
                result_batch_num, result_batch_pairs, results = future.result()
                batch_results[result_batch_num] = (result_batch_pairs, results)
                
                if results:
                    duplicates_in_batch = sum(1 for r in results if r.get("is_duplicate", False))
                    print(f"   Batch {result_batch_num + 1}/{total_batches}: Done ({duplicates_in_batch} duplicates found)")
                else:
                    print(f"   Batch {result_batch_num + 1}/{total_batches}: FAILED")
            except Exception as e:
                print(f"   Batch {batch_num + 1}/{total_batches}: ERROR - {e}")
    
    # Reassemble results in order
    for batch_num in range(total_batches):
        if batch_num in batch_results:
            batch_pairs, results = batch_results[batch_num]
            if results:
                for idx, result in enumerate(results):
                    pair_i, pair_j = batch_pairs[idx]
                    result["row_a"] = pair_i
                    result["row_b"] = pair_j
                    result["row_a_id"] = get_id(df.iloc[pair_i])
                    result["row_b_id"] = get_id(df.iloc[pair_j])
                    result["row_a_risk"] = str(df.iloc[pair_i].get("Risk Title", "N/A"))
                    result["row_b_risk"] = str(df.iloc[pair_j].get("Risk Title", "N/A"))
                    result["row_a_app"] = str(df.iloc[pair_i].get("Application/System", "N/A"))
                    result["row_b_app"] = str(df.iloc[pair_j].get("Application/System", "N/A"))
                    result["process"] = process_name
                    all_results.append(result)
    
    # Filter to duplicates only
    duplicates = [r for r in all_results if r.get("is_duplicate", False)]
    
    print(f"\n>> Results for {process_name}:")
    print(f"   Pairs analyzed: {len(all_results)}")
    print(f"   Duplicates found: {len(duplicates)}")
    print(f"   Unique pairs: {len(all_results) - len(duplicates)}")
    
    return duplicates


def get_id(row):
    """Get identifier for a row"""
    ids = []
    for col in ["Risk Id", "Control Id"]:
        if col in row.index and pd.notna(row[col]):
            ids.append(str(row[col]))
    return " / ".join(ids) if ids else f"Row {row.name}"


def create_duplicates_excel(all_duplicates, process_dfs, compare_cols, output_path):
    """Create Excel with duplicate pairs stacked vertically, grouped by process"""
    
    rows = []
    pair_counter = 0
    
    for process_name, duplicates in all_duplicates.items():
        df = process_dfs[process_name]
        available_cols = [c for c in compare_cols if c in df.columns]
        
        for dup in duplicates:
            pair_counter += 1
            row_a = df.iloc[dup["row_a"]]
            row_b = df.iloc[dup["row_b"]]
            
            # Record A
            row_data_a = {
                "Pair #": pair_counter,
                "Process": process_name,
                "Confidence": dup.get("confidence", "N/A"),
                "Record": "A",
                "Row #": dup["row_a"],
            }
            for col in available_cols:
                if col != "Process":
                    row_data_a[col] = row_a[col] if pd.notna(row_a[col]) else ""
            row_data_a["LLM Reasoning"] = dup.get("reasoning", "")
            row_data_a["Risk Match"] = "Yes" if dup.get("risk_match", False) else "No"
            row_data_a["Control Match"] = "Yes" if dup.get("control_mechanism_match", False) else "No"
            row_data_a["LLM Recommendation"] = dup.get("recommendation", "")
            row_data_a["Keep This Record? (Yes/No)"] = ""
            rows.append(row_data_a)
            
            # Record B
            row_data_b = {
                "Pair #": pair_counter,
                "Process": process_name,
                "Confidence": dup.get("confidence", "N/A"),
                "Record": "B",
                "Row #": dup["row_b"],
            }
            for col in available_cols:
                if col != "Process":
                    row_data_b[col] = row_b[col] if pd.notna(row_b[col]) else ""
            row_data_b["LLM Reasoning"] = dup.get("why_controls_differ", "") if not dup.get("control_mechanism_match", True) else ""
            row_data_b["Risk Match"] = ""
            row_data_b["Control Match"] = ""
            row_data_b["LLM Recommendation"] = ""
            row_data_b["Keep This Record? (Yes/No)"] = ""
            rows.append(row_data_b)
    
    if not rows:
        print("No duplicates to write to Excel.")
        return None
    
    df_duplicates = pd.DataFrame(rows)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_duplicates.to_excel(writer, sheet_name="Duplicate Pairs", index=False)
        
        ws = writer.sheets["Duplicate Pairs"]
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Colors
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        record_a_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        record_b_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        decision_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        reasoning_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        high_conf = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        med_conf = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        low_conf = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
        pair_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        thick_top_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="thin"),
        )
        
        # Format header
        for col_idx in range(1, len(df_duplicates.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        
        # Format data rows
        total_cols = len(df_duplicates.columns)
        
        for row_idx in range(2, len(df_duplicates) + 2):
            data_row_idx = row_idx - 2
            record_type = df_duplicates.iloc[data_row_idx]["Record"]
            confidence = df_duplicates.iloc[data_row_idx]["Confidence"]
            is_first = (record_type == "A")
            
            for col_idx in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_name = df_duplicates.columns[col_idx - 1]
                
                cell.border = thick_top_border if is_first else thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                if col_name == "Confidence":
                    if confidence == "HIGH":
                        cell.fill = high_conf
                    elif confidence == "MEDIUM":
                        cell.fill = med_conf
                    else:
                        cell.fill = low_conf
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif col_name in ["LLM Reasoning", "LLM Recommendation", "Risk Match", "Control Match"]:
                    cell.fill = reasoning_fill
                elif col_name == "Keep This Record? (Yes/No)":
                    cell.fill = decision_fill
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.font = Font(bold=True)
                elif col_name in ["Pair #", "Record", "Row #"]:
                    cell.fill = record_a_fill if record_type == "A" else record_b_fill
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    if col_name in ["Pair #", "Record"]:
                        cell.font = pair_font
                else:
                    cell.fill = record_a_fill if record_type == "A" else record_b_fill
        
        # Column widths
        for col_idx in range(1, total_cols + 1):
            col_letter = get_column_letter(col_idx)
            col_name = df_duplicates.columns[col_idx - 1]
            
            if col_name in ["Risk Description", "Control Description", "Control Objective"]:
                ws.column_dimensions[col_letter].width = 45
            elif col_name in ["LLM Reasoning"]:
                ws.column_dimensions[col_letter].width = 50
            elif col_name in ["Risk Title", "SubProcess", "LLM Recommendation"]:
                ws.column_dimensions[col_letter].width = 30
            elif col_name == "Keep This Record? (Yes/No)":
                ws.column_dimensions[col_letter].width = 22
            elif col_name in ["Pair #", "Record", "Row #", "Confidence", "Risk Match", "Control Match"]:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 20
        
        # Row heights
        for row_idx in range(2, len(df_duplicates) + 2):
            ws.row_dimensions[row_idx].height = 80
        
        ws.freeze_panes = "F2"
    
    return df_duplicates


# =====================================================
# MAIN EXECUTION
# =====================================================

def main():
    """Main execution function for deduplication analysis"""
    global client
    client = create_llm_client()

    print("=" * 70)
    print("RCM DEDUPLICATION ENGINE - LLM VERSION")
    print("Powered by Azure OpenAI")
    print("=" * 70)

    print(f"\nConfiguration:")
    print(f"  Provider: Azure OpenAI (API Key: ...{OPENAI_API_KEY[-8:]})")
    print(f"  Deployment: {OPENAI_MODEL}")
    print(f"  Input: {RCM_INPUT}")
    print(f"  Input mode: {'Folder (per-process files)' if INPUT_IS_FOLDER else 'Single file with Process column'}")
    print(f"  Batch size: {BATCH_SIZE} pairs per API call")
    print(f"  Parallel workers: {MAX_WORKERS}")
    print(f"  Output folder: {OUTPUT_FOLDER}")
    print(f"  Output Excel: {OUTPUT_EXCEL_NAME}.xlsx")


    # Step 1: Load RCM data
    print("\n" + "=" * 70)
    print("STEP 1: LOADING RCM DATA")
    print("=" * 70)

    process_dfs = {}

    if INPUT_IS_FOLDER:
        # Load all Excel files from folder
        excel_files = glob.glob(os.path.join(RCM_INPUT, "*.xlsx"))
        print(f"Found {len(excel_files)} Excel files in folder")

        for filepath in excel_files:
            filename = os.path.basename(filepath)
            process_name = os.path.splitext(filename)[0]
            df = pd.read_excel(filepath)
            process_dfs[process_name] = df
            print(f"  Loaded: {filename} ({len(df)} rows)")
    else:
        # Single file - split by Process column
        df_all = pd.read_excel(RCM_INPUT)
        print(f"Loaded: {os.path.basename(RCM_INPUT)} ({len(df_all)} rows)")

        if "Process" in df_all.columns:
            for process_name, group_df in df_all.groupby("Process"):
                process_dfs[process_name] = group_df.reset_index(drop=True)
                print(f"  Process: {process_name} ({len(group_df)} rows)")
        else:
            process_dfs["All"] = df_all
            print(f"  No 'Process' column found - treating as single RCM")

    print(f"\nTotal processes: {len(process_dfs)}")
    print(f"Total records: {sum(len(df) for df in process_dfs.values())}")


    # Step 2: Process each RCM
    print("\n" + "=" * 70)
    print("STEP 2: ANALYZING EACH PROCESS RCM")
    print("=" * 70)

    all_duplicates = {}
    total_api_calls = 0
    total_duplicates = 0

    for process_name, df in process_dfs.items():
        if len(df) < 2:
            print(f"\nSkipping {process_name} (only {len(df)} record)")
            continue

        duplicates = process_single_rcm(df, process_name)
        all_duplicates[process_name] = duplicates
        total_duplicates += len(duplicates)

        # Estimate API calls made
        candidate_pairs, _, _, _, _ = pre_filter_pairs(df)
        batches = (len(candidate_pairs) + BATCH_SIZE - 1) // BATCH_SIZE
        total_api_calls += batches


    # Step 3: Display results
    print("\n" + "=" * 70)
    print("STEP 3: RESULTS")
    print("=" * 70)

    for process_name, duplicates in all_duplicates.items():
        if duplicates:
            print(f"\n{process_name}:")
            for dup in duplicates:
                conf = dup.get("confidence", "N/A")
                rec = dup.get("recommendation", "N/A")
                print(f"  [{conf}] Row {dup['row_a']} <-> Row {dup['row_b']}")
                print(f"    {dup['row_a_risk']} <-> {dup['row_b_risk']}")
                print(f"    App: {dup['row_a_app']} / {dup['row_b_app']}")
                print(f"    Reason: {dup.get('reasoning', 'N/A')}")
                print(f"    Recommendation: {rec}")
        else:
            print(f"\n{process_name}: No duplicates found")


    # Step 4: Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)

    print(f"\nProcesses analyzed: {len(process_dfs)}")
    print(f"Total records: {sum(len(df) for df in process_dfs.values())}")
    print(f"API calls made: ~{total_api_calls}")
    print(f"Total duplicates found: {total_duplicates}")

    for process_name, duplicates in all_duplicates.items():
        df = process_dfs[process_name]
        print(f"  {process_name}: {len(df)} records, {len(duplicates)} duplicates")


    # Step 5: Save results
    print("\n" + "=" * 70)
    print("SAVING RESULTS")
    print("=" * 70)

    if total_duplicates > 0:
        excel_output = f"{OUTPUT_FOLDER}/{OUTPUT_EXCEL_NAME}.xlsx"
        create_duplicates_excel(all_duplicates, process_dfs, COMPARE_COLUMNS, excel_output)
        print(f"Excel saved: {excel_output}")
        print(f"  - {total_duplicates} duplicate pairs")
        print(f"  - Record A (blue) / Record B (orange) stacked vertically")
        print(f"  - LLM reasoning (green) for each pair")
        print(f"  - Confidence: RED=High, YELLOW=Medium, GREEN=Low")
    else:
        print("No duplicates found - no Excel output created.")

    # Save JSON
    json_output_data = {
        "config": {
            "input": RCM_INPUT,
            "provider": "openai",
            "deployment": OPENAI_MODEL,
            "batch_size": BATCH_SIZE,
            "max_workers": MAX_WORKERS,
            "input_mode": "folder" if INPUT_IS_FOLDER else "single_file",
        },
        "summary": {
            "processes_analyzed": len(process_dfs),
            "total_records": sum(len(df) for df in process_dfs.values()),
            "total_duplicates": total_duplicates,
            "api_calls": total_api_calls,
        },
        "results_by_process": {}
    }

    for process_name, duplicates in all_duplicates.items():
        json_output_data["results_by_process"][process_name] = {
            "records": len(process_dfs[process_name]),
            "duplicates": len(duplicates),
            "pairs": duplicates
        }

    json_output = f"{OUTPUT_FOLDER}/{OUTPUT_JSON_NAME}.json"
    with open(json_output, "w") as f:
        json.dump(json_output_data, f, indent=2, default=str)
    print(f"JSON saved: {json_output}")


    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)
    if total_duplicates > 0:
        print(f"\nOpen the Excel to review: {OUTPUT_FOLDER}/{OUTPUT_EXCEL_NAME}.xlsx")
        print(f"\nExcel Layout:")
        print(f"  - Grouped by process")
        print(f"  - Record A (blue) on top, Record B (orange) below")
        print(f"  - LLM Reasoning column (green) explains WHY it's a duplicate")
        print(f"  - Confidence: RED=High, YELLOW=Medium, GREEN=Low")
        print(f"  - Fill 'Keep This Record? (Yes/No)' column")
        print(f"\nInstructions:")
        print(f"  1. Open {OUTPUT_EXCEL_NAME}.xlsx")
        print(f"  2. Review LLM reasoning for each pair")
        print(f"  3. Type 'Yes' next to the record to KEEP")
        print(f"  4. Type 'No' next to the record to REMOVE")
    else:
        print("\nNo duplicates found across any process RCM.")
        print("Your RCM is clean!")

if __name__ == "__main__":
    main()
