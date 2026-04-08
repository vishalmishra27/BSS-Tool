"""
OnGround Check — Final Version
===============================

LOGIC:
  Policy PDFs → What processes/controls SHOULD exist (process-level mandate)
  SOP PDFs    → How controls are actually performed (attribute-level detail)

  Both accept MULTIPLE files. All pages are pooled per category.

FLOW:
  1. Load RCM Excel (all rows)
  2. For each RCM control → LLM extracts 7 control attributes from the record
  3. Pool all Policy PDFs → LLM checks if each control's process is documented
  4. Pool all SOP PDFs   → LLM extracts actual control attributes per control
  5. Compare: RCM-extracted vs SOP-extracted (weighted match %)
  6. Output Excel with 2 sheets:
       Sheet 1 "OnGround Check":
         Original RCM → RCM Extracted → Policy Check → SOP Extracted
         → SOP Source → Comparison → Analyst Override → Final Status
       Sheet 2 "Legend & Instructions"
  7. Output JSON backup

USAGE:
  from onground_check import OnGroundCheck, Config

  Config.OPENAI_API_KEY   = "your-key"
  Config.OPENAI_MODEL     = "gpt-4o-mini"

  checker = OnGroundCheck(
      rcm_path     = "RCM.xlsx",
      policy_paths = ["Policy_Finance.pdf", "Policy_IT.pdf", "Policy_HR.pdf"],
      sop_paths    = ["SOP_JournalEntry.pdf", "SOP_AP.pdf", "SOP_Access.pdf"],
      out_excel    = "Results.xlsx",
  )
  checker.run()

pip install pandas pdfplumber openpyxl requests
"""

import re
import os
import json
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import sys as _sys
_di_dir = str(Path(__file__).resolve().parent.parent)
if _di_dir not in _sys.path:
    _sys.path.insert(0, _di_dir)
from Document_Intelligence import parse_document as _di_parse_document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openai import AzureOpenAI
from engines.config import (
    AZURE_OPENAI_ENDPOINT as _CENTRAL_ENDPOINT,
    AZURE_OPENAI_API_KEY as _CENTRAL_API_KEY,
    AZURE_OPENAI_API_VERSION as _CENTRAL_API_VERSION,
    AZURE_OPENAI_DEPLOYMENT as _CENTRAL_DEPLOYMENT,
    OPENAI_API_KEY as _CENTRAL_OPENAI_KEY,
    OPENAI_MODEL as _CENTRAL_OPENAI_MODEL,
)


# ============================================================================
# CONFIG
# ============================================================================

class Config:
    # ── File Paths ──────────────────────────────────────────────────
    RCM_EXCEL_PATH = r"/Users/rishi/Downloads/Sample_Data/output.xlsx"

    # Multiple policies: what processes/controls SHOULD exist
    POLICY_PDF_PATHS = [
        r"/Users/rishi/Downloads/Sample_Data/Policy_Procure_to_Pay.pdf"
    ]

    # Multiple SOPs: how controls are actually performed
    SOP_PDF_PATHS = [
        r"/Users/rishi/Downloads/Sample_Data/SOP_Procure_to_Pay.pdf"
    ]

    OUTPUT_EXCEL = r"/Users/rishi/Downloads/Results/Re2.xlsx"
    OUTPUT_JSON = r"/Users/rishi/Downloads/Results/Re2.json"

    # ── Azure OpenAI — from central engines/config.py ────────────────
    AZURE_OPENAI_ENDPOINT = _CENTRAL_ENDPOINT
    AZURE_OPENAI_API_KEY = _CENTRAL_API_KEY
    AZURE_OPENAI_API_VERSION = _CENTRAL_API_VERSION
    AZURE_OPENAI_DEPLOYMENT_NAME = _CENTRAL_DEPLOYMENT
    OPENAI_API_KEY = _CENTRAL_OPENAI_KEY
    OPENAI_MODEL = _CENTRAL_OPENAI_MODEL

    # ── Parallel Processing ─────────────────────────────────────────
    # Number of concurrent LLM calls (adjust based on API limits)
    MAX_WORKERS = 5

    # ── 7 Control Attributes ────────────────────────────────────────
    ATTRIBUTES = [
        'frequency', 'owner', 'threshold',
        'time_period', 'system', 'nature', 'control_type',
    ]

    # ── Parallel Processing ─────────────────────────────────────────
    # Number of concurrent API calls (adjust based on rate limits)
    MAX_WORKERS = 5

    LABELS = {
        'frequency': 'Frequency', 'owner': 'Owner',
        'threshold': 'Threshold', 'time_period': 'Time Period',
        'system': 'System', 'nature': 'Nature',
        'control_type': 'Control Type',
    }

    # ── RCM Column Names (as they appear in your Excel) ────────────
    RCM_COLUMNS = [
        'Control Id', 'Process', 'SubProcess', 'Control Objective',
        'Risk Id', 'Risk Title', 'Risk Description',
        'Control Description', 'Control Owner', 'Control Rating',
        'Nature of Control', 'Control Type', 'Control Frequency',
        'Application/System', 'risk_level',
    ]


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class RCMControl:
    raw: Dict[str, str] = field(default_factory=dict)
    control_id: str = ""
    process: str = ""
    subprocess: str = ""
    control_description: str = ""
    control_owner: str = ""
    nature_of_control: str = ""
    control_type: str = ""
    control_frequency: str = ""
    application_system: str = ""
    rcm_extracted: Dict[str, str] = field(default_factory=dict)


@dataclass
class PolicyCheck:
    control_id: str = ""
    is_documented: str = ""
    policy_reference: str = ""
    source_file: str = ""
    source_page: int = 0


@dataclass
class SOPExtraction:
    control_id: str = ""
    extracted: Dict[str, str] = field(default_factory=dict)
    source_file: str = ""
    source_page: int = 0


@dataclass
class ComparisonResult:
    control_id: str = ""
    matches: Dict[str, str] = field(default_factory=dict)
    match_pct: float = 0.0
    auto_status: str = "Fail"
    gaps: List[str] = field(default_factory=list)


# ============================================================================
# LLM CLIENT — with retry logic for timeouts
# ============================================================================

class LLMClient:
    MAX_RETRIES = 3
    RETRY_DELAY = 5  # seconds

    def __init__(self):
        self.enabled = bool(Config.OPENAI_API_KEY)

    def extract_json(self, prompt: str, system_msg: str = "") -> Dict:
        if not self.enabled:
            raise RuntimeError(
                "LLM not configured — set Config.OPENAI_API_KEY"
            )
        client = AzureOpenAI(
            api_key=Config.OPENAI_API_KEY,
            azure_endpoint=Config.AZURE_OPENAI_ENDPOINT,
            api_version=Config.AZURE_OPENAI_API_VERSION,
        )

        msgs = []
        if system_msg:
            msgs.append({"role": "system", "content": system_msg})
        msgs.append({"role": "user", "content": prompt})

        last_error = None
        for attempt in range(self.MAX_RETRIES):
            try:
                response = client.chat.completions.create(
                    model=Config.OPENAI_MODEL,
                    messages=msgs,
                    max_completion_tokens=800,
                )
                raw = response.choices[0].message.content or ""
                clean = re.sub(r'^```(?:json)?\n?|\n?```$', '', raw.strip())
                m = re.search(r'\{[^{}]*\}', clean, re.DOTALL)
                return json.loads(m.group()) if m else {}

            except Exception as e:
                last_error = e
                if attempt < self.MAX_RETRIES - 1:
                    import time
                    print(f"        [API Error, retrying in {self.RETRY_DELAY}s... "
                          f"attempt {attempt + 2}/{self.MAX_RETRIES}]")
                    time.sleep(self.RETRY_DELAY)
                continue

        # All retries failed
        raise last_error


# ============================================================================
# PDF LOADER — pools pages from multiple files
# ============================================================================

class PDFLoader:
    """Load and pool pages from multiple PDFs under one category."""

    @staticmethod
    def load(paths: List[str], label: str = "PDF") -> List[Dict]:
        pages = []
        for p in paths:
            if not os.path.exists(p):
                print(f"      ⚠ Not found: {p}")
                continue
            result = _di_parse_document(p)
            if result.full_text:
                # Split into per-page sections if available
                page_sections = [s for s in result.sections if s.page is not None]
                if page_sections:
                    page_map: Dict[int, List[str]] = {}
                    for s in page_sections:
                        page_map.setdefault(s.page, []).append(s.text)
                    for pn in sorted(page_map):
                        text = "\n".join(page_map[pn]).strip()
                        if text:
                            pages.append({'file': Path(p).name, 'page': pn, 'text': text})
                else:
                    pages.append({'file': Path(p).name, 'page': 1, 'text': result.full_text})
                page_count = result.metadata.page_count or 1
                print(f"      ✓ {Path(p).name} — {page_count} pages")
            else:
                print(f"      ⚠ No text extracted: {Path(p).name}")
        print(f"      Total: {len(pages)} {label} pages pooled "
              f"from {len(paths)} file(s)")
        return pages


# ============================================================================
# PAGE RANKER — shared relevance scoring
# ============================================================================

class PageRanker:
    """Score and rank pages by relevance to a control."""

    KEYWORDS = {
        'journal', 'entry', 'reconcil', 'approv', 'review', 'payment',
        'invoice', 'vendor', 'access', 'procure', 'segregat', 'authoriz',
        'inventory', 'disbursement', 'payroll', 'receipt', 'three-way',
        'match', 'terminated', 'user', 'posting',
    }

    @staticmethod
    def rank(pages: List[Dict], rcm: RCMControl, top_n: int = 3
             ) -> List[Dict]:
        proc_words = {
            w.lower() for w in rcm.process.split() if len(w) > 3
        }
        sub_words = {
            w.lower() for w in rcm.subprocess.split() if len(w) > 3
        }
        desc_words = {
            w.lower()
            for w in rcm.control_description.split()
            if len(w) > 3
        }
        desc_lo = rcm.control_description.lower()

        scored = []
        for pg in pages:
            tl = pg['text'].lower()
            s = sum(3 for w in proc_words if w in tl)
            s += sum(2 for w in sub_words if w in tl)
            s += sum(1 for w in desc_words if w in tl)
            s += sum(4 for k in PageRanker.KEYWORDS
                     if k in desc_lo and k in tl)
            if s > 0:
                scored.append((s, pg))

        scored.sort(key=lambda x: -x[0])
        return [p for _, p in scored[:top_n]]


# ============================================================================
# RCM EXTRACTOR
# ============================================================================

class RCMExtractor:
    SYS = (
        "You extract control attributes from RCM records. "
        "Return only valid JSON, nothing else."
    )

    ATTR_JSON = (
        '{"frequency":"","owner":"","threshold":"",'
        '"time_period":"","system":"","nature":"","control_type":""}'
    )

    def __init__(self, llm: LLMClient):
        self.llm = llm

    def load(self, path: str) -> List[RCMControl]:
        """Load RCM Excel and extract attributes in parallel."""
        df = pd.read_excel(path)

        # First pass: create RCMControl objects
        controls = []
        for _, row in df.iterrows():
            raw = {col: str(row.get(col, '')) for col in Config.RCM_COLUMNS}
            c = RCMControl(
                raw=raw,
                control_id=raw.get('Control Id', ''),
                process=raw.get('Process', ''),
                subprocess=raw.get('SubProcess', ''),
                control_description=raw.get('Control Description', ''),
                control_owner=raw.get('Control Owner', ''),
                nature_of_control=raw.get('Nature of Control', ''),
                control_type=raw.get('Control Type', ''),
                control_frequency=raw.get('Control Frequency', ''),
                application_system=raw.get('Application/System', ''),
            )
            controls.append(c)

        # Parallel extraction
        print(f"      Extracting {len(controls)} controls in parallel "
              f"(workers={Config.MAX_WORKERS})...")

        def extract_one(c):
            c.rcm_extracted = self._extract(c)
            return c

        with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
            futures = {executor.submit(
                extract_one, c): i for i, c in enumerate(controls)}
            done = 0
            for future in as_completed(futures):
                c = future.result()
                done += 1
                print(
                    f"        [{done}/{len(controls)}] {c.control_id}: {c.rcm_extracted}")

        return controls

    def _extract(self, c: RCMControl) -> Dict[str, str]:
        prompt = f"""Extract 7 control attributes from this RCM record.

FULL RCM RECORD:
  Control ID:          {c.control_id}
  Process:             {c.process}
  SubProcess:          {c.subprocess}
  Control Objective:   {c.raw.get('Control Objective','')}
  Risk ID:             {c.raw.get('Risk Id','')}
  Risk Title:          {c.raw.get('Risk Title','')}
  Risk Description:    {c.raw.get('Risk Description','')}
  Control Description: {c.control_description}
  Control Owner:       {c.control_owner}
  Control Rating:      {c.raw.get('Control Rating','')}
  Nature of Control:   {c.nature_of_control}
  Control Type:        {c.control_type}
  Control Frequency:   {c.control_frequency}
  Application/System:  {c.application_system}

Extract from the COMPLETE record (columns + description text):
  1. frequency    — How often? (Daily/Weekly/Monthly/Quarterly/Annually/Per Occurrence)
  2. owner        — Who performs/approves?
  3. threshold    — Dollar amount? (e.g. "$50,000")
  4. time_period  — Time requirement? (e.g. "5 business days")
  5. system       — System/application?
  6. nature       — Preventive or Detective
  7. control_type — Manual or Automated

Return ONLY this JSON:
{self.ATTR_JSON}"""

        try:
            return self.llm.extract_json(prompt, self.SYS)
        except Exception as e:
            print(f"      [LLM Error {c.control_id}] {e}")
            return {
                'frequency': c.control_frequency,
                'owner': c.control_owner,
                'threshold': '',
                'time_period': '',
                'system': c.application_system,
                'nature': c.nature_of_control,
                'control_type': c.control_type,
            }


# ============================================================================
# POLICY CHECKER — pools all policy PDFs, checks process documentation
# ============================================================================

class PolicyChecker:
    SYS = (
        "You are a compliance analyst. Check if a control's process "
        "is documented in the policy. Return only valid JSON."
    )

    def __init__(self, llm: LLMClient):
        self.llm = llm
        self.pages: List[Dict] = []

    def load(self, paths: List[str]):
        self.pages = PDFLoader.load(paths, "Policy")

    def check(self, rcm: RCMControl) -> PolicyCheck:
        result = PolicyCheck(control_id=rcm.control_id)
        ranked = PageRanker.rank(self.pages, rcm)
        if not ranked:
            result.is_documented = "No"
            return result

        result.source_file = ranked[0]['file']
        result.source_page = ranked[0]['page']
        text = "\n---\n".join(p['text'] for p in ranked[:2])

        prompt = f"""Check if this control's process is documented in the policy.

CONTROL FROM RCM:
  Control ID:   {rcm.control_id}
  Process:      {rcm.process} / {rcm.subprocess}
  Description:  {rcm.control_description}

POLICY TEXT (from {ranked[0]['file']}):
{text}

Answer:
  1. is_documented — Does the policy mention this process/control? (Yes / No / Partial)
  2. policy_reference — Which section/paragraph? (brief quote or section number)

Return ONLY JSON:
{{"is_documented":"","policy_reference":""}}"""

        try:
            data = self.llm.extract_json(prompt, self.SYS)
            result.is_documented = data.get('is_documented', 'No')
            result.policy_reference = data.get('policy_reference', '')
        except Exception as e:
            print(f"      [LLM Error {rcm.control_id}] {e}")
            result.is_documented = "Error"
        return result


# ============================================================================
# SOP EXTRACTOR — pools all SOP PDFs, extracts attributes per control
# ============================================================================

class SOPExtractor:
    SYS = (
        "You are a compliance analyst. Extract control attributes "
        "from SOP documents. Only extract what the SOP states. "
        "Return only valid JSON."
    )

    ATTR_JSON = (
        '{"frequency":"","owner":"","threshold":"",'
        '"time_period":"","system":"","nature":"","control_type":""}'
    )

    def __init__(self, llm: LLMClient):
        self.llm = llm
        self.pages: List[Dict] = []

    def load(self, paths: List[str]):
        self.pages = PDFLoader.load(paths, "SOP")

    def extract(self, rcm: RCMControl) -> SOPExtraction:
        result = SOPExtraction(control_id=rcm.control_id)
        ranked = PageRanker.rank(self.pages, rcm)
        if not ranked:
            return result

        result.source_file = ranked[0]['file']
        result.source_page = ranked[0]['page']
        text = "\n---\n".join(p['text'] for p in ranked[:2])

        prompt = f"""Extract control attributes from SOP text for ONE specific control.

THE CONTROL (from RCM):
  Control ID:   {rcm.control_id}
  Process:      {rcm.process} / {rcm.subprocess}
  Description:  {rcm.control_description}

SOP TEXT (from {ranked[0]['file']}):
{text}

Extract what the SOP says about THIS specific control:
  1. frequency    — How often per the SOP?
  2. owner        — Who performs/approves per the SOP?
  3. threshold    — Dollar threshold in the SOP?
  4. time_period  — Time requirement in the SOP?
  5. system       — System/application in the SOP?
  6. nature       — Preventive or Detective per the SOP?
  7. control_type — Manual or Automated per the SOP?

IMPORTANT: Only extract what relates to THIS control. Use "" if not found.

Return ONLY this JSON:
{self.ATTR_JSON}"""

        try:
            result.extracted = self.llm.extract_json(prompt, self.SYS)
        except Exception as e:
            print(f"      [LLM Error {rcm.control_id}] {e}")
        return result


# ============================================================================
# COMPARATOR — simple equal comparison (no weights)
# ============================================================================

class Comparator:
    """Compare RCM extracted vs SOP extracted — each attribute counts equally."""

    def compare(self, rcm_attrs: Dict, sop_attrs: Dict
                ) -> Tuple[Dict[str, str], float]:
        matches = {}
        total = 0
        matched = 0

        for attr in Config.ATTRIBUTES:
            rv = str(rcm_attrs.get(attr, '') or '').strip()
            sv = str(sop_attrs.get(attr, '') or '').strip()

            if not rv and not sv:
                # Neither source has this attribute
                matches[attr] = "Not Found in Both"
            elif not rv:
                # SOP has it but RCM extraction doesn't
                matches[attr] = "Not Found in RCM"
                total += 1
            elif not sv:
                # RCM has it but SOP doesn't
                matches[attr] = "Not Found in SOP"
                total += 1
            elif self._match(rv, sv):
                # Both have it and they match
                matches[attr] = "Match"
                total += 1
                matched += 1
            else:
                # Both have it but they don't match
                matches[attr] = "Mismatch"
                total += 1

        pct = round(matched / total * 100, 1) if total else 0
        return matches, pct

    @staticmethod
    def _match(a: str, b: str) -> bool:
        al, bl = a.lower().strip(), b.lower().strip()
        if al == bl or al in bl or bl in al:
            return True
        # Try numeric comparison for thresholds like "$50,000"
        na = re.search(r'[\d,]+', a.replace('$', ''))
        nb = re.search(r'[\d,]+', b.replace('$', ''))
        if na and nb:
            try:
                return (
                    int(na.group().replace(',', ''))
                    == int(nb.group().replace(',', ''))
                )
            except ValueError:
                pass
        return False


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

class ExcelOutput:
    """Two sheets: OnGround Check (data + review) and Legend."""

    # Fonts
    F = Font(name="Arial", size=10)
    HF = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    B = Border(*(Side('thin'),) * 4)

    # Header fills
    RCM_RAW_H = PatternFill("solid", fgColor="1F4E79")   # Dark blue
    RCM_EXT_H = PatternFill("solid", fgColor="2E75B6")   # Blue
    POL_CHK_H = PatternFill("solid", fgColor="7030A0")   # Purple
    SOP_EXT_H = PatternFill("solid", fgColor="548235")   # Green
    SRC_H = PatternFill("solid", fgColor="404040")   # Gray
    CMP_H = PatternFill("solid", fgColor="BF8F00")   # Gold
    REV_H = PatternFill("solid", fgColor="C00000")   # Red

    # Cell fills
    MATCH_C = PatternFill("solid", fgColor="C6EFCE")
    MISMATCH_C = PatternFill("solid", fgColor="FFC7CE")
    NOTFOUND_C = PatternFill("solid", fgColor="FFEB9C")
    INPUT_C = PatternFill("solid", fgColor="FFF2CC")

    def generate(self, rcm_list, policy_checks, sop_exts, comps,
                 policy_files, sop_files, path):
        wb = Workbook()
        self._main_sheet(wb, rcm_list, policy_checks, sop_exts, comps)
        wb.save(path)

    # ── Sheet 1: OnGround Check ─────────────────────────────────────

    def _main_sheet(self, wb, rcm_list, policy_checks, sop_exts, comps):
        ws = wb.active
        ws.title = "OnGround Check"

        # Header groups in order
        groups = [
            (Config.RCM_COLUMNS, self.RCM_RAW_H),
            (
                [f"RCM: {Config.LABELS[a]}" for a in Config.ATTRIBUTES],
                self.RCM_EXT_H,
            ),
            (
                ['Policy Documented?', 'Policy Reference',
                 'Policy Source File', 'Policy Page'],
                self.POL_CHK_H,
            ),
            (
                [f"SOP: {Config.LABELS[a]}" for a in Config.ATTRIBUTES],
                self.SOP_EXT_H,
            ),
            (['SOP Source File', 'SOP Page'], self.SRC_H),
            (
                [f"{Config.LABELS[a]} Match" for a in Config.ATTRIBUTES],
                self.CMP_H,
            ),
            (
                [f"Override: {Config.LABELS[a]}" for a in Config.ATTRIBUTES],
                self.REV_H,
            ),
            (['Analyst Remarks'], self.REV_H),
        ]

        # Write headers and track group start columns
        col = 1
        g_starts = []
        for hdrs, fill in groups:
            g_starts.append(col)
            for h in hdrs:
                c = ws.cell(1, col, h)
                c.fill, c.font, c.border = fill, self.HF, self.B
                c.alignment = Alignment(
                    horizontal='center', wrap_text=True, vertical='center'
                )
                col += 1

        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 45

        raw_s, ext_s, pol_s, sop_s, src_s, cmp_s, ovr_s, rem_s = g_starts

        # ── Data rows ───────────────────────────────────────────────

        wrap_cols = set()
        for i, cn in enumerate(Config.RCM_COLUMNS):
            if any(k in cn for k in ('Description', 'Objective')):
                wrap_cols.add(raw_s + i)

        for ri, rcm in enumerate(rcm_list, 2):
            pc = policy_checks.get(rcm.control_id, PolicyCheck())
            sop = sop_exts.get(rcm.control_id, SOPExtraction())
            cmp = comps.get(rcm.control_id, ComparisonResult())

            row = []

            # 1. Original RCM
            for cn in Config.RCM_COLUMNS:
                row.append(rcm.raw.get(cn, ''))

            # 2. RCM Extracted
            for a in Config.ATTRIBUTES:
                row.append(rcm.rcm_extracted.get(a, ''))

            # 3. Policy Check
            row += [
                pc.is_documented, pc.policy_reference,
                pc.source_file, pc.source_page or '',
            ]

            # 4. SOP Extracted
            for a in Config.ATTRIBUTES:
                row.append(sop.extracted.get(a, ''))

            # 5. SOP Source
            row += [sop.source_file, sop.source_page or '']

            # 6. Comparison (match results)
            for a in Config.ATTRIBUTES:
                row.append(cmp.matches.get(a, 'N/A'))

            # 7. Override columns (empty × 7 - analyst enters corrected values)
            row += [''] * len(Config.ATTRIBUTES)

            # 8. Analyst Remarks (empty × 1)
            row += ['']

            # Write
            for ci, v in enumerate(row, 1):
                cell = ws.cell(ri, ci, v)
                cell.font, cell.border = self.F, self.B
                cell.alignment = Alignment(
                    vertical='center', wrap_text=(ci in wrap_cols)
                )

                # Policy documented color
                if ci == pol_s:
                    if v == "Yes":
                        cell.fill = self.MATCH_C
                    elif v == "No":
                        cell.fill = self.MISMATCH_C
                    elif v == "Partial":
                        cell.fill = self.NOTFOUND_C

                # Comparison match colors
                if cmp_s <= ci < cmp_s + len(Config.ATTRIBUTES):
                    if v == "Match":
                        cell.fill = self.MATCH_C
                    elif v == "Mismatch":
                        cell.fill = self.MISMATCH_C
                    elif isinstance(v, str) and v.startswith("Not Found"):
                        cell.fill = self.NOTFOUND_C

                # Override + Remarks cells (analyst input)
                if ci >= ovr_s:
                    cell.fill = self.INPUT_C

        # ── Column widths ───────────────────────────────────────────

        for i, cn in enumerate(Config.RCM_COLUMNS):
            c = raw_s + i
            w = 42 if cn in wrap_cols or 'Description' in cn or 'Objective' in cn else 15
            ws.column_dimensions[get_column_letter(c)].width = w

        for c in range(ext_s, ext_s + len(Config.ATTRIBUTES)):
            ws.column_dimensions[get_column_letter(c)].width = 16
        for c in range(pol_s, pol_s + 4):
            ws.column_dimensions[get_column_letter(c)].width = 20
        for c in range(sop_s, sop_s + len(Config.ATTRIBUTES)):
            ws.column_dimensions[get_column_letter(c)].width = 16
        for c in range(src_s, src_s + 2):
            ws.column_dimensions[get_column_letter(c)].width = 18
        for c in range(cmp_s, cmp_s + len(Config.ATTRIBUTES)):
            ws.column_dimensions[get_column_letter(c)].width = 15
        for c in range(ovr_s, ovr_s + len(Config.ATTRIBUTES)):
            ws.column_dimensions[get_column_letter(c)].width = 18
        ws.column_dimensions[get_column_letter(rem_s)].width = 45

    # ── Sheet 2: Legend ──────────────────────────────────────────────


# ============================================================================
# MAIN ENGINE
# ============================================================================

class OnGroundCheck:
    def __init__(
        self,
        rcm_path=None,
        policy_paths=None,
        sop_paths=None,
        out_excel=None,
        out_json=None,
    ):
        self.rcm_path = rcm_path or Config.RCM_EXCEL_PATH
        self.policy_paths = policy_paths or Config.POLICY_PDF_PATHS
        self.sop_paths = sop_paths or Config.SOP_PDF_PATHS
        self.out_excel = out_excel or Config.OUTPUT_EXCEL
        self.out_json = out_json or Config.OUTPUT_JSON

        self.llm = LLMClient()
        self.rcm_ext = RCMExtractor(self.llm)
        self.pol_chk = PolicyChecker(self.llm)
        self.sop_ext = SOPExtractor(self.llm)
        self.comp = Comparator()
        self.xl = ExcelOutput()

        self.rcm_list:      List[RCMControl] = []
        self.policy_checks: Dict[str, PolicyCheck] = {}
        self.sop_exts:      Dict[str, SOPExtraction] = {}
        self.results:       Dict[str, ComparisonResult] = {}

    def run(self):
        n_pol = len(self.policy_paths)
        n_sop = len(self.sop_paths)

        print("\n" + "=" * 65)
        print("   OnGround Check — LLM Assessment (Final)")
        print("=" * 65)
        print(f"   RCM:      {self.rcm_path}")
        print(f"   Policies: {n_pol} file(s)")
        print(f"   SOPs:     {n_sop} file(s)")

        # ── 1. RCM ──────────────────────────────────────────────────
        print(f"\n[1/5] Loading RCM + LLM extraction...")
        self.rcm_list = self.rcm_ext.load(self.rcm_path)
        print(f"      ✓ {len(self.rcm_list)} controls extracted")

        # ── 2. POLICY CHECK ─────────────────────────────────────────
        print(f"\n[2/5] Loading {n_pol} Policy PDF(s) + checking "
              f"process documentation...")
        self.pol_chk.load(self.policy_paths)

        print(f"      Checking {len(self.rcm_list)} controls in parallel "
              f"(workers={Config.MAX_WORKERS})...")

        def check_policy(rcm):
            return rcm.control_id, self.pol_chk.check(rcm)

        with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
            futures = [executor.submit(check_policy, rcm)
                       for rcm in self.rcm_list]
            done = 0
            for future in as_completed(futures):
                cid, pc = future.result()
                self.policy_checks[cid] = pc
                done += 1
                ref = f"  ({pc.policy_reference[:50]})" if pc.policy_reference else ""
                print(f"        [{done}/{len(self.rcm_list)}] {cid}: "
                      f"{pc.is_documented} ← {pc.source_file}{ref}")

        # ── 3. SOP EXTRACTION ───────────────────────────────────────
        print(f"\n[3/5] Loading {n_sop} SOP PDF(s) + LLM extraction...")
        self.sop_ext.load(self.sop_paths)

        print(f"      Extracting {len(self.rcm_list)} controls in parallel "
              f"(workers={Config.MAX_WORKERS})...")

        def extract_sop(rcm):
            return rcm.control_id, self.sop_ext.extract(rcm)

        with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
            futures = [executor.submit(extract_sop, rcm)
                       for rcm in self.rcm_list]
            done = 0
            for future in as_completed(futures):
                cid, ext = future.result()
                self.sop_exts[cid] = ext
                done += 1
                tag = f"{ext.source_file} p{ext.source_page}" if ext.source_file else "no match"
                print(f"        [{done}/{len(self.rcm_list)}] {cid}: "
                      f"{ext.extracted}  ← {tag}")

        # ── 4. COMPARE ──────────────────────────────────────────────
        print(f"\n[4/5] Comparing RCM-extracted vs SOP-extracted...")
        for rcm in self.rcm_list:
            sop = self.sop_exts.get(
                rcm.control_id, SOPExtraction()
            )
            matches, pct = self.comp.compare(
                rcm.rcm_extracted, sop.extracted
            )

            gaps = []
            for attr, result in matches.items():
                lbl = Config.LABELS.get(attr, attr)
                if result == "Not Found in Both":
                    gaps.append(
                        f"{lbl}: not found in RCM extraction or SOP extraction")
                elif result == "Not Found in RCM":
                    gaps.append(f"{lbl}: not found in RCM extraction")
                elif result == "Not Found in SOP":
                    gaps.append(f"{lbl}: not found in SOP extraction")
                elif result == "Mismatch":
                    rv = rcm.rcm_extracted.get(attr, '')
                    sv = sop.extracted.get(attr, '')
                    gaps.append(f"{lbl}: RCM='{rv}' vs SOP='{sv}'")

            # Count matches/mismatches for console
            n_match = sum(1 for v in matches.values() if v == "Match")
            n_mis = sum(1 for v in matches.values() if v == "Mismatch")
            n_nf = sum(1 for v in matches.values()
                       if v.startswith("Not Found"))

            self.results[rcm.control_id] = ComparisonResult(
                control_id=rcm.control_id, matches=matches,
                match_pct=pct, auto_status="", gaps=gaps,
            )
            print(f"        {rcm.control_id}: {n_match} Match, "
                  f"{n_mis} Mismatch, {n_nf} Not Found")

        # ── 5. OUTPUT ───────────────────────────────────────────────
        print(f"\n[5/5] Generating output...")
        self.xl.generate(
            self.rcm_list, self.policy_checks, self.sop_exts,
            self.results, self.policy_paths, self.sop_paths,
            self.out_excel,
        )
        print(f"      ✓ Excel: {self.out_excel}")

        # JSON
        jdata = {
            "generated": datetime.now().isoformat(),
            "sources": {
                "rcm": self.rcm_path,
                "policies": [str(p) for p in self.policy_paths],
                "sops": [str(p) for p in self.sop_paths],
            },
            "controls": [],
        }
        for rcm in self.rcm_list:
            pc = self.policy_checks.get(rcm.control_id, PolicyCheck())
            sop = self.sop_exts.get(rcm.control_id, SOPExtraction())
            c = self.results.get(rcm.control_id, ComparisonResult())
            jdata["controls"].append({
                "control_id": rcm.control_id,
                "rcm_raw": rcm.raw,
                "rcm_extracted": rcm.rcm_extracted,
                "policy_check": {
                    "documented": pc.is_documented,
                    "reference": pc.policy_reference,
                    "file": pc.source_file,
                    "page": pc.source_page,
                },
                "sop_extracted": sop.extracted,
                "sop_source": {
                    "file": sop.source_file,
                    "page": sop.source_page,
                },
                "comparison": c.matches,
                "match_pct": c.match_pct,
                "auto_status": c.auto_status,
                "gaps": c.gaps,
            })
        with open(self.out_json, 'w') as f:
            json.dump(jdata, f, indent=2)
        print(f"      ✓ JSON: {self.out_json}")

        # ── SUMMARY ─────────────────────────────────────────────────
        t = len(self.rcm_list)
        pol = sum(1 for c in self.policy_checks.values()
                  if c.is_documented == "Yes")

        print(f"\n{'─' * 60}")
        print(f"  Controls processed:  {t}")
        print(f"  Policy Documented:   {pol}/{t}")
        print(f"{'─' * 60}")
        print(f"\n{'=' * 65}")
        print("  ✅ Done — open Excel → review → add overrides if needed")
        print(f"{'=' * 65}\n")


# ============================================================================
if __name__ == "__main__":
    OnGroundCheck().run()
