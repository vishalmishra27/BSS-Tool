"""
RCM Control Testing -- Test of Operating Effectiveness (TOE)
=============================================================
Multiple samples per control. Evaluates whether controls actually
OPERATED correctly across transactions during the audit period.

Usage:
    tester = RCMControlTester(...)
    toe_bank = load_toe_evidence_folder("evidence_toe/")
    results = tester.test_all_toe(toe_bank)
    tester.export_toe_workpaper(results, "toe_workpaper.xlsx")
"""

from __future__ import annotations

# Load centralised config (pushes API keys into env vars before anything else)
try:
    import config as _config  # noqa: F401
except ImportError:
    pass  # config.py is optional — falls back to env vars

import json
import time
import os
import re
import math
import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from typing import Optional
import requests


# ===========================================================================
#  DATA MODELS
# ===========================================================================

@dataclass
class RCMRow:
    process: str
    subprocess: str
    risk_id: str
    risk_description: str
    risk_level: str
    control_id: str
    control_description: str
    control_objective: str = ""
    control_owner: str = ""
    nature_of_control: str = ""
    control_type: str = ""
    control_frequency: str = ""
    application_system: str = ""
    risk_title: str = ""
    count_of_samples: str = ""


@dataclass
class SupportingDocument:
    """A supporting file referenced by the sample evidence (SOP, policy, config, etc.)."""
    filename: str
    content: str
    doc_type: str = ""   # e.g. "Policy", "Config Export", "Screenshot Description"


@dataclass
class SampleEvidence:
    """One sample for a control -- walkthrough evidence + supporting documents."""
    sample_id: str
    description: str
    source_document: str = ""
    supporting_docs: list = None   # list[SupportingDocument]

    def __post_init__(self):
        if self.supporting_docs is None:
            self.supporting_docs = []

# ===========================================================================


# ===========================================================================
#  TOE DATA MODELS
# ===========================================================================

@dataclass
class TOESampleResult:
    """Result of testing one sample for operating effectiveness."""
    control_id: str
    sample_id: str
    result: str                    # PASS / FAIL
    operated_effectively: str      # Yes / No
    control_performed: str         # Yes / No
    timely_execution: str          # Yes / No / N/A
    accurate_execution: str        # Yes / No
    authorized_performer: str      # Yes / No / N/A
    evidence_sufficient: str       # Yes / No
    remarks: str
    deviation_details: str         # specific deviation, or "None"
    raw_evidence: str = ""
    sample_details: dict = None    # {"PR Number": "PR-2024...", "Amount": "$18K", ...}
    attribute_results: dict = None # {"A": "Yes", "B": "No", ...}
    attribute_reasoning: dict = None # {"A": "PO approved by FM...", "B": "Missing second approval..."}
    validator_corrected: bool = False  # True if validator changed the result
    validator_details: str = ""       # What the validator changed (for audit trail)


@dataclass
class ControlSchema:
    """LLM-generated testing schema for a control -- defines what to extract and test."""
    control_id: str
    worksteps: list                # ["1. Observed...", "2. Verified...", ...]
    attributes: list               # [{"id":"A","name":"...","description":"..."}, ...]
    sample_columns: list           # [{"key":"pr_number","header":"PR Number"}, ...]


@dataclass
class TOEControlResult:
    """Aggregated TOE result for one control across all samples."""
    control_id: str
    risk_id: str
    control_type: str
    nature_of_control: str
    control_frequency: str

    # Aggregated metrics
    total_samples: int
    passed_samples: int
    failed_samples: int
    deviation_rate: float          # 0.0 to 1.0

    # Overall conclusion
    operating_effectiveness: str   # "Effective" / "Effective with Exceptions" / "Not Effective"
    deficiency_type: str           # None / Control Deficiency / Significant Deficiency / Material Weakness
    overall_remarks: str

    # Individual results
    sample_results: list           # list[TOESampleResult]

    # Traceability
    evaluation_timestamp: str = ""
    schema: ControlSchema = None   # LLM-generated testing schema




# ===========================================================================
#  DEFICIENCY CLASSIFICATION THRESHOLDS (configurable per engagement)
# ===========================================================================
# These deviation-rate thresholds determine deficiency severity.
# AS 2201 does not prescribe fixed percentages — these are professional
# judgment defaults aligned with common Big 4 practice.  Adjust per
# engagement risk appetite or firm methodology.

CONTROL_DEFICIENCY_MAX_RATE = 0.10     # deviation rate <= 10%  → Control Deficiency
SIGNIFICANT_DEFICIENCY_MAX_RATE = 0.20 # deviation rate <= 20%  → Significant Deficiency
                                        # deviation rate > 20%   → Material Weakness


# ===========================================================================
#  TOE SYSTEM PROMPTS
# ===========================================================================

TOE_SYSTEM_PROMPT_MANUAL = """You are a senior compliance auditor performing a Test of Operating
Effectiveness (TOE) for a MANUAL control.

TOE is DIFFERENT from Test of Design (TOD):
  - TOD asks: "Is the control DESIGNED adequately?"
  - TOE asks: "Did the control actually OPERATE CORRECTLY for this specific transaction?"

You are given ONE transaction sample. Evaluate whether the control was executed 
correctly for THIS transaction -- not whether the control is well-designed in general.

## Evaluate These Five Questions for THIS Transaction:

### 1. Was the control PERFORMED?
- Is there evidence that the control activity actually happened?
- Was the required action (approval, review, verification, reconciliation) completed?
- If the control was NOT performed at all -> FAIL, deviation = "Control not performed"

### 2. Was it performed by the AUTHORIZED person?
- Did the right level of authority execute/approve?
- Was segregation of duties maintained?
- If wrong person performed it -> FAIL

### 3. Was it performed TIMELY?
- Was the control executed within the expected timeframe?
- For per-occurrence controls: was it done before the next step proceeded?
- For periodic controls: was the review done within the expected period?
- Late execution = potential deviation depending on significance

### 4. Was it performed ACCURATELY?
- Did the control achieve the correct outcome for this transaction?
- Were all required steps completed (not just some)?
- Were the right amounts, quantities, thresholds applied?
- Partial completion = deviation

### 5. Is there SUFFICIENT EVIDENCE of performance?
- Can you see documented proof that the control operated?
- Timestamps, signatures, system logs, approval records?
- If no evidence exists -> FAIL (cannot confirm control operated)

## CRITICAL: Evaluating the TRANSACTION, Not the Design
You already know the control is designed (TOD was done separately). Your job is 
to check if the control WORKED for this specific transaction.

GOOD TOE evaluation: "PR-20240315 was approved by both required approvers with 
timestamps in SAP. Control operated as intended for this transaction."

BAD TOE evaluation: "The approval process could be improved by adding formal 
sign-off sheets." <- That's a DESIGN observation, not an operating effectiveness finding.

## Operating Effectiveness Assessment
- **Yes**: The control operated correctly for this transaction. All five questions 
  are satisfactory. No deviation.
- **No**: The control did NOT operate correctly. At least one deviation found.

## Deviation Classification
A deviation means the control did NOT work as intended for this specific transaction:
- Control was not performed at all
- Wrong person performed it (SoD violation, insufficient authority)
- Performed too late (after the risk event could have materialized)
- Performed inaccurately (wrong amount, incomplete steps, partial execution)
- No evidence of performance (cannot confirm it happened)

## MANDATORY CONSISTENCY RULES
1. result = "PASS" -> operated_effectively = "Yes", deviation_details = "None"
2. result = "FAIL" -> operated_effectively = "No", deviation_details must describe the specific deviation
3. If control_performed = "No" -> result MUST be "FAIL"
4. If evidence_sufficient = "No" -> result MUST be "FAIL"
5. If timely_execution = "No" -> result MUST be "FAIL"
6. If accurate_execution = "No" -> result MUST be "FAIL"
7. If authorized_performer = "No" -> result MUST be "FAIL"
8. If operated_effectively = "No" -> result MUST be "FAIL"

## ATTRIBUTE CONSISTENCY (CRITICAL)
9.  Each attribute MUST be answered with EXACTLY one of: "Yes", "No", or "N/A" — no other values allowed
10. Do NOT write qualifiers like "Partially", "Yes with observations", "Mostly" — these are INVALID
    - If there is ANY doubt or partial compliance → attribute = "No"
    - "N/A" means this attribute genuinely does NOT APPLY to this transaction type
    - "N/A" does NOT mean "I cannot find evidence" — that is "No" (insufficient evidence)
11. If ANY attribute_results value = "No" → result MUST be "FAIL"
12. If result = "PASS" → ALL attribute_results MUST be "Yes" or "N/A" (never "No")
13. If result = "FAIL" → remarks MUST describe what went wrong (NEVER say "control operated effectively" for a FAIL sample)
14. For EACH attribute, provide your reasoning in attribute_reasoning explaining WHY you assessed it Yes/No/N/A
    with specific evidence references (document numbers, dates, amounts, approvers)
15. If ALL attribute_results values are "Yes" or "N/A" (none are "No") → result MUST be "PASS"
    DO NOT mark result = "FAIL" when every attribute is satisfied or not applicable.
    If the control was performed and no issues were found → that is a PASS.

## YOUR ROLE AS THE DECISION MAKER
You are the sole evaluator. Your result, attribute assessments, and reasoning ARE the audit judgment.
Consider the FULL CONTEXT — control description, risk, all attributes, and evidence holistically.
Your attribute assessments and overall result must tell a consistent story.

## CRITICAL EVALUATION RULES — READ CAREFULLY (COMMON MISTAKES)

### Rule A: Summary Language Satisfies Individual Attributes
If the evidence contains summary statements like "all items verified", "all 6 items completed",
"full screening performed", "checklist completed — all steps done", or "signed off after full review",
this SATISFIES all individual attributes that are part of that process.
DO NOT mark an attribute as "No" just because the evidence does not use the exact attribute name.
Instead, ask: "Is this attribute covered by the summary statement?"
EXAMPLE:
  - Evidence: "Checklist VEN-DDC-0495 completed: all 6 items verified. Signed by Lead."
  - Attribute "Compliance Certification" → "Yes" (covered by "all 6 items verified" + sign-off)
  - Evidence: "Full screening: D&B score 84, compliance clear, site visit done, ISO verified."
  - Attribute "Financial Review Completed" → "Yes" (D&B score IS the financial health review)
  - Attribute "References Checked" → "Yes" (covered by "Full screening" + "All documented")
  - IMPORTANT: A D&B score IS a financial health review. "Compliance clear" IS regulatory
    compliance validation. "Site visit done" IS a reference/verification check. "ISO verified"
    IS a compliance check. Do NOT require a separate document called "financial health review
    report" when the evidence shows the D&B financial score was obtained.
  - HOWEVER: If the D&B score is BELOW the policy threshold (e.g., evidence says "D&B score 32,
    policy requires >60"), then the financial review FAILED — mark the financial attribute as "No"
    with reason "D&B score X is below the required threshold of Y." A low score means the review
    was done but the vendor FAILED it — that is a control failure, not a pass.

### Rule B: Conditional Attributes — MANDATORY DECISION TREE
Some attributes test what should happen IF a condition arises (e.g., "discrepancies logged",
"variances documented", "exceptions escalated", "discrepancies addressed"). 

** BEFORE marking any conditional attribute, you MUST follow this decision tree: **

  STEP 1: Does the evidence show that discrepancies/variances/exceptions ACTUALLY EXISTED?
    - "shortfall of 15 units" → YES, discrepancies existed → go to Step 2
    - "qty confirmed, quality passed" → NO discrepancies existed → MARK N/A (done)
    - "delivery note matched PO, qty 150 confirmed" → NO variances existed → MARK N/A (done)
    - "all items matched" → NO discrepancies existed → MARK N/A (done)
    - "6 banking changes dual approved" with no mention of issues → NO discrepancies → MARK N/A (done)

  STEP 2 (only if discrepancies/variances EXISTED): Were they properly logged/documented?
    - "discrepancy report created" → YES → MARK "Yes"
    - "no discrepancy report despite shortfall" → NO → MARK "No" (TRUE FAIL)

  CRITICAL: "No" means "the condition existed AND was not handled."
            "N/A" means "the condition did not exist."
            These are COMPLETELY DIFFERENT. Mixing them up is the #1 error.

  WRONG: Marking "Discrepancy Logged = No" with reasoning "no discrepancies were logged 
         because all items matched" → This is N/A, not No!
  RIGHT: Marking "Discrepancy Logged = N/A" with reasoning "no discrepancies found, all 
         items matched PO and packing slip"

### Rule E: If ALL Attributes Are Yes or N/A → Result MUST Be PASS
This is a HARD LOGICAL RULE with ZERO EXCEPTIONS:
  - If EVERY attribute you assessed is "Yes" or "N/A" → result MUST be "PASS"
  - You are FORBIDDEN from marking result = "FAIL" when no attribute is "No"
  - You MUST have at least one attribute = "No" to justify a FAIL result

Think about it: if you said every testable aspect is satisfied ("Yes") or doesn't apply ("N/A"),
there is NOTHING wrong with this sample. A FAIL with all-Yes/N/A is a logical contradiction.

WRONG: attrs = {Inspection: Yes, Discrepancy Logged: N/A, Resolution: N/A} → result = "FAIL"
       (Why? There's nothing wrong! Every attribute is fine.)
RIGHT: attrs = {Inspection: Yes, Discrepancy Logged: N/A, Resolution: N/A} → result = "PASS"
       (The inspection was done, no discrepancies existed, control worked correctly.)

### Rule C: Preventive Controls That Block = PASS
For PREVENTIVE controls, if the evidence shows the control BLOCKED or REJECTED a bad transaction,
that means the control WORKED. The block IS the success.
EXAMPLE:
  - Control: "System blocks requisitions exceeding budget"
  - Evidence: "SAP blocked: 'Budget exceeded.' Revised to $35K — accepted."
  - Result → PASS. The control did exactly what it should: it blocked the overspend.
  - Do NOT mark FAIL because the original amount was rejected — that's the control working.
ANOTHER EXAMPLE:
  - Control: "System prevents PO creation without approved requisition"
  - Evidence: "SAP hard error: 'PR reference mandatory.' PO blocked."
  - Result → PASS. The system prevented the unauthorized PO.

### Rule D: Test Only What the Control Says — No Scope Creep
Evaluate ONLY the control as described. Do NOT invent additional requirements beyond the control text.
EXAMPLE:
  - Control: "banking changes require independent verification via call-back procedure"
  - Evidence: "4 banking changes — all dual approved with timestamps"
  - If evidence shows verification was done (dual approval with timestamps), the control operated.
  - Do NOT fail because the evidence doesn't use the exact words "call-back procedure."
  - The substance of the control (independent verification) was met.
ANOTHER EXAMPLE:
  - Control: "AP Manager reviews weekly aging report to identify and prioritize past-due invoices"
  - Evidence: "AP Manager reviewed 9:30 AM. 6 prioritized, 2 disputed."
  - The control is about REVIEW and PRIORITIZATION — not about paying every invoice.
  - Result → PASS if the review was performed and items were prioritized.

## MANDATORY SELF-CHECK (Do this before writing your final JSON)
After you determine each attribute value:
  1. Count how many attributes = "No". If zero → result MUST be "PASS".
  2. For each "N/A": verify the condition truly did not arise (e.g., zero discrepancies existed).
  3. For each "No": verify the issue ACTUALLY existed but went unhandled. 
     If the issue didn't exist → change to "N/A".
  4. Confirm result is consistent: all Yes/N/A → PASS; any No → FAIL.

## Rules
- Evaluate THIS TRANSACTION only -- not the control's overall design
- Reference specific evidence: transaction numbers, dates, amounts, approvers
- Be factual and precise about what you observed
- If the control worked correctly for this transaction -> PASS, no deviation
- If there is ANY deviation -> FAIL with specific details
- Respond ONLY with valid JSON"""


TOE_SYSTEM_PROMPT_AUTOMATED = """You are a senior compliance auditor performing a Test of Operating
Effectiveness (TOE) for an AUTOMATED control.

TOE is DIFFERENT from Test of Design (TOD):
  - TOD asks: "Is the system rule CONFIGURED correctly?"
  - TOE asks: "Did the system rule actually WORK correctly for this specific transaction?"

You are given ONE transaction sample. Evaluate whether the automated control 
operated correctly for THIS transaction.

## Evaluate These Five Questions for THIS Transaction:

### 1. Was the automated control PERFORMED (triggered)?
- Did the system rule fire when it should have?
- Was the validation/check/block/flag executed?
- If the rule should have triggered but didn't -> FAIL

### 2. Was the outcome CORRECT?
- Did the system produce the right result for this transaction?
- Block when it should block? Allow when it should allow? Flag correctly?
- Wrong outcome = deviation

### 3. Was it TIMELY?
- Did the control operate at the right point in the process?
- Was there any delay that allowed the risk to materialize?

### 4. Were any OVERRIDES properly handled?
- If an override or exception occurred, was it properly authorized?
- Was the override reason documented and justified?
- Unauthorized override = deviation

### 5. Is there SUFFICIENT EVIDENCE?
- System logs, error messages, approval records, screenshots?
- Can the auditor independently verify the control operated?

## CRITICAL: Evaluating the TRANSACTION, Not the Configuration
You already know the system is configured correctly (TOD was done). Your job is 
to verify it WORKED for this specific transaction.

GOOD: "SAP correctly blocked PO creation without PR reference. Hard error logged."
BAD: "The system could benefit from additional fuzzy matching logic." <- That's design.

## For Automated Controls: ITGC Reliance
If the automated control operated correctly for the sample tested AND you have 
no evidence of system changes during the period, the control can generally be 
considered effective. Automated controls are consistent -- if they work once, they 
work every time (assuming ITGC are effective).

## MANDATORY CONSISTENCY RULES
1. result = "PASS" -> operated_effectively = "Yes", deviation_details = "None"
2. result = "FAIL" -> operated_effectively = "No", deviation_details must describe the deviation
3. If control_performed = "No" -> result MUST be "FAIL"
4. If evidence_sufficient = "No" -> result MUST be "FAIL"
5. If timely_execution = "No" -> result MUST be "FAIL"
6. If accurate_execution = "No" -> result MUST be "FAIL"
7. If authorized_performer = "No" -> result MUST be "FAIL"
8. If operated_effectively = "No" -> result MUST be "FAIL"

## ATTRIBUTE CONSISTENCY (CRITICAL)
9.  Each attribute MUST be answered with EXACTLY one of: "Yes", "No", or "N/A" — no other values allowed
10. Do NOT write qualifiers like "Partially", "Yes with observations", "Mostly" — these are INVALID
    - If there is ANY doubt or partial compliance → attribute = "No"
    - "N/A" means this attribute genuinely does NOT APPLY to this transaction type
    - "N/A" does NOT mean "I cannot find evidence" — that is "No" (insufficient evidence)
11. If ANY attribute_results value = "No" → result MUST be "FAIL"
12. If result = "PASS" → ALL attribute_results MUST be "Yes" or "N/A" (never "No")
13. If result = "FAIL" → remarks MUST describe what went wrong (NEVER say "control operated effectively" for a FAIL sample)
14. For EACH attribute, provide your reasoning in attribute_reasoning explaining WHY you assessed it Yes/No/N/A
    with specific evidence references (system logs, transaction IDs, error messages)
15. If ALL attribute_results values are "Yes" or "N/A" (none are "No") → result MUST be "PASS"
    DO NOT mark result = "FAIL" when every attribute is satisfied or not applicable.

## YOUR ROLE AS THE DECISION MAKER
You are the sole evaluator. Your result, attribute assessments, and reasoning ARE the audit judgment.
Consider the FULL CONTEXT — control description, risk, all attributes, and evidence holistically.
Your attribute assessments and overall result must tell a consistent story.

## CRITICAL EVALUATION RULES — READ CAREFULLY (COMMON MISTAKES)

### Rule A: Summary Language Satisfies Individual Attributes
If the evidence contains summary statements like "all checks passed", "all validations completed",
"system performed all required checks", this SATISFIES individual attributes that are part of that process.
DO NOT mark an attribute as "No" just because the evidence does not use the exact attribute name.
Instead, ask: "Is this attribute covered by the summary statement?"

### Rule B: Conditional Attributes — MANDATORY DECISION TREE
Some attributes test what happens IF a condition arises (e.g., "exception flagged",
"override documented", "tolerance breach escalated").

** BEFORE marking any conditional attribute, follow this decision tree: **
  STEP 1: Did the exception/condition ACTUALLY OCCUR in this transaction?
    - Evidence says "matched within tolerance" → NO exception → MARK "N/A" (done)
    - Evidence says "all validations passed" → NO exception → MARK "N/A" (done)
    - Evidence says "tolerance exceeded by 5%" → YES exception occurred → go to Step 2
  STEP 2 (only if condition existed): Was it properly handled?
    - "Exception flagged for review" → MARK "Yes"
    - "No flag despite breach" → MARK "No" (TRUE FAIL)

  "No" = the condition EXISTED but was NOT handled.
  "N/A" = the condition did NOT exist.

### Rule E: If ALL Attributes Are Yes or N/A → Result MUST Be PASS
This is a HARD LOGICAL RULE with ZERO EXCEPTIONS:
  - If EVERY attribute is "Yes" or "N/A" → result MUST be "PASS"
  - You are FORBIDDEN from marking result = "FAIL" when no attribute is "No"
  - You MUST have at least one attribute = "No" to justify a FAIL result
  - N/A is NOT a failure condition. It means the attribute does not apply.

### Rule C: Preventive Controls That Block = PASS
For PREVENTIVE automated controls, if the system BLOCKED or REJECTED a bad transaction,
that means the control WORKED. The block IS the success.
EXAMPLE:
  - Control: "System blocks PO creation without approved requisition"
  - Evidence: "SAP hard error: 'PR reference mandatory.' PO blocked."
  - Result → PASS. The system prevented the unauthorized PO — control operated correctly.
  - Do NOT mark FAIL because the transaction was rejected — that's the control working.
ANOTHER EXAMPLE:
  - Control: "System duplicate detection flags matching POs"
  - Evidence: "PO-89101 flagged as potential duplicate. Required justification."
  - Result → PASS. The system detected the duplicate — that's exactly what it should do.

### Rule D: Test Only What the Control Says — No Scope Creep
Evaluate ONLY the control as described. Do NOT invent additional requirements beyond the control text.
EXAMPLE:
  - Control: "Automated payment scheduling system processes approved invoices based on payment terms"
  - Evidence: "Aging report generated. AP Manager reviewed. 6 prioritized, 2 disputed. Processed next day."
  - The control is about the scheduling system generating the report and processing payments.
  - If the system ran correctly and the review happened → PASS.

## MANDATORY SELF-CHECK (Do this before writing your final JSON)
After you determine each attribute value:
  1. Count how many attributes = "No". If zero → result MUST be "PASS".
  2. For each "N/A": verify the condition truly did not arise.
  3. For each "No": verify the issue ACTUALLY existed but went unhandled. 
     If the issue didn't exist → change to "N/A".

## Rules
- Evaluate THIS TRANSACTION only -- not configuration quality
- Reference specific evidence: transaction numbers, system logs, error messages
- For automated controls, one successful operation with ITGC reliance is typically sufficient
- Respond ONLY with valid JSON"""


# ===========================================================================
#  SCHEMA GENERATION PROMPT
# ===========================================================================

SCHEMA_GENERATION_PROMPT = """You are a senior Big 4 auditor designing a workpaper template for Test of Operating Effectiveness (TOE).

Given a control, you must define:

## 1. Worksteps Performed (3-5 steps)
These are the AUDIT PROCEDURES you performed to test whether the control operated effectively. Write them as past-tense numbered steps that describe what the auditor physically did.

Good examples:
- "1. Obtained the purchase requisition from SAP and verified it exists in the system."
- "2. Verified that Department Head approval was recorded in the SAP workflow with user ID and timestamp."
- "3. Confirmed that the Finance Manager approval preceded the PO creation date."

Bad examples (too vague):
- "1. Checked the control." <- no specifics
- "2. Reviewed documentation." <- what documentation?

## 2. Attributes (A, B, C...) -- as many as the control genuinely requires (typically 3-5)
Each attribute is a SPECIFIC, TESTABLE condition that maps directly to a requirement in the control description. For each sample, the auditor will mark Yes, No, or N/A.

*** THINK LIKE AN EXPERIENCED HUMAN AUDITOR — THIS IS THE MOST IMPORTANT SECTION ***
Generate ONE attribute for EACH distinct testable requirement in the control description. The number
of attributes should match the complexity of the control — simple controls may need 2-3, complex
controls may need 4-6. Do NOT artificially cap or pad the count.

**MANDATORY COVERAGE CHECK — DO THIS FIRST:**
Before writing any attributes, list every distinct requirement stated in the control description.
Each requirement MUST have at least one attribute covering it. If you have requirements without
attributes, add attributes. If a requirement is genuinely part of another requirement, they can
share one attribute — but only if no audit coverage is lost.

**Scope rules — CRITICAL:**
- ONLY test what the control description explicitly states. If the control description does not
  mention a requirement, do NOT create an attribute for it.
- Do NOT invent evidence sources, document names, system names, or report names that are not
  mentioned in the control description or RCM details. Only reference documents/systems that
  the control description explicitly names.
- Do NOT add attributes for general audit best practices that the control does not require.

**Consolidation rules — apply carefully, NEVER at the cost of coverage:**
- CONSOLIDATE APPROVAL CHAINS: If a control requires multiple approvers (Legal + Manager + CFO),
  do NOT create one attribute per approver. Create ONE attribute like "Approval per DOA" that checks
  all required approvals per the Delegation of Authority matrix were obtained.
  BAD:  "Legal Approval" + "Manager Approval" + "CFO Approval" (3 separate attrs for one approval chain)
  GOOD: "Approval per DOA" — All required approvals obtained per DOA matrix (1 consolidated attr)
- TEST OUTCOMES, NOT PROCESS STEPS: Focus on what the control achieves, not each sub-step.
  BAD:  "CFO Approval Evidenced" + "Within DOA Limit" + "Approval Before Payment" + "Amount Matches" (4 process steps)
  GOOD: "Payment Approval" + "Payment After Approval" (2 outcome checks that cover everything)
- INCLUDE LINKAGE FOR FINANCIAL CONTROLS: For provisions, journal entries, reconciliations, or
  expense recordings, include a "Linkage" attribute that tests traceability back to source documents.
  EXAMPLE: "Linkage Verified" — Transaction traces back to supporting source documents (GRN, invoice, contract).
- PRESERVE STANDARD AUDIT TERMINOLOGY: If the control description uses specific audit terms
  (e.g., "authorized", "verified", "reconciled", "approved"), preserve those EXACT terms in
  attribute names and descriptions. Do NOT rephrase established audit concepts into different
  wording — the meaning can shift subtly and change what is actually being tested.
  BAD:  Control says "Authorized Purchase Return" → attribute named "Return Documentation Reviewed"
  GOOD: Control says "Authorized Purchase Return" → attribute named "Authorized Purchase Return"
- WRITE EVIDENCE-ORIENTED DESCRIPTIONS: Say what to LOOK FOR in the evidence, not what the rule says.
  BAD:  "Signed vendor agreement and required KYC documents attached before vendor creation"
  GOOD: "Vendor master creation in system is supported by VRF and KYC documents"
- NEVER consolidate two genuinely independent audit dimensions just to reduce attribute count.
  If the control tests physical verification AND system recording AND approval, those are 3
  separate attributes — do not merge them.

**Critical rules for attributes:**
- Each attribute must test ONE specific aspect of the control
- Together, all attributes must FULLY COVER the control description -- if all attributes are "Yes", the control operated effectively for that sample
- Attributes should be MUTUALLY EXCLUSIVE -- don't overlap
- Name must be SHORT (3-6 words) for use as a column header
- Description should explain what evidence to look for in the supporting documents

**IMPORTANT — Avoiding bad attributes:**
- Do NOT create attributes for CONDITIONAL actions unless the condition always applies.
  BAD: "Discrepancy Logged" (only applies when discrepancies exist — if goods match PO, this is N/A)
  BAD: "Variance Escalated" (only applies when variances exist — if delivery matches PO, this is N/A)
  BAD: "Discrepancies Addressed" (only applies when discrepancies exist)
  BETTER: "Inspection Performed" (always applies) or "Match Verified" (always applies)
- If you MUST include a conditional attribute (because the control explicitly requires logging
  discrepancies, escalating variances, etc.), you MUST follow these TWO rules:
  RULE 1: The attribute DESCRIPTION must explicitly state when N/A applies:
    BAD description:  "Discrepancies were logged in the system"
    GOOD description: "Discrepancies were logged in the system. N/A if no discrepancies existed (all items matched PO)."
    BAD description:  "Variances documented and escalated to procurement"
    GOOD description: "Variances documented and escalated. N/A if delivery matched PO with no variances."
  RULE 2: The attribute NAME should hint at its conditional nature:
    BETTER: "Discrepancy Logged (if any)" or "Variance Escalated (if any)"
  If a control says "discrepancies are logged and routed to procurement", create:
    1: "Inspection Completed" — Physical inspection was performed (always applies)
    2: "Discrepancy Logged (if any)" — Any discrepancies found were logged. N/A if zero discrepancies.
    3: "Resolution Routed (if any)" — Discrepancies routed to procurement. N/A if zero discrepancies.
  NOT:
    1: "Inspection Completed"
    2: "Discrepancy Logged" — (missing N/A guidance = evaluator will mark No when no discrepancies exist)
- For FINANCIAL SCREENING / VENDOR VETTING attributes:
  *** THIS IS CRITICAL — GET THIS RIGHT ***
  The following are EQUIVALENT forms of financial health review:
    - D&B (Dun & Bradstreet) score or report = financial health review
    - Credit report = financial health review
    - Financial stability assessment = financial health review
  The following are EQUIVALENT forms of compliance validation:
    - "Compliance clear" = regulatory compliance validated
    - ISO certification verified = compliance check done
  The following are EQUIVALENT forms of reference/verification:
    - Site visit = reference check / physical verification
    - Trade references obtained = reference check
  DO NOT create an attribute called "Financial Review Completed" with description requiring a
  separate "financial health review report." Instead:
    GOOD: "Financial Health Reviewed" — "D&B score obtained OR credit report OR financial assessment documented"
    BAD:  "Financial Review Completed" — "Financial health review report present and signed"
  If the evidence says "Full screening: D&B score 76, compliance clear, site visit done, ISO verified"
  then ALL screening attributes are satisfied.
  NOTE: If the control or evidence mentions a MINIMUM THRESHOLD for scores (e.g., "D&B must be >60",
  "policy requires score above 60"), the attribute description should include this threshold so the
  evaluator can check whether the score PASSES or FAILS the requirement. A score below the threshold
  means the financial review was done but FAILED — that is a control failure.
- Do NOT break the control into more granular attributes than the evidence can support.
  If the control says "complete a standardized checklist including X, Y, and Z", do NOT create
  separate attributes for X, Y, and Z if evidence typically confirms the checklist holistically 
  (e.g., "all items verified"). Instead, use "Checklist Completed" as one attribute.
- For PREVENTIVE controls that BLOCK bad transactions: the primary attribute should test whether
  the blocking/rejection mechanism worked — NOT whether the transaction succeeded.
  EXAMPLE for "System blocks PO without approved PR":
  1: "Block Triggered" — System rejected the unauthorized attempt
  2: "Valid PR Linked" — Successful PO linked to approved PR
  NOT: "PO Created Successfully" — the block PREVENTING creation is the success.

Good example for "CFO approval for recurring expenses" (simple control — 3 attributes):
  1: "CFO Approval" -- Recording expenses are approved by the CFO as per defined authority limits
  2: "Liability Linked to Approved Invoice" -- Expense liability recorded in the system is linked to the approved invoice
  3: "Payment After Approval" -- Payment is processed only after the required approval is obtained

Good example for "Vendor creation based on agreement and KYC" (3 attributes):
  1: "VRF and KYC Availability" -- Vendor master creation is supported by a completed Vendor Registration Form and KYC documents
  2: "Authorized Creation and Approval" -- Vendor master creation and approval are performed by authorized personnel in line with defined authority
  3: "Vendor Classification" -- Vendor is created by authorized users and classified correctly as per approval details

Good example for "Purchase Order creation" (4 attributes — complex control needs more):
  1: "Vendor Sourcing Evidence" -- Vendor selection is supported by quotations or market assessment
  2: "Vendor Selection Support" -- Vendor is reviewed and approved by appropriate management
  3: "PO Approved as per Authority" -- Purchase Order is approved by authorized personnel before issuance
  4: "Alignment with Company Policy" -- PO aligns with company procurement policy and exceptions are documented

Bad example — separate attribute per approver (do NOT do this):
  1: "Legal Approval Recorded" <- merge approval chain into one DOA check
  2: "Accounts Manager Approval" <- separate attribute per approver is wrong
  3: "CFO Sign-off" <- yet another approver as separate attribute

Bad example — inventing requirements not in control description:
  1: "Audit Trail Maintained" <- control description doesn't mention audit trail
  2: "System Access Restricted" <- control description doesn't mention access controls

## 3. Sample Columns (3-6 columns)
These are the TRANSACTION DETAILS extracted from each sample's evidence -- the key identifiers and facts an auditor would record. They are NOT Yes/No -- they are descriptive values.

Good examples: "PR Number", "PO Amount", "Vendor Name", "Approval Date", "Invoice #"
Bad examples: "Status" (too vague), "Result" (that's the test outcome, not sample detail)

## RESPOND ONLY WITH JSON:
{
  "worksteps": ["1. ...", "2. ...", "3. ..."],
  "attributes": [
    {"id": "1", "name": "Short Name", "description": "Specific evidence to verify"},
    {"id": "2", "name": "Short Name", "description": "Specific evidence to verify"}
  ],
  "sample_columns": [
    {"key": "snake_case_key", "header": "Column Header"},
    {"key": "snake_case_key2", "header": "Column Header 2"}
  ]
}"""


SCHEMA_FEW_SHOT_EXAMPLES = """
## FEW-SHOT EXAMPLES — Human auditor style (attribute count matches control complexity: 2-3 for simple, 4-5 for complex)

Example A — Vendor master creation (3 attributes — preserving audit terminology from control)
Control: "Vendor master creation based on agreement and KYC documents"
Human auditors use 3 attributes matching the control's stated requirements:
{
    "worksteps": [
        "1. Verified vendor master creation is supported by a completed Vendor Registration Form and KYC documents.",
        "2. Confirmed vendor master creation and approval are performed by authorized personnel in line with defined authority.",
        "3. Checked that vendor is created by authorized users and classified correctly based on the agreement and nature of services."
    ],
    "attributes": [
        {"id": "1", "name": "VRF and KYC Availability", "description": "Vendor master creation is supported by a completed Vendor Registration Form and KYC documents."},
        {"id": "2", "name": "Authorized Creation and Approval", "description": "Vendor master creation and approval are performed by authorized personnel in line with defined authority."},
        {"id": "3", "name": "Vendor Classification", "description": "Vendor is created by authorized users and classified correctly as per approval details."}
    ],
    "sample_columns": [
        {"key": "vendor_code", "header": "Vendor Code"},
        {"key": "vendor_name", "header": "Vendor Name"},
        {"key": "creation_date", "header": "Creation Date"},
        {"key": "created_by", "header": "Created By"}
    ]
}

Example B — CFO payment approval (simple control — 3 attributes)
Control: "CFO approval for recurring expenses"
Human auditors use 3 attributes:
{
    "worksteps": [
        "1. Obtained payment voucher and supporting invoice for the selected sample.",
        "2. Verified recording expenses are approved by the CFO as per defined authority limits.",
        "3. Confirmed expense liability recorded in the system is linked to the approved invoice.",
        "4. Verified payment is processed only after the required approval is obtained."
    ],
    "attributes": [
        {"id": "1", "name": "CFO Approval", "description": "Recording expenses are approved by the CFO as per defined authority limits."},
        {"id": "2", "name": "Liability Linked to Approved Invoice", "description": "Expense liability recorded in the system is linked to the approved invoice."},
        {"id": "3", "name": "Payment After Approval", "description": "Payment is processed only after the required approval is obtained."}
    ],
    "sample_columns": [
        {"key": "invoice_number", "header": "Invoice Number"},
        {"key": "payment_amount", "header": "Payment Amount"},
        {"key": "approval_date", "header": "Approval Date"},
        {"key": "payment_date", "header": "Payment Date"}
    ]
}

Example C — Three-way matching and approval (consolidated — 2 attributes, not 5)
Control: "Three-way matching and CFO approval"
NOTE: The tool previously generated 5 attrs (Supporting Docs Present, Invoice Rate ≤ PO Rate, Invoice Qty ≤ GRN Qty, User Dept Approval, CFO Approval Recorded).
Human auditors consolidate into 2:
{
    "worksteps": [
        "1. Obtained PO, GRN, and invoice for the selected transaction.",
        "2. Verified department personnel approved the invoice after three-way match.",
        "3. Confirmed all approvals are as per DOA matrix."
    ],
    "attributes": [
        {"id": "1", "name": "Department Approval", "description": "Department personnel approved the invoice after verifying three-way match of PO, GRN, and invoice amounts."},
        {"id": "2", "name": "DOA Compliance", "description": "All approvals obtained as per Delegation of Authority matrix with evidence of sign-off."}
    ],
    "sample_columns": [
        {"key": "po_number", "header": "PO Number"},
        {"key": "grn_number", "header": "GRN Number"},
        {"key": "invoice_number", "header": "Invoice Number"},
        {"key": "invoice_amount", "header": "Invoice Amount"},
        {"key": "approver_name", "header": "Approver Name"}
    ]
}

Example D — Expense provision recording (3 attributes — using control's own terminology)
Control: "Expense provision recording"
Human auditors use 3 attributes:
{
    "worksteps": [
        "1. Verified provisions are calculated based on the nature of expense and available information.",
        "2. Confirmed provision amounts are supported by documents and calculation workings.",
        "3. Checked provision entries recorded are traceable to supporting documents."
    ],
    "attributes": [
        {"id": "1", "name": "Basis for Provision", "description": "Provisions are calculated based on the nature of expense and available information."},
        {"id": "2", "name": "Supporting Documents and Calculation", "description": "Provision amounts are supported by documents and calculation workings."},
        {"id": "3", "name": "Provision Linked to Accounting Records", "description": "Provision entries recorded are traceable to supporting source documents."}
    ],
    "sample_columns": [
        {"key": "provision_ref", "header": "Provision Reference"},
        {"key": "grn_invoice_ref", "header": "GRN/Invoice Ref"},
        {"key": "provision_amount", "header": "Provision Amount"},
        {"key": "posting_date", "header": "Posting Date"}
    ]
}

Example E — Purchase Order creation (4 attributes — complex control with multiple requirements)
Control: "Purchase Order creation"
Human auditors use 4 attributes to cover all requirements:
{
    "worksteps": [
        "1. Verified vendor selection is supported by quotations or market assessment.",
        "2. Confirmed vendor is reviewed and approved by appropriate management.",
        "3. Checked Purchase Order is approved by authorized personnel before issuance.",
        "4. Verified exceptions are aligned with company procurement policy."
    ],
    "attributes": [
        {"id": "1", "name": "Vendor Sourcing Evidence", "description": "Vendor selection is supported by quotations or market assessment."},
        {"id": "2", "name": "Vendor Selection Support", "description": "Vendor is reviewed and approved by appropriate management."},
        {"id": "3", "name": "PO Approved as per Authority", "description": "Purchase Order is approved by authorized personnel before issuance."},
        {"id": "4", "name": "Alignment with Company Policy", "description": "Exceptions aligned with company procurement policy and documented."}
    ],
    "sample_columns": [
        {"key": "pr_number", "header": "PR Number"},
        {"key": "po_number", "header": "PO Number"},
        {"key": "vendor_name", "header": "Vendor Name"},
        {"key": "po_amount", "header": "PO Amount"},
        {"key": "approval_date", "header": "Approval Date"}
    ]
}

Example F — Goods receipt and GRN approval (3 attributes — physical + system + approval)
Control: "Goods receipt and GRN approval"
NOTE: The tool previously generated 5 attrs (PO Referenced, System GRN, Security Verified, Dept Approved, Qty Matches).
Human auditors consolidate into 3:
{
    "worksteps": [
        "1. Obtained GRN and verified it reflects accurate PO values (quantities, descriptions).",
        "2. Confirmed security verified goods received and quantities match delivery documentation.",
        "3. Verified user department approved the GRN with evidence of sign-off."
    ],
    "attributes": [
        {"id": "1", "name": "GRN Reflects PO", "description": "GRN verified to reflect accurate PO values — quantities and item descriptions match the purchase order."},
        {"id": "2", "name": "Security Verification", "description": "Security verified goods received and quantities match delivery documentation."},
        {"id": "3", "name": "Department Approval", "description": "User department approved the GRN with documented evidence of sign-off."}
    ],
    "sample_columns": [
        {"key": "po_number", "header": "PO Number"},
        {"key": "grn_number", "header": "GRN Number"},
        {"key": "grn_date", "header": "GRN Date"},
        {"key": "approved_by", "header": "Approved By"}
    ]
}

Example G — Legal review and approval of agreements (3 attributes — DOA consolidated)
Control: "Legal review and approval of agreements"
NOTE: The tool previously generated 4 attrs (Legal Review, Unfavorable Terms, Signatory per DOA, Pre-Execution).
Human auditors consolidate into 3:
{
    "worksteps": [
        "1. Obtained evidence of formal legal review for the selected agreement.",
        "2. Verified agreement was approved by signatory with appropriate DOA authority.",
        "3. Confirmed all required signatures were obtained before contract execution start date."
    ],
    "attributes": [
        {"id": "1", "name": "Legal Review Documented", "description": "Evidence of formal legal review documented and attributable to the legal department."},
        {"id": "2", "name": "DOA Signatory Verified", "description": "Agreement approved by authorized signatory per DOA matrix with appropriate authority level."},
        {"id": "3", "name": "Pre-Execution Signatures", "description": "All required signatures obtained before contract start date."}
    ],
    "sample_columns": [
        {"key": "agreement_ref", "header": "Agreement Reference"},
        {"key": "vendor_name", "header": "Vendor/Counterparty"},
        {"key": "signatory_name", "header": "Signatory Name"},
        {"key": "execution_date", "header": "Execution Date"}
    ]
}

Example H — Preventive automated block (3 attributes)
Control: "System blocks PO without approved PR"
{
    "worksteps": [
        "1. Reperformed transaction path to test behavior when mandatory prerequisite was missing.",
        "2. Verified system generated hard-stop message and blocked unauthorized transaction.",
        "3. Verified successful transaction only after prerequisite was satisfied.",
        "4. Reviewed system logs to confirm event timestamps and user actions."
    ],
    "attributes": [
        {"id": "1", "name": "Block Triggered", "description": "System blocks transaction when prerequisite rule is violated, with error message/log evidence."},
        {"id": "2", "name": "Valid Reference Linked", "description": "Successful transaction includes valid prerequisite reference (e.g., approved PR ID) in system record."},
        {"id": "3", "name": "Audit Trail Retained", "description": "System log retains user ID, timestamp, and action outcome for blocked and successful attempts."}
    ],
    "sample_columns": [
        {"key": "transaction_id", "header": "Transaction ID"},
        {"key": "error_code", "header": "Error Code"},
        {"key": "reference_id", "header": "Reference ID"},
        {"key": "event_timestamp", "header": "Event Timestamp"}
    ]
}

Example I — Board authorization for bank account opening/closure (3 tool attrs → 2 human attrs)
Control: "Board resolution and authorization for opening or closing bank accounts"
NOTE: The tool previously generated 3 attrs (Board Authorization Evidenced, Resolution Details Align to Action, Bank Account Action Linked to Resolution).
Human auditors consolidate into 2:
{
    "worksteps": [
        "1. Obtained certified Board Resolution and meeting minutes for the selected bank account action.",
        "2. Verified board resolution was passed prior to opening or closure of the bank account.",
        "3. Confirmed evidence of submission of board resolution to the bank by compliance officer."
    ],
    "attributes": [
        {"id": "1", "name": "Board Resolution for Bank Account", "description": "Board resolution was passed prior to opening and closure of bank account, evidenced by certified Board Resolution and meeting minutes."},
        {"id": "2", "name": "Submission of Board Resolution to Bank", "description": "Evidence of submission of board resolution by compliance officer prior to opening/closure of bank account (email trail)."}
    ],
    "sample_columns": [
        {"key": "bank_account_number", "header": "Bank Account Number"},
        {"key": "bank_name", "header": "Bank Name"},
        {"key": "resolution_date", "header": "Resolution Date"},
        {"key": "action_type", "header": "Action (Open/Close)"}
    ]
}

Example J — Monthly bank reconciliation (3 tool attrs → 3 human attrs, renamed for clarity)
Control: "Monthly bank reconciliation statement preparation and review"
NOTE: The tool previously generated 3 attrs (Bank-to-Ledger Reconciliation, Reconciling Items Supported and Posted, Preparation and Review Evidenced).
Human auditors use 3 but with clearer, action-oriented names:
{
    "worksteps": [
        "1. Verified bank reconciliation statement exists for the bank account and period.",
        "2. Checked that unreconciled items (e.g., stale cheques, bank charges) are posted accurately.",
        "3. Confirmed BRS is prepared by Assistant Manager and reviewed by AVP Finance or CFO."
    ],
    "attributes": [
        {"id": "1", "name": "Existence of Bank Reconciliation Statements", "description": "Bank reconciliation statements exist for the bank account for the selected period."},
        {"id": "2", "name": "Accurate Posting of Unreconciled Items", "description": "Unreconciled items (e.g., stale cheques, bank charges, timing differences) are identified and posted accurately in the books."},
        {"id": "3", "name": "Preparation and Review of Bank Reconciliation Statement", "description": "BRS is prepared by Assistant Manager and reviewed by AVP Finance or CFO, evidenced by name, signature/initials, and date."}
    ],
    "sample_columns": [
        {"key": "bank_account", "header": "Bank Account"},
        {"key": "brs_month", "header": "BRS Month"},
        {"key": "prepared_by", "header": "Prepared By"},
        {"key": "reviewed_by", "header": "Reviewed By"}
    ]
}

Example K — Physical cash verification and certificate (2 tool attrs → 3 human attrs — tool UNDER-decomposed)
Control: "Physical cash count reconciliation and year-end cash balance certificate"
NOTE: The tool previously generated only 2 attrs (Cash Reconciliation Performed, Year-End Cash Certificate Approval).
Human auditors expand to 3 — the tool missed location coverage:
{
    "worksteps": [
        "1. Confirmed physical cash count matches Odoo system balance for the location and period.",
        "2. Verified cash balance certificate is prepared and approved by Cashier and Admin Head.",
        "3. Ensured all locations' cash balances are included in the verification."
    ],
    "attributes": [
        {"id": "1", "name": "Physical Cash Verification with Odoo Balance", "description": "Physical cash count matches Odoo system balance, with documented comparison and evidence of resolution for any differences."},
        {"id": "2", "name": "Cash Balance Certificate Preparation and Approval", "description": "Year-end cash balance certificate prepared based on reconciled balances and approved by Cashier and Admin Head with sign-off evidence."},
        {"id": "3", "name": "Inclusion of All Locations in Cash Verification", "description": "All locations' cash balances are included in the verification — no location omitted."}
    ],
    "sample_columns": [
        {"key": "location", "header": "Location"},
        {"key": "cash_count_date", "header": "Count Date"},
        {"key": "physical_balance", "header": "Physical Balance"},
        {"key": "system_balance", "header": "Odoo Balance"}
    ]
}

Example L — Segregation of duties in cash handling (2 tool attrs → 2 human attrs, renamed)
Control: "Segregation of duties between cash handling, posting, and verification"
NOTE: The tool previously generated 2 attrs (Segregation of Duties Evidenced, Independent Cash Verification).
Human auditors use 2 but focus on role separation and system accuracy:
{
    "worksteps": [
        "1. Verified roles are clearly separated between cashier, accountant, and verifier.",
        "2. Reviewed Odoo entries for accuracy and proper posting by distinct personnel."
    ],
    "attributes": [
        {"id": "1", "name": "Segregation of Duties in Cash Handling", "description": "Cash handling (cashier), posting in Odoo (accountant), and verification are performed by different individuals, evidenced by distinct names and user IDs."},
        {"id": "2", "name": "Accuracy of Cash Entries in Odoo", "description": "Cash entries in Odoo are accurate and properly posted, reviewed against cash vouchers and verification reports."}
    ],
    "sample_columns": [
        {"key": "transaction_ref", "header": "Transaction Ref"},
        {"key": "cashier_name", "header": "Cashier"},
        {"key": "posted_by", "header": "Posted By"},
        {"key": "verified_by", "header": "Verified By"}
    ]
}

Example M — Payment voucher with supporting documents and cancellation (3 tool attrs → 3 human attrs, system posting added)
Control: "Cash payment authorization with adequate supporting documents and prevention of duplicate payments"
NOTE: The tool previously generated 3 attrs (Payment Authorization per DOA, Adequate Original Supporting, Originals Cancelled to Prevent Reuse).
Human auditors use 3 but replace cancellation check with system posting verification:
{
    "worksteps": [
        "1. Verified payment vouchers are approved before disbursement per DOA matrix.",
        "2. Checked supporting documents are adequate and substantiate the payment.",
        "3. Confirmed Odoo entry shows posting with document reference for traceability."
    ],
    "attributes": [
        {"id": "1", "name": "Approval of Payment Vouchers", "description": "Payment vouchers are approved by authorized personnel before disbursement, per DOA matrix."},
        {"id": "2", "name": "Adequacy of Supporting Documents for Payments", "description": "Payment vouchers supported by adequate and relevant original documents (invoice, bill, receipt) that substantiate the payment."},
        {"id": "3", "name": "System Posting of Payments with Document Reference", "description": "Odoo entry shows payment posting with document reference for traceability and audit trail."}
    ],
    "sample_columns": [
        {"key": "voucher_number", "header": "Voucher Number"},
        {"key": "payment_amount", "header": "Payment Amount"},
        {"key": "approved_by", "header": "Approved By"},
        {"key": "odoo_posting_ref", "header": "Odoo Posting Ref"}
    ]
}

Example N — Cash transaction recording with document linkage (3 tool attrs → 4 human attrs — tool missed system consistency check)
Control: "Cash transaction recording with source document linkage and manager review"
NOTE: The tool previously generated 3 attrs (Source Document Linkage, Manager Review and Approval, Daily Recording Timeliness).
Human auditors expand to 4 — adding system consistency verification:
{
    "worksteps": [
        "1. Verified valid supporting documents exist for the cash transaction.",
        "2. Checked timely recording in system (same day or next working day).",
        "3. Confirmed entries reviewed and approved by Finance Manager.",
        "4. Validated Odoo entries match supporting documents in amount and date."
    ],
    "attributes": [
        {"id": "1", "name": "Supporting Documents for Cash Transactions", "description": "Cash receipt or payment supported by appropriate source documents (cash voucher, receipt, bill, withdrawal slip)."},
        {"id": "2", "name": "Timely Recording of Cash Transactions", "description": "Cash transactions recorded in the cash book on transaction date or next working day; delays supported by documented justification."},
        {"id": "3", "name": "Review and Approval of Odoo Entries", "description": "Cash entries reviewed and approved by Manager – Finance and Accounts, evidenced by signature or system user ID with date."},
        {"id": "4", "name": "Consistency of Odoo Entries with Supporting Documents", "description": "Odoo entries match supporting documents in amount, date, and transaction details."}
    ],
    "sample_columns": [
        {"key": "cash_book_ref", "header": "Cash Book Ref"},
        {"key": "transaction_date", "header": "Transaction Date"},
        {"key": "recording_date", "header": "Recording Date"},
        {"key": "approved_by", "header": "Approved By"}
    ]
}

Example O — Long-outstanding cheques review and corrective action (3 tool attrs → 3 human attrs, clearer naming)
Control: "Identification, review, and corrective action for long-outstanding cheques"
NOTE: The tool previously generated 3 attrs (Long-Outstanding Cheques Identified, Review Evidenced, Corrective Action and Traceability).
Human auditors use 3 with clearer, specific names:
{
    "worksteps": [
        "1. Identified cheques outstanding for more than 90 days through ageing analysis.",
        "2. Verified evidence of management review of long-outstanding cheques.",
        "3. Confirmed corrective actions are documented for cheques outstanding beyond three months."
    ],
    "attributes": [
        {"id": "1", "name": "Identification of Long Outstanding Cheques", "description": "Cheques outstanding for more than 90 days clearly identified through ageing analysis."},
        {"id": "2", "name": "Management Review of Long Outstanding Cheques", "description": "Documented evidence of management review of long-outstanding cheques exists."},
        {"id": "3", "name": "Corrective Action on Long Outstanding Cheques", "description": "Corrective actions for cheques outstanding beyond three months are documented and traceable."}
    ],
    "sample_columns": [
        {"key": "cheque_number", "header": "Cheque Number"},
        {"key": "cheque_date", "header": "Cheque Date"},
        {"key": "days_outstanding", "header": "Days Outstanding"},
        {"key": "corrective_action", "header": "Corrective Action"}
    ]
}

Example P — Bank guarantee tracking and monitoring (3 tool attrs → 4 human attrs — tool missed maintenance check)
Control: "Bank guarantee tracking, expiry monitoring, and access restriction"
NOTE: The tool previously generated 3 attrs (BG Tracking Completeness and Traceability, Expiry Monitoring and Action, Restricted Access).
Human auditors expand to 4 — separating existence from accuracy and adding maintenance:
{
    "worksteps": [
        "1. Verified existence of bank guarantee tracking sheet with complete details.",
        "2. Checked accuracy of bank guarantee records against source documents.",
        "3. Confirmed monitoring of bank guarantees nearing expiry dates.",
        "4. Ensured tracking sheet is maintained and updated regularly."
    ],
    "attributes": [
        {"id": "1", "name": "Existence of Bank Guarantee Tracking Sheet", "description": "Bank guarantee tracking sheet exists with complete details for all outstanding BGs."},
        {"id": "2", "name": "Accuracy of Bank Guarantee Records", "description": "BG tracking records are accurate — amounts, dates, and beneficiary details match underlying BG documents."},
        {"id": "3", "name": "Monitoring of Bank Guarantees Nearing Expiry", "description": "BG expiry dates are actively monitored and timely action taken for renewals or releases."},
        {"id": "4", "name": "Maintenance of Bank Guarantee Tracking Sheet", "description": "Tracking sheet is maintained and updated regularly with current status of each bank guarantee."}
    ],
    "sample_columns": [
        {"key": "bg_number", "header": "BG Number"},
        {"key": "bg_amount", "header": "BG Amount"},
        {"key": "expiry_date", "header": "Expiry Date"},
        {"key": "beneficiary", "header": "Beneficiary"}
    ]
}

Example Q — Interest recalculation and supervisory review (3 tool attrs → 3 human attrs, ERP linkage added)
Control: "Interest recalculation, supervisory review, and Odoo posting verification"
NOTE: The tool previously generated 3 attrs (Interest Recalculation Verified, Supervisory Review Evidenced, Odoo Entry and Audit Trail).
Human auditors use 3 with focus on FD terms and accounting linkage:
{
    "worksteps": [
        "1. Verified interest calculation against FD terms and recalculated independently.",
        "2. Checked supervisory review of recalculation is evidenced.",
        "3. Confirmed accounting entry in Odoo is linked and verified against recalculation."
    ],
    "attributes": [
        {"id": "1", "name": "Interest Calculation & FD Terms Verified", "description": "Interest recalculated independently and agrees with FD terms (rate, tenure, principal)."},
        {"id": "2", "name": "Supervisory Review Evidenced", "description": "Recalculation reviewed by supervisor with documented evidence of sign-off."},
        {"id": "3", "name": "Accounting Entry Linked & Verified", "description": "Interest entry posted in Odoo is linked to the recalculation and verified for accuracy."}
    ],
    "sample_columns": [
        {"key": "fd_number", "header": "FD Number"},
        {"key": "principal_amount", "header": "Principal Amount"},
        {"key": "interest_rate", "header": "Interest Rate"},
        {"key": "interest_amount", "header": "Interest Amount"}
    ]
}
"""


SCHEMA_CRITIC_PROMPT = """You are a controls schema quality reviewer who thinks like an experienced human auditor.
You will receive a control context and a DRAFT schema JSON.

Your job:
1) Ensure COMPLETE COVERAGE — every distinct requirement in the control description has an attribute.
2) Remove only truly redundant attributes (e.g., separate attributes per approver in one approval chain).
3) Preserve standard audit terminology from the control description — do NOT rephrase.
4) Return ONLY valid JSON with keys: worksteps, attributes, sample_columns.

**MANDATORY COVERAGE CHECK (do this FIRST):**
- List every distinct requirement from the control description.
- Verify each requirement is covered by at least one attribute.
- If a requirement is MISSING, ADD an attribute for it.
- Do NOT remove an attribute if it is the ONLY one covering a requirement.

**Quality checks:**
- APPROVAL CHAIN DECOMPOSITION: If you see separate attributes per approver (e.g., "Legal Approval",
  "Manager Approval", "CFO Approval"), consolidate into ONE "Approval per DOA" attribute.
- SCOPE: Remove attributes that test requirements NOT stated in the control description.
  If the control doesn't mention access controls, don't test access controls.
- WORDING: Attribute names and descriptions should use the same terminology as the control
  description. Do not rephrase "Authorized Purchase Return" into "Return Documentation Reviewed".
- MISSING LINKAGE: For financial controls (provisions, journal entries, reconciliations), ensure
  there is a linkage/traceability attribute connecting to source documents.
- DESCRIPTIONS MUST BE EVIDENCE-ORIENTED: Say what to look for in the evidence, not what the rule
  requires. "Supported by VRF and KYC documents" not "Required KYC attached before creation."
- Do NOT reference document names, system names, or report names that are not mentioned in the
  control description.

**Quality bar:**
- 3-5 concrete worksteps, past tense, evidence-oriented
- As many attributes as the control needs (typically 3-5), mutually exclusive, non-generic, measurable
- Conditional attributes MUST include explicit N/A guidance
- If control implies thresholds/limits, include threshold checks in descriptions
- 3-6 sample columns with transaction-identifying fields (not result/status)
- Avoid generic names: Properly Approved, Review Done, Evidence Sufficient, Authorization Obtained
"""


def _schema_quality_issues(schema: ControlSchema, rcm: RCMRow) -> list[str]:
        """Return blocking quality issues for a generated schema."""
        issues = []
        attrs = schema.attributes or []
        steps = schema.worksteps or []
        cols = schema.sample_columns or []

        if not (3 <= len(steps) <= 5):
                issues.append(f"worksteps count is {len(steps)} (expected 3-5)")
        if not (2 <= len(attrs) <= 7):
                issues.append(f"attributes count is {len(attrs)} (expected 2-7)")
        if not (3 <= len(cols) <= 6):
                issues.append(f"sample_columns count is {len(cols)} (expected 3-6)")

        generic_markers = (
                "properly approved", "review done", "reviewed documentation", "evidence sufficient",
                "authorization obtained", "control performed", "status check", "checked control",
        )
        evidence_markers = (
                "timestamp", "date", "user", "id", "log", "report", "workflow", "approval",
                "amount", "number", "threshold", "limit", "reference",
                "linkage", "trace", "source", "verified", "documented", "validated", "comply", "doa",
                "supported", "restricted", "obtained", "maintained",
                "processed", "recorded", "classified", "match", "matched", "marked",
        )
        conditional_keywords = ("discrep", "variance", "exception", "escalat", "resolution", "mismatch", "deviation", "dispute")

        for i, a in enumerate(attrs, 1):
                name = str(a.get("name", "")).strip()
                desc = str(a.get("description", "")).strip()
                name_l = name.lower()
                desc_l = desc.lower()

                if not name or len(name.split()) < 2:
                        issues.append(f"attribute {i} name is too short/unclear")
                if any(g in name_l for g in generic_markers):
                        issues.append(f"attribute {i} has generic name '{name}'")
                if len(desc) < 35:
                        issues.append(f"attribute {i} description is too brief")
                if not any(m in desc_l for m in evidence_markers):
                        issues.append(f"attribute {i} lacks evidence-anchor wording")

                is_conditional = any(k in name_l or k in desc_l for k in conditional_keywords)
                if is_conditional and "n/a" not in desc_l:
                        issues.append(f"attribute {i} is conditional but missing explicit N/A guidance")

        # Overlap check (name token similarity)
        stop = {"the", "and", "with", "for", "from", "before", "after", "if", "any"}
        token_sets = []
        for a in attrs:
                toks = {t for t in re.findall(r"[a-z0-9]+", str(a.get("name", "")).lower()) if t not in stop}
                token_sets.append(toks)
        for i in range(len(token_sets)):
                for j in range(i + 1, len(token_sets)):
                        a, b = token_sets[i], token_sets[j]
                        if not a or not b:
                                continue
                        jaccard = len(a & b) / max(1, len(a | b))
                        if jaccard >= 0.75:
                                issues.append(f"attributes {i+1} and {j+1} appear overlapping")

        # Threshold-aware check
        ctext = " ".join([
                str(rcm.control_description or ""),
                str(rcm.control_objective or ""),
        ]).lower()
        threshold_hint = any(x in ctext for x in [">", "<", "threshold", "limit", "minimum", "maximum", "within", "%"])
        if threshold_hint:
                threshold_in_attrs = any(
                        any(x in str(a.get("description", "")).lower() for x in ["threshold", "limit", "minimum", "maximum", ">", "<", "%"])
                        for a in attrs
                )
                if not threshold_in_attrs:
                        issues.append("control suggests thresholds/limits but attributes do not mention threshold checks")

        return issues


def _estimate_tokens(text: str) -> int:
    """Rough token estimate: ~4 characters per token for English text."""
    return len(text) // 4


# Token threshold for warning (does NOT truncate -- just logs a heads-up).
# With gpt-4o / gpt-4o-mini (128K context), prompts up to ~100K tokens are fine.
# With gpt-35-turbo (16K context), anything above ~12K may fail.
PROMPT_WARNING_THRESHOLD = 100000  # tokens -- for 128K context models (gpt-4o / gpt-4o-mini)


def _split_chunks(text: str, size: int = 2500, overlap: int = 300) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    chunks = []
    i = 0
    while i < len(text):
        j = min(len(text), i + size)
        chunks.append(text[i:j])
        if j >= len(text):
            break
        i = max(j - overlap, i + 1)
    return chunks


def _azure_embed(texts: list[str]) -> list[list[float]]:
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "").strip()
        or os.getenv("AZURE_OPENAI_EMBEDDING_MODEL", "").strip()
        or os.getenv("OPENAI_EMBEDDING_MODEL", "").strip()
        or "text-embedding-ada-002"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()
    if not endpoint or not api_key or not texts:
        return []

    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/embeddings?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}
    payload = {"input": texts}
    resp = requests.post(url, headers=headers, json=payload, timeout=90)
    resp.raise_for_status()
    data = resp.json().get("data", [])
    return [d.get("embedding", []) for d in data]


def _cos(a: list[float], b: list[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(y * y for y in b))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)


def _write_embedding_log(
    engine: str, control_id: str, query: str,
    chunks: list[str], scored: list[tuple[int, float]],
    keep_idx: list[int], final_text: str,
    sample_index: int = 0, total_samples: int = 0,
    original_text_len: int = 0,
) -> None:
    """Write a detailed embedding selection log file."""
    log_dir = _get_embedding_log_dir()
    if log_dir is None:
        return
    try:
        from datetime import datetime, timezone
        ts = datetime.now(timezone.utc)
        ts_str = ts.strftime("%Y%m%d_%H%M%S")
        safe_id = re.sub(r"[^\w.\-]", "_", control_id)
        sample_tag = f"_s{sample_index}" if sample_index else ""
        log_path = log_dir / f"{ts_str}_{safe_id}{sample_tag}_{engine}.log"

        sep = "═" * 70
        thin = "─" * 70
        lines: list[str] = []

        lines.append(sep)
        lines.append("EMBEDDING SELECTION LOG")
        lines.append(sep)
        lines.append(f"Timestamp:          {ts.isoformat()}")
        lines.append(f"Engine:             {engine}")
        lines.append(f"Control ID:         {control_id}")
        if total_samples:
            lines.append(f"Sample:             {sample_index} of {total_samples}")
        lines.append(f"Original Text Len:  {original_text_len:,} chars")
        lines.append(f"Chunk Count:        {len(chunks)}")
        lines.append(f"Chunks Selected:    {len(keep_idx)} of {len(chunks)}")
        lines.append("")

        lines.append(thin)
        lines.append("EMBEDDING QUERY (from RCM fields)")
        lines.append(thin)
        lines.append(query)
        lines.append("")

        lines.append(thin)
        lines.append("CHUNK SCORES (sorted by similarity)")
        lines.append(thin)
        for rank, (chunk_idx, score) in enumerate(scored, 1):
            selected_marker = " ← SELECTED" if chunk_idx in keep_idx else ""
            preview = chunks[chunk_idx][:80].replace("\n", " ").strip()
            lines.append(f"Rank {rank:3d}: Chunk {chunk_idx + 1:3d} (score: {score:.4f}){selected_marker}  |  {preview}...")
        lines.append("")

        lines.append(thin)
        lines.append(f"SELECTED CHUNKS CONTENT ({len(keep_idx)} chunks)")
        lines.append(thin)
        for idx in keep_idx:
            lines.append(f"[CHUNK {idx + 1}]")
            lines.append(chunks[idx])
            lines.append("")

        lines.append(sep)
        lines.append("FINAL EVIDENCE TEXT SENT TO LLM")
        lines.append(sep)
        lines.append(final_text)
        lines.append("")
        lines.append(sep)
        lines.append(f"END OF LOG — {control_id} ({engine})")
        lines.append(sep)

        log_path.write_text("\n".join(lines), encoding="utf-8")
    except Exception:
        pass


def _write_embedding_fallback_log(
    engine: str, control_id: str, reason: str,
    text_len: int, text_preview: str,
) -> None:
    """Log when embedding selection was skipped and full text was used directly."""
    log_dir = _get_embedding_log_dir()
    if log_dir is None:
        return
    try:
        from datetime import datetime, timezone
        ts = datetime.now(timezone.utc)
        ts_str = ts.strftime("%Y%m%d_%H%M%S")
        safe_id = re.sub(r"[^\w.\-]", "_", control_id)
        log_path = log_dir / f"{ts_str}_{safe_id}_{engine}_FULLTEXT.log"

        sep = "═" * 70
        thin = "─" * 70
        lines: list[str] = [
            sep,
            "EMBEDDING BYPASS LOG — FULL TEXT USED",
            sep,
            f"Timestamp:          {ts.isoformat()}",
            f"Engine:             {engine}",
            f"Control ID:         {control_id}",
            f"Reason:             {reason}",
            f"Text Length:        {text_len:,} chars",
            "",
            thin,
            "EVIDENCE TEXT SENT TO LLM (full, no embedding selection)",
            thin,
            text_preview,
            "",
            sep,
            f"END OF LOG — {control_id} ({engine})",
            sep,
        ]
        log_path.write_text("\n".join(lines), encoding="utf-8")
    except Exception:
        pass


def _prepare_evidence_for_prompt(rcm: RCMRow, sample: SampleEvidence) -> str:
    """Create retrieval-style evidence context via embeddings; fallback to full text."""
    text = (sample.description or "").strip()
    if not text:
        return text

    use_embed = os.getenv("EVIDENCE_EMBEDDINGS", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not use_embed:
        _write_embedding_fallback_log("TOE", rcm.control_id, "EVIDENCE_EMBEDDINGS disabled", len(text), text)
        return text

    try:
        chunks = _split_chunks(text)
        if len(chunks) <= 1:
            _write_embedding_fallback_log("TOE", rcm.control_id, f"Text fits in 1 chunk ({len(text)} chars) — no splitting needed", len(text), text)
            return text
        parts = [rcm.control_id, rcm.risk_id, rcm.control_description,
                 rcm.control_objective, rcm.control_type, rcm.nature_of_control]
        query = " ".join(p for p in parts if p)
        vecs = _azure_embed([query] + chunks)
        if len(vecs) != len(chunks) + 1:
            _write_embedding_fallback_log("TOE", rcm.control_id, f"Embedding API returned wrong vector count (expected {len(chunks)+1}, got {len(vecs)})", len(text), text)
            return text
        qv, cvs = vecs[0], vecs[1:]
        scored = sorted(((i, _cos(qv, v)) for i, v in enumerate(cvs)), key=lambda x: x[1], reverse=True)
        k = int(os.getenv("EVIDENCE_TOP_K", "24"))
        keep_idx = sorted(i for i, _ in scored[:max(1, min(k, len(chunks)))])
        selected = [f"[CHUNK {i+1}]\n{chunks[i]}" for i in keep_idx]
        final_text = "\n\n".join(selected).strip() or text

        _write_embedding_log(
            engine="TOE", control_id=rcm.control_id, query=query,
            chunks=chunks, scored=scored, keep_idx=keep_idx,
            final_text=final_text, original_text_len=len(text),
        )

        return final_text
    except Exception as exc:
        _write_embedding_fallback_log("TOE", rcm.control_id, f"Exception during embedding: {exc}", len(text), text)
        return text

# ===========================================================================
#  TOE PROMPT BUILDER
# ===========================================================================

def build_toe_prompt(rcm: RCMRow, sample: SampleEvidence,
                     sample_index: int = 1, total_samples: int = 1,
                     schema: ControlSchema = None) -> str:
    """Build TOE evaluation prompt for one sample of one control."""

    # Build supporting documents section
    supporting_section = ""
    if sample.supporting_docs:
        doc_blocks = []
        for doc in sample.supporting_docs:
            label = f" ({doc.doc_type})" if doc.doc_type else ""
            doc_blocks.append(f"""### {doc.filename}{label}
{doc.content}""")
        supporting_section = f"""

## SUPPORTING DOCUMENTS ({len(sample.supporting_docs)} file{"s" if len(sample.supporting_docs) > 1 else ""})
{chr(10).join(doc_blocks)}"""

    # Build schema-specific sections if schema provided
    attr_section = ""
    attr_json = ""
    attr_reasoning_json = ""
    detail_json = ""
    if schema and schema.attributes:
        attr_lines = []
        for attr in schema.attributes:
            attr_lines.append(f"  - **Attribute {attr['id']}** -- {attr['name']}: {attr['description']}")
        attr_section = f"""

## CONTROL-SPECIFIC TESTING ATTRIBUTES
Evaluate each attribute below for THIS specific transaction.
For EACH attribute, you must:
  1. Decide: Yes / No / N/A (STRICT — no other values allowed)
  2. Explain WHY in attribute_reasoning (cite specific evidence)

{chr(10).join(attr_lines)}

**ATTRIBUTE VALUE RULES (STRICT — ONLY these three values allowed):**
- **"Yes"** = Evidence CONFIRMS this aspect operated correctly
- **"No"**  = Evidence shows this was NOT performed, incomplete, deficient, or evidence is missing
- **"N/A"** = This attribute genuinely does NOT APPLY to this transaction
  IMPORTANT: "I cannot find evidence" is "No", NOT "N/A".
  IMPORTANT: "The condition did not arise" IS "N/A", NOT "No". Example:
    - "Discrepancy Logged" when evidence says "all items matched, zero discrepancies" → "N/A"
    - "Variance Documented" when evidence says "delivery matched PO, qty confirmed" → "N/A"
    - "Discrepancies Addressed" when evidence says "no discrepancies found" → "N/A"
    "No" means the issue EXISTED but was NOT handled. "N/A" means the issue DID NOT EXIST.

**DO NOT USE**: "Partially", "Yes with observations", "Mostly", "Partially Compliant",
"Yes - see notes", or ANY qualified answer. If there is ANY doubt → "No".

**HOLISTIC ASSESSMENT — YOUR JUDGMENT MATTERS:**
You are the auditor. Consider ALL attributes together alongside the control description,
risk context, and full evidence. Your attribute assessments + result + remarks must form
a CONSISTENT, COHERENT professional judgment for this transaction.

**CRITICAL — COMMON MISTAKES TO AVOID:**
- If evidence says "all items verified", "full screening done", "checklist completed — all steps done",
  that SATISFIES individual attributes covered by that summary. Do NOT mark "No" just because the
  exact attribute name is not literally in the evidence text.
- A D&B (Dun & Bradstreet) score IS a financial health review. If evidence says "D&B score 84" that
  means the financial review was completed. Do NOT mark "Financial Review Completed = No" when a D&B
  score is present. Similarly: "compliance clear" = compliance validated; "site visit done" = reference check done.
- HOWEVER: If the D&B score is BELOW the policy threshold (e.g., "D&B score 32, policy requires >60"),
  the financial review FAILED — mark the financial attribute as "No" because the vendor did not meet
  the required standard. A low D&B score = review done but FAILED = control failure.
- **CONDITIONAL ATTRIBUTES (MOST COMMON MISTAKE):**
  If an attribute tests a CONDITIONAL action (e.g., "discrepancies logged", "variances documented",
  "exceptions flagged", "discrepancies addressed") and the condition DOES NOT ARISE (no discrepancies
  exist, everything matched, no variances found), mark that attribute **"N/A"** — NEVER "No".
  "No" means "the issue existed but was not handled." "N/A" means "the issue did not exist."
  When evidence says "qty confirmed, quality passed" — there are ZERO discrepancies → "N/A".
- For PREVENTIVE controls: if the system BLOCKED or REJECTED a bad transaction, that means the
  control WORKED → result = "PASS". The block IS the success, not a failure.
- Test ONLY the control as described. Do not invent requirements beyond the control description.

** MANDATORY SELF-CHECK — DO THIS BEFORE WRITING YOUR FINAL ANSWER: **
After you determine each attribute value, STOP and check:
  Step 1: Count how many attributes are "No". 
  Step 2: If count = 0 (all are "Yes" or "N/A") → result MUST be "PASS". No exceptions.
  Step 3: If count > 0 → result MUST be "FAIL", and remarks must explain which attribute failed and why.
  Step 4: For any attribute you marked "N/A", verify: is this truly a condition that didn't arise?
          (e.g., no discrepancies existed → N/A is correct)
  Step 5: For any attribute you marked "No", verify: did the issue ACTUALLY exist but go unhandled?
          If the issue didn't exist at all, change to "N/A".

**CONSISTENCY RULES:**
1. If ANY attribute = "No" → result MUST be "FAIL" and operated_effectively MUST be "No"
2. If result = "PASS" → ALL attributes MUST be "Yes" or "N/A" (never "No")
3. If result = "FAIL" → at least ONE attribute MUST be "No" (identify which aspect failed)
4. Remarks MUST be consistent with the result — do NOT write "control operated effectively" when result is FAIL
5. **ABSOLUTE RULE**: If ALL attributes are "Yes" or "N/A" (NONE are "No") → result MUST be "PASS".
   N/A is NOT a failure. If inspection done, all items matched, no discrepancies to log → PASS."""

        # Build attribute_results and attribute_reasoning JSON fragments
        attr_entries = ', '.join(f'"{a["id"]}": "Yes or No or N/A"' for a in schema.attributes)
        attr_json = f',\n  "attribute_results": {{{attr_entries}}}'
        reason_entries = ', '.join(f'"{a["id"]}": "1-2 sentence explanation citing specific evidence"' for a in schema.attributes)
        attr_reasoning_json = f',\n  "attribute_reasoning": {{{reason_entries}}}'

    if schema and schema.sample_columns:
        detail_entries = ', '.join(f'"{c["header"]}": "extract from evidence"' for c in schema.sample_columns)
        detail_json = f',\n  "sample_details": {{{detail_entries}}}'

    evidence_for_prompt = _prepare_evidence_for_prompt(rcm, sample)
    file_list = [s.strip() for s in str(sample.source_document or "").split(",") if s.strip()]
    file_manifest = "\n".join(f"- {f}" for f in file_list) if file_list else "- (No file names provided)"

    return f"""Test the OPERATING EFFECTIVENESS of this control for the following transaction.
Did the control actually work correctly for this specific sample?

## RCM DETAILS
- **Process**: {rcm.process} > {rcm.subprocess}
{"- **Control Objective**: " + rcm.control_objective + chr(10) if rcm.control_objective else ""}- **Risk**: {rcm.risk_id}
- **Control ID**: {rcm.control_id}
- **Control Description**: {rcm.control_description}
- **Nature**: {rcm.nature_of_control} | **Type**: {rcm.control_type} | **Frequency**: {rcm.control_frequency}

## SAMPLE {sample_index} of {total_samples}
- **Sample ID**: {sample.sample_id}
- **Source Document**: {sample.source_document or "Not specified"}
- **Evidence File Count**: {len(file_list)}
- **Evidence File Manifest**:
{file_manifest}
- **Transaction Evidence**:
{evidence_for_prompt}{supporting_section}{attr_section}

## IMPORTANT EVIDENCE-USAGE INSTRUCTION
- Use ALL listed evidence files holistically; do not rely on only one file.
- If multiple document IDs/numbers/dates exist, capture representative combined values (comma-separated) in sample_details.
- If a specific detail cannot be found in any file, set that detail to "Not extracted from evidence".

## REQUIRED JSON OUTPUT
{{
  "result": "PASS or FAIL",
  "operated_effectively": "Yes or No",
  "control_performed": "Yes or No",
  "timely_execution": "Yes or No or N/A",
  "accurate_execution": "Yes or No",
  "authorized_performer": "Yes or No or N/A",
  "evidence_sufficient": "Yes or No",
  "remarks": "2-4 sentence assessment of operating effectiveness for THIS transaction. Reference specific evidence (doc numbers, dates, names).",
  "deviation_details": "If FAIL: describe the SPECIFIC deviation in detail (what went wrong, which document, which approver, what date). If PASS: None"{attr_json}{attr_reasoning_json}{detail_json}
}}"""


def _is_low_information_evidence(text: str) -> tuple[bool, str]:
    """Check if evidence text is completely empty. All other cases are sent to the LLM."""
    if not (text or "").strip():
        return True, "empty evidence text"
    return False, ""



# ===========================================================================
#  TOE VALIDATOR
# ===========================================================================

# Failure-indicating phrases that should not appear in remarks when result=PASS
_FAIL_INDICATORS = (
    "[fail]", "not effective", "did not operate", "failed to",
    "deviation identified", "deviation found", "control failure",
    "not performed", "not executed", "unauthorized",
    "insufficient evidence", "no evidence",
)

# Pass-indicating phrases that contradict a FAIL result
_PASS_INDICATORS = (
    "control is effective", "operated effectively", "control operated correctly",
    "no deviation", "no exceptions", "operated as intended",
    "no issues identified", "control worked correctly",
    # Additional LLM phrasings that cause contradictions
    "operating effectively", "effectively operated", "effectively performed",
    "control was effective", "control is operating", "control worked as",
    "operating correctly", "correctly operated", "control functioned correctly",
    "control functioned as", "performed effectively", "executed correctly",
    "executed effectively", "no deficiencies", "no findings",
    "control operated as designed", "operated as designed",
    "control is functioning", "functioning correctly", "functioning effectively",
    "all attributes met", "all criteria met", "all requirements met",
    "control passed", "sample passed", "test passed",
)

import re as _re


def _is_negative(val: str) -> bool:
    """
    Robust check for negative/failing values from LLM output.
    The LLM returns MANY variants beyond a simple "No":
      - "No", "No.", "NO", "no"
      - "No - Finance Manager did not approve"
      - "No, the control was not performed"
      - "N" (abbreviation)
      - "Not met", "Not performed", "Not effective"
      - "Failed", "Fail"
      - "Non-compliant", "Deficient", "Unsatisfactory", "Incomplete"
      - "Negative", "Inadequate", "Insufficient"
      - "Partial", "Partially met", "Unmet"
      - "Missing", "Absent", "Rejected", "Denied", "Expired"
      - "Exception", "Deviation", "Overdue", "Late"
    Does NOT treat N/A, Unable to Assess, blank, or affirmative as negative.
    """
    if not val or not val.strip():
        return False
    v = val.strip().lower().rstrip(".")
    # Exclude N/A variants and affirmative values
    if v in ("n/a", "na", "not applicable", "unable to assess", "yes",
             "true", "ok", "pass", "passed", "effective", "met", "compliant",
             "satisfactory", "adequate", "sufficient", "complete"):
        return False
    # Exact negative matches (comprehensive)
    if v in ("no", "n", "fail", "failed", "false", "negative",
             "non-compliant", "noncompliant", "deficient", "unsatisfactory",
             "inadequate", "insufficient", "incomplete", "not met",
             "not effective", "not compliant", "not performed",
             "not executed", "not timely", "not authorized",
             "not sufficient", "not adequate",
             # --- Additional LLM variants (the gap values) ---
             "unmet", "partial", "partially", "partially met",
             "partially compliant", "partially performed",
             "missing", "absent", "rejected", "denied",
             "overdue", "late", "expired", "lapsed",
             "exception", "exception noted", "exception found",
             "exception identified", "exceptions noted",
             "deviation", "deviation found", "deviation noted",
             "deviation identified", "deficiency noted",
             "conditional", "pending", "outstanding",
             "unable to verify", "unable to confirm",
             "unable to determine", "could not verify",
             "could not confirm", "cannot confirm",
             "not evidenced", "not documented", "not observed",
             "not confirmed", "not verified", "not available",
             "not found", "not present", "not provided",
             "no evidence", "no documentation",
             ):
        return True
    # Starts with "no " or "no," or "no-" (LLM adds explanations: "No - reason")
    if v.startswith("no ") or v.startswith("no,") or v.startswith("no-") or v.startswith("no;"):
        return True
    # Starts with "not " (catches "not met", "not performed timely", etc.)
    if v.startswith("not "):
        return True
    # Starts with "fail" (catches "failed - reason", "failure noted")
    if v.startswith("fail"):
        return True
    # "non-" prefix (catches "non-compliant", "non-conforming")
    if v.startswith("non-") or v.startswith("non "):
        return True
    # "un" prefix (catches "unmet", "unverified", "unauthorized", "unapproved")
    if v.startswith("un"):
        return True
    # "partial" prefix (catches "partially met", "partial compliance")
    if v.startswith("partial"):
        return True
    # "missing" / "absent" prefix
    if v.startswith("missing") or v.startswith("absent"):
        return True
    # "delay" prefix (catches "delayed but completed", "delayed execution")
    # Audit standard: late/delayed execution is ALWAYS a deviation regardless of eventual completion
    if v.startswith("delay"):
        return True
    # "untimely" prefix
    if v.startswith("untimely"):
        return True
    # "exception" / "deviation" / "deficiency" anywhere (catches "exception: late approval")
    if any(w in v for w in ("exception", "deviation", "deficiency")):
        return True
    # "unable" prefix (catches "unable to verify documentation")
    if v.startswith("unable"):
        return True
    # "could not" / "cannot" prefix
    if v.startswith("could not") or v.startswith("cannot") or v.startswith("can't"):
        return True
    return False


# -- Neutral values (N/A, blank, unable to assess) --
_NEUTRAL_VALUES = frozenset({
    "n/a", "na", "not applicable", "unable to assess", "",
})


def _is_neutral(val: str) -> bool:
    """Check if a value is neutral (N/A, blank, unable to assess)."""
    if not val or not val.strip():
        return True
    return val.strip().lower().rstrip(".") in _NEUTRAL_VALUES


def _is_attr_fail(val: str) -> bool:
    """
    Conservative attribute failure check per audit testing methodology.

    For attribute results specifically, we apply the conservative principle:
    if the LLM returned a value that is NOT clearly positive and NOT neutral,
    it must be treated as a failure. This catches ALL unknown/ambiguous LLM
    responses that _is_negative might miss.

    Why: In control testing, an attribute either passed or it didn't. There is no
    middle ground. If the LLM can't clearly say "Yes"/"Met"/"Compliant", the
    auditor must flag it. Silent passes on ambiguous data = audit failure.

    Returns True if the attribute should be counted as FAILED.
    """
    if not val or not val.strip():
        return False  # Blank = not evaluated, not a failure
    v = val.strip().lower().rstrip(".")
    # Neutral values are NOT failures (N/A means control didn't apply)
    if v in _NEUTRAL_VALUES:
        return False
    # If it's clearly positive, it's not a failure
    if _is_positive(val):
        return False
    # If it's clearly negative, it IS a failure
    if _is_negative(val):
        return True
    # CONSERVATIVE CATCH-ALL: Value is not positive, not neutral, not recognized
    # Per audit standard: ambiguous attribute = flag as failure for auditor review
    # Examples caught here: any future LLM response we haven't seen yet
    return True


def _is_positive(val: str) -> bool:
    """
    Robust check for positive/passing values from LLM output.
    Catches: "Yes", "Yes.", "YES", "Y", "Yes - verified in SAP",
             "Met", "Passed", "Compliant", "Effective", "True", etc.
    """
    if not val or not val.strip():
        return False
    v = val.strip().lower().rstrip(".")
    # Exact positive matches
    if v in ("yes", "y", "true", "pass", "passed", "ok", "effective",
             "met", "compliant", "satisfactory", "adequate", "sufficient",
             "complete", "confirmed", "verified", "approved"):
        return True
    # Starts with "yes " or "yes," (LLM adds explanations: "Yes - verified")
    if v.startswith("yes ") or v.startswith("yes,") or v.startswith("yes-") or v.startswith("yes;"):
        return True
    return False

def _strip_pass_language(text: str) -> str:
    """Remove pass-indicating phrases from text to prevent contradictions.
    Used when FAIL sample text incorrectly contains pass language."""
    if not text:
        return text
    result = text
    for phrase in _PASS_INDICATORS:
        result = _re.sub(_re.escape(phrase), "control deviation noted", result, flags=_re.IGNORECASE)
    # Also strip raw result tags
    result = _re.sub(r'\[FAIL\]|\[PASS\]|\[OK\]|\[ERROR\]', '', result, flags=_re.IGNORECASE)
    return result.strip()


def _sanitize_remarks(remarks: str, is_fail: bool, deviation: str) -> str:
    """
    Clean up LLM remarks to prevent contradictions in the workpaper.
    
    Problems this solves:
    - LLM puts literal "[FAIL]" or "[PASS]" tags in remarks text
    - LLM says "control is effective" but sample is actually FAIL
    - LLM says "control failed" but sample is actually PASS
    - Deviation text copied from contradictory remarks still has pass language
    """
    if not remarks:
        return remarks
    
    # Step 1: Strip any raw result tags the LLM embedded
    cleaned = _re.sub(r'\[FAIL\]|\[PASS\]|\[OK\]|\[ERROR\]', '', remarks, flags=_re.IGNORECASE).strip()
    
    # Step 2: If sample is FAIL but remarks contain pass-indicating language,
    # replace with deviation details (the contradiction confuses reviewers)
    if is_fail:
        lower = cleaned.lower()
        has_contradiction = any(phrase in lower for phrase in _PASS_INDICATORS)
        if has_contradiction:
            # Use deviation details as the primary remark instead
            # Also strip any tags from the deviation text
            clean_dev = _re.sub(r'\[FAIL\]|\[PASS\]|\[OK\]|\[ERROR\]', '', deviation or '',
                                flags=_re.IGNORECASE).strip()
            if clean_dev and clean_dev.lower() not in ("none", "n/a", "see remarks", ""):
                # CRITICAL: Check if deviation text ALSO contains pass language.
                # This happens when Rule 5 copies remarks->deviation before sanitization.
                dev_lower = clean_dev.lower()
                dev_has_contradiction = any(phrase in dev_lower for phrase in _PASS_INDICATORS)
                if dev_has_contradiction:
                    # Deviation text is ALSO contradictory — strip pass phrases from it
                    clean_dev = _strip_pass_language(clean_dev)
                    if clean_dev and clean_dev.lower() not in ("none", "n/a", "see remarks", ""):
                        cleaned = f"Deviation: {clean_dev}"
                    else:
                        cleaned = "Control did not operate effectively for this sample -- see attribute results."
                else:
                    cleaned = f"Deviation: {clean_dev}"
            else:
                cleaned = "Control did not operate effectively for this sample -- see attribute results."
    
    return cleaned


def validate_toe_sample(eval_result: dict) -> dict:
    """
    Validate LLM output for LOGICAL CONSISTENCY ONLY.

    ┌────────────────────────────────────────────────────────────────────┐
    │  PHILOSOPHY: The LLM is the auditor. It saw the control, risk,   │
    │  evidence, and attributes. Its judgment is PRIMARY.               │
    │                                                                    │
    │  This function ONLY catches HARD LOGICAL CONTRADICTIONS:           │
    │  - LLM said attribute = "No" but result = "PASS"                  │
    │  - LLM said standard field = "No" but result = "PASS"             │
    │  - LLM said result = "ERROR"                                       │
    │  - LLM said result = "FAIL" but gave no deviation details          │
    │                                                                    │
    │  This function does NOT re-interpret ambiguous values.             │
    │  If the LLM says "Partially Compliant" and result = "PASS",       │
    │  the LLM made a professional judgment — we trust it.              │
    └────────────────────────────────────────────────────────────────────┘
    """
    r = dict(eval_result)
    corrections = []

    result = r.get("result", "").strip()
    deviation = r.get("deviation_details", "").strip()
    _dev_lower = deviation.lower().rstrip(".")
    deviation_empty = _dev_lower in ("", "none", "n/a", "none.", "no deviation",
                                      "no deviations", "none identified",
                                      "no deviation identified", "no deviations identified",
                                      "no deviation found", "no deviations found",
                                      "no exceptions", "no exceptions noted",
                                      "no issues", "no issues identified",
                                      "not applicable", "see remarks")
    # Also treat deviation as empty if it contains PASS-indicating language
    if not deviation_empty and _dev_lower:
        if any(phrase in _dev_lower for phrase in _PASS_INDICATORS):
            deviation_empty = True

    # ── Rule 0: ERROR result → FAIL ─────────────────────────────────
    if result == "ERROR":
        r["result"] = "FAIL"
        r["operated_effectively"] = "No"
        if deviation_empty:
            r["deviation_details"] = "API evaluation error -- manual review required"
        corrections.append("CORRECTED: ERROR result converted to FAIL for manual review")

    # ── Rule 1: Standard field explicitly negative + result PASS → contradiction ──
    # These fields should be strict Yes/No. If the LLM said "No" for any of them
    # but still marked PASS, that's a logical contradiction we must fix.
    _STANDARD_FIELDS = [
        ("control_performed", "control not performed"),
        ("evidence_sufficient", "insufficient evidence"),
        ("timely_execution", "control not timely"),
        ("accurate_execution", "inaccurate execution"),
        ("authorized_performer", "unauthorized performer"),
        ("operated_effectively", "operated_effectively=No"),
    ]
    for field_name, reason in _STANDARD_FIELDS:
        field_val = r.get(field_name, "").strip()
        if _is_negative(field_val) and r.get("result") != "FAIL":
            corrections.append(f"CORRECTED: result {r.get('result')}->FAIL ({reason})")
            r["result"] = "FAIL"
            r["operated_effectively"] = "No"

    # ── Rule 2: Attribute explicitly "No" + result PASS → contradiction ──
    # The prompt forces strict Yes/No/N/A. If the LLM answered "No" for any
    # attribute but still said PASS, that's a contradiction.
    # For NON-STANDARD values (e.g. "Partially"), we normalize first.
    attr_results = r.get("attribute_results")
    failing_attrs = []
    if attr_results and isinstance(attr_results, dict):
        for aid, val in attr_results.items():
            normalized = _normalize_attr_value(str(val))
            if normalized == "No":
                failing_attrs.append((aid, str(val)))
    if failing_attrs and r.get("result") != "FAIL":
        attr_desc = "; ".join(f"Attribute {a}={v}" for a, v in failing_attrs)
        corrections.append(f"CORRECTED: result {r.get('result')}->FAIL ({attr_desc})")
        r["result"] = "FAIL"
        r["operated_effectively"] = "No"
        r["deviation_details"] = attr_desc

    # ── Rule 3: PASS → ensure operated=Yes, no contradictory deviation ──
    if r.get("result") == "PASS":
        if r.get("operated_effectively") != "Yes":
            corrections.append(f"CORRECTED: operated {r.get('operated_effectively')}->Yes (PASS)")
            r["operated_effectively"] = "Yes"
        # Re-check deviation after possible corrections
        deviation = r.get("deviation_details", "").strip()
        _dev_chk = deviation.lower().rstrip(".")
        dev_empty = _dev_chk in ("", "none", "n/a", "none.", "no deviation", "no deviations",
                                   "none identified", "not applicable", "see remarks")
        if not dev_empty and deviation:
            if any(phrase in _dev_chk for phrase in _PASS_INDICATORS):
                dev_empty = True
        if not dev_empty:
            # Has meaningful deviation text but marked PASS → flip to FAIL
            corrections.append(f"CORRECTED: result PASS->FAIL (deviation found: {deviation[:60]})")
            r["result"] = "FAIL"
            r["operated_effectively"] = "No"

    # ── Rule 5: ALL attrs Yes/N/A + ALL std fields positive + result FAIL ──
    # This catches a HARD LOGICAL CONTRADICTION: the LLM's detailed attribute
    # assessments all say the control is fine, but it still said FAIL.
    # This happens when the LLM is confused about conditional attributes
    # (e.g., "Discrepancy Logged = N/A" but thinks that means FAIL).
    # Trust the DETAILED assessment (attributes) over the SUMMARY (result).
    # NOTE: We exclude operated_effectively from the check because it always
    # mirrors the result field — it's not an independent evidence-based field.

    # ── Rule 6 (pre-pass for Rule 5): Fix conditional attrs marked "No" that
    # should be "N/A" based on the LLM's OWN reasoning confirming the
    # triggering condition didn't arise. ──
    # Example: LLM marks "Discrepancy Logged = No" with reasoning
    # "no discrepancies were found, all items matched" → should be N/A.
    # But if reasoning says "despite shortfall of 15 units" → keep No.
    _CONDITIONAL_ATTR_KEYWORDS = (
        "discrepanc", "variance", "escalat", "exception", "resolution",
        "deviation", "mismatch", "dispute",
    )
    _CONDITION_NOT_ARISE_PHRASES = (
        "no discrepanc", "no variance", "no issue", "no exception",
        "no mismatch", "everything matched", "all items matched",
        "all matched", "matched the po", "matched po", "zero variance",
        "zero discrepanc", "none were found", "none were identified",
        "none existed", "not applicable", "inspection found no",
        "quality passed", "qty confirmed", "no deviation",
    )
    _REAL_DEVIATION_PHRASES = (
        "despite", "shortfall", "short by", "defective", "but not logged",
        "was identified but", "was found but", "existed but", "occurred but",
        "were identified but", "gap of", "missing",
    )
    
    attr_results_current = r.get("attribute_results")
    attr_reasoning = r.get("attribute_reasoning")
    if (attr_results_current and isinstance(attr_results_current, dict) and
        attr_reasoning and isinstance(attr_reasoning, dict)):
        for aid, val in list(attr_results_current.items()):
            if _normalize_attr_value(str(val)) != "No":
                continue
            reason = str(attr_reasoning.get(aid, "")).lower()
            aid_lower = aid.lower()
            
            # Check if this attr is a conditional type (by name or reasoning keywords)
            is_conditional = any(kw in aid_lower or kw in reason for kw in _CONDITIONAL_ATTR_KEYWORDS)
            if not is_conditional:
                continue
            
            # Check if reasoning says the condition didn't arise
            condition_absent = any(phrase in reason for phrase in _CONDITION_NOT_ARISE_PHRASES)
            if not condition_absent:
                continue
            
            # Safety check: make sure it's NOT a real deviation that went unhandled
            real_deviation = any(phrase in reason for phrase in _REAL_DEVIATION_PHRASES)
            if real_deviation:
                continue
            
            # Override No → N/A
            attr_results_current[aid] = "N/A"
            corrections.append(
                f"CORRECTED: Attr {aid} No->N/A (LLM reasoning confirms condition "
                f"didn't arise: '{reason[:60]}...')"
            )
    
    # Now run Rule 5 with potentially corrected attributes
    if r.get("result") == "FAIL":
        all_attrs_ok = True
        any_evidence_field_negative = False
        
        # Check attributes (using post-Rule-6 corrected values)
        attr_results_check = r.get("attribute_results")
        if attr_results_check and isinstance(attr_results_check, dict):
            for aid, val in attr_results_check.items():
                normalized = _normalize_attr_value(str(val))
                if normalized == "No":
                    all_attrs_ok = False
                    break
        
        # Check EVIDENCE-BASED standard fields only (not operated_effectively which mirrors result)
        _EVIDENCE_FIELDS = [
            ("control_performed", "control not performed"),
            ("evidence_sufficient", "insufficient evidence"),
            ("timely_execution", "control not timely"),
            ("accurate_execution", "inaccurate execution"),
            ("authorized_performer", "unauthorized performer"),
        ]
        for field_name, reason in _EVIDENCE_FIELDS:
            field_val = r.get(field_name, "").strip()
            if _is_negative(field_val):
                any_evidence_field_negative = True
                break
        
        if all_attrs_ok and not any_evidence_field_negative:
            corrections.append(
                "CORRECTED: result FAIL->PASS (all attributes Yes/N/A, all evidence fields positive — "
                "LLM's detailed assessment contradicts its summary result)"
            )
            r["result"] = "PASS"
            r["operated_effectively"] = "Yes"
            r["deviation_details"] = "None"
            
            # Build clean PASS remarks from attribute reasoning instead of keeping
            # the LLM's confused FAIL text (which says things like "control did not
            # operate effectively because discrepancy logging was N/A")
            attr_reasoning = r.get("attribute_reasoning")
            attr_results_final = r.get("attribute_results")
            if attr_reasoning and isinstance(attr_reasoning, dict) and attr_results_final:
                # Extract the "Yes" attribute reasoning as the positive PASS narrative
                yes_parts = []
                na_parts = []
                for aid, val in attr_results_final.items():
                    norm = _normalize_attr_value(str(val))
                    reason = attr_reasoning.get(aid, "")
                    if not reason:
                        continue
                    reason = reason.rstrip(".")
                    if norm == "Yes":
                        yes_parts.append(reason)
                    elif norm == "N/A":
                        na_parts.append(reason)
                
                if yes_parts:
                    pass_rmk = ". ".join(yes_parts[:2])  # Use first 2 Yes reasons
                    if na_parts:
                        pass_rmk += ". " + na_parts[0]  # Add first N/A context
                    r["remarks"] = pass_rmk + ". Control operated effectively."
                else:
                    r["remarks"] = "Control operated effectively for this transaction. All attributes satisfied or not applicable."
            else:
                r["remarks"] = "Control operated effectively for this transaction. All attributes satisfied or not applicable."

    # ── Rule 4: FAIL → ensure operated=No, must have deviation ──
    if r.get("result") == "FAIL":
        if r.get("operated_effectively") != "No":
            corrections.append(f"CORRECTED: operated {r.get('operated_effectively')}->No (FAIL)")
            r["operated_effectively"] = "No"
        # Ensure we have deviation details
        deviation = r.get("deviation_details", "").strip()
        _dev_chk = deviation.lower().rstrip(".")
        dev_empty = _dev_chk in ("", "none", "n/a", "none.", "no deviation", "no deviations",
                                   "none identified", "not applicable", "see remarks")
        if not dev_empty and _dev_chk:
            if any(phrase in _dev_chk for phrase in _PASS_INDICATORS):
                dev_empty = True
        if dev_empty:
            # FAIL with no deviation details → pull from remarks or attribute_reasoning
            attr_reasoning = r.get("attribute_reasoning")
            if failing_attrs and attr_reasoning and isinstance(attr_reasoning, dict):
                # Use LLM's own reasoning for failing attributes
                reasons = []
                for aid, _ in failing_attrs:
                    reason_text = attr_reasoning.get(aid, "")
                    if reason_text:
                        reasons.append(f"Attr {aid}: {reason_text}")
                if reasons:
                    r["deviation_details"] = "; ".join(reasons)
                    corrections.append("CORRECTED: built deviation from LLM attribute_reasoning")
            if r.get("deviation_details", "").strip().lower() in ("", "none", "n/a"):
                # Fallback to remarks
                remarks_text = r.get("remarks", "").strip()
                clean_rmk = _re.sub(r'\[FAIL\]|\[PASS\]|\[OK\]|\[ERROR\]', '', remarks_text,
                                    flags=_re.IGNORECASE).strip()
                if clean_rmk and clean_rmk.lower() not in ("", "none", "n/a"):
                    clean_rmk = _strip_pass_language(clean_rmk)
                    if clean_rmk and clean_rmk.lower() not in ("", "none", "n/a"):
                        r["deviation_details"] = clean_rmk
                        corrections.append("CORRECTED: copied sanitized remarks into deviation_details")
                    else:
                        r["deviation_details"] = "Deviation identified -- manual review required for details"
                        corrections.append("CORRECTED: added deviation_details placeholder")
                else:
                    r["deviation_details"] = "Deviation identified -- manual review required for details"
                    corrections.append("CORRECTED: added deviation_details placeholder")

    # Sanitize remarks for contradictions
    final_is_fail = r.get("result") == "FAIL"
    final_deviation = r.get("deviation_details", "")
    r["remarks"] = _sanitize_remarks(r.get("remarks", ""), final_is_fail, final_deviation)

    if corrections:
        correction_note = " [AUTO-CORRECTED: " + "; ".join(corrections) + "]"
        r["remarks"] = r.get("remarks", "") + correction_note

    return r


def _enforce_schema_output_quality(eval_result: dict, schema: ControlSchema | None, source_document: str = "") -> dict:
    """Enforce schema/output consistency to prevent blank PASS rows in workpapers."""
    r = dict(eval_result or {})

    if not isinstance(r.get("attribute_results"), dict):
        r["attribute_results"] = {}
    if not isinstance(r.get("attribute_reasoning"), dict):
        r["attribute_reasoning"] = {}
    if not isinstance(r.get("sample_details"), dict):
        r["sample_details"] = {}

    attr_results = {str(k): v for k, v in r.get("attribute_results", {}).items()}
    attr_reasoning = {str(k): v for k, v in r.get("attribute_reasoning", {}).items()}
    sample_details = {str(k): v for k, v in r.get("sample_details", {}).items()}

    missing_attr_ids = []
    if schema and schema.attributes:
        by_name = {str(k).strip().lower(): v for k, v in attr_results.items()}

        for a in schema.attributes:
            aid = str(a.get("id", "")).strip()
            aname = str(a.get("name", "")).strip().lower()
            if not aid:
                continue

            val = attr_results.get(aid)
            if val in (None, "") and aname:
                val = by_name.get(aname)

            if val in (None, ""):
                missing_attr_ids.append(aid)
                attr_results[aid] = "No"
                attr_reasoning[aid] = "Missing attribute assessment in model output; treated as No for audit safety."
            else:
                norm = _normalize_attr_value(str(val))
                if norm in ("Yes", "No", "N/A"):
                    attr_results[aid] = norm
                else:
                    attr_results[aid] = "No"
                    existing = str(attr_reasoning.get(aid, "")).strip()
                    attr_reasoning[aid] = (
                        (existing + " ") if existing else ""
                    ) + "Ambiguous attribute value normalized to No for audit safety."

    all_detail_missing = False
    if schema and schema.sample_columns:
        ci_lookup = {k.strip().lower(): v for k, v in sample_details.items()}
        normalized_details = {}
        for col in schema.sample_columns:
            hdr = str(col.get("header", "")).strip()
            key = str(col.get("key", "")).strip()
            val = sample_details.get(hdr)
            if val in (None, "") and key:
                val = sample_details.get(key)
            if val in (None, "") and hdr:
                val = ci_lookup.get(hdr.lower())
            if val in (None, "") and key:
                val = ci_lookup.get(key.lower())
            if val in (None, "") and (hdr.lower() == "evidence files reviewed" or key.lower() == "evidence_files_reviewed"):
                val = source_document or "Not extracted from evidence"
            normalized_details[hdr] = val if val not in (None, "") else "Not extracted from evidence"

        sample_details = normalized_details
        all_detail_missing = all(str(v).strip() in ("", "Not extracted from evidence") for v in sample_details.values())

    all_na = False
    if schema and schema.attributes and attr_results:
        vals = [
            _normalize_attr_value(str(attr_results.get(str(a.get("id", "")).strip(), "")))
            for a in schema.attributes if str(a.get("id", "")).strip()
        ]
        vals = [v for v in vals if v != "AMBIGUOUS"]
        all_na = bool(vals) and all(v == "N/A" for v in vals)

    force_fail_reasons = []
    if missing_attr_ids:
        force_fail_reasons.append(f"missing attribute assessments: {', '.join(missing_attr_ids)}")
    if all_detail_missing and schema and schema.sample_columns:
        force_fail_reasons.append("sample detail extraction is blank for all configured columns")
    if all_na and str(r.get("result", "")).upper() == "PASS":
        force_fail_reasons.append("all attributes are N/A; cannot conclude effective operation from this sample")

    if force_fail_reasons:
        r["result"] = "FAIL"
        r["operated_effectively"] = "No"
        r["evidence_sufficient"] = "No"
        r["deviation_details"] = "; ".join(force_fail_reasons)
        base_rmk = str(r.get("remarks", "")).strip()
        if base_rmk:
            base_rmk += " "
        r["remarks"] = base_rmk + "[AUTO-CORRECTED: output quality guardrail triggered]"

    r["attribute_results"] = attr_results
    r["attribute_reasoning"] = attr_reasoning
    r["sample_details"] = sample_details
    return r


def _normalize_attr_value(val: str) -> str:
    """
    Normalize LLM attribute value to strict Yes/No/N/A.

    The prompt forces strict Yes/No/N/A but LLMs sometimes add qualifiers.
    This normalizes: "Yes - verified" → "Yes", "No - missing" → "No", etc.

    Returns:
      "Yes"       - if value is clearly affirmative
      "No"        - if value is clearly negative
      "N/A"       - if value is not-applicable
      "AMBIGUOUS" - if value doesn't clearly fit → trust the LLM's overall result
    """
    if not val or not val.strip():
        return "AMBIGUOUS"
    v = val.strip().lower().rstrip(".")

    # N/A variants
    if v in ("n/a", "na", "not applicable"):
        return "N/A"

    # Explicit No (with possible explanation appended)
    if v == "no" or v.startswith("no ") or v.startswith("no,") or v.startswith("no-") or v.startswith("no;"):
        return "No"

    # Explicit Yes (with possible explanation appended)
    if v == "yes" or v.startswith("yes ") or v.startswith("yes,") or v.startswith("yes-") or v.startswith("yes;"):
        return "Yes"

    # "Not" prefix → No
    if v.startswith("not ") or v.startswith("not-"):
        return "No"

    # "Fail" / "failed" → No
    if v.startswith("fail"):
        return "No"

    # "Missing" / "absent" → No
    if v.startswith("missing") or v.startswith("absent"):
        return "No"

    # For anything else (e.g., "Partially Compliant", "Mostly", "Acceptable"):
    # The LLM had the full context and decided on an overall result.
    # We trust the LLM's holistic judgment rather than regex-interpreting this value.
    return "AMBIGUOUS"



# ===========================================================================
#  TOE AGGREGATION LOGIC
# ===========================================================================

def is_sample_fail(sr) -> bool:
    """
    ┌──────────────────────────────────────────────────────────────┐
    │  SINGLE SOURCE OF TRUTH for whether a sample is a failure.  │
    │  EVERY pass/fail decision in the codebase MUST call this.   │
    │  Nothing else should independently decide pass vs fail.     │
    └──────────────────────────────────────────────────────────────┘

    LLM-TRUSTING ARCHITECTURE:
      The LLM is the primary decision maker. It saw the control, risk,
      evidence, and attributes together in one call. Its result field
      is the authoritative judgment.

    A sample FAILS if ANY of these conditions is true:
      1. The LLM/validator marked result as "FAIL" or "ERROR"
         (This is the PRIMARY check — the LLM's judgment)
      2. Any attribute normalizes to "No" (safety net for contradictions
         the validator might have missed)
      3. Any standard field (control_performed, timely_execution, etc.)
         is explicitly negative (safety net)

    A sample does NOT fail just because an attribute has an ambiguous value
    like "Partially Compliant" — the LLM considered that in context.
    """
    # PRIMARY: Trust the LLM's result field
    if sr.result in ("FAIL", "ERROR"):
        return True

    # SAFETY NET: attribute explicitly "No" but result somehow still "PASS"
    if sr.attribute_results and isinstance(sr.attribute_results, dict):
        for val in sr.attribute_results.values():
            if _normalize_attr_value(str(val)) == "No":
                return True

    # SAFETY NET: standard field explicitly negative
    for field_name in ("operated_effectively", "control_performed", "timely_execution",
                       "accurate_execution", "authorized_performer",
                       "evidence_sufficient"):
        field_val = getattr(sr, field_name, "")
        if _is_negative(str(field_val)):
            return True

    return False


def verify_sample_integrity(sr) -> list:
    """
    Integrity check: verify that sr.result agrees with is_sample_fail().

    If the validator ran correctly, sr.result=="FAIL" should ALWAYS mean
    is_sample_fail()==True, and sr.result=="PASS" should mean False.
    Any disagreement is a data integrity issue.

    Returns list of error strings (empty = all good).
    """
    errors = []
    actual_fail = is_sample_fail(sr)

    # Case 1: result says PASS but is_sample_fail says FAIL
    #   This means the validator MISSED a failure. Critical bug.
    if sr.result == "PASS" and actual_fail:
        reasons = []
        if sr.attribute_results and isinstance(sr.attribute_results, dict):
            for aid, val in sr.attribute_results.items():
                if _normalize_attr_value(str(val)) == "No":
                    reasons.append(f"attr {aid}={val}")
        for fn in ("operated_effectively", "control_performed", "timely_execution",
                    "accurate_execution", "authorized_performer", "evidence_sufficient"):
            fv = getattr(sr, fn, "")
            if _is_negative(str(fv)):
                reasons.append(f"{fn}={fv}")
        errors.append(
            f"INTEGRITY VIOLATION [{sr.control_id}/{sr.sample_id}]: "
            f"result='PASS' but is_sample_fail()=True. "
            f"Validator missed: {'; '.join(reasons)}. "
            f"This sample will be counted as FAIL (safety net caught it)."
        )

    # Case 2: result says FAIL but is_sample_fail says PASS
    #   This should be impossible — is_sample_fail checks sr.result first.
    #   If this happens, there's a code logic error.
    if sr.result == "FAIL" and not actual_fail:
        errors.append(
            f"LOGIC ERROR [{sr.control_id}/{sr.sample_id}]: "
            f"result='FAIL' but is_sample_fail()=False. "
            f"This should be impossible — investigate code logic."
        )

    return errors


def verify_control_integrity(r) -> list:
    """
    Integrity check: verify that aggregate counts match is_sample_fail() recount.

    The aggregate stores r.failed_samples (computed at aggregation time).
    This function re-counts by calling is_sample_fail() on each sample
    and verifies the counts match.

    Returns list of error strings (empty = all good).
    """
    errors = []

    # Recount using is_sample_fail
    recount_fail = sum(1 for sr in r.sample_results if is_sample_fail(sr))
    recount_pass = sum(1 for sr in r.sample_results if not is_sample_fail(sr))

    if recount_fail != r.failed_samples:
        errors.append(
            f"COUNT MISMATCH [{r.control_id}]: "
            f"aggregate says {r.failed_samples} failed, "
            f"recount says {recount_fail} failed. "
            f"The summary sheet will show {r.failed_samples} but "
            f"individual sheet will have {recount_fail} [FAIL] markers."
        )

    if recount_pass != r.passed_samples:
        errors.append(
            f"COUNT MISMATCH [{r.control_id}]: "
            f"aggregate says {r.passed_samples} passed, "
            f"recount says {recount_pass} passed."
        )

    if recount_fail + recount_pass != r.total_samples:
        errors.append(
            f"TOTAL MISMATCH [{r.control_id}]: "
            f"recount_fail({recount_fail}) + recount_pass({recount_pass}) "
            f"!= total_samples({r.total_samples})"
        )

    # Verify each sample's integrity too
    for sr in r.sample_results:
        errors.extend(verify_sample_integrity(sr))

    return errors

def aggregate_toe_results(
    control_id: str,
    rcm: RCMRow,
    sample_results: list,    # list[TOESampleResult]
) -> TOEControlResult:
    """
    Aggregate individual sample results into a control-level TOE conclusion.

    IMPORTANT: Before counting, this function verifies that sr.result agrees
    with is_sample_fail() for every sample. If the validator missed a failure,
    the safety net in is_sample_fail() catches it, and we CORRECT sr.result
    here so that the stored field, the counting, and the sheet markers all
    agree. This eliminates the "two truth sources" problem.

    Deficiency classification by deviation rate:
      0 deviations                                                -> Effective / None
      >0% and <= CONTROL_DEFICIENCY_MAX_RATE (default 10%)        -> Control Deficiency
      >10% and <= SIGNIFICANT_DEFICIENCY_MAX_RATE (default 20%)   -> Significant Deficiency
      >SIGNIFICANT_DEFICIENCY_MAX_RATE (default 20%)              -> Material Weakness
      All samples fail                                            -> Material Weakness
    """
    # ── INTEGRITY GUARANTEE ──────────────────────────────────────
    # Ensure sr.result agrees with is_sample_fail() for EVERY sample.
    # If is_sample_fail() says FAIL but sr.result says PASS, that means
    # the validator missed something (the safety net caught it).
    # We CORRECT sr.result here so the stored field matches the truth.
    for s in sample_results:
        actual_fail = is_sample_fail(s)
        if actual_fail and s.result != "FAIL":
            # Safety net caught what the validator missed — fix it
            s.result = "FAIL"
            s.operated_effectively = "No"
            if not s.remarks:
                s.remarks = ""
            s.remarks += " [AUTO-CORRECTED at aggregation: is_sample_fail() safety net]"
        elif not actual_fail and s.result == "FAIL":
            # This should be impossible — but if it happens, trust is_sample_fail
            # (is_sample_fail checks sr.result first, so this case can't arise
            # unless someone modified sr.result after validation)
            pass  # Leave as FAIL — conservative
    # ─────────────────────────────────────────────────────────────

    total = len(sample_results)
    passed = sum(1 for s in sample_results if not is_sample_fail(s))
    failed = total - passed
    deviation_rate = failed / total if total > 0 else 0.0

    # Collect deviation details
    deviations = [s.deviation_details for s in sample_results
                  if is_sample_fail(s) and s.deviation_details
                  and s.deviation_details.lower() not in ("none", "n/a", "see remarks")]

    # -- Determine overall effectiveness and deficiency ----------------
    if total == 0:
        # Cannot conclude effectiveness without testing any samples
        effectiveness = "Not Effective"
        deficiency = "Significant Deficiency"
        overall_remarks = (
            "NO SAMPLES TESTED -- cannot conclude on operating effectiveness. "
            "Auditor must obtain and test required samples before concluding."
        )

    elif failed == 0:
        effectiveness = "Effective"
        deficiency = "None"
        overall_remarks = (
            f"Control operated effectively across all {total} samples tested. "
            f"No deviations identified. Operating effectiveness confirmed."
        )

    elif deviation_rate <= CONTROL_DEFICIENCY_MAX_RATE:
        # 0% < rate <= threshold  ->  Control Deficiency
        effectiveness = "Effective with Exceptions"
        deficiency = "Control Deficiency"
        dev_text = deviations[0] if deviations else "See sample details"
        overall_remarks = (
            f"{failed} deviation(s) in {total} samples tested (rate: {deviation_rate:.1%}). "
            f"Deviation appears isolated and does not indicate a systematic failure. "
            f"Deviation: {'; '.join(deviations[:3]) if deviations else dev_text}"
        )

    elif deviation_rate <= SIGNIFICANT_DEFICIENCY_MAX_RATE:
        # threshold < rate <= threshold  ->  Significant Deficiency
        effectiveness = "Not Effective"
        deficiency = "Significant Deficiency"
        overall_remarks = (
            f"{failed} deviation(s) in {total} samples tested (rate: {deviation_rate:.1%}). "
            f"Multiple deviations indicate the control is not operating effectively. "
            f"Deviations: {'; '.join(deviations[:3]) if deviations else 'See sample details'}"
        )

    else:
        # rate > threshold  ->  Material Weakness
        effectiveness = "Not Effective"
        deficiency = "Material Weakness"
        overall_remarks = (
            f"{failed} deviation(s) in {total} samples tested (rate: {deviation_rate:.1%}). "
            f"High deviation rate indicates a pervasive failure in control operation. "
            f"Deviations: {'; '.join(deviations[:3]) if deviations else 'See sample details'}"
        )

    # Special case: all samples fail
    if failed == total and total > 0:
        deficiency = "Material Weakness"
        overall_remarks = (
            f"ALL {total} samples failed. Control is not operating at all. "
            f"This represents a material weakness requiring immediate remediation. "
            f"Deviations: {'; '.join(deviations[:3]) if deviations else 'See sample details'}"
        )

    # Special context: automated control with failures
    # Automated controls are binary -- if they work once, they work
    # every time (assuming effective ITGCs). ANY failure in an automated control
    # suggests a system change or ITGC issue -- escalate severity accordingly.
    if rcm.control_type.strip().lower() == "automated" and failed > 0:
        if deficiency in ("None", "Control Deficiency"):
            # Automated failure = at minimum Significant Deficiency
            deficiency = "Significant Deficiency"
            effectiveness = "Not Effective"
            overall_remarks = (
                f"{failed} deviation(s) in {total} automated control samples "
                f"(rate: {deviation_rate:.1%}). "
                f"Automated controls are expected to operate consistently. "
                f"Any failure indicates a potential system configuration change or "
                f"ITGC deficiency during the period. "
                f"Verify change management and access controls (ITGCs). "
                f"Deviations: {'; '.join(deviations[:3]) if deviations else 'See sample details'}"
            )
        else:
            overall_remarks += (
                f" NOTE: This is an automated control -- verify ITGC (change management, "
                f"access controls) to determine if system configuration was altered during the period."
            )

    # Warn if any samples have ALL attributes as N/A (shouldn't be in population)
    na_samples = []
    for s in sample_results:
        if s.attribute_results and isinstance(s.attribute_results, dict):
            vals = [str(v).strip().lower() for v in s.attribute_results.values()]
            if vals and all(v in ("n/a", "na", "not applicable", "") for v in vals):
                na_samples.append(s.sample_id)
    if na_samples:
        overall_remarks += (
            f" WARNING: {len(na_samples)} sample(s) had all attributes marked N/A "
            f"({', '.join(na_samples[:3])}{'...' if len(na_samples) > 3 else ''}). "
            f"These transactions may not belong in the test population -- "
            f"auditor should verify sampling methodology."
        )

    return TOEControlResult(
        control_id=control_id,
        risk_id=rcm.risk_id,
        control_type=rcm.control_type,
        nature_of_control=rcm.nature_of_control,
        control_frequency=rcm.control_frequency,
        total_samples=total,
        passed_samples=passed,
        failed_samples=failed,
        deviation_rate=deviation_rate,
        operating_effectiveness=effectiveness,
        deficiency_type=deficiency,
        overall_remarks=overall_remarks,
        sample_results=sample_results,
        evaluation_timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
    )



# ===========================================================================
#  AZURE OPENAI CALLER
# ===========================================================================

def _fix_schema_attributes(schema: ControlSchema) -> None:
    """Post-process LLM-generated schema to fix known-bad attribute descriptions.

    Even with strong prompt guidance, the schema LLM sometimes generates attribute
    descriptions that cause the evaluation LLM to false-positive. This function
    patches the two most common patterns:

    1. FINANCIAL SCREENING: Description requires a literal "financial health review
       report" when D&B scores, credit reports, etc. are equivalent evidence.
    2. CONDITIONAL ACTIONS: Description fails to state when N/A applies (e.g.,
       "Discrepancy Logged" without mentioning N/A when no discrepancies exist).
    """
    if not schema or not schema.attributes:
        return

    for attr in schema.attributes:
        desc = attr.get("description", "")
        name = attr.get("name", "")
        desc_lower = desc.lower()
        name_lower = name.lower()

        # --- Fix 1: Financial screening attributes ---
        # If attribute mentions "financial health review" or "financial review"
        # but doesn't mention D&B, inject the equivalence
        if ("financial" in desc_lower and ("review" in desc_lower or "health" in desc_lower)
                and "d&b" not in desc_lower and "dun" not in desc_lower
                and "credit report" not in desc_lower):
            # Append D&B equivalence guidance
            if not desc.endswith("."):
                desc += "."
            desc += (" A D&B score, credit report, or financial stability assessment"
                     " counts as a completed financial health review."
                     " Do NOT require a separate document called 'financial health review report'."
                     " HOWEVER: If the D&B score is below the policy threshold (e.g., score 32"
                     " when policy requires >60), mark as 'No' — the review was done but FAILED.")
            attr["description"] = desc

        # --- Fix 2: Conditional attributes missing N/A guidance ---
        conditional_keywords = [
            "discrepanc", "variance", "escalat", "exception", "resolution",
            "mismatch", "deviation", "dispute",
        ]
        is_conditional = any(kw in name_lower or kw in desc_lower for kw in conditional_keywords)

        if is_conditional and "n/a" not in desc_lower:
            if not desc.endswith("."):
                desc += "."
            # Determine the right N/A clause based on keyword
            if "discrepanc" in name_lower or "discrepanc" in desc_lower:
                desc += " Mark N/A if no discrepancies existed (all items matched)."
            elif "variance" in name_lower or "variance" in desc_lower:
                desc += " Mark N/A if no variances existed (delivery matched PO)."
            elif "escalat" in name_lower or "escalat" in desc_lower:
                desc += " Mark N/A if no issues existed requiring escalation."
            elif "exception" in name_lower or "exception" in desc_lower:
                desc += " Mark N/A if no exceptions occurred."
            else:
                desc += " Mark N/A if the triggering condition did not arise."
            attr["description"] = desc


class AzureOpenAIEvaluator:
    def __init__(self, api_key, model="gpt-5.2-chat", max_tokens=4096,
                 retry_attempts=2, retry_delay=2.0,
                 azure_endpoint=None, azure_api_key=None, azure_deployment=None,
                 azure_api_version=None):
        if not azure_endpoint:
            raise ValueError("AZURE endpoint is required for AzureOpenAIEvaluator")

        self.api_key = azure_api_key or api_key
        self.model = azure_deployment or model
        self.api_url = (
            f"{azure_endpoint.rstrip('/')}/openai/deployments/{self.model}"
            f"/chat/completions?api-version={azure_api_version or '2024-12-01-preview'}"
        )
        self.auth_header = {"api-key": self.api_key}
        self.max_tokens = max_tokens
        self.retry_attempts = retry_attempts
        self.retry_delay = retry_delay

    def _chat_json(self, messages: list, max_tokens: int = 1024, timeout: int = 60) -> dict:
        headers = {**self.auth_header, "Content-Type": "application/json"}
        payload = {
            "model": self.model,
            "messages": messages,
            "max_completion_tokens": max_tokens,
            "response_format": {"type": "json_object"},
        }
        resp = requests.post(self.api_url, headers=headers, json=payload, timeout=timeout)
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"]
        clean = text.strip().removeprefix("```json").removesuffix("```").strip()
        return json.loads(clean)

    def generate_schema(self, rcm: RCMRow) -> ControlSchema:
        """Generate control-specific testing schema via LLM with critic + quality gate."""
        # --- Retrieve similar controls from training library (if enabled) ---
        library_examples = ""
        try:
            from engines.config import ENABLE_ATTRIBUTE_LIBRARY
            if ENABLE_ATTRIBUTE_LIBRARY:
                from engines.AttributeLibrary import get_library
                lib = get_library()
                similar = lib.retrieve(rcm.control_description, top_k=3)
                if similar:
                    library_examples = lib.format_for_prompt(similar)
                    print(f"  [LIB] {rcm.control_id}: found {len(similar)} similar controls in training library")
        except Exception as e:
            print(f"  [!] AttributeLibrary retrieval skipped: {e}")

        user_prompt = f"""Define the testing framework for this control:

**Control ID**: {rcm.control_id}
**Control Description**: {rcm.control_description}
**Control Type**: {rcm.control_type}
**Nature**: {rcm.nature_of_control}
**Frequency**: {rcm.control_frequency}
**Process**: {rcm.process} > {rcm.subprocess}
{"**Control Objective**: " + rcm.control_objective if rcm.control_objective else ""}

{library_examples}

{SCHEMA_FEW_SHOT_EXAMPLES}"""

        try:
            # Pass 1: draft schema
            draft = self._chat_json([
                {"role": "system", "content": SCHEMA_GENERATION_PROMPT},
                {"role": "user", "content": user_prompt},
            ], max_tokens=2500, timeout=90)

            # Pass 2: critic/refiner schema
            critic_user = f"""Control context:
Control ID: {rcm.control_id}
Control Description: {rcm.control_description}
Control Objective: {rcm.control_objective or ''}
Control Type: {rcm.control_type}
Nature: {rcm.nature_of_control}
Frequency: {rcm.control_frequency}
Process: {rcm.process} > {rcm.subprocess}

DRAFT SCHEMA JSON:
{json.dumps(draft, ensure_ascii=False, indent=2)}

{SCHEMA_FEW_SHOT_EXAMPLES}

Return an improved schema JSON only."""
            refined = self._chat_json([
                {"role": "system", "content": SCHEMA_CRITIC_PROMPT},
                {"role": "user", "content": critic_user},
            ], max_tokens=2500, timeout=90)

            schema = ControlSchema(
                control_id=rcm.control_id,
                worksteps=refined.get("worksteps", []),
                attributes=refined.get("attributes", []),
                sample_columns=refined.get("sample_columns", []),
            )
            if not any(str(c.get("key", "")).strip().lower() == "evidence_files_reviewed" for c in (schema.sample_columns or [])):
                schema.sample_columns.append({"key": "evidence_files_reviewed", "header": "Evidence Files Reviewed"})
            _fix_schema_attributes(schema)

            # Quality gate + one repair regeneration
            issues = _schema_quality_issues(schema, rcm)
            if issues:
                print(f"  [QA] {rcm.control_id}: schema quality issues detected ({len(issues)}). Regenerating once...")
                repair_prompt = f"""The previous schema has quality issues.
Fix ALL issues and return full corrected JSON only.

Control context:
Control ID: {rcm.control_id}
Control Description: {rcm.control_description}
Control Objective: {rcm.control_objective or ''}
Control Type: {rcm.control_type}
Nature: {rcm.nature_of_control}
Frequency: {rcm.control_frequency}
Process: {rcm.process} > {rcm.subprocess}

Issues to fix:
- """ + "\n- ".join(issues) + f"""

Current schema:
{json.dumps({'worksteps': schema.worksteps, 'attributes': schema.attributes, 'sample_columns': schema.sample_columns}, ensure_ascii=False, indent=2)}

{SCHEMA_FEW_SHOT_EXAMPLES}
"""
                repaired = self._chat_json([
                    {"role": "system", "content": SCHEMA_CRITIC_PROMPT},
                    {"role": "user", "content": repair_prompt},
                ], max_tokens=2500, timeout=90)
                repaired_schema = ControlSchema(
                    control_id=rcm.control_id,
                    worksteps=repaired.get("worksteps", []),
                    attributes=repaired.get("attributes", []),
                    sample_columns=repaired.get("sample_columns", []),
                )
                if not any(str(c.get("key", "")).strip().lower() == "evidence_files_reviewed" for c in (repaired_schema.sample_columns or [])):
                    repaired_schema.sample_columns.append({"key": "evidence_files_reviewed", "header": "Evidence Files Reviewed"})
                _fix_schema_attributes(repaired_schema)
                repaired_issues = _schema_quality_issues(repaired_schema, rcm)
                if len(repaired_issues) < len(issues):
                    schema = repaired_schema

            return schema
        except Exception as e:
            print(f"  [!]  Schema generation failed for {rcm.control_id}: {e}")
            return ControlSchema(control_id=rcm.control_id, worksteps=[], attributes=[], sample_columns=[])

    def evaluate_toe(self, rcm: RCMRow, sample: SampleEvidence,
                     sample_index: int = 1, total_samples: int = 1,
                     schema: ControlSchema = None) -> dict:
        """Evaluate one TOE sample via Azure OpenAI."""
        base_prompt = (
            TOE_SYSTEM_PROMPT_AUTOMATED if rcm.control_type.strip().lower() == "automated"
            else TOE_SYSTEM_PROMPT_MANUAL
        )
        # Inject process context from RCM so the LLM knows the business cycle
        process_ctx = rcm.process.strip() if rcm.process else ""
        if process_ctx:
            system_prompt = base_prompt.replace(
                "for a MANUAL control.",
                f"for a MANUAL control in the {process_ctx} cycle."
            ).replace(
                "for an AUTOMATED control.",
                f"for an AUTOMATED control in the {process_ctx} cycle."
            )
        else:
            system_prompt = base_prompt
        user_prompt = build_toe_prompt(rcm, sample, sample_index, total_samples, schema=schema)

        prompt_tokens = _estimate_tokens(user_prompt)
        if prompt_tokens > PROMPT_WARNING_THRESHOLD:
            print(f"  [!]  {rcm.control_id}/{sample.sample_id}: prompt is ~{prompt_tokens:,} tokens")

        headers = {**self.auth_header, "Content-Type": "application/json"}
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "max_completion_tokens": self.max_tokens,
            "response_format": {"type": "json_object"},
        }

        last_error = None
        for attempt in range(1, self.retry_attempts + 1):
            try:
                resp = requests.post(self.api_url, headers=headers, json=payload, timeout=60)
                resp.raise_for_status()
                text = resp.json()["choices"][0]["message"]["content"]
                clean = text.strip().removeprefix("```json").removesuffix("```").strip()
                return json.loads(clean)
            except requests.exceptions.HTTPError:
                last_error = f"HTTP {resp.status_code}: {resp.text[:300]}"
                if resp.status_code == 429:
                    time.sleep(self.retry_delay * attempt * 2)
                elif resp.status_code >= 500:
                    time.sleep(self.retry_delay * attempt)
                else:
                    break
            except (requests.RequestException, json.JSONDecodeError, KeyError, IndexError) as e:
                last_error = str(e)
                if attempt < self.retry_attempts:
                    time.sleep(self.retry_delay * attempt)

        return {
            "result": "ERROR", "operated_effectively": "Unable to Assess",
            "control_performed": "Unable to Assess", "timely_execution": "Unable to Assess",
            "accurate_execution": "Unable to Assess", "authorized_performer": "Unable to Assess",
            "evidence_sufficient": "Unable to Assess",
            "remarks": f"LLM evaluation failed: {last_error}",
            "deviation_details": "Evaluation error -- manual review required",
        }

    def re_evaluate_toe(self, rcm: RCMRow, sample: SampleEvidence,
                         schema: ControlSchema, first_pass: dict,
                         issues: list, sample_index: int = 1,
                         total_samples: int = 1) -> dict:
        """Re-evaluate a TOE sample via LLM self-correction (second API call).

        Uses a multi-turn conversation:
          - System: QC reviewer prompt
          - User 1: Original evidence + control context
          - Assistant 1: First-pass JSON (the LLM's own output)
          - User 2: Specific QC feedback about what went wrong

        This lets the LLM see its own work, understand the specific mistakes,
        and produce a corrected assessment with proper attributes, result, and remarks.
        """
        # Build the original context (same as first pass)
        user_prompt_original = build_toe_prompt(rcm, sample, sample_index, total_samples, schema=schema)

        # Build the QC feedback message
        issues_text = "\n".join(f"  {i+1}. {issue}" for i, issue in enumerate(issues))
        qc_feedback = f"""## QC REVIEW — ISSUES FOUND IN YOUR ASSESSMENT

The following logical errors were identified in your assessment above:

{issues_text}

## INSTRUCTIONS:
Please provide a CORRECTED assessment that fixes ALL the issues listed above.
- Keep everything that was correct in your original assessment
- Fix ONLY the identified issues
- Ensure attributes, result, and remarks are all internally consistent
- Write professional remarks that accurately describe what the evidence shows
- For PASS results, write remarks explaining why the control operated effectively
- For FAIL results, write remarks identifying the specific deviation

Respond with the complete corrected JSON (same schema as your original response)."""

        # Multi-turn conversation: original context → first pass → QC feedback
        messages = [
            {"role": "system", "content": RE_EVALUATION_SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt_original},
            {"role": "assistant", "content": json.dumps(first_pass, indent=2)},
            {"role": "user", "content": qc_feedback},
        ]

        headers = {**self.auth_header, "Content-Type": "application/json"}
        payload = {
            "model": self.model,
            "messages": messages,
            "max_completion_tokens": self.max_tokens,
            "response_format": {"type": "json_object"},
        }

        last_error = None
        for attempt in range(1, self.retry_attempts + 1):
            try:
                resp = requests.post(self.api_url, headers=headers, json=payload, timeout=90)
                resp.raise_for_status()
                text = resp.json()["choices"][0]["message"]["content"]
                clean = text.strip().removeprefix("```json").removesuffix("```").strip()
                return json.loads(clean)
            except requests.exceptions.HTTPError:
                last_error = f"HTTP {resp.status_code}: {resp.text[:300]}"
                if resp.status_code == 429:
                    time.sleep(self.retry_delay * attempt * 2)
                elif resp.status_code >= 500:
                    time.sleep(self.retry_delay * attempt)
                else:
                    break
            except (requests.RequestException, json.JSONDecodeError, KeyError, IndexError) as e:
                last_error = str(e)
                if attempt < self.retry_attempts:
                    time.sleep(self.retry_delay * attempt)

        # If re-evaluation API call fails, return original first pass (don't lose work)
        return first_pass



# ===========================================================================
#  LLM SELF-CORRECTION: Detect → Re-evaluate Loop
# ===========================================================================

def detect_evaluation_issues(eval_result: dict) -> list:
    """Detect logical contradictions in an LLM evaluation result.

    Returns a list of human-readable issue descriptions. Empty list = no issues.
    This does NOT correct anything — it just identifies problems for re-evaluation.
    """
    issues = []
    r = eval_result
    attr_results = r.get("attribute_results")
    attr_reasoning = r.get("attribute_reasoning")

    if not attr_results or not isinstance(attr_results, dict):
        return issues

    # --- Issue 1: All attrs Yes/N/A but result is FAIL ---
    has_no_attr = False
    no_attrs = {}
    for aid, val in attr_results.items():
        norm = _normalize_attr_value(str(val))
        if norm == "No":
            has_no_attr = True
            no_attrs[aid] = val

    if not has_no_attr and r.get("result") == "FAIL":
        issues.append(
            "LOGICAL CONTRADICTION: You marked every attribute as 'Yes' or 'N/A' (none are 'No'), "
            "but your overall result is 'FAIL'. This is impossible — if no attribute failed, "
            "the result MUST be 'PASS'. Please re-evaluate: either change result to 'PASS', "
            "or identify which specific attribute should be 'No' and why."
        )

    # --- Issue 2: Conditional attr marked 'No' but reasoning says condition didn't arise ---
    _CONDITIONAL_KW = ("discrepanc", "variance", "escalat", "exception", "resolution",
                       "mismatch", "deviation", "dispute")
    _ABSENT_PHRASES = ("no discrepanc", "no variance", "no issue", "no exception",
                       "no mismatch", "everything matched", "all items matched",
                       "all matched", "matched the po", "matched po", "zero variance",
                       "zero discrepanc", "none were found", "none were identified",
                       "inspection found no", "quality passed", "qty confirmed",
                       "no deviation", "no issues", "all items verified")
    _REAL_DEVIATION = ("despite", "shortfall", "short by", "defective", "but not logged",
                       "was identified but", "was found but", "existed but", "occurred but",
                       "gap of", "missing")

    if attr_reasoning and isinstance(attr_reasoning, dict):
        for aid, val in attr_results.items():
            if _normalize_attr_value(str(val)) != "No":
                continue
            reason = str(attr_reasoning.get(aid, "")).lower()
            aid_lower = aid.lower()

            is_conditional = any(kw in aid_lower or kw in reason for kw in _CONDITIONAL_KW)
            if not is_conditional:
                continue

            condition_absent = any(phrase in reason for phrase in _ABSENT_PHRASES)
            if not condition_absent:
                continue

            real_deviation = any(phrase in reason for phrase in _REAL_DEVIATION)
            if real_deviation:
                continue

            issues.append(
                f"INCORRECT N/A vs NO for attribute '{aid}': You marked this attribute 'No', "
                f"but your own reasoning says the triggering condition did NOT arise "
                f"(e.g., '{reason[:80]}...'). "
                f"'No' means 'the condition existed but was not handled.' "
                f"'N/A' means 'the condition did not exist.' "
                f"Since your reasoning confirms the condition didn't exist, this should be 'N/A', not 'No'. "
                f"Please re-evaluate this attribute."
            )

    # --- Issue 3: D&B score present in evidence but financial review marked No ---
    # BUT: if D&B score is below policy threshold, "No" is CORRECT — skip the issue
    if attr_reasoning and isinstance(attr_reasoning, dict):
        for aid, val in attr_results.items():
            if _normalize_attr_value(str(val)) != "No":
                continue
            reason = str(attr_reasoning.get(aid, "")).lower()
            aid_lower = aid.lower()

            is_financial = "financial" in aid_lower or "financial" in reason
            if not is_financial:
                continue

            # Check if evidence mentions D&B score
            evidence_text = str(r.get("remarks", "") + " " + str(attr_reasoning.get(aid, ""))).lower()
            raw_ev = str(r.get("_raw_evidence", "")).lower()  # Will be passed in
            all_text = evidence_text + " " + raw_ev

            if "d&b" in all_text or "dun" in all_text or "credit report" in all_text:
                # Check if the D&B score is below threshold — if so, "No" is correct
                below_threshold = any(phrase in all_text for phrase in (
                    "below threshold", "below the required", "below policy",
                    "critical risk", "policy requires", "does not meet",
                    "failed financial", "insufficient score"))
                if not below_threshold:
                    issues.append(
                        f"D&B SCORE = FINANCIAL REVIEW for attribute '{aid}': You marked this 'No' "
                        f"saying no financial health review report was found. However, a D&B (Dun & Bradstreet) "
                        f"score IS a financial health review — that's the industry-standard financial health "
                        f"assessment tool. If the evidence mentions a D&B score, credit report, or financial "
                        f"stability assessment, then the financial review WAS completed. "
                        f"Please re-evaluate this attribute."
                    )

    # --- Issue 3b: D&B score BELOW threshold but financial attribute marked Yes ---
    # This catches cases like S29 where D&B=32, policy requires >60, but LLM said "Yes"
    if attr_reasoning and isinstance(attr_reasoning, dict):
        for aid, val in attr_results.items():
            normalized = _normalize_attr_value(str(val))
            if normalized != "Yes":
                continue
            reason = str(attr_reasoning.get(aid, "")).lower()
            aid_lower = aid.lower()

            is_financial = ("financial" in aid_lower or "financial" in reason
                           or "health" in aid_lower)
            if not is_financial:
                continue

            # Check evidence for D&B below threshold indicators
            evidence_text = str(r.get("remarks", "") + " " + reason).lower()
            raw_ev = str(r.get("_raw_evidence", "")).lower()
            all_text = evidence_text + " " + raw_ev

            below_threshold = any(phrase in all_text for phrase in (
                "below threshold", "below the required", "below policy",
                "critical risk", "policy requires >", "does not meet minimum",
                "score 32", "score 28", "score 15", "score 20",  # common low scores
                "no exception approval", "activated with note"))

            if below_threshold and ("d&b" in all_text or "financial" in all_text):
                issues.append(
                    f"D&B SCORE BELOW THRESHOLD for attribute '{aid}': You marked this 'Yes' "
                    f"but the evidence indicates the D&B/financial score is BELOW the required policy "
                    f"threshold. A low score means the review was done but the vendor FAILED it — "
                    f"this is a control failure. Mark this attribute 'No' and set result to 'FAIL'. "
                    f"Please re-evaluate this attribute."
                )

    # --- Issue 4: Result PASS but has attr = No (opposite direction) ---
    if has_no_attr and r.get("result") == "PASS":
        no_list = ", ".join(f"'{k}'" for k in no_attrs)
        issues.append(
            f"CONTRADICTION: Result is 'PASS' but attribute(s) {no_list} are marked 'No'. "
            f"If any attribute failed, the result should be 'FAIL'. "
            f"Please re-evaluate: either the attributes should be 'Yes'/'N/A', or the result should be 'FAIL'."
        )

    return issues


RE_EVALUATION_SYSTEM_PROMPT = """You are a senior QC (Quality Control) reviewer on a Big 4 audit engagement.

Your job is to REVIEW AND CORRECT a junior auditor's Test of Operating Effectiveness (TOE) assessment.
The junior auditor made specific logical errors that have been identified. You must fix these errors
while preserving everything that was correct in the original assessment.

## YOUR TASK:
1. Read the original evidence and control description
2. Read the junior auditor's assessment (their JSON output)
3. Read the specific QC issues identified
4. Produce a CORRECTED assessment that fixes the identified issues

## KEY RULES:
- A D&B (Dun & Bradstreet) score IS a financial health review. Do NOT require a separate document.
- HOWEVER: If the D&B score is BELOW the policy threshold (e.g., "D&B score 32, policy requires >60"),
  then the financial review FAILED — mark the financial attribute "No" because the vendor did not meet
  the required standard. A low score = review done but FAILED = control failure.
- "Compliance clear" IS regulatory compliance validation.
- "Site visit done" IS a reference/physical verification check.
- If a CONDITIONAL attribute's triggering condition did NOT arise (e.g., no discrepancies existed, 
  everything matched, zero variances), that attribute MUST be "N/A" — NEVER "No".
  "No" = the condition existed but was not handled (a TRUE failure).
  "N/A" = the condition did not exist (NOT a failure).
- If ALL attributes are "Yes" or "N/A", the result MUST be "PASS". No exceptions.
- If ANY attribute is "No", the result MUST be "FAIL".
- Remarks must be consistent with the result and explain what was observed.

## RESPOND ONLY WITH CORRECTED JSON (same schema as the original)."""


# ===========================================================================
#  RCM LOADER
# ===========================================================================

COLUMN_MAP = {
    "Process": "process", "Sub Process": "subprocess", "SubProcess": "subprocess",
    "Control Objective": "control_objective",
    "Risk ID": "risk_id", "Risk Id": "risk_id",
    "Risk Title": "risk_title", "Risk Description": "risk_description",
    "Risk Level": "risk_level", "risk_level": "risk_level",
    "Control ID": "control_id", "Control Id": "control_id",
    "Control Description": "control_description",
    "Control Owner": "control_owner",
    "Nature of Control": "nature_of_control", "Control Type": "control_type",
    "Control Frequency": "control_frequency", "Application/System": "application_system",
    "Count of Samples": "count_of_samples", "count_of_samples": "count_of_samples",
}


def load_rcm(filepath, sheet_name=None):
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"RCM file not found: {filepath}")
    from engines.rcm_reader import smart_read_excel

    # 1. Read original (all columns, original names) — kept for output summary
    try:
        original_df = smart_read_excel(str(path), sheet_name=sheet_name, normalize_columns=False)
    except Exception:
        original_df = pd.read_excel(path, sheet_name=sheet_name or 0)

    # 2. Read normalised (mapped column names) — used by engines
    try:
        df, _info = smart_read_excel(str(path), sheet_name=sheet_name, normalize_columns=True)
    except Exception:
        df = pd.read_excel(path, sheet_name=sheet_name or 0)
    df.rename(columns=COLUMN_MAP, inplace=True)

    rows = []
    for _, row in df.iterrows():
        rows.append(RCMRow(**{col: str(row.get(col, "")) for col in RCMRow.__dataclass_fields__}))

    # Build control_id → original DataFrame row index mapping
    orig_row_map = {r.control_id: i for i, r in enumerate(rows)}

    return rows, original_df, orig_row_map




# ===========================================================================
#  EVIDENCE LOADER -- reads one subfolder per control
# ===========================================================================
#
#  Folder structure:
#
#    evidence/
#    |--- C-P2P-001/
#    |   |--- evidence.txt              <- walkthrough (has ---EVIDENCE--- marker)
#    |   |--- Procurement_SOP_v1.2.pdf  <- supporting doc (PDF)
#    |   |--- SAP_Config_Export.xlsx     <- supporting doc (Excel)
#    |   `--- Approval_Matrix.docx      <- supporting doc (Word)
#    |--- C-P2P-002/
#    |   |--- evidence.txt
#    |   `--- Budget_Policy.pptx        <- supporting doc (PowerPoint)
#    `--- ...
#
#  Supported formats for text extraction:
#    All formats supported by Document_Intelligence module:
#    .pdf .docx .doc .xlsx .xls .xlsm .csv .tsv .pptx .msg .eml
#    .png .jpg .jpeg .tiff .tif .bmp (OCR via Azure Document Intelligence)
#    .txt .md .csv .log (plain text — handled directly)
# ===========================================================================

import sys as _sys
_di_dir = str(Path(__file__).resolve().parent.parent)
if _di_dir not in _sys.path:
    _sys.path.insert(0, _di_dir)
from Document_Intelligence import extract_text as _di_extract_text
from Document_Intelligence import get_embedding_log_dir as _get_embedding_log_dir
from Document_Intelligence import set_parse_phase as _set_parse_phase


def _llm_refine_extraction(content: str, filename: str, doc_type: str) -> str:
    """Optionally refine extracted document text with LLM for higher parsing accuracy."""
    text = (content or "").strip()
    if not text:
        return content

    enabled = os.getenv("LLM_EVIDENCE_PARSING", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return content

    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or "gpt-5.2-chat"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()

    if not endpoint or not api_key:
        return content

    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}

    try:
        chunk_chars = int(os.getenv("LLM_PARSER_CHUNK_CHARS", "50000"))
        overlap = int(os.getenv("LLM_PARSER_CHUNK_OVERLAP", "500"))
        chunks = []
        i = 0
        while i < len(text):
            j = min(len(text), i + chunk_chars)
            chunks.append(text[i:j])
            if j >= len(text):
                break
            i = max(j - overlap, i + 1)

        refined_chunks = []
        for idx, ch in enumerate(chunks, 1):
            payload = {
                "model": deployment,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a forensic document parser. Clean extraction noise while preserving all facts. "
                            "Never invent. Keep dates, amounts, IDs, approvers, headers, and key workflow language. "
                            "Return JSON only."
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"Document: {filename}\nType: {doc_type}\nChunk: {idx}/{len(chunks)}\n\n"
                            "Normalize formatting and OCR artifacts. Do not drop material evidence details.\n\n"
                            f"Extracted text:\n{ch}"
                        ),
                    },
                ],
                "max_completion_tokens": 16384,
                "response_format": {"type": "json_object"},
            }
            resp = requests.post(url, headers=headers, json=payload, timeout=120)
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"]
            data = json.loads(raw.strip().removeprefix("```json").removesuffix("```").strip())
            refined = str(data.get("clean_text") or data.get("parsed_text") or data.get("text") or "").strip()
            if refined:
                refined_chunks.append(refined)

        merged = "\n\n".join(refined_chunks).strip()
        if merged:
            return merged
        return content
    except Exception:
        return content


# -- File-type registry ---------------------------------------------------

# Plain text extensions -- only these can contain the ---EVIDENCE--- marker
_TEXT_EXTS = {".txt", ".md", ".csv", ".log"}

# NOTE: All format-specific extractors (_extract_pdf, _extract_docx, _extract_xlsx,
# _extract_pptx, _extract_image_ocr, _extract_eml, _extract_msg, _azure_docint_extract,
# _azure_vision_describe_image) have been replaced by Document_Intelligence module.


def _extract_file(filepath: Path) -> tuple[str, str, bool]:
    """
    Extract text content from any supported file.

    Delegates to Document_Intelligence for all structured/binary formats.
    Plain text files are read directly.

    Returns: (content_text, doc_type_label, extraction_succeeded)
    """
    ext = filepath.suffix.lower()

    # Plain text files — read directly (fast, no external parser needed)
    if ext in _TEXT_EXTS:
        try:
            content = filepath.read_text(encoding="utf-8", errors="replace").strip()
            return content, _infer_doc_type(filepath.name), True
        except Exception:
            return "", "Text File", False

    # All other formats — delegate to Document_Intelligence
    text, doc_type, ok = _di_extract_text(str(filepath))

    return text, doc_type, ok


def _infer_doc_type(filename: str) -> str:
    """Infer document type from filename keywords."""
    name = filename.lower()
    if any(k in name for k in ("sop", "procedure", "policy", "manual", "guideline")):
        return "Policy/SOP"
    if any(k in name for k in ("config", "setting", "param", "spro")):
        return "System Configuration"
    if any(k in name for k in ("log", "trail", "audit")):
        return "Audit Log"
    if any(k in name for k in ("screenshot", "screen", "capture")):
        return "Screenshot/Capture"
    if any(k in name for k in ("report", "summary", "dashboard")):
        return "Report"
    if any(k in name for k in ("email", "approval", "sign", "ticket")):
        return "Approval/Communication"
    if any(k in name for k in ("checklist", "form", "template")):
        return "Checklist/Form"
    if any(k in name for k in ("matrix", "role", "access")):
        return "Access/Role Matrix"
    if any(k in name for k in ("invoice", "receipt", "payment", "credit")):
        return "Transaction Document"
    return "Supporting Document"


def _score_text_file_for_toe_sample(filepath: Path) -> float:
    """Heuristic score for classifying a text file as TOE sample narrative vs supporting doc."""
    name = filepath.stem.lower()
    score = 0.0

    # Filename cues
    if re.search(r"\b(sample|evidence|walkthrough|transaction)\b", name):
        score += 3.0
    if re.search(r"\b(email|mail|approval|policy|sop|report|matrix|config|log|audit|screenshot)\b", name):
        score -= 2.0

    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")
        head = "\n".join(text.splitlines()[:40]).lower()

        # Email-like headers indicate supporting evidence, not sample narrative
        if re.search(r"^\s*(from|to|subject|cc|bcc|date)\s*:", head, flags=re.MULTILINE):
            score -= 4.0

        # TOE sample metadata cues
        meta_hits = 0
        for k in ("sample id:", "source document:", "control id:"):
            if k in head:
                meta_hits += 1
        if meta_hits:
            score += 2.0 + 0.5 * meta_hits

        if any(k in head for k in ("transaction", "control", "performed", "tested", "sample")):
            score += 1.0
    except Exception:
        pass

    return score


def _llm_classify_text_role(filepath: Path, mode: str = "toe") -> tuple[str, float]:
    """Classify text file role using LLM: returns (role, confidence).

    role is one of: "sample", "supporting", "unknown"
    """
    enabled = os.getenv("LLM_ROLE_CLASSIFIER", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return "unknown", 0.0

    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or "gpt-5.2-chat"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()
    if not endpoint or not api_key:
        return "unknown", 0.0

    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")[:8000]
    except Exception:
        return "unknown", 0.0

    prompt_mode = "TOE sample" if mode == "toe" else "TOD walkthrough"
    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}
    payload = {
        "model": deployment,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Classify file role for audit evidence. "
                    "Return JSON {role, confidence}. role must be sample|supporting|unknown."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Target role type: {prompt_mode}.\n"
                    f"Filename: {filepath.name}\n\n"
                    "sample means primary transaction narrative to test control; "
                    "supporting means email/policy/report/log/screenshot/reference evidence.\n\n"
                    f"Content:\n{text}"
                ),
            },
        ],
        "max_completion_tokens": 300,
        "response_format": {"type": "json_object"},
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=60)
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"]
        data = json.loads(raw.strip().removeprefix("```json").removesuffix("```").strip())
        role = str(data.get("role", "unknown")).strip().lower()
        if role not in {"sample", "supporting", "unknown"}:
            role = "unknown"
        conf = float(data.get("confidence", 0.0) or 0.0)
        conf = max(0.0, min(1.0, conf))
        return role, conf
    except Exception:
        return "unknown", 0.0


# -- Evidence folder loader -----------------------------------------------


# ===========================================================================
#  TOE EVIDENCE LOADER
# ===========================================================================

_SAMPLE_RE = re.compile(r"^sample[_\- ]?\d+", re.IGNORECASE)


def _extract_evidence_file(ef, format_stats, errors, context_prefix):
    """Extract text from a single evidence file. Returns (block_str, ok)."""
    ext = ef.suffix.lower()
    format_stats[ext] = format_stats.get(ext, 0) + 1

    txt, _, ok = _extract_file(ef)

    if not (txt or "").strip():
        txt = f"[No extractable content in {ef.stem}]"
        ok = False

    if not ok:
        errors.append(f"  [!]  {context_prefix}/{ef.stem}: extraction may be incomplete")

    return f"[EVIDENCE FILE: {ef.stem}]\n{txt}", ok


def load_toe_evidence_folder(folder_path: str, include_control_ids: set[str] | None = None,
                             pre_extract_cache: dict | None = None) -> dict:
    """
    Load TOE evidence -- multiple samples per control.

    Supports three folder layouts (auto-detected per control):

    Mode A — Sample subfolders (multiple evidence files per sample):
      evidence_toe/
        C-P2P-001/
          sample1/              <- subfolder per sample
            invoice.pdf
            approval_email.txt
          sample2/
            invoice.pdf

    Mode B — Sample-named files (one file per sample):
      evidence_toe/
        C-P2P-001/
          sample_1.txt          <- filename starts with sample_N / sample-N / sampleN
          sample_2.pdf
          sample_3.docx

    Mode C — Fallback (each file = 1 sample):
      evidence_toe/
        C-P2P-001/
          invoice.pdf           <- becomes sample 1
          approval_email.txt    <- becomes sample 2
          report.docx           <- becomes sample 3

    Args:
        folder_path: Root TOE evidence path
        include_control_ids: optional set of control IDs to parse. If provided,
            only matching control folders are parsed (case-insensitive), which
            avoids unnecessary extraction for controls that should be skipped
            (e.g., non-PASS controls after TOD).
        pre_extract_cache: Optional dict[abs_file_path_str -> (content, doc_type, ok)]
            from evidence validation (Phase 2). When provided, files that were
            already extracted via Document Intelligence will be read from cache
            instead of re-extracting.

    Returns dict[control_id -> list[SampleEvidence]]
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"TOE evidence folder not found: {folder_path}")

    toe_bank: dict[str, list] = {}
    errors: list[str] = []
    format_stats: dict[str, int] = {}

    # Pre-built extraction cache from evidence validation (Phase 2)
    _pre_cache: dict[str, tuple[str, str, bool]] = pre_extract_cache or {}
    cache_hits = 0

    control_dirs = sorted(d for d in folder.iterdir() if d.is_dir())
    if not control_dirs:
        raise FileNotFoundError(
            f"No subfolders found in {folder_path}. "
            f"Expected one folder per control (e.g. C-P2P-001/)."
        )

    include_upper = None
    if include_control_ids is not None:
        include_upper = {str(cid).strip().upper() for cid in include_control_ids if str(cid).strip()}

    skipped_by_filter = 0

    # ── Phase 1: Collect all evidence files that need DI extraction ──
    all_extract_files: list[Path] = []

    def _norm_key(p: Path) -> str:
        """Normalize path to match evidence_validator cache keys."""
        return os.path.normcase(os.path.normpath(str(p.resolve())))

    for control_dir in control_dirs:
        control_id = control_dir.name
        if include_upper is not None and control_id.strip().upper() not in include_upper:
            continue

        sample_dirs = sorted(d for d in control_dir.iterdir() if d.is_dir())
        if sample_dirs:
            for sample_dir in sample_dirs:
                for f in sample_dir.iterdir():
                    if f.is_file() and f.name.lower() not in ("readme.txt", "readme.md"):
                        if f.suffix.lower() not in _TEXT_EXTS:
                            abs_key = _norm_key(f)
                            if abs_key in _pre_cache:
                                cache_hits += 1
                            else:
                                all_extract_files.append(f)
        else:
            for f in control_dir.iterdir():
                if f.is_file() and f.name.lower() not in ("readme.txt", "readme.md"):
                    if f.suffix.lower() not in _TEXT_EXTS:
                        abs_key = _norm_key(f)
                        if abs_key in _pre_cache:
                            cache_hits += 1
                        else:
                            all_extract_files.append(f)

    if _pre_cache:
        print(f"[TOE Evidence] Reusing {cache_hits} cached extractions from evidence validation")

    # ── Phase 2: Parallel extraction of ONLY uncached non-text files ──
    parse_workers = int(os.getenv("EVIDENCE_PARSE_WORKERS", "8"))
    extract_cache: dict[Path, tuple[str, str, bool]] = {}
    if all_extract_files:
        print(f"[TOE Evidence] Extracting {len(all_extract_files)} new files via Document Intelligence")
        with ThreadPoolExecutor(max_workers=parse_workers) as executor:
            future_to_path = {executor.submit(_extract_file, fp): fp for fp in all_extract_files}
            for future in as_completed(future_to_path):
                fp = future_to_path[future]
                try:
                    extract_cache[fp] = future.result()
                except Exception as e:
                    extract_cache[fp] = (f"[Extraction error for {fp.name}: {e}]", "Unknown", False)
    elif cache_hits > 0:
        print(f"[TOE Evidence] All files served from cache — no Document Intelligence calls needed")

    # ── Phase 3: Assemble evidence (sequential, uses cached results) ──
    def _extract_evidence_file_cached(ef, fmt_stats, errs, context_prefix):
        """Extract text from a single evidence file using cache. Returns (block_str, ok).

        Checks the pre-built validation cache (_pre_cache) first, then the
        local extract_cache from Phase 2, to avoid redundant DI calls.
        """
        ext = ef.suffix.lower()
        fmt_stats[ext] = fmt_stats.get(ext, 0) + 1

        # Always check pre-cache first (from evidence validation Phase 2)
        abs_key = _norm_key(ef)
        if abs_key in _pre_cache:
            txt, _, ok = _pre_cache[abs_key]
        elif ext in _TEXT_EXTS:
            try:
                txt = ef.read_text(encoding="utf-8", errors="replace").strip()
                ok = True
            except Exception:
                txt = ""
                ok = False
        else:
            txt, _, ok = extract_cache.get(ef, ("", "Unknown", False))

        if not (txt or "").strip():
            txt = f"[No extractable content in {ef.stem}]"
            ok = False

        if not ok:
            errs.append(f"  [!]  {context_prefix}/{ef.stem}: extraction may be incomplete")

        return f"[EVIDENCE FILE: {ef.stem}]\n{txt}", ok

    for control_dir in control_dirs:
        control_id = control_dir.name

        if include_upper is not None and control_id.strip().upper() not in include_upper:
            skipped_by_filter += 1
            continue

        sample_dirs = sorted(d for d in control_dir.iterdir() if d.is_dir())
        all_files = sorted(
            f for f in control_dir.iterdir()
            if f.is_file() and f.name.lower() not in ("readme.txt", "readme.md")
        )

        if sample_dirs:
            # ── Mode A: sample subfolders ──
            samples_list = []
            for sample_dir in sample_dirs:
                sample_name = sample_dir.name
                sample_files = sorted(
                    f for f in sample_dir.iterdir()
                    if f.is_file() and f.name.lower() not in ("readme.txt", "readme.md")
                )

                if not sample_files:
                    errors.append(f"  [!]  {control_id}/{sample_name}/: empty sample folder, skipped")
                    continue

                evidence_blocks = []
                for ef in sample_files:
                    block, _ = _extract_evidence_file_cached(ef, format_stats, errors, f"{control_id}/{sample_name}")
                    evidence_blocks.append(block)

                samples_list.append(SampleEvidence(
                    sample_id=f"{control_id}_{sample_name}",
                    description="\n\n".join(evidence_blocks),
                    source_document=", ".join(f.name for f in sample_files),
                    supporting_docs=[],
                ))

            if samples_list:
                toe_bank[control_id] = samples_list
            else:
                errors.append(f"  [!]  {control_id}/: all sample subfolders empty, skipped")

        elif all_files:
            # Check if files follow sample_N naming pattern
            sample_files = [f for f in all_files if _SAMPLE_RE.match(f.stem)]

            if sample_files:
                # ── Mode B: sample-named files, each file = 1 sample ──
                samples_list = []
                for ef in sample_files:
                    block, _ = _extract_evidence_file_cached(ef, format_stats, errors, control_id)
                    samples_list.append(SampleEvidence(
                        sample_id=f"{control_id}_{ef.stem}",
                        description=block,
                        source_document=ef.name,
                        supporting_docs=[],
                    ))

                # Warn about non-sample files that are being ignored
                non_sample = [f for f in all_files if not _SAMPLE_RE.match(f.stem)]
                if non_sample:
                    names = ", ".join(f.name for f in non_sample)
                    errors.append(f"  [!]  {control_id}/: non-sample files ignored: {names}")

                toe_bank[control_id] = samples_list
            else:
                # ── Mode C: fallback — each file = 1 sample ──
                samples_list = []
                for idx, ef in enumerate(all_files, 1):
                    block, _ = _extract_evidence_file_cached(ef, format_stats, errors, control_id)
                    samples_list.append(SampleEvidence(
                        sample_id=f"{control_id}_sample{idx}",
                        description=block,
                        source_document=ef.name,
                        supporting_docs=[],
                    ))
                toe_bank[control_id] = samples_list
        else:
            errors.append(f"  [!]  {control_id}/: no evidence files found, skipped")

    # -- Print summary ------------------------------------------------
    total_files = sum(format_stats.values())
    total_samples = sum(len(s) for s in toe_bank.values())
    _EXT_LABELS = {".txt": "text", ".md": "text", ".csv": "csv", ".log": "log",
                   ".pdf": "pdf", ".docx": "docx", ".xlsx": "xlsx", ".doc": "doc",
                   ".xls": "xls", ".pptx": "pptx", ".ppt": "ppt",
                   ".png": "image", ".jpg": "image", ".jpeg": "image",
                   ".tiff": "image", ".tif": "image", ".bmp": "image",
                   ".eml": "email", ".msg": "email"}

    print(f"[TOE Evidence Loaded] {len(toe_bank)} controls, {total_samples} total samples from {folder_path}")
    print(f"[Evidence Files]      {total_files} files across {len(toe_bank)} controls")
    if format_stats:
        label_counts: dict[str, int] = {}
        for ext, n in format_stats.items():
            label = _EXT_LABELS.get(ext, ext.lstrip("."))
            label_counts[label] = label_counts.get(label, 0) + n
        fmt = ", ".join(f"{lbl}: {n}" for lbl, n in sorted(label_counts.items()))
        print(f"[File Formats]        {fmt}")

    for cid in sorted(toe_bank):
        n_samples = len(toe_bank[cid])
        print(f"  [DIR] {cid}: {n_samples} sample{'s' if n_samples != 1 else ''}")

    if include_upper is not None:
        print(f"[Filter]              Parsed controls: {len(toe_bank)} | Skipped by filter: {skipped_by_filter}")

    if errors:
        for e in errors:
            print(e)

    return toe_bank




# ===========================================================================
#  TOE ENGINE
# ===========================================================================

def _build_remediation(r) -> tuple:
    """
    Build Remedial Actions and Remediation Workplan based on TOE test results.

    Logic:
    - None deficiency -> "No remediation required"
    - Control Deficiency -> targeted fix for the specific failure
    - Significant Deficiency -> process-level improvements + monitoring
    - Material Weakness -> systemic overhaul + escalation to management

    Returns (remedial_actions: str, remediation_workplan: str)
    """
    if r.deficiency_type == "None":
        return ("No remediation required -- control operating effectively.",
                "N/A -- no deficiencies identified.")

    # -- Collect failure patterns --
    fail_remarks = []
    fail_attrs = {}     # attribute_id -> count of "No"
    for sr in r.sample_results:
        if is_sample_fail(sr):
            if sr.remarks:
                fail_remarks.append(sr.remarks)
            if sr.attribute_results and isinstance(sr.attribute_results, dict):
                for aid, val in sr.attribute_results.items():
                    if _normalize_attr_value(str(val)) == "No":
                        fail_attrs[aid] = fail_attrs.get(aid, 0) + 1

    # -- Get attribute names for the failing attributes --
    attr_lookup = {}
    if r.schema and r.schema.attributes:
        for a in r.schema.attributes:
            attr_lookup[a["id"]] = a["name"]

    # Top failing attributes sorted by count
    top_fails = sorted(fail_attrs.items(), key=lambda x: -x[1])
    top_fail_names = [f"{attr_lookup.get(aid, 'Attr '+aid)} ({cnt} failures)" for aid, cnt in top_fails[:3]]

    rate = r.deviation_rate
    ctrl_type = r.control_type if hasattr(r, 'control_type') else ""
    is_manual = "automated" not in ctrl_type.lower() if ctrl_type else True

    # -- Build remedial actions based on deficiency severity --
    actions = []
    workplan = []

    if r.deficiency_type == "Control Deficiency":
        actions.append(f"1. Investigate root cause of the {r.failed_samples} failed sample(s) ({rate:.1%} deviation rate).")
        if top_fail_names:
            actions.append(f"2. Address specific attribute failures: {'; '.join(top_fail_names)}.")
        if is_manual:
            actions.append(f"3. Provide refresher training to control owner on execution requirements.")
            actions.append(f"4. Implement additional supervisory review for the next testing cycle.")
        else:
            actions.append(f"3. Review system configuration to close the gap that allowed the exception.")
            actions.append(f"4. Perform regression testing after configuration fix.")
        workplan.append(f"Week 1-2: Root cause analysis and documentation of corrective action.")
        workplan.append(f"Week 2-3: Implement corrective measures and update SOPs if needed.")
        workplan.append(f"Week 3-4: Verify corrective action through targeted re-testing.")
        workplan.append(f"Ongoing: Enhanced monitoring for one quarter post-remediation.")

    elif r.deficiency_type == "Significant Deficiency":
        actions.append(f"1. Escalate to process owner -- {r.failed_samples} failures ({rate:.1%} deviation).")
        if top_fail_names:
            actions.append(f"2. Systematic failure in: {'; '.join(top_fail_names)}.")
        if is_manual:
            actions.append(f"3. Redesign control procedure with preventive checkpoints and mandatory sign-off gates.")
            actions.append(f"4. Implement compensating controls (independent secondary review) until remediation complete.")
            actions.append(f"5. Mandatory re-training for all personnel involved in control execution.")
        else:
            actions.append(f"3. Engage IT to review and strengthen system validation rules.")
            actions.append(f"4. Add detective controls (exception reports, automated alerts) to catch bypasses.")
            actions.append(f"5. Perform full population scan for similar undetected exceptions.")
        workplan.append(f"Week 1: Escalation to process owner and Audit Committee notification.")
        workplan.append(f"Week 1-3: Root cause analysis -- people, process, or technology.")
        workplan.append(f"Week 3-6: Redesign and implement enhanced control procedures / system fixes.")
        workplan.append(f"Week 6-8: Re-testing of remediated control (minimum 10 samples).")
        workplan.append(f"Quarter 2+: Monthly exception reporting to management.")

    elif r.deficiency_type == "Material Weakness":
        actions.append(f"1. IMMEDIATE ESCALATION to Senior Management and Audit Committee.")
        actions.append(f"2. {r.failed_samples}/{r.total_samples} samples failed ({rate:.1%}) -- control fundamentally not operating.")
        if top_fail_names:
            actions.append(f"3. Critical attribute failures: {'; '.join(top_fail_names)}.")
        if is_manual:
            actions.append(f"4. Suspend current procedure; implement emergency compensating controls immediately.")
            actions.append(f"5. Complete redesign -- consider automating manual checkpoints.")
            actions.append(f"6. Full population review to quantify financial impact of control failures.")
            actions.append(f"7. Engage internal audit / external consultants to validate redesigned control.")
        else:
            actions.append(f"4. Emergency system patch / configuration remediation by IT.")
            actions.append(f"5. Manual compensating control immediately while system fix is deployed.")
            actions.append(f"6. Full-period population analysis to identify all bypass instances.")
            actions.append(f"7. Independent IT review to validate system fix effectiveness.")
        workplan.append(f"Day 1-3: Emergency escalation, compensating controls activated.")
        workplan.append(f"Week 1-2: Full population impact assessment and financial exposure quantification.")
        workplan.append(f"Week 2-4: Root cause analysis and remediation design.")
        workplan.append(f"Week 4-8: Implementation with UAT / parallel testing.")
        workplan.append(f"Week 8-10: Comprehensive re-testing under new control.")
        workplan.append(f"Quarter 2-4: Weekly exception reports, monthly mgmt review, quarterly re-test.")

    return ("\n".join(actions), "\n".join(workplan))


# ===========================================================================

class RCMControlTester:
    """
    Control Testing Engine -- Test of Operating Effectiveness (TOE).

    Usage:
        tester = RCMControlTester(...)
        toe_bank = load_toe_evidence_folder("evidence_toe/")
        results = tester.test_all_toe(toe_bank)
        tester.export_toe_workpaper(results, "toe_workpaper.xlsx")
    """

    def __init__(self, rcm_path=None, openai_api_key=None, openai_model="gpt-5.2-chat",
                 azure_endpoint=None, azure_api_key=None, azure_deployment=None,
                 azure_api_version="2024-12-01-preview", sheet_name=None,
                 original_rcm_df=None, normalized_rcm_df=None):
        """
        Args:
            rcm_path:          Path to RCM file. Required when starting fresh (direct TOE).
            normalized_rcm_df: Pre-normalised DataFrame (canonical column names). When provided,
                               skips file I/O entirely — used in the TOD → Sampling → TOE pipeline
                               so the mapping doesn't run twice.
            original_rcm_df:   Original RCM DataFrame (all original column names) for the
                               output summary sheet. If not provided, load_rcm reads it from file.
        """
        if normalized_rcm_df is not None:
            # Pipeline path: data already normalised (e.g. coming from TOD / Sampling).
            df = normalized_rcm_df.copy()
            df.rename(columns=COLUMN_MAP, inplace=True)
            self.rcm_rows = []
            for _, row in df.iterrows():
                self.rcm_rows.append(RCMRow(**{col: str(row.get(col, "")) for col in RCMRow.__dataclass_fields__}))
            self.orig_row_map = {r.control_id: i for i, r in enumerate(self.rcm_rows)}
            self.original_rcm_df = original_rcm_df
            print(f"[RCM Loaded] {len(self.rcm_rows)} controls (pre-normalised)")
        else:
            # Direct path: read from file (full normalisation pipeline).
            if not rcm_path:
                raise ValueError("Either rcm_path or normalized_rcm_df must be provided.")
            self.rcm_rows, loaded_orig_df, self.orig_row_map = load_rcm(rcm_path, sheet_name)
            self.original_rcm_df = original_rcm_df if original_rcm_df is not None else loaded_orig_df
            print(f"[RCM Loaded] {len(self.rcm_rows)} controls from {rcm_path}")

        self.rcm_lookup = {r.control_id: r for r in self.rcm_rows}
        self.evaluator = AzureOpenAIEvaluator(
            api_key=openai_api_key or azure_api_key,
            model=openai_model or azure_deployment,
            azure_endpoint=azure_endpoint, azure_api_key=azure_api_key,
            azure_deployment=azure_deployment, azure_api_version=azure_api_version,
        )
        provider = "Azure OpenAI" if azure_endpoint else "OpenAI"
        model_name = openai_model or azure_deployment
        print(f"[Mode] Test of Operating Effectiveness (TOE) -- multiple samples per control\n")

    def list_controls(self):
        return pd.DataFrame([{
            "Control ID": r.control_id, "Risk ID": r.risk_id,
            "Risk Title": r.risk_title, "Nature": r.nature_of_control,
            "Type": r.control_type, "Frequency": r.control_frequency,
        } for r in self.rcm_rows])

    # ===================================================================
    #  TOE METHODS
    # ===================================================================

    def _evaluate_toe_sample(self, control_id: str, sample: SampleEvidence,
                             sample_index: int, total_samples: int,
                             schema: ControlSchema = None) -> TOESampleResult:
        """Evaluate one TOE sample with self-correction loop (thread-safe).

        Flow:
          1. First pass: Normal LLM evaluation
          2. Detect: Check for logical contradictions
          3. Re-evaluate: If issues found, send back to LLM with specific feedback
          4. Validate: Light safety net (should rarely trigger after re-evaluation)
        """
        rcm = self.rcm_lookup.get(control_id)
        if not rcm:
            raise ValueError(f"Control '{control_id}' not found.")

        # Guardrail: skip LLM when evidence is too weak/unusable
        low_info, low_reason = _is_low_information_evidence(sample.description)
        if low_info:
            attr_results = {}
            attr_reasoning = {}
            if schema and schema.attributes:
                for a in schema.attributes:
                    aid = str(a.get("id", "")).strip()
                    if not aid:
                        continue
                    attr_results[aid] = "No"
                    attr_reasoning[aid] = f"Insufficient usable evidence for assessment ({low_reason})."

            sample_details = {}
            if schema and schema.sample_columns:
                for c in schema.sample_columns:
                    hdr = str(c.get("header", "")).strip()
                    if hdr:
                        sample_details[hdr] = "Not extracted from evidence"

            return TOESampleResult(
                control_id=control_id,
                sample_id=sample.sample_id,
                result="FAIL",
                operated_effectively="No",
                control_performed="No",
                timely_execution="N/A",
                accurate_execution="No",
                authorized_performer="N/A",
                evidence_sufficient="No",
                remarks=f"Insufficient or non-usable evidence for reliable evaluation ({low_reason}).",
                deviation_details="Evidence insufficient/irrelevant for this sample; manual review required.",
                raw_evidence=sample.description or "",
                sample_details=sample_details,
                attribute_results=attr_results,
                attribute_reasoning=attr_reasoning,
                validator_corrected=True,
                validator_details="Pre-LLM evidence-quality guardrail",
            )

        # ── Phase 1: First-pass LLM evaluation ──
        eval_result = self.evaluator.evaluate_toe(rcm, sample, sample_index, total_samples, schema=schema)

        # Inject raw evidence for issue detection (D&B check)
        eval_result["_raw_evidence"] = sample.description

        # Track the evaluation journey
        was_re_evaluated = False
        re_eval_issues = []

        # ── Phase 2: Detect contradictions ──
        issues = detect_evaluation_issues(eval_result)

        if issues:
            re_eval_issues = issues
            # ── Phase 3: Re-evaluate via second LLM call ──
            corrected = self.evaluator.re_evaluate_toe(
                rcm, sample, schema, eval_result, issues,
                sample_index=sample_index, total_samples=total_samples
            )

            if corrected and corrected.get("result") != "ERROR":
                was_re_evaluated = True
                corrected["_raw_evidence"] = sample.description
                eval_result = corrected

                # Check if re-evaluation actually fixed the issues
                remaining_issues = detect_evaluation_issues(eval_result)
                if remaining_issues:
                    re_eval_issues.extend(
                        [f"(STILL AFTER RE-EVAL) {i}" for i in remaining_issues]
                    )

        # Clean up internal field
        eval_result.pop("_raw_evidence", None)

        # ── Phase 4: Light validation (safety net) ──
        eval_result = validate_toe_sample(eval_result)
        eval_result = _enforce_schema_output_quality(eval_result, schema, source_document=sample.source_document or "")

        # Build audit trail
        correction_details = ""
        if was_re_evaluated:
            correction_details = (
                f"Re-evaluated via LLM self-correction ({len(re_eval_issues)} issue(s) detected). "
                f"Issues: {'; '.join(i[:80] for i in re_eval_issues[:3])}"
            )

        return TOESampleResult(
            control_id=control_id,
            sample_id=sample.sample_id,
            result=eval_result.get("result", "ERROR"),
            operated_effectively=eval_result.get("operated_effectively", "Unable to Assess"),
            control_performed=eval_result.get("control_performed", "Unable to Assess"),
            timely_execution=eval_result.get("timely_execution", "Unable to Assess"),
            accurate_execution=eval_result.get("accurate_execution", "Unable to Assess"),
            authorized_performer=eval_result.get("authorized_performer", "Unable to Assess"),
            evidence_sufficient=eval_result.get("evidence_sufficient", "Unable to Assess"),
            remarks=eval_result.get("remarks", ""),
            deviation_details=eval_result.get("deviation_details", ""),
            raw_evidence=sample.description or "",
            sample_details=eval_result.get("sample_details"),
            attribute_results=eval_result.get("attribute_results"),
            attribute_reasoning=eval_result.get("attribute_reasoning"),
            validator_corrected=was_re_evaluated,
            validator_details=correction_details,
        )

    def test_all_toe(self, toe_bank: dict, max_workers: int = 5, pre_schemas: dict = None, progress_callback=None, tod_results: list = None) -> list:
        """
        Run TOE for all controls -- parallel API calls for ALL phases:
          Phase 1: Generate schemas (parallel -- 1 API call per control)
          Phase 2: Evaluate all samples (parallel -- 1 API call per sample)
                   + Detect contradictions → Re-evaluate via second LLM call
                   + Light validation safety net

        Each sample may use 1 or 2 API calls depending on whether the detector
        finds logical contradictions that need LLM self-correction.

        Args:
            toe_bank: dict[control_id -> list[SampleEvidence]] from load_toe_evidence_folder()
            max_workers: parallel API workers (default 5)
            tod_results: optional list of TOD results. Controls with result == 'FAIL' or
                         'NO EVIDENCE' will be skipped and given placeholder results.

        Returns:
            list[TOEControlResult] -- one aggregated result per control
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed

        import random

        _set_parse_phase("TOE")

        # -- Build sets from TOD results --
        tod_pass_cids = set()
        tod_failed_cids = set()
        tod_no_evidence_cids = set()
        if tod_results is not None:
            for tr in tod_results:
                if tr.result == "PASS":
                    tod_pass_cids.add(tr.control_id)
                elif tr.result == "FAIL":
                    tod_failed_cids.add(tr.control_id)
                elif tr.result == "NO EVIDENCE":
                    tod_no_evidence_cids.add(tr.control_id)

        # Filter to controls in RCM
        valid = {}
        for cid, samples in toe_bank.items():
            if cid.startswith("_"):
                print(f"  [!]  Skipping template: {cid}")
                continue
            if cid not in self.rcm_lookup:
                print(f"  [!]  Skipping unknown control: {cid}")
                continue
            # Skip controls that failed TOD (even if evidence exists)
            if cid in tod_failed_cids:
                print(f"  [SKIP] {cid}: TOD Failed — skipping TOE evaluation")
                continue
            # Skip controls that had no evidence in TOD
            if cid in tod_no_evidence_cids:
                print(f"  [SKIP] {cid}: TOD had no evidence — skipping TOE evaluation")
                continue
            # Strict gate: if TOD results are provided, TOE runs ONLY for TOD PASS controls
            if tod_results is not None and cid not in tod_pass_cids:
                print(f"  [SKIP] {cid}: Not PASS in TOD — skipping TOE evaluation")
                continue
            valid[cid] = samples

        # Apply count_of_samples from RCM: randomly select if folder has more
        for cid in list(valid):
            rcm = self.rcm_lookup[cid]
            try:
                expected = int(rcm.count_of_samples)
            except (ValueError, TypeError, AttributeError):
                continue  # no valid count_of_samples — use all available
            available = valid[cid]
            if len(available) > expected > 0:
                selected = random.sample(available, expected)
                print(f"  [SAMPLE] {cid}: randomly selected {expected} of {len(available)} available samples")
                valid[cid] = selected

        # Count total API calls
        total_samples = sum(len(s) for s in valid.values())
        total_controls = len(valid)
        total_api_calls = total_controls + total_samples  # schemas + samples

        print("=" * 70)
        print(f"  TEST OF OPERATING EFFECTIVENESS (TOE)")
        print(f"  Controls to evaluate: {total_controls}")
        print(f"  Total samples: {total_samples}")
        print(f"  Total API calls: {total_api_calls} ({total_controls} schemas + {total_samples} samples)")
        print(f"  Parallel workers: {max_workers}")
        print("=" * 70)
        print()

        # Notify caller of total counts — use total_controls as the progress
        # denominator so the bar advances per control, not per sample.
        if progress_callback:
            progress_callback("init", 0, total_controls, {"total_controls": total_controls, "total_samples": total_samples})

        # -- Phase 1: Generate schemas (or use pre-approved ones) --
        schemas = {}  # control_id -> ControlSchema
        schema_start = time.time()

        if pre_schemas:
            # Use pre-approved schemas from user review
            fallback_count = 0
            for cid in sorted(valid):
                if cid in pre_schemas:
                    schemas[cid] = pre_schemas[cid]
                else:
                    # Fallback: generate schema for controls not in pre_schemas
                    schemas[cid] = self.evaluator.generate_schema(self.rcm_lookup[cid])
                    fallback_count += 1
            schema_elapsed = time.time() - schema_start
            print(f"  Phase 1: Using {len(pre_schemas)} pre-approved schemas"
                  + (f" (+{fallback_count} generated)" if fallback_count else ""))
            print(f"  Phase 1 complete in {schema_elapsed:.1f}s")
            print()
            if progress_callback:
                progress_callback("schema_done", 0, total_controls, {"phase": "schemas_ready"})
        else:
            print(f"  Phase 1: Generating {total_controls} testing schemas in parallel...")
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_cid = {}
                for cid in sorted(valid):
                    rcm = self.rcm_lookup[cid]
                    future = executor.submit(self.evaluator.generate_schema, rcm)
                    future_to_cid[future] = cid

                schema_done = 0
                for future in as_completed(future_to_cid):
                    cid = future_to_cid[future]
                    schema_done += 1
                    try:
                        schema = future.result()
                        schemas[cid] = schema
                        n_attrs = len(schema.attributes)
                        n_cols = len(schema.sample_columns)
                        print(f"    [{schema_done:>2}/{total_controls}] [OK] {cid}: {n_attrs} attributes, {n_cols} detail columns")
                    except Exception as e:
                        print(f"    [{schema_done:>2}/{total_controls}] [!]  {cid}: schema failed ({e}) -- using defaults")
                        schemas[cid] = ControlSchema(control_id=cid, worksteps=[], attributes=[], sample_columns=[])
                    # Report schema-generation progress so the bar isn't stuck at 0%
                    if progress_callback:
                        progress_callback("schema", schema_done, total_controls, {"control_id": cid, "phase": "schema"})

            schema_elapsed = time.time() - schema_start
            print(f"  Phase 1 complete: {total_controls} schemas in {schema_elapsed:.1f}s")
            print()
            if progress_callback:
                progress_callback("schema_done", 0, total_controls, {"phase": "schemas_ready"})

        # -- Phase 2: Evaluate all samples in PARALLEL -------------------
        print(f"  Phase 2: Evaluating {total_samples} samples in parallel...")
        tasks = []
        for cid in sorted(valid):
            samples_list = valid[cid]
            for idx, sample in enumerate(samples_list, 1):
                tasks.append((cid, sample, idx, len(samples_list)))

        # Execute all samples in parallel
        sample_results_map = {}   # control_id -> list[TOESampleResult]
        completed = 0
        eval_start = time.time()

        # Track per-control expected sample counts so we know when a control is fully done
        samples_per_control = {cid: len(samps) for cid, samps in valid.items()}
        samples_done_per_control: dict = {cid: 0 for cid in valid}
        controls_completed = 0

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {}
            for cid, sample, idx, total in tasks:
                future = executor.submit(
                    self._evaluate_toe_sample, cid, sample, idx, total,
                    schema=schemas.get(cid)
                )
                future_map[future] = (cid, sample.sample_id, idx, total)

            for future in as_completed(future_map):
                cid, sid, idx, total = future_map[future]
                completed += 1
                try:
                    result = future.result()
                    if cid not in sample_results_map:
                        sample_results_map[cid] = []
                    sample_results_map[cid].append(result)

                    icon = "[OK]" if not is_sample_fail(result) else "[FAIL]"
                    re_tag = " [RE-EVAL]" if result.validator_corrected else ""
                    print(f"  [{completed:>3}/{total_samples}] {icon} {cid} / {sid} | {'PASS' if not is_sample_fail(result) else 'FAIL'}{re_tag}")
                    sample_result_str = "PASS" if not is_sample_fail(result) else "FAIL"
                except Exception as e:
                    print(f"  [{completed:>3}/{total_samples}] [!]  {cid} / {sid} | ERROR: {e}")
                    # Create an ERROR result so the sample is NOT silently dropped
                    error_result = TOESampleResult(
                        control_id=cid, sample_id=sid,
                        result="ERROR", operated_effectively="No",
                        control_performed="Unable to Assess",
                        timely_execution="Unable to Assess",
                        accurate_execution="Unable to Assess",
                        authorized_performer="Unable to Assess",
                        evidence_sufficient="Unable to Assess",
                        remarks=f"Evaluation failed: {str(e)[:200]}",
                        deviation_details="Evaluation error -- manual review required",
                    )
                    if cid not in sample_results_map:
                        sample_results_map[cid] = []
                    sample_results_map[cid].append(error_result)
                    sample_result_str = "ERROR"

                # Track control-level completion for progress reporting
                samples_done_per_control[cid] = samples_done_per_control.get(cid, 0) + 1
                control_just_finished = samples_done_per_control[cid] >= samples_per_control.get(cid, 1)
                if control_just_finished:
                    controls_completed += 1

                if progress_callback:
                    progress_callback("eval", controls_completed, total_controls, {
                        "control_id": cid, "sample_id": sid,
                        "result": sample_result_str,
                        "samples_done": completed, "samples_total": total_samples,
                    })

        eval_elapsed = time.time() - eval_start
        total_elapsed = time.time() - schema_start
        
        # Count re-evaluations
        re_eval_count = 0
        for samples in sample_results_map.values():
            for s in samples:
                if s.validator_corrected:
                    re_eval_count += 1
        
        total_api_calls_actual = total_controls + total_samples + re_eval_count
        print(f"\n  Phase 2 complete: {total_samples} samples in {eval_elapsed:.1f}s")
        if re_eval_count > 0:
            print(f"  Re-evaluations: {re_eval_count} samples sent back to LLM for self-correction")
        print(f"  Total elapsed: {total_elapsed:.1f}s ({total_api_calls_actual} API calls, {max_workers} workers)")
        print()

        # Aggregate per control
        control_results = []
        for cid in sorted(valid):
            rcm = self.rcm_lookup[cid]
            samples = sample_results_map.get(cid, [])
            # Sort by sample_id (natural sort: S1, S2, ..., S10 not S1, S10, S2)
            import re as _re_sort
            def _natural_key(s):
                return [int(t) if t.isdigit() else t.lower() for t in _re_sort.split(r'(\d+)', s.sample_id)]
            samples.sort(key=_natural_key)
            agg = aggregate_toe_results(cid, rcm, samples)
            agg.schema = schemas.get(cid)

            # Validate sample count vs RCM expected
            try:
                expected = int(rcm.count_of_samples)
                actual = agg.total_samples
                if actual < expected:
                    print(f"  [!]  {cid}: SAMPLE SHORTFALL -- tested {actual} of {expected} expected. "
                          f"Auditor must document reason for reduced sample size.")
                    agg.overall_remarks += (
                        f" NOTE: Only {actual} of {expected} expected samples were tested."
                    )
            except (ValueError, TypeError):
                pass  # count_of_samples not a valid number

            control_results.append(agg)

        # -- Append placeholder results for TOD-failed controls --
        for cid in sorted(tod_failed_cids):
            rcm = self.rcm_lookup.get(cid)
            if not rcm:
                continue
            placeholder = TOEControlResult(
                control_id=cid,
                risk_id=rcm.risk_id,
                control_type=rcm.control_type,
                nature_of_control=rcm.nature_of_control,
                control_frequency=rcm.control_frequency,
                total_samples=0,
                passed_samples=0,
                failed_samples=0,
                deviation_rate=0.0,
                operating_effectiveness="TOD Failed",
                deficiency_type="TOD Failed — TOE not required",
                overall_remarks="Test of Design (TOD) failed for this control. TOE was not performed. The design deficiency must be remediated before operating effectiveness can be assessed.",
                sample_results=[],
                evaluation_timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                schema=pre_schemas.get(cid) if pre_schemas else None,
            )
            control_results.append(placeholder)
            print(f"  [TOD-FAIL] {cid}: TOD Failed — placeholder TOE result added")

        # -- Append placeholder results for TOD-no-evidence controls --
        for cid in sorted(tod_no_evidence_cids):
            rcm = self.rcm_lookup.get(cid)
            if not rcm:
                continue
            placeholder = TOEControlResult(
                control_id=cid,
                risk_id=rcm.risk_id,
                control_type=rcm.control_type,
                nature_of_control=rcm.nature_of_control,
                control_frequency=rcm.control_frequency,
                total_samples=0,
                passed_samples=0,
                failed_samples=0,
                deviation_rate=0.0,
                operating_effectiveness="No Evidence",
                deficiency_type="Unable to Assess",
                overall_remarks="No evidence was provided for this control in TOD. TOE was not performed.",
                sample_results=[],
                evaluation_timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                schema=None,
            )
            control_results.append(placeholder)
            print(f"  [NO-EVID] {cid}: No TOD evidence — placeholder TOE result added")

        # -- Append placeholder results for RCM controls with NO TOE evidence folder --
        evaluated_cids = set(r.control_id for r in control_results)
        no_toe_evidence_cids = sorted(
            cid for cid in self.rcm_lookup
            if cid not in evaluated_cids
        )
        for cid in no_toe_evidence_cids:
            rcm = self.rcm_lookup[cid]
            placeholder = TOEControlResult(
                control_id=cid,
                risk_id=rcm.risk_id,
                control_type=rcm.control_type,
                nature_of_control=rcm.nature_of_control,
                control_frequency=rcm.control_frequency,
                total_samples=0,
                passed_samples=0,
                failed_samples=0,
                deviation_rate=0.0,
                operating_effectiveness="No Evidence",
                deficiency_type="Unable to Assess",
                overall_remarks="Evidence folder not found — no TOE evidence was provided for this control. Manual review required.",
                sample_results=[],
                evaluation_timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                schema=None,
            )
            control_results.append(placeholder)
            print(f"  [SKIP] {cid}: TOE evidence folder not found — placeholder result added")
        if no_toe_evidence_cids:
            print(f"  {len(no_toe_evidence_cids)} control(s) had no TOE evidence folder — placeholder results added")

        # Summary
        effective = sum(1 for r in control_results if r.operating_effectiveness == "Effective")
        exceptions = sum(1 for r in control_results
                         if r.operating_effectiveness == "Effective with Exceptions")
        not_effective = sum(1 for r in control_results
                           if r.operating_effectiveness == "Not Effective")
        tod_failed_count = sum(1 for r in control_results if r.operating_effectiveness == "TOD Failed")
        no_evidence_count = sum(1 for r in control_results if r.operating_effectiveness == "No Evidence")

        print(f"\n{'='*70}")
        print(f"  TOE COMPLETE in {total_elapsed:.1f}s")
        print(f"  Controls: {len(control_results)} (RCM universe) | Evaluated: {total_controls} | Samples: {total_samples}")
        print(f"  [OK] Effective: {effective} | [!] Exceptions: {exceptions} | [FAIL] Not Effective: {not_effective}")
        if tod_failed_count:
            print(f"  [TOD-FAIL] Skipped (TOD Failed): {tod_failed_count}")
        if no_evidence_count:
            print(f"  [NO-EVID] Skipped (No Evidence): {no_evidence_count}")
        print(f"{'='*70}")

        return control_results

    def export_toe_workpaper(self, results: list, output_path: str,
                             toe_bank: dict = None, company_name: str = "",
                             prepared_by: str = "", reviewed_by: str = ""):
        """
        Export TOE results as a full workpaper Excel file:
          - Sheet 1: Control Summary
          - Sheet 2+: One sheet per control in KPMG workpaper format
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment

        wb = Workbook()

        # -- Shared styles --------------------------------------------
        hfont = Font(bold=True, color="FFFFFF", size=10)
        bfont = Font(bold=True, size=10)
        nfont = Font(size=10)
        sfont = Font(size=9)
        bdr = Border(left=Side("thin"), right=Side("thin"),
                     top=Side("thin"), bottom=Side("thin"))
        wrap = Alignment(wrap_text=True, vertical="top")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Colours matched to KPMG workpaper format
        fill_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")   # dark navy -- table headers
        fill_label = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")    # medium blue -- label cells
        fill_attr = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # light green -- attribute legend
        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # green -- Yes / Pass
        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # red  -- No / Fail
        fill_red_banner = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # solid red -- Not Effective
        fill_green_banner = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") # solid green -- Effective
        fill_amber_banner = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid") # amber -- Effective with Exceptions

        # ==============================================================
        #  SHEET 1: Summary (reuse existing logic inline)
        # ==============================================================
        ws_sum = wb.active
        ws_sum.title = "Control Summary"
        fill_rcm = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        fill_result = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_agg = PatternFill(start_color="483698", end_color="483698", fill_type="solid")
        exc_fill = PatternFill(start_color="FFEB9C", fill_type="solid")      # yellow -- Control Deficiency
        sd_fill = PatternFill(start_color="F4B084", fill_type="solid")       # orange -- Significant Deficiency
        mw_fill = PatternFill(start_color="FF4444", fill_type="solid")       # dark red -- Material Weakness
        fill_remediation = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

        # -- COSO Mapping: run LLM classification for all controls --
        from CosoMapper import (
            map_coso_for_controls,
            FINANCIAL_ASSERTION_HEADERS, IFC_COMPONENT_HEADERS, COSO_COLUMN_HEADERS,
            N_ASSERTION_COLS, N_IFC_COLS, N_COSO_COLS,
            write_assertion_headers, write_assertion_data_row, set_assertion_column_widths,
            write_ifc_headers, write_ifc_data_row, set_ifc_column_widths,
            write_coso_headers, write_coso_data_row, set_coso_column_widths,
        )
        try:
            coso_map = map_coso_for_controls(self.rcm_rows, self.evaluator)
            print(f"[COSO] Mapped {len(coso_map)} controls to COSO principles")
        except Exception as e:
            print(f"[COSO] Mapping failed ({e}) — COSO columns will be empty")
            coso_map = {}
        n_assert = N_ASSERTION_COLS
        n_ifc = N_IFC_COLS
        n_coso = N_COSO_COLS

        # -- Build dynamic summary headers: Row 1 = group headers, Row 2 = column headers --
        orig_cols = list(self.original_rcm_df.columns) if self.original_rcm_df is not None else []
        n_orig = len(orig_cols)
        assert_start_col = n_orig + 1
        ifc_start_col = assert_start_col + n_assert
        coso_start_col = ifc_start_col + n_ifc
        sep_col = coso_start_col + n_coso
        result_start_col = sep_col + 1

        result_headers = [
            ("Detailed Testing Steps", fill_rcm),
            ("Required Samples", fill_agg), ("Samples Tested", fill_agg),
            ("Passed / Total", fill_agg),
            ("Failed / Total", fill_agg), ("Deviation Rate", fill_agg),
            ("Validator Overrides", fill_agg), ("Deficiency Type", fill_result),
            ("Failed Sample Details", fill_result), ("Overall Remarks", fill_result),
            ("Remedial Actions", fill_remediation),
        ]
        n_result = len(result_headers)
        result_end_col = result_start_col + n_result - 1

        # ── Row 1: Merged group headers ──
        fill_group_rcm = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
        fill_group_assert = PatternFill(start_color="005EB8", end_color="005EB8", fill_type="solid")
        fill_group_ifc = PatternFill(start_color="483698", end_color="483698", fill_type="solid")
        fill_group_coso = PatternFill(start_color="009A44", end_color="009A44", fill_type="solid")
        fill_group_result = PatternFill(start_color="470A68", end_color="470A68", fill_type="solid")
        group_font = Font(bold=True, color="FFFFFF", size=12)

        def _merge_group(label, start, count, fill):
            if count > 0:
                ws_sum.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + count - 1)
                cell = ws_sum.cell(row=1, column=start, value=label)
                cell.font, cell.fill, cell.border = group_font, fill, bdr
                cell.alignment = Alignment(horizontal="center", vertical="center")

        _merge_group("RCM", 1, n_orig, fill_group_rcm)
        _merge_group("Financial Assertions", assert_start_col, n_assert, fill_group_assert)
        _merge_group("IFC Components", ifc_start_col, n_ifc, fill_group_ifc)
        _merge_group("COSO", coso_start_col, n_coso, fill_group_coso)

        fill_sep = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        ws_sum.cell(row=1, column=sep_col).fill = fill_sep
        ws_sum.column_dimensions[get_column_letter(sep_col)].width = 3

        _merge_group("Results", result_start_col, n_result, fill_group_result)

        # ── Row 2: Actual column headers ──
        for c, col_name in enumerate(orig_cols, 1):
            cell = ws_sum.cell(row=2, column=c, value=col_name)
            cell.font, cell.fill, cell.border, cell.alignment = hfont, fill_rcm, bdr, center

        fill_assert_hdr = PatternFill(start_color="005EB8", end_color="005EB8", fill_type="solid")
        write_assertion_headers(ws_sum, assert_start_col, header_row=2,
                                font=hfont, fill=fill_assert_hdr, border=bdr, alignment=center)

        fill_ifc_hdr = PatternFill(start_color="483698", end_color="483698", fill_type="solid")
        write_ifc_headers(ws_sum, ifc_start_col, header_row=2,
                          font=hfont, fill=fill_ifc_hdr, border=bdr, alignment=center)

        fill_coso_yes = PatternFill(start_color="D5F0DC", end_color="D5F0DC", fill_type="solid")
        fill_coso_hdr = PatternFill(start_color="009A44", end_color="009A44", fill_type="solid")
        write_coso_headers(ws_sum, coso_start_col, header_row=2,
                           font=hfont, fill=fill_coso_hdr, border=bdr, alignment=center)

        ws_sum.cell(row=2, column=sep_col).fill = fill_sep

        for c, (h, f) in enumerate(result_headers, result_start_col):
            cell = ws_sum.cell(row=2, column=c, value=h)
            cell.font, cell.fill, cell.border, cell.alignment = hfont, f, bdr, center

        # Deficiency type column index (for conditional coloring)
        deficiency_col = result_start_col + 7  # 8th result column = Deficiency Type

        # -- Write data rows (start from row 3 due to group header in row 1) --
        for ri, r in enumerate(results, 3):
            rcm = self.rcm_lookup.get(r.control_id)

            # Original RCM columns
            orig_idx = self.orig_row_map.get(r.control_id)
            if orig_idx is not None and self.original_rcm_df is not None:
                orig_row = self.original_rcm_df.iloc[orig_idx]
                for c, col_name in enumerate(orig_cols, 1):
                    val = orig_row.get(col_name, "")
                    cell = ws_sum.cell(row=ri, column=c, value=str(val) if pd.notna(val) else "")
                    cell.alignment, cell.border = wrap, bdr

            # Financial Assertions
            fill_assert_yes = PatternFill(start_color="DAEAF7", end_color="DAEAF7", fill_type="solid")
            write_assertion_data_row(ws_sum, ri, assert_start_col,
                                     mapping=coso_map.get(r.control_id),
                                     border=bdr, alignment=wrap, fill_yes=fill_assert_yes)

            # IFC Components
            fill_ifc_yes = PatternFill(start_color="E8E0F3", end_color="E8E0F3", fill_type="solid")
            write_ifc_data_row(ws_sum, ri, ifc_start_col,
                               mapping=coso_map.get(r.control_id),
                               border=bdr, alignment=wrap, fill_yes=fill_ifc_yes)

            # COSO columns
            write_coso_data_row(ws_sum, ri, coso_start_col,
                                mapping=coso_map.get(r.control_id),
                                border=bdr, alignment=wrap, fill_yes=fill_coso_yes)

            # Separator column (dark grey for every row)
            ws_sum.cell(row=ri, column=sep_col).fill = fill_sep

            # Result values
            fd = []
            for sr in r.sample_results:
                if is_sample_fail(sr):
                    d = sr.deviation_details if sr.deviation_details and sr.deviation_details.lower() not in ("none", "n/a", "see remarks") else sr.remarks
                    fd.append(f"{sr.sample_id}: {d}")
            testing_steps = ""
            if r.schema and r.schema.worksteps:
                testing_steps = "\n".join(r.schema.worksteps)
            remedial_actions, _ = _build_remediation(r)
            sum_corrected = sum(1 for sr in r.sample_results
                                if getattr(sr, 'validator_corrected', False))
            required_samples = ""
            if rcm and rcm.count_of_samples:
                required_samples = rcm.count_of_samples
            result_vals = [
                testing_steps,
                required_samples, r.total_samples,
                f"{r.passed_samples}/{r.total_samples}",
                f"{r.failed_samples}/{r.total_samples}", f"{r.deviation_rate:.1%}",
                sum_corrected if sum_corrected > 0 else "",
                r.deficiency_type, "; ".join(fd) if fd else "None", r.overall_remarks,
                remedial_actions,
            ]
            for c, v in enumerate(result_vals, sep_col + 1):
                cell = ws_sum.cell(row=ri, column=c, value=v)
                cell.alignment, cell.border = wrap, bdr

            # Warning highlight when samples tested < required samples
            if required_samples:
                try:
                    req_int = int(float(str(required_samples)))
                    if r.total_samples < req_int:
                        fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        # Highlight the "Samples Tested" cell (3rd result column)
                        samples_tested_col = sep_col + 1 + 2
                        warn_cell = ws_sum.cell(row=ri, column=samples_tested_col)
                        warn_cell.fill = fill_warn
                        warn_cell.comment = Comment(
                            f"Warning: {r.total_samples} samples tested, but {req_int} required",
                            "Control Testing Agent",
                        )
                except (ValueError, TypeError):
                    pass

            # Conditional color on Deficiency Type
            dc = ws_sum.cell(row=ri, column=deficiency_col)
            if r.deficiency_type == "None": dc.fill = fill_pass
            elif r.deficiency_type == "Control Deficiency": dc.fill = exc_fill
            elif r.deficiency_type == "Significant Deficiency": dc.fill = sd_fill
            else: dc.fill = mw_fill

        # Auto-size: original cols, COSO cols, result cols
        for i in range(1, n_orig + 1):
            ws_sum.column_dimensions[get_column_letter(i)].width = 20
        set_assertion_column_widths(ws_sum, assert_start_col)
        set_ifc_column_widths(ws_sum, ifc_start_col)
        set_coso_column_widths(ws_sum, coso_start_col)
        result_widths = [45, 16, 14, 14, 14, 12, 14, 22, 60, 60, 55]
        for i, w in enumerate(result_widths):
            ws_sum.column_dimensions[get_column_letter(sep_col + 1 + i)].width = w

        # ==============================================================
        #  PER-CONTROL WORKPAPER SHEETS
        # ==============================================================

        for r in results:
            # Skip placeholder results — no individual sheet for controls
            # with missing evidence or TOD failures
            if r.operating_effectiveness in ("No Evidence", "TOD Failed"):
                continue

            rcm = self.rcm_lookup.get(r.control_id)
            if not rcm:
                continue
            schema = r.schema

            # Sheet name: control ID (max 31 chars for Excel)
            sheet_name = r.control_id[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Helper to write a label-value pair
            def label_val(row, label, value, merge_to=8):
                lc = ws.cell(row=row, column=1, value=label)
                lc.font = bfont
                lc.fill = fill_label
                lc.border = bdr
                vc = ws.cell(row=row, column=2, value=value)
                vc.font = nfont
                vc.alignment = wrap
                vc.border = bdr
                if merge_to > 2:
                    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=merge_to)

            # -- Header section ----------------------------------------
            row = 1
            if company_name:
                ws.cell(row=row, column=1, value=company_name).font = Font(bold=True, size=14)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                row += 2
            else:
                row = 1

            label_val(row, "Process:", rcm.process)
            row += 1
            label_val(row, "Sub Process:", rcm.subprocess)
            row += 1
            label_val(row, "Control No:", rcm.control_id)
            row += 1
            label_val(row, "Control Description:", rcm.control_description)
            row += 1
            label_val(row, "Control Type:", f"{rcm.control_type} | {rcm.nature_of_control}")
            row += 1
            label_val(row, "Control Frequency:", rcm.control_frequency)
            row += 1
            label_val(row, "Control Owner:", rcm.control_owner)
            row += 1
            label_val(row, "RCM Sample Size:", rcm.count_of_samples)
            row += 1
            label_val(row, "Sample Size:", str(r.total_samples))
            row += 1

            # -- Worksteps ---------------------------------------------
            if schema and schema.worksteps:
                ws_text = "\n".join(schema.worksteps)
                label_val(row, "Worksteps Performed:", ws_text)
                # Set row height for wrapped text
                ws.row_dimensions[row].height = max(30, 15 * len(schema.worksteps))
                row += 1

            row += 1  # blank row

            # -- Attributes legend -------------------------------------
            if schema and schema.attributes:
                # Header: "Attribute" in col A+B merged, "Attribute to be tested" spanning rest
                ws.cell(row=row, column=1, value="Attribute").font = bfont
                ws.cell(row=row, column=1).fill = fill_attr
                ws.cell(row=row, column=1).border = bdr
                ws.cell(row=row, column=1).alignment = center
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.cell(row=row, column=2).fill = fill_attr
                ws.cell(row=row, column=2).border = bdr
                ws.cell(row=row, column=3, value="Attribute to be tested").font = bfont
                ws.cell(row=row, column=3).fill = fill_attr
                ws.cell(row=row, column=3).border = bdr
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
                for cc in range(3, 11):
                    ws.cell(row=row, column=cc).fill = fill_attr
                    ws.cell(row=row, column=cc).border = bdr
                row += 1
                for attr in schema.attributes:
                    ws.cell(row=row, column=1, value=attr["id"]).font = bfont
                    ws.cell(row=row, column=1).alignment = center
                    ws.cell(row=row, column=1).border = bdr
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    ws.cell(row=row, column=2).border = bdr
                    desc_text = f"{attr['name']}: {attr['description']}" if attr.get('description') else attr['name']
                    ws.cell(row=row, column=3, value=desc_text).font = nfont
                    ws.cell(row=row, column=3).alignment = wrap
                    ws.cell(row=row, column=3).border = bdr
                    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
                    for cc in range(3, 11):
                        ws.cell(row=row, column=cc).border = bdr
                    row += 1
                row += 1  # blank row

            # -- Testing table --------------------------------------
            row += 1  # blank row before table
            table_start_row = row

            # -- Column mapping --
            sr_no_col = 2          # B: Sr. No.
            sample_no_col = 3      # C: Sample No.
            detail_start_col = 4   # D onwards: Testing Sample Details

            sample_col_keys = []
            if schema and schema.sample_columns:
                for sc in schema.sample_columns:
                    sample_col_keys.append(sc["header"])
            else:
                sample_col_keys = ["Source Document"]
            n_detail_cols = len(sample_col_keys)
            detail_end_col = detail_start_col + n_detail_cols - 1

            attr_start_col = detail_end_col + 1
            attr_ids = []
            attr_names = []
            if schema and schema.attributes:
                for attr in schema.attributes:
                    attr_ids.append(attr["id"])
                    attr_names.append(attr["name"])
            n_attr_cols = len(attr_ids)
            attr_end_col = attr_start_col + n_attr_cols - 1 if n_attr_cols > 0 else attr_start_col - 1

            remarks_col = attr_end_col + 1
            # NEW: "Attribute Analysis" column — shows LLM reasoning for each attribute
            analysis_col = remarks_col
            remarks_col = analysis_col + 1
            wpref_col = remarks_col + 1
            last_col = wpref_col

            # -- Header Row 1 & Row 2 --
            h1 = row
            h2 = row + 1

            # B: "Sr. No." merged 2 rows
            c = ws.cell(row=h1, column=sr_no_col, value="Sr. No.")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            ws.merge_cells(start_row=h1, start_column=sr_no_col, end_row=h2, end_column=sr_no_col)

            # C+D...N: "Testing Sample Details" merged across sample_no + detail cols (row 1)
            tsd_start = sample_no_col
            tsd_end = detail_end_col
            c = ws.cell(row=h1, column=tsd_start, value="Testing Sample Details")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            if tsd_end > tsd_start:
                ws.merge_cells(start_row=h1, start_column=tsd_start, end_row=h1, end_column=tsd_end)

            # Row 2 sub-headers: "Sample No." then individual detail columns
            c = ws.cell(row=h2, column=sample_no_col, value="Sample No.")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            for idx_sc, sc_header in enumerate(sample_col_keys):
                col_num = detail_start_col + idx_sc
                c = ws.cell(row=h2, column=col_num, value=sc_header)
                c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center

            # Attribute columns: Row 1 = "Attribute A", Row 2 = short name
            attr_hfont = Font(bold=True, color="000000", size=10)   # black bold on green bg
            attr_sfont = Font(bold=True, color="000000", size=9)    # black bold on green bg
            for idx_a, (aid, aname) in enumerate(zip(attr_ids, attr_names)):
                col_num = attr_start_col + idx_a
                c1 = ws.cell(row=h1, column=col_num, value=f"Attribute {aid}")
                c1.font = attr_hfont; c1.fill = fill_attr; c1.border = bdr; c1.alignment = center
                c2 = ws.cell(row=h2, column=col_num, value=aname)
                c2.font = attr_sfont; c2.fill = fill_attr; c2.border = bdr
                c2.alignment = center

            # "Attribute Analysis" merged 2 rows — LLM reasoning for each attribute
            c = ws.cell(row=h1, column=analysis_col, value="Attribute Analysis")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            ws.merge_cells(start_row=h1, start_column=analysis_col, end_row=h2, end_column=analysis_col)

            # "Remarks" merged 2 rows
            c = ws.cell(row=h1, column=remarks_col, value="Remarks")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            ws.merge_cells(start_row=h1, start_column=remarks_col, end_row=h2, end_column=remarks_col)

            # "Workpaper Reference" merged 2 rows
            c = ws.cell(row=h1, column=wpref_col, value="Workpaper Reference")
            c.font = hfont; c.fill = fill_header; c.border = bdr; c.alignment = center
            ws.merge_cells(start_row=h1, start_column=wpref_col, end_row=h2, end_column=wpref_col)

            # Ensure borders on all header cells
            for col_num in range(2, last_col + 1):
                for hr in (h1, h2):
                    ws.cell(row=hr, column=col_num).border = bdr

            row = h2 + 1  # data starts after 2-row header

            # -- Sample data rows --------------------------------------
            data_start_row = row
            for idx, sr in enumerate(r.sample_results, 1):
                # B: Sr. No.
                ws.cell(row=row, column=sr_no_col, value=idx).border = bdr
                ws.cell(row=row, column=sr_no_col).alignment = center

                # C: Sample No. (S1, S15, Q1-2024, etc.)
                ws.cell(row=row, column=sample_no_col, value=sr.sample_id).border = bdr
                ws.cell(row=row, column=sample_no_col).alignment = center
                ws.cell(row=row, column=sample_no_col).font = nfont

                # D+: Detail columns
                if schema and schema.sample_columns:
                    for idx_sc, sc in enumerate(schema.sample_columns):
                        col_num = detail_start_col + idx_sc
                        val = ""
                        if sr.sample_details and isinstance(sr.sample_details, dict):
                            hdr = str(sc.get("header", ""))
                            key = str(sc.get("key", ""))
                            val = sr.sample_details.get(hdr, "")
                            if val in (None, "") and key:
                                val = sr.sample_details.get(key, "")
                            if val in (None, ""):
                                ci = {str(k).strip().lower(): v for k, v in sr.sample_details.items()}
                                if hdr:
                                    val = ci.get(hdr.lower(), "")
                                if val in (None, "") and key:
                                    val = ci.get(key.lower(), "")
                        cell = ws.cell(row=row, column=col_num, value=val)
                        cell.border = bdr; cell.alignment = wrap; cell.font = nfont
                else:
                    ws.cell(row=row, column=detail_start_col, value=sr.raw_evidence[:80] if sr.raw_evidence else "").border = bdr

                # Attribute result columns — values from LLM, with reasoning as comments
                for idx_a, aid in enumerate(attr_ids):
                    col_num = attr_start_col + idx_a
                    val = ""
                    if sr.attribute_results and isinstance(sr.attribute_results, dict):
                        val = sr.attribute_results.get(aid, "")
                    cell = ws.cell(row=row, column=col_num, value=val)
                    cell.border = bdr; cell.alignment = center; cell.font = nfont
                    # Cell coloring uses normalized value — LLM-trusting
                    norm_val = _normalize_attr_value(str(val))
                    if norm_val == "Yes":
                        cell.fill = fill_pass
                    elif norm_val == "No":
                        cell.fill = fill_fail
                    # N/A and AMBIGUOUS: no color (trust LLM's overall result)

                    # Add LLM reasoning as cell comment (hover to see WHY)
                    if sr.attribute_reasoning and isinstance(sr.attribute_reasoning, dict):
                        reason_text = sr.attribute_reasoning.get(aid, "")
                        if reason_text and str(reason_text).strip().lower() not in ("", "n/a", "none"):
                            cell.comment = Comment(
                                f"LLM Assessment:\n{reason_text}",
                                "TOE Evaluator", width=350, height=120
                            )

                # NEW: Attribute Analysis column — all LLM reasoning in one cell
                analysis_parts = []
                if sr.attribute_reasoning and isinstance(sr.attribute_reasoning, dict):
                    attr_name_map = {}
                    if schema and schema.attributes:
                        attr_name_map = {a["id"]: a["name"] for a in schema.attributes}
                    for aid in attr_ids:
                        aval = sr.attribute_results.get(aid, "?") if sr.attribute_results else "?"
                        areason = sr.attribute_reasoning.get(aid, "")
                        aname = attr_name_map.get(aid, f"Attr {aid}")
                        if areason and str(areason).strip().lower() not in ("", "n/a", "none"):
                            analysis_parts.append(f"{aid} ({aname}): {aval} — {areason}")
                        else:
                            analysis_parts.append(f"{aid} ({aname}): {aval}")
                analysis_text = "\n".join(analysis_parts) if analysis_parts else ""
                ac = ws.cell(row=row, column=analysis_col, value=analysis_text)
                ac.border = bdr; ac.alignment = wrap; ac.font = nfont

                # Determine if this sample is a failure (unified logic)
                is_fail = is_sample_fail(sr)

                # Remarks — build from LLM's reasoning (single source, no duplication)
                rmk_parts = []
                if is_fail:
                    rmk_parts.append("[FAIL]")
                    
                    # For FAIL samples: use LLM attribute reasoning as PRIMARY source
                    has_attr_reasoning = False
                    fail_reasons = []
                    
                    # List failing attributes with LLM reasoning
                    if sr.attribute_results and isinstance(sr.attribute_results, dict):
                        if schema and schema.attributes:
                            attr_name_map = {a["id"]: a["name"] for a in schema.attributes}
                        else:
                            attr_name_map = {}
                        for aid, val in sr.attribute_results.items():
                            if _normalize_attr_value(str(val)) == "No":
                                aname = attr_name_map.get(aid, f"Attribute {aid}")
                                reason = ""
                                if sr.attribute_reasoning and isinstance(sr.attribute_reasoning, dict):
                                    reason = sr.attribute_reasoning.get(aid, "")
                                if reason:
                                    fail_reasons.append(f"{aname}: {reason.rstrip('.')}")
                                    has_attr_reasoning = True
                                else:
                                    fail_reasons.append(f"{aname} = {val}")
                    
                    # Only add standard field labels if NOT already covered by attribute reasoning
                    if not has_attr_reasoning:
                        for fn, label in (("control_performed", "Control not performed"),
                                          ("timely_execution", "Not timely"),
                                          ("accurate_execution", "Inaccurate"),
                                          ("authorized_performer", "Unauthorized"),
                                          ("evidence_sufficient", "Insufficient evidence")):
                            fv = getattr(sr, fn, "")
                            if _is_negative(str(fv)):
                                fail_reasons.append(label)
                    
                    if fail_reasons:
                        rmk_parts.append("Deviation: " + "; ".join(fail_reasons) + ".")
                    
                    # Only append deviation_details as FALLBACK when no attribute reasoning exists
                    if not has_attr_reasoning and not fail_reasons:
                        if sr.deviation_details and sr.deviation_details.lower() not in (
                                "none", "n/a", "see remarks", ""):
                            clean_dev = _strip_pass_language(sr.deviation_details)
                            if clean_dev and clean_dev.lower() not in ("none", "n/a", ""):
                                rmk_parts.append(clean_dev)
                
                else:
                    # PASS sample — use LLM's remarks directly
                    if sr.remarks and sr.remarks.strip().lower() not in ("none", "n/a", "ok"):
                        clean_pass_rmk = sr.remarks
                        # Strip ALL internal audit trail tags (not for workpaper)
                        # Tags: [AUTO-CORRECTED: ...], [VALIDATOR: ...], [ATTR FIX: ...]
                        import re as _re
                        clean_pass_rmk = _re.sub(r'\s*\[(?:AUTO-CORRECTED|VALIDATOR|ATTR FIX)[^\]]*\]', '', clean_pass_rmk).strip()
                        if clean_pass_rmk:
                            rmk_parts.append(clean_pass_rmk)
                
                rmk_text = " ".join(rmk_parts) if rmk_parts else ""
                # Final cleanup: fix double periods and trailing whitespace
                while ".." in rmk_text:
                    rmk_text = rmk_text.replace("..", ".")
                rmk_text = rmk_text.strip()
                cell = ws.cell(row=row, column=remarks_col, value=rmk_text)
                cell.border = bdr; cell.alignment = wrap; cell.font = nfont
                if is_fail:
                    cell.font = Font(size=10, color="C00000")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif getattr(sr, 'validator_corrected', False):
                    # Validator-corrected PASS: italic blue font + light blue fill
                    cell.font = Font(size=10, italic=True, color="1F4E79")
                    cell.fill = PatternFill(start_color="DAEAF7", end_color="DAEAF7", fill_type="solid")
                    # Add hover comment explaining the correction
                    correction_note = getattr(sr, 'validator_details', '')
                    if correction_note:
                        cell.comment = Comment(
                            f"Validator Override:\n{correction_note}\n\n"
                            "The LLM's detailed attribute assessments all confirmed the control "
                            "operated correctly, but its summary result contradicted this. "
                            "The validator corrected the result based on the attribute-level evidence.",
                            "TOE Validator", width=400, height=160
                        )
                    # Also style the sample ID cell with italic to flag it at a glance
                    sample_cell = ws.cell(row=row, column=sample_no_col)
                    sample_cell.font = Font(size=10, italic=True, color="1F4E79")

                # Workpaper Reference -- unique per sample for audit trail
                wp_ref = f"{r.control_id}.T.{sr.sample_id}"
                ws.cell(row=row, column=wpref_col, value=wp_ref).border = bdr
                ws.cell(row=row, column=wpref_col).alignment = center
                ws.cell(row=row, column=wpref_col).font = nfont

                row += 1

            data_end_row = row - 1

            # -- Summary totals row after testing data --------------------
            row_summary = row
            ws.cell(row=row_summary, column=sr_no_col, value="TOTAL").font = bfont
            ws.cell(row=row_summary, column=sr_no_col).alignment = center
            ws.cell(row=row_summary, column=sr_no_col).border = bdr
            ws.cell(row=row_summary, column=sr_no_col).fill = fill_label
            total_text = (f"Samples: {r.total_samples} | "
                          f"Passed: {r.passed_samples} | "
                          f"Failed: {r.failed_samples} | "
                          f"Deviation Rate: {r.deviation_rate:.1%}")
            # Count validator-corrected samples for transparency
            corrected_count = sum(1 for sr in r.sample_results 
                                  if getattr(sr, 'validator_corrected', False))
            if corrected_count > 0:
                total_text += f" | Validator Overrides: {corrected_count}"
            tc = ws.cell(row=row_summary, column=sample_no_col, value=total_text)
            tc.font = bfont; tc.alignment = wrap; tc.border = bdr; tc.fill = fill_label
            ws.merge_cells(start_row=row_summary, start_column=sample_no_col,
                           end_row=row_summary, end_column=last_col)
            for cc in range(sample_no_col, last_col + 1):
                ws.cell(row=row_summary, column=cc).border = bdr
                ws.cell(row=row_summary, column=cc).fill = fill_label
            row += 1

            # Legend row for validator-corrected samples (if any)
            if corrected_count > 0:
                legend_row = row
                legend_cell = ws.cell(row=legend_row, column=sample_no_col,
                    value=("Legend: Rows in italic blue text with light blue fill indicate samples where the "
                           "validator corrected the LLM's result. The LLM's detailed attribute assessments confirmed "
                           "the control operated correctly, but its summary result contradicted this. "
                           "Hover over the Remarks cell for details."))
                legend_cell.font = Font(size=9, italic=True, color="1F4E79")
                legend_cell.alignment = wrap
                ws.merge_cells(start_row=legend_row, start_column=sample_no_col,
                               end_row=legend_row, end_column=last_col)
                row += 1

            data_end_row = row - 1

            # -- "Testing:" label in column A, merged vertically across entire table --
            testing_cell = ws.cell(row=table_start_row, column=1, value="Testing:")
            testing_cell.font = bfont
            testing_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            testing_cell.border = bdr
            if data_end_row > table_start_row:
                ws.merge_cells(start_row=table_start_row, start_column=1,
                               end_row=data_end_row, end_column=1)
            for rr in range(table_start_row, data_end_row + 1):
                ws.cell(row=rr, column=1).border = bdr

            row += 1  # blank row

            # -- Conclusion banner -------------------------------------
            conc_text = r.operating_effectiveness
            if r.deficiency_type != "None":
                conc_text += f" -- {r.deficiency_type}"
            conc_fill = fill_green_banner if r.operating_effectiveness == "Effective" else \
                        fill_amber_banner if r.operating_effectiveness == "Effective with Exceptions" else \
                        fill_red_banner

            ws.cell(row=row, column=1, value="Conclusion:").font = bfont
            ws.cell(row=row, column=1).border = bdr
            conc_cell = ws.cell(row=row, column=2, value=conc_text)
            conc_cell.font = Font(bold=True, color="FFFFFF", size=11)
            conc_cell.fill = conc_fill
            conc_cell.border = bdr
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 2

            # -- Sign-off section --------------------------------------
            ws.cell(row=row, column=1, value="Prepared by").font = bfont
            ws.cell(row=row, column=1).border = bdr
            ws.cell(row=row, column=2, value=prepared_by).border = bdr
            ws.cell(row=row, column=4, value="Reviewed by").font = bfont
            ws.cell(row=row, column=4).border = bdr
            ws.cell(row=row, column=5, value=reviewed_by).border = bdr
            row += 1
            ws.cell(row=row, column=1, value="Date").font = bfont
            ws.cell(row=row, column=1).border = bdr
            ws.cell(row=row, column=2, value=time.strftime("%d-%b-%Y")).border = bdr
            ws.cell(row=row, column=4, value="Date").font = bfont
            ws.cell(row=row, column=4).border = bdr
            ws.cell(row=row, column=5, value="").border = bdr

            # -- Column widths -----------------------------------------
            ws.column_dimensions["A"].width = 14   # "Testing:" label
            ws.column_dimensions["B"].width = 8    # Sr. No.
            ws.column_dimensions["C"].width = 12   # Sample No.
            for c_idx in range(detail_start_col, detail_end_col + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 22
            for c_idx in range(attr_start_col, attr_end_col + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 20
            ws.column_dimensions[get_column_letter(analysis_col)].width = 55
            ws.column_dimensions[get_column_letter(remarks_col)].width = 55
            ws.column_dimensions[get_column_letter(wpref_col)].width = 18

        # ============================================================
        # INTEGRITY VERIFICATION — audit tick-mark equivalent
        # Verify all counts and markers are internally consistent
        # before signing off on the workpaper.
        # ============================================================
        integrity_errors = []
        for r in results:
            ctrl_errors = verify_control_integrity(r)
            integrity_errors.extend(ctrl_errors)

        if integrity_errors:
            print(f"\n[!] INTEGRITY WARNINGS ({len(integrity_errors)}):")
            for e in integrity_errors:
                print(f"    {e}")
            print(f"    Workpaper saved but contains inconsistencies -- auditor must review.\n")
        else:
            print(f"[Integrity] All {len(results)} controls verified: "
                  f"aggregate counts match is_sample_fail() recount, "
                  f"all sample result fields consistent.")

        wb.save(output_path)
        print(f"[Export] TOE workpaper saved to {output_path} ({len(results)} control sheets)")

    def generate_toe_report(self, results: list) -> str:
        """Generate text summary report for TOE results."""
        lines = []
        lines.append(f"{'='*75}")
        lines.append(f"  TEST OF OPERATING EFFECTIVENESS (TOE) -- SUMMARY REPORT")
        lines.append(f"  Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"{'='*75}")

        for r in results:
            rcm = self.rcm_lookup.get(r.control_id)
            risk = rcm.risk_title if rcm else ""

            icon = "[OK]" if r.operating_effectiveness == "Effective" else \
                   "[!]" if r.operating_effectiveness == "Effective with Exceptions" else "[FAIL]"

            lines.append(f"\n{'-'*75}")
            lines.append(f"  {r.control_id} | {risk}")
            lines.append(f"  {r.control_type} | {r.nature_of_control} | {r.control_frequency}")
            lines.append(f"  {icon} TOE: {r.operating_effectiveness} | "
                         f"Samples: {r.passed_samples}/{r.total_samples} passed | "
                         f"Deviation Rate: {r.deviation_rate:.1%}")
            lines.append(f"     Deficiency: {r.deficiency_type}")
            lines.append(f"     {r.overall_remarks}")

            # Show individual sample results
            for sr in r.sample_results:
                s_icon = "[FAIL]" if is_sample_fail(sr) else "[OK]"
                lines.append(f"     {s_icon} {sr.sample_id}: {'FAIL' if is_sample_fail(sr) else 'PASS'}")
                if sr.deviation_details and sr.deviation_details.lower() not in ("none", "n/a"):
                    lines.append(f"        Deviation: {sr.deviation_details}")

        effective = sum(1 for r in results if r.operating_effectiveness == "Effective")
        exceptions = sum(1 for r in results
                         if r.operating_effectiveness == "Effective with Exceptions")
        not_effective = sum(1 for r in results
                           if r.operating_effectiveness == "Not Effective")
        total_samples = sum(r.total_samples for r in results)
        total_deviations = sum(r.failed_samples for r in results)
        mw = sum(1 for r in results if r.deficiency_type == "Material Weakness")
        sd = sum(1 for r in results if r.deficiency_type == "Significant Deficiency")

        lines.append(f"\n{'='*75}")
        lines.append(f"  OVERALL")
        lines.append(f"  Controls Evaluated:     {len(results)}")
        lines.append(f"  Total Samples Tested:   {total_samples}")
        lines.append(f"  Total Deviations:       {total_deviations}")
        lines.append(f"  [OK] Effective:            {effective}")
        lines.append(f"  [!] Effective w/ Exceptions: {exceptions}")
        lines.append(f"  [FAIL] Not Effective:        {not_effective}")
        if mw: lines.append(f"  [X] Material Weaknesses: {mw}")
        if sd: lines.append(f"  [!] Significant Deficiencies: {sd}")
        lines.append(f"{'='*75}")
        return "\n".join(lines)




# ===========================================================================
#  DEMO RUNNER
# ===========================================================================

if __name__ == "__main__":

    # ===================================================================
    #  CONFIGURATION -- Edit these values
    # ===================================================================

    # Azure OpenAI — from central engines/config.py
    from engines.config import (
        AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY,
        AZURE_OPENAI_API_VERSION, AZURE_OPENAI_DEPLOYMENT as AZURE_OPENAI_DEPLOYMENT_NAME,
    )
    OPENAI_MODEL     = AZURE_OPENAI_DEPLOYMENT_NAME

    RCM_PATH         = "RCM_With_Samples.xlsx"
    EVIDENCE_PATH    = "evidence_toe"    # folder with sample .txt files per control
    MAX_WORKERS      = 5                 # parallel API calls

    # Workpaper metadata (for per-control sheets)
    COMPANY_NAME     = ""                # e.g. "Acme Corporation"
    PREPARED_BY      = ""                # e.g. "Nisarg Thakkar"
    REVIEWED_BY      = ""                # e.g. "Nimisha Jain"

    # ===================================================================
    #  RUN
    # ===================================================================

    # Step 1: Load TOE evidence
    toe_bank = load_toe_evidence_folder(EVIDENCE_PATH)

    # Step 2: Initialize tester
    tester = RCMControlTester(
        rcm_path=RCM_PATH,
        openai_api_key=AZURE_OPENAI_API_KEY,
        openai_model=OPENAI_MODEL,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        azure_api_key=AZURE_OPENAI_API_KEY,
        azure_deployment=OPENAI_MODEL,
        azure_api_version=AZURE_OPENAI_API_VERSION,
    )

    # Step 3: Run TOE (includes schema generation + sample evaluation)
    results = tester.test_all_toe(toe_bank, max_workers=MAX_WORKERS)

    # Step 4: Export full workpaper (Control Summary + per-control sheets)
    tester.export_toe_workpaper(
        results, "toe_workpaper.xlsx", toe_bank=toe_bank,
        company_name=COMPANY_NAME, prepared_by=PREPARED_BY, reviewed_by=REVIEWED_BY,
    )

    # Step 5: Generate report
    report = tester.generate_toe_report(results)
    print(report)
    with open("toe_report.txt", "w") as f:
        f.write(report)
    print(f"\n[Report] Saved to toe_report.txt")

    # Step 6: Summary table
    print(f"\n\nTOE Summary:")
    print(f"{'Control ID':>11s} {'Samples':>8s} {'Passed':>7s} "
          f"{'Failed':>7s} {'Dev Rate':>9s} "
          f"{'Effectiveness':<28s} {'Deficiency Type'}")
    for r in results:
        print(f" {r.control_id:>10s} {r.total_samples:>8d} "
              f"{r.passed_samples:>7d} {r.failed_samples:>7d} "
              f"{r.deviation_rate:>8.1%} "
              f"{r.operating_effectiveness:<28s} {r.deficiency_type}")
