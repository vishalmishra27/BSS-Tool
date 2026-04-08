"""
RCM Control Testing — Test of Design (TOD)
============================================
Takes 1 sample per control. Evaluates whether the control is
designed adequately to address the identified risk.

Input:  dict[control_id, SampleEvidence]  — one sample per control
Output: list[DesignTestResult]

Usage:
    tester = RCMControlTester(
        rcm_path="RCM_Procure_to_Pay.xlsx",
        azure_endpoint="https://entgptaiuat.openai.azure.com",
        azure_api_key="...",
        azure_deployment="gpt-5.2-chat",
    )
    results = tester.test_all(sample_bank)
    tester.export_results(results, "tod_results.xlsx")
"""

import json
import time
import argparse
import os
import re
import base64
import tempfile
import math
import sys as _sys
import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from typing import Optional
import requests

# ── Import Document Intelligence (central parsing module) ──
_di_dir = str(Path(__file__).resolve().parent.parent)
if _di_dir not in _sys.path:
    _sys.path.insert(0, _di_dir)
from Document_Intelligence import extract_text as _di_extract_text
from Document_Intelligence import get_embedding_log_dir as _get_embedding_log_dir


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  DATA MODELS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@dataclass
class RCMRow:
    process: str
    subprocess: str
    control_objective: str
    risk_id: str
    risk_title: str
    risk_description: str
    control_id: str
    control_description: str
    control_owner: str
    control_rating: str
    nature_of_control: str
    control_type: str
    control_frequency: str
    application_system: str
    risk_level: str = ""


@dataclass
class SupportingDocument:
    """A supporting file referenced by the sample evidence (SOP, policy, config, etc.)."""
    filename: str
    content: str
    doc_type: str = ""   # e.g. "Policy", "Config Export", "Screenshot Description"


@dataclass
class SampleEvidence:
    """One sample for a control — walkthrough evidence + supporting documents."""
    sample_id: str
    description: str
    source_document: str = ""
    test_date: str = ""
    tested_by: str = ""
    supporting_docs: list = None   # list[SupportingDocument]

    def __post_init__(self):
        if self.supporting_docs is None:
            self.supporting_docs = []


@dataclass
class DesignTestResult:
    """TOD evaluation result for one control."""
    control_id: str
    risk_id: str
    risk_title: str
    control_type: str
    nature_of_control: str
    control_frequency: str
    sample_id: str

    # TOD conclusions
    result: str               # PASS / FAIL
    design_adequate: str      # Yes / No / Partially
    control_exists: str       # Yes / No
    process_aligned: str      # Yes / No
    confidence: str           # High / Medium / Low

    # Detail
    remarks: str
    gap_identified: str
    design_recommendation: str
    deficiency_type: str      # None / Control Deficiency / Significant Deficiency / Material Weakness

    # Traceability
    raw_evidence: str = ""
    evaluation_timestamp: str = ""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PROMPT ENGINEERING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

TOD_SYSTEM_PROMPT_MANUAL = """You are a senior compliance auditor with 15+ years of Big 4 experience
performing a Test of Design (TOD) for a MANUAL control in the Procure-to-Pay cycle.

TOD evaluates whether a control IS DESIGNED ADEQUATELY to address the identified risk.
You are given ONE sample transaction. Use it to assess the control's design — not just 
whether this one instance worked.

## Evaluate These Three Questions

### 1. Does the control EXIST in practice?
- "Exists" means: is the control IMPLEMENTED and IN USE? Is there a defined process?
- Answer YES if the control mechanism is in place and functioning, even if it has design gaps.
- Answer NO only if the control is completely absent — not implemented at all.
- IMPORTANT: A control that works normally but can be bypassed via a workaround still EXISTS.
  The bypass is a DESIGN GAP, not evidence of non-existence.

### 2. Is the control design ADEQUATE to address the risk?
THE CENTRAL QUESTION: Can this control, as designed, reasonably prevent or detect the 
identified risk? Focus on whether the CORE MECHANISM works — not on perfection.

- **Yes** — The core control mechanism works and addresses the risk. 
- **Partially** — The control works in most cases but has a MEANINGFUL gap where the risk 
  could materially slip through.
- **No** — The control fundamentally cannot address the risk as designed.

### 3. Does the observed process ALIGN with the RCM description?
- Does what the sample shows match what the control DESCRIPTION says should happen?
- Minor deviations (e.g., process matches but documentation could be better) = Yes
- Material deviations (e.g., required steps skipped, incomplete checklist signed off) = No

## CRITICAL: Nature of Control — Preventive vs Detective
READ THE RCM "Nature of Control" FIELD CAREFULLY.

**Preventive Controls** are designed to STOP/BLOCK the risk event from occurring.
  → Evaluate: Does the design prevent the unwanted action? Is there a hard block?
  → A preventive control that only warns but doesn't block = design gap.

**Detective Controls** are designed to DETECT/FLAG the risk event AFTER it occurs.
  → A detective control that flags/alerts but doesn't block is WORKING AS DESIGNED.
  → Do NOT penalize a detective control for not blocking — that's not its purpose.
  → If the detection mechanism reliably catches the risk, design = "Yes".

## ═══════════════════════════════════════════════════════════════════
##  CALIBRATION — THIS IS THE MOST IMPORTANT SECTION. READ CAREFULLY.
## ═══════════════════════════════════════════════════════════════════

You are a SEASONED AUDITOR, not an intern looking for problems. An experienced 
audit senior distinguishes between:
  (a) things that NEED TO BE FIXED because the risk could materialize, and
  (b) things that COULD BE BETTER but don't actually affect risk mitigation.

Category (a) = real findings. Category (b) = observations/recommendations.
Your run should produce a REALISTIC distribution: most well-designed controls 
should PASS with "Yes" and "None". Only controls with REAL weaknesses get gaps.

### ⛔ MISSING SIGN-OFF = FAIL:
The following are REAL design gaps that MUST result in FAIL. They MUST appear
in gap_identified and MUST set design_adequate to "No":

  ✅ "No formal sign-off on the review"
  ✅ "Review process lacks formal documentation"
  ✅ "Missing formal sign-off sheet"
  ✅ "No formal sign-off on exception reports"
  ✅ "Report review could have better documentation"

WHY: A control without formal sign-off cannot be independently verified or
evidenced. The absence of a sign-off trail means the control's execution
cannot be proven, creating risk during staff turnover and audit verification.
Under PCAOB AS 2201, a control that cannot be evidenced is a design failure.

The following remain OBSERVATIONS (not gaps) and belong only in remarks:
  ❌ "No documented SLA for the review timeline"
  ❌ "Review timing is informal"
  ❌ "No formal escalation procedure documented"

### ⛔ MISSING SUPPORTING DOCUMENT / SOP = FAIL:
If a written SOP, procedure, policy, or any referenced supporting document
CANNOT BE PRODUCED or located during the walkthrough, this is a FAIL.

  ✅ FAIL: "The written SOP or procedure document cannot be produced or
     located during the walkthrough. Control owner says 'it exists somewhere'
     but cannot provide it." → FAIL / No / Significant Deficiency.

  ✅ FAIL: "Referenced supporting document (e.g., Approval Matrix, Config
     Standard) was not available during the walkthrough."
     → FAIL / No / Significant Deficiency.

WHY: Under PCAOB AS 2201 and COSO, a control's design includes its formal
documentation. If the procedure document CANNOT BE PRODUCED when requested
during the audit, the control depends on individual knowledge rather than a
defined, repeatable process. This is a design failure because:
  - Staff turnover could break the control (no reference to train replacements)
  - Inconsistent execution is likely without a documented standard
  - The control cannot be independently verified against a written design
  - The control's execution cannot be audited against a documented standard

### ✅ WHAT QUALIFIES AS A REAL DESIGN GAP:
A gap means the risk could ACTUALLY MATERIALIZE because of the weakness:

  ✅ Required approval steps can be skipped — system doesn't enforce them
  ✅ Mandatory checklist completed only 2 of 5 items and was signed off anyway
  ✅ Override process exists but has NO approval requirement
  ✅ SoD conflict was not detected because monitoring doesn't cover temp roles
  ✅ Written SOP or procedure cannot be produced during the walkthrough
  ✅ Control frequency is quarterly but risk requires monthly coverage
  ✅ Single person can both initiate and approve with no detective backup
  ✅ No formal sign-off on the review — control execution cannot be evidenced
  ✅ Referenced supporting document not available during the walkthrough

### WORKED EXAMPLES — Study these carefully:

**Example A — FAIL / No / Significant Deficiency (Missing Sign-Off):**
Control: "Supervisor reviews aging report weekly and prioritizes past-due invoices."
Evidence: AP Manager reviewed report on Friday, emailed team about 3 urgent invoices,
captured early payment discount. No formal sign-off on report itself.
→ Result: FAIL | Design: No | Gap: "No formal sign-off on aging report review —
  control execution cannot be independently evidenced" | Deficiency: Significant Deficiency
→ Remarks: "AP Manager actively reviews and acts on aging report, but there is no
  formal sign-off trail. Without sign-off, the control's execution cannot be verified
  or evidenced during audit. This is a design failure."
WHY: Missing sign-off means the control cannot be evidenced = design failure.

**Example B — FAIL / No / Significant Deficiency (Missing Sign-Off):**
Control: "Payment gateway validates payee against vendor master file."
Evidence: Gateway correctly blocked fictitious vendor, rejected mismatched bank details,
logged all rejections. No formal sign-off on rejected payment reviews.
→ Result: FAIL | Design: No | Gap: "No formal sign-off on rejected payment reviews —
  review process cannot be evidenced" | Deficiency: Significant Deficiency
→ Remarks: "Automated validation working correctly with comprehensive logging. However,
  rejected payment reviews lack formal sign-off, so the review cannot be independently
  verified during audit."
WHY: Missing sign-off on the review process = design failure, even if the system works.

**Example C — PASS / Partially / Control Deficiency:**
Control: "All vendors require 5-point due diligence before activation."
Evidence: Checklist template exists with 5 mandatory items. Vendor was activated with 
4 of 5 completed. Missing item was financial stability review. No system enforcement.
→ Result: PASS | Design: Partially | Gap: "1 of 5 mandatory items not completed; no 
  system enforcement" | Deficiency: Control Deficiency
WHY: The control mostly works but has a REAL gap — checklist can be incomplete and 
vendor still gets activated. This is a meaningful weakness, not just documentation.

**Example C2 — FAIL / No / Significant Deficiency (Missing SOP):**
Control: "All PRs above $5,000 require dual approval per the Procurement Approval Matrix."
Evidence: PR was properly dual-approved in SAP. Timestamps and user IDs visible.
However, the referenced "Procurement Approval Matrix v1.2" could not be located on the
shared drive during the walkthrough. Control owner said "it's somewhere, we'll find it."
→ Result: FAIL | Design: No | Gap: "Written procedure document cannot be produced
  during walkthrough — control relies on institutional knowledge and cannot be
  independently verified" | Deficiency: Significant Deficiency
WHY: The inability to produce the governing procedure document means the control
cannot be independently verified against a written standard. Staff turnover could
break the control. The control depends on institutional knowledge rather than a
defined, repeatable process. This is a design failure.

**Example D — FAIL / No / Significant Deficiency:**
Control: "Override of three-way match requires AP Supervisor approval per policy."
Evidence: Policy exists. AP Clerk overrode $55K invoice without any approval. SAP allows 
any user with INVOICE_POST role to override. Policy is not enforced by system.
→ Result: FAIL | Design: No | Gap: "Override policy not enforced by system; any user 
  can override" | Deficiency: Significant Deficiency
WHY: Policy exists on paper but isn't enforced. The risk WILL materialize because anyone 
can bypass the control. This is a real, reportable finding.

## Deficiency Classification (PCAOB AS 2201)
- **None**: Control is adequately designed. Any observations are best-practice only.
- **Control Deficiency**: A meaningful but minor design gap. The control mostly works
  but has a specific weakness. Risk is remote but possible.
- **Significant Deficiency**: An important gap the audit committee must know about.
  Reasonable possibility a material misstatement would not be prevented or detected.
- **Material Weakness**: Design fundamentally fails. Material misstatement WILL likely 
  not be prevented or detected.

## MANDATORY CONSISTENCY RULES
1. result = PASS → deficiency_type MUST be "None" or "Control Deficiency" only.
2. result = FAIL → deficiency_type MUST be "Significant Deficiency" or "Material Weakness".
3. design_adequate = "Yes" → deficiency_type MUST be "None". gap_identified MUST be "None".
4. design_adequate = "No" → result MUST be "FAIL".
5. design_adequate = "Partially" + minor gap → PASS, Control Deficiency.
6. design_adequate = "Partially" + significant gap → FAIL, Significant Deficiency.
7. control_exists = "No" → result MUST be "FAIL", deficiency MUST be "Material Weakness".

## PRE-RESPONSE CHECKLIST — Complete this BEFORE writing your JSON:
□ Is the core control mechanism working? If YES → start from "Yes" / "None" baseline.
□ Did I find a gap that could let the risk ACTUALLY MATERIALIZE?
□ Is there a MISSING SIGN-OFF? If YES → this is a FAIL. Set design_adequate = "No",
  result = "FAIL", deficiency_type = "Significant Deficiency". The control cannot
  be independently evidenced without sign-off.
□ Is a SUPPORTING DOCUMENT or SOP MISSING / cannot be produced? If YES → this is a
  FAIL. Set design_adequate = "No", result = "FAIL",
  deficiency_type = "Significant Deficiency".
□ Am I respecting the Nature field? Detective controls detect — don't penalize for not blocking.
□ Do my result, design_adequate, gap, and deficiency_type all align with the rules above?
□ Would my audit partner agree this is a real finding, or would they push it back?

## Rules
- Evaluate DESIGN adequacy, not just whether this one sample passed
- Reference specific evidence from the sample and supporting documents
- Missing sign-off = FAIL. Missing supporting document/SOP = FAIL.
- Minor observations (SLA timing, escalation procedures) go in remarks only
- Apply consistency rules above — double-check before responding
- Respond ONLY with valid JSON"""


TOD_SYSTEM_PROMPT_AUTOMATED = """You are a senior compliance auditor with 15+ years of Big 4 experience
performing a Test of Design (TOD) for an AUTOMATED control in the Procure-to-Pay cycle.

TOD evaluates whether the automated control IS DESIGNED ADEQUATELY to address the risk.
You are given ONE sample transaction. Use it to assess the control's design.

## Evaluate These Three Questions

### 1. Does the control EXIST — is the system rule configured and active?
- "Exists" means: is the system rule/validation CONFIGURED and ACTIVE in the system?
- Answer YES if the rule is in place and functioning under normal conditions, even if 
  it has design gaps or can be bypassed via workarounds.
- Answer NO only if the system rule is NOT configured or NOT active at all.
- IMPORTANT: A system rule that works normally but can be bypassed via emergency access 
  or role changes still EXISTS. The bypass path is a DESIGN GAP, not non-existence.

### 2. Is the control design ADEQUATE?
THE CENTRAL QUESTION: Does this automated control, as configured, reasonably prevent 
or detect the identified risk? Focus on whether the CORE SYSTEM RULE works.

- **Yes** — The system rule is correctly configured and addresses the risk. The core 
  automated mechanism works as intended.
- **Partially** — The system rule works but has a MEANINGFUL gap where the risk could 
  actually materialize.
- **No** — The automated control fundamentally cannot address the risk as configured.

### 3. Does the observed process ALIGN with the RCM description?
- Does the system behavior in this sample match the control description?
- If the CORE system rule works as described but has additional gaps (e.g., override 
  process weakness), the core process IS aligned — note the gap separately.

## CRITICAL: Nature of Control — Preventive vs Detective
READ THE RCM "Nature of Control" FIELD CAREFULLY.

**Preventive Controls** are designed to BLOCK the risk event.
  → The system should STOP the unwanted action (hard block, rejection, error).
  → A preventive control that only warns/flags but doesn't block = design gap.
  → Evaluate the BLOCKING mechanism and any override/bypass paths.

**Detective Controls** are designed to DETECT/FLAG the risk event after it occurs.
  → A detective control that generates warnings, alerts, or exception reports is 
    WORKING AS DESIGNED — do NOT penalize it for not blocking.
  → If the detection mechanism reliably catches the risk, design = "Yes".
  → A detective control where the review process is informal is STILL FUNCTIONING 
    if the detection and flagging mechanism works correctly.

## ═══════════════════════════════════════════════════════════════════
##  CALIBRATION — THIS IS THE MOST IMPORTANT SECTION. READ CAREFULLY.
## ═══════════════════════════════════════════════════════════════════

You are a SEASONED AUDITOR, not an intern looking for problems. An experienced 
audit partner distinguishes between:
  (a) things that NEED TO BE FIXED because the risk could materialize, and
  (b) things that COULD BE BETTER but don't actually affect risk mitigation.

Category (a) = real findings. Category (b) = observations/recommendations.
Your run should produce a REALISTIC distribution: most well-designed automated controls 
should PASS with "Yes" and "None". Only controls with REAL weaknesses get gaps.

### ⛔ MISSING SIGN-OFF = FAIL:
The following are REAL design gaps that MUST result in FAIL. They MUST appear
in gap_identified and MUST set design_adequate to "No":

  ✅ "No formal sign-off on exception report reviews"
  ✅ "Review of system alerts lacks formal documentation"
  ✅ "Missing formal sign-off sheet on automated reports"
  ✅ "Rejected payment reviews not formally documented"
  ✅ "Suspense queue review lacks sign-off"

WHY: A control without formal sign-off cannot be independently verified or
evidenced. Even when the automated system rule works correctly, the human
review layer around automated outputs MUST have formal sign-off. Without it,
the control's execution cannot be proven during audit. This is a design failure.

The following remain OBSERVATIONS (not gaps) and belong only in remarks:
  ❌ "No documented SLA for exception review timeline"
  ❌ "Report review process could be more formal"
  ❌ "No formal escalation procedure for flagged items"

### ⛔ MISSING SUPPORTING DOCUMENT / PROCEDURE = FAIL:
If a written procedure, configuration documentation, policy, or any referenced
supporting document CANNOT BE PRODUCED during the walkthrough, this is a FAIL.

  ✅ FAIL: "The written procedure or configuration documentation for this
     automated control cannot be produced during the walkthrough. The team says
     'it exists somewhere' but cannot provide it."
     → FAIL / No / Significant Deficiency.

  ✅ FAIL: "Referenced supporting document (e.g., Configuration Standard,
     Policy Document) was not available during the walkthrough."
     → FAIL / No / Significant Deficiency.

WHY: If the governing procedure, configuration standard, or policy document
CANNOT BE PRODUCED when requested during the audit, the control's design
cannot be independently verified. This creates risk during staff turnover and
makes the control non-auditable. This is a design failure.

### ✅ WHAT QUALIFIES AS A REAL DESIGN GAP FOR AUTOMATED CONTROLS:
A gap means the risk could ACTUALLY MATERIALIZE because of the weakness:

  ✅ Override/bypass process has NO approval requirement at all
  ✅ Preventive control only WARNS instead of BLOCKING (and it should block per its nature)
  ✅ Detection logic has significant blind spots (e.g., exact match only, no fuzzy logic)
  ✅ System rule can be bypassed via emergency access WITHOUT any compensating control
  ✅ SoD enforcement broken by temporary role assignments with no monitoring
  ✅ Policy exists but system does NOT enforce it — any user can override
  ✅ Thresholds are set materially too loose for the risk level
  ✅ No formal sign-off on review of system outputs — control execution cannot be evidenced
  ✅ Written procedure or configuration documentation cannot be produced during walkthrough

### WORKED EXAMPLES — Study these carefully:

**Example A — FAIL / No / Significant Deficiency (Missing Sign-Off):**
Control (Detective): "SAP generates daily exception report for GRs not entered within 24 hours."
Evidence: Report auto-generated at 6 AM, flagged 3 late GRs. Supervisor reviewed and
followed up via email. No formal sign-off on the report itself.
→ Result: FAIL | Design: No | Gap: "No formal sign-off on exception report review —
  control execution cannot be independently evidenced" | Deficiency: Significant Deficiency
→ Remarks: "Detection mechanism working as designed and supervisor actively follows up.
  However, the review lacks formal sign-off, so the control's execution cannot be
  verified or evidenced during audit. This is a design failure."
WHY: Missing sign-off means the control cannot be evidenced = design failure.

**Example B — PASS / Yes / None:**
Control (Preventive): "SAP blocks PO creation without approved PR reference."
Evidence: Hard error generated when creating PO without PR. No user override possible. 
Config confirms mandatory linkage.
→ Result: PASS | Design: Yes | Gap: None | Deficiency: None
→ Remarks: "Strong preventive control. System enforces PR-PO linkage with no user-level 
  override capability."
WHY: Automated block works exactly as designed. No weakness found.

**Example C — FAIL / No / Significant Deficiency (Missing Sign-Off):**
Control (Preventive): "Payment gateway validates payee against vendor master."
Evidence: Gateway blocked fictitious vendor, rejected mismatched bank details, logged
all rejections. No formal sign-off on rejected payment review.
→ Result: FAIL | Design: No | Gap: "No formal sign-off on rejected payment reviews —
  review process cannot be evidenced" | Deficiency: Significant Deficiency
→ Remarks: "Automated validation comprehensive and functioning with all rejections
  properly logged. However, rejected payment reviews lack formal sign-off, so the
  review cannot be independently verified during audit."
WHY: Missing sign-off on the review process = design failure, even if the system works.

**Example D — PASS / Partially / Control Deficiency:**
Control (Detective): "SAP flags duplicate POs based on vendor + amount + date."
Evidence: System correctly flagged PO-89050 as potential duplicate of PO-88955. 
However, detection only matches EXACT amounts — $10,000 vs $10,001 would not be flagged.
→ Result: PASS | Design: Partially | Gap: "Duplicate detection uses exact match only; 
  near-duplicates not flagged" | Deficiency: Control Deficiency
WHY: The detection works but has a meaningful blind spot. Near-duplicates could slip through.

**Example E — FAIL / No / Significant Deficiency:**
Control (Preventive): "Three-way match override requires AP Supervisor approval."
Evidence: Policy document states tiered approvals. AP Clerk overrode $55K invoice with 
no approval. SAP allows any INVOICE_POST user to override with free-text reason.
→ Result: FAIL | Design: No | Gap: "Policy not enforced by system; any user can override 
  without approval" | Deficiency: Significant Deficiency
WHY: Policy exists on paper but system doesn't enforce it. The risk WILL materialize.

**Example F — FAIL / No / Significant Deficiency:**
Control (Preventive): "SoD enforcement prevents same user from creating and approving POs."
Evidence: System blocks self-approval under normal conditions. But emergency access granted 
via IT ticket without business manager approval, creating SoD conflict. No automated 
monitoring for temporary role assignments.
→ Result: FAIL | Design: No | Gap: "Emergency access bypasses SoD without proper approval; 
  no monitoring for temporary role conflicts" | Deficiency: Significant Deficiency
WHY: The normal control works, but the bypass path has no gate and no detection.

## Deficiency Classification (PCAOB AS 2201)
- **None**: Control is adequately designed. Any observations are best-practice only.
- **Control Deficiency**: A meaningful but minor design gap. The control mostly works 
  but has a specific weakness. Risk is remote but possible.
- **Significant Deficiency**: An important gap the audit committee must know about. 
  Reasonable possibility a material misstatement would not be prevented or detected.
- **Material Weakness**: Design fundamentally fails. Material misstatement WILL likely 
  not be prevented or detected.

## MANDATORY CONSISTENCY RULES
1. result = PASS → deficiency_type MUST be "None" or "Control Deficiency" only.
2. result = FAIL → deficiency_type MUST be "Significant Deficiency" or "Material Weakness".
3. design_adequate = "Yes" → deficiency_type MUST be "None". gap_identified MUST be "None".
4. design_adequate = "No" → result MUST be "FAIL".
5. design_adequate = "Partially" + minor gap → PASS, Control Deficiency.
6. design_adequate = "Partially" + significant gap → FAIL, Significant Deficiency.
7. control_exists = "No" → result MUST be "FAIL", deficiency MUST be "Material Weakness".

## PRE-RESPONSE CHECKLIST — Complete this BEFORE writing your JSON:
□ Does the core system rule work correctly? If YES → start from "Yes" / "None" baseline.
□ Did I find a gap where the risk could ACTUALLY MATERIALIZE?
□ Is there a MISSING SIGN-OFF on review of system outputs? If YES → this is a FAIL.
  Set design_adequate = "No", result = "FAIL", deficiency_type = "Significant Deficiency".
  The control cannot be independently evidenced without sign-off.
□ Is a SUPPORTING DOCUMENT, SOP, or configuration documentation MISSING / cannot be
  produced? If YES → this is a FAIL. Set design_adequate = "No", result = "FAIL",
  deficiency_type = "Significant Deficiency".
□ Am I respecting the Nature field? Detective controls detect, preventive controls prevent.
□ If it's a detective control and the detection works → design = "Yes" unless there's
  a real blind spot in what it detects OR sign-off is missing.
□ Do my result, design_adequate, gap, and deficiency_type all align with the rules?
□ Would my audit partner agree this is a real finding, or would they send it back?

## Rules
- Evaluate DESIGN and CONFIGURATION, not transaction volume
- RESPECT the Nature of Control field
- Reference specific system evidence from the sample and supporting documents
- Missing sign-off = FAIL. Missing supporting document/SOP = FAIL.
- Minor observations (SLA timing, escalation procedures) go in remarks only
- Apply consistency rules — double-check before responding
- Respond ONLY with valid JSON"""


def _estimate_tokens(text: str) -> int:
    """Rough token estimate: ~4 characters per token for English text."""
    return len(text) // 4


# Token threshold for warning (does NOT truncate — just logs a heads-up).
# With gpt-4o / gpt-5.2-chat (128K context), prompts up to ~100K tokens are fine.
# With gpt-35-turbo (16K context), anything above ~12K may fail.
PROMPT_WARNING_THRESHOLD = 100000  # tokens — for 128K context models (gpt-4o / gpt-5.2-chat)


def _split_chunks(text: str, size: int = 2500, overlap: int = 300) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    chunks = []
    i = 0
    while i < len(text):
        j = min(len(text), i + size)
        chunks.append(text[i:j])
        if j >= len(text):
            break
        i = max(j - overlap, i + 1)
    return chunks


def _azure_embed(texts: list[str]) -> list[list[float]]:
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "").strip()
        or os.getenv("AZURE_OPENAI_EMBEDDING_MODEL", "").strip()
        or os.getenv("OPENAI_EMBEDDING_MODEL", "").strip()
        or "text-embedding-ada-002"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()
    if not endpoint or not api_key or not texts:
        return []

    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/embeddings?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}
    payload = {"input": texts}
    resp = requests.post(url, headers=headers, json=payload, timeout=90)
    resp.raise_for_status()
    data = resp.json().get("data", [])
    return [d.get("embedding", []) for d in data]


def _cos(a: list[float], b: list[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(y * y for y in b))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)


def _write_embedding_log(
    engine: str, control_id: str, query: str,
    chunks: list[str], scored: list[tuple[int, float]],
    keep_idx: list[int], final_text: str,
    original_text_len: int = 0,
) -> None:
    """Write a detailed embedding selection log file."""
    log_dir = _get_embedding_log_dir()
    if log_dir is None:
        return
    try:
        from datetime import datetime, timezone
        ts = datetime.now(timezone.utc)
        ts_str = ts.strftime("%Y%m%d_%H%M%S")
        safe_id = re.sub(r"[^\w.\-]", "_", control_id)
        log_path = log_dir / f"{ts_str}_{safe_id}_{engine}.log"

        sep = "═" * 70
        thin = "─" * 70
        lines: list[str] = []

        lines.append(sep)
        lines.append("EMBEDDING SELECTION LOG")
        lines.append(sep)
        lines.append(f"Timestamp:          {ts.isoformat()}")
        lines.append(f"Engine:             {engine}")
        lines.append(f"Control ID:         {control_id}")
        lines.append(f"Original Text Len:  {original_text_len:,} chars")
        lines.append(f"Chunk Count:        {len(chunks)}")
        lines.append(f"Chunks Selected:    {len(keep_idx)} of {len(chunks)}")
        lines.append("")

        lines.append(thin)
        lines.append("EMBEDDING QUERY (from RCM fields)")
        lines.append(thin)
        lines.append(query)
        lines.append("")

        lines.append(thin)
        lines.append("CHUNK SCORES (sorted by similarity)")
        lines.append(thin)
        for rank, (chunk_idx, score) in enumerate(scored, 1):
            selected_marker = " ← SELECTED" if chunk_idx in keep_idx else ""
            preview = chunks[chunk_idx][:80].replace("\n", " ").strip()
            lines.append(f"Rank {rank:3d}: Chunk {chunk_idx + 1:3d} (score: {score:.4f}){selected_marker}  |  {preview}...")
        lines.append("")

        lines.append(thin)
        lines.append(f"SELECTED CHUNKS CONTENT ({len(keep_idx)} chunks)")
        lines.append(thin)
        for idx in keep_idx:
            lines.append(f"[CHUNK {idx + 1}]")
            lines.append(chunks[idx])
            lines.append("")

        lines.append(sep)
        lines.append("FINAL EVIDENCE TEXT SENT TO LLM")
        lines.append(sep)
        lines.append(final_text)
        lines.append("")
        lines.append(sep)
        lines.append(f"END OF LOG — {control_id} ({engine})")
        lines.append(sep)

        log_path.write_text("\n".join(lines), encoding="utf-8")
    except Exception:
        pass


def _write_embedding_fallback_log(
    engine: str, control_id: str, reason: str,
    text_len: int, text_preview: str,
) -> None:
    """Log when embedding selection was skipped and full text was used directly."""
    log_dir = _get_embedding_log_dir()
    if log_dir is None:
        return
    try:
        from datetime import datetime, timezone
        ts = datetime.now(timezone.utc)
        ts_str = ts.strftime("%Y%m%d_%H%M%S")
        safe_id = re.sub(r"[^\w.\-]", "_", control_id)
        log_path = log_dir / f"{ts_str}_{safe_id}_{engine}_FULLTEXT.log"

        sep = "═" * 70
        thin = "─" * 70
        lines: list[str] = [
            sep,
            "EMBEDDING BYPASS LOG — FULL TEXT USED",
            sep,
            f"Timestamp:          {ts.isoformat()}",
            f"Engine:             {engine}",
            f"Control ID:         {control_id}",
            f"Reason:             {reason}",
            f"Text Length:        {text_len:,} chars",
            "",
            thin,
            "EVIDENCE TEXT SENT TO LLM (full, no embedding selection)",
            thin,
            text_preview,
            "",
            sep,
            f"END OF LOG — {control_id} ({engine})",
            sep,
        ]
        log_path.write_text("\n".join(lines), encoding="utf-8")
    except Exception:
        pass


def _prepare_evidence_for_prompt(rcm: RCMRow, sample: SampleEvidence) -> str:
    """Create retrieval-style evidence context via embeddings; fallback to full text."""
    text = (sample.description or "").strip()
    if not text:
        return text

    use_embed = os.getenv("EVIDENCE_EMBEDDINGS", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not use_embed:
        _write_embedding_fallback_log("TOD_RCM", rcm.control_id, "EVIDENCE_EMBEDDINGS disabled", len(text), text)
        return text

    try:
        chunks = _split_chunks(text)
        if len(chunks) <= 1:
            _write_embedding_fallback_log("TOD_RCM", rcm.control_id, f"Text fits in 1 chunk ({len(text)} chars) — no splitting needed", len(text), text)
            return text
        query = (
            f"{rcm.control_id} {rcm.risk_id} {rcm.risk_title} {rcm.control_description} "
            f"{rcm.control_objective} {rcm.control_type} {rcm.nature_of_control}"
        )
        vecs = _azure_embed([query] + chunks)
        if len(vecs) != len(chunks) + 1:
            _write_embedding_fallback_log("TOD_RCM", rcm.control_id, f"Embedding API returned wrong vector count (expected {len(chunks)+1}, got {len(vecs)})", len(text), text)
            return text
        qv, cvs = vecs[0], vecs[1:]
        scored = sorted(((i, _cos(qv, v)) for i, v in enumerate(cvs)), key=lambda x: x[1], reverse=True)
        k = int(os.getenv("EVIDENCE_TOP_K", "24"))
        keep_idx = sorted(i for i, _ in scored[:max(1, min(k, len(chunks)))])
        selected = [f"[CHUNK {i+1}]\n{chunks[i]}" for i in keep_idx]
        final_text = "\n\n".join(selected).strip() or text

        _write_embedding_log(
            engine="TOD_RCM", control_id=rcm.control_id, query=query,
            chunks=chunks, scored=scored, keep_idx=keep_idx,
            final_text=final_text, original_text_len=len(text),
        )

        return final_text
    except Exception as exc:
        _write_embedding_fallback_log("TOD_RCM", rcm.control_id, f"Exception during embedding: {exc}", len(text), text)
        return text


def build_tod_prompt(rcm: RCMRow, sample: SampleEvidence) -> str:
    """
    Build the TOD evaluation prompt. All evidence and supporting documents
    are included in FULL — nothing is truncated or omitted.
    """

    # Build supporting documents section
    supporting_section = ""
    if sample.supporting_docs:
        doc_blocks = []
        for doc in sample.supporting_docs:
            label = f" ({doc.doc_type})" if doc.doc_type else ""
            doc_blocks.append(f"""### {doc.filename}{label}
{doc.content}""")
        supporting_section = f"""

## SUPPORTING DOCUMENTS ({len(sample.supporting_docs)} file{"s" if len(sample.supporting_docs) > 1 else ""})
The following documents were provided as supporting evidence. Use them to verify the control's design, confirm policy existence, and identify gaps.

{chr(10).join(doc_blocks)}"""

    evidence_for_prompt = _prepare_evidence_for_prompt(rcm, sample)

    return f"""Perform a Test of Design (TOD) evaluation for this control using the sample evidence and any supporting documents provided.

## RCM DETAILS
- **Process**: {rcm.process} > {rcm.subprocess}
- **Control Objective**: {rcm.control_objective}
- **Risk ID**: {rcm.risk_id} | **Risk Title**: {rcm.risk_title}
- **Risk Level**: {rcm.risk_level or "Not specified"}
- **Risk Description**: {rcm.risk_description}
- **Control ID**: {rcm.control_id}
- **Control Description**: {rcm.control_description}
- **Control Owner**: {rcm.control_owner}
- **Nature**: {rcm.nature_of_control} | **Type**: {rcm.control_type} | **Frequency**: {rcm.control_frequency}
- **System**: {rcm.application_system}

## SAMPLE EVIDENCE
- **Sample ID**: {sample.sample_id}
- **Source Document**: {sample.source_document or "Not specified"}
- **Test Date**: {sample.test_date or "Not specified"}
- **Tested By**: {sample.tested_by or "Not specified"}
- **Evidence**:
{evidence_for_prompt}{supporting_section}

## REQUIRED JSON OUTPUT
{{
  "result": "PASS or FAIL",
  "design_adequate": "Yes or No or Partially",
  "control_exists": "Yes or No",
  "process_aligned": "Yes or No",
  "confidence": "High or Medium or Low",
  "remarks": "3-5 sentence TOD analysis referencing specific evidence and supporting documents.",
  "gap_identified": "Specific design gap, or None",
  "design_recommendation": "If gap found, how to fix. If no gap, N/A",
  "deficiency_type": "None | Control Deficiency | Significant Deficiency | Material Weakness"
}}"""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  POST-PROCESSING VALIDATOR
#  Catches and corrects logical contradictions in LLM output
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def validate_and_correct(eval_result: dict, rcm: RCMRow) -> dict:
    """
    Apply hard consistency rules to LLM output. Fixes contradictions.
    Returns corrected dict with 'corrections' list documenting any changes.

    Rules enforce PCAOB AS 2201 hierarchy:
      control_exists=No  → FAIL + Material Weakness
      design_adequate=No → FAIL + Significant Deficiency (minimum)
      design_adequate=Yes → PASS + None (no gap)
      PASS → None or Control Deficiency only
      FAIL → Significant Deficiency or Material Weakness only
    """
    r = dict(eval_result)  # copy
    corrections = []

    result = r.get("result", "")
    design = r.get("design_adequate", "")
    exists = r.get("control_exists", "")
    deficiency = r.get("deficiency_type", "")
    gap = r.get("gap_identified", "").strip()

    # Normalize gap — treat generic non-gaps as empty
    gap_is_empty = gap.lower() in ("", "none", "n/a", "none.", "no gap", "no gaps",
                                    "no gap identified", "none identified")

    # ── Rule 1: control_exists=No → FAIL + Material Weakness ─────────────
    if exists == "No":
        if result != "FAIL":
            corrections.append(f"CORRECTED: result {result}→FAIL (control does not exist)")
            r["result"] = "FAIL"
        if deficiency != "Material Weakness":
            corrections.append(f"CORRECTED: deficiency {deficiency}→Material Weakness (control does not exist)")
            r["deficiency_type"] = "Material Weakness"
        if design != "No":
            corrections.append(f"CORRECTED: design_adequate {design}→No (control does not exist)")
            r["design_adequate"] = "No"

    # ── Rule 2: design_adequate=Yes → deficiency=None, gap=None ──────────
    if r.get("design_adequate") == "Yes":
        if r.get("deficiency_type") not in ("None", ""):
            corrections.append(f"CORRECTED: deficiency {r['deficiency_type']}→None (design=Yes means no deficiency)")
            r["deficiency_type"] = "None"
        if not gap_is_empty:
            # LLM said Yes but also identified a gap — this is contradictory.
            # The gap is likely an observation, not a real deficiency. Clear it.
            corrections.append(
                f"CORRECTED: gap_identified moved to remarks as observation "
                f"(design=Yes means gap should be None; original gap: {gap[:80]})"
            )
            # Preserve the observation in remarks
            obs_note = f" [Observation: {gap}]"
            r["remarks"] = r.get("remarks", "") + obs_note
            r["gap_identified"] = "None"
        if r.get("result") != "PASS":
            corrections.append(f"CORRECTED: result {r['result']}→PASS (design=Yes)")
            r["result"] = "PASS"
        # Also clean up recommendation
        if r.get("design_recommendation", "").strip().lower() not in ("", "n/a", "none"):
            r["design_recommendation"] = "N/A"

    # ── Rule 3: design_adequate=No → must be FAIL ───────────────────────
    if r.get("design_adequate") == "No" and r.get("result") != "FAIL":
        corrections.append(f"CORRECTED: result {r['result']}→FAIL (design=No)")
        r["result"] = "FAIL"

    # ── Rule 4: PASS cannot have Significant Deficiency or MW ────────────
    if r.get("result") == "PASS" and r.get("deficiency_type") in ("Significant Deficiency", "Material Weakness"):
        if r.get("design_adequate") == "Partially":
            corrections.append(
                f"CORRECTED: deficiency {r['deficiency_type']}→Control Deficiency "
                f"(PASS cannot have {r['deficiency_type']}; design=Partially → Control Deficiency)"
            )
            r["deficiency_type"] = "Control Deficiency"
        else:
            corrections.append(
                f"CORRECTED: deficiency {r['deficiency_type']}→None "
                f"(PASS + design=Yes cannot have {r['deficiency_type']})"
            )
            r["deficiency_type"] = "None"

    # ── Rule 5: FAIL cannot have None or just Control Deficiency ─────────
    if r.get("result") == "FAIL" and r.get("deficiency_type") in ("None", "Control Deficiency", ""):
        corrections.append(
            f"CORRECTED: deficiency {r.get('deficiency_type', 'None')}→Significant Deficiency "
            f"(FAIL requires at least Significant Deficiency)"
        )
        r["deficiency_type"] = "Significant Deficiency"

    # ── Rule 6: PASS + Yes + no real gap → deficiency must be None ───────
    gap_after = r.get("gap_identified", "").strip()
    gap_after_empty = gap_after.lower() in ("", "none", "n/a", "none.", "no gap",
                                             "no gaps", "no gap identified", "none identified")
    if r.get("result") == "PASS" and r.get("design_adequate") == "Yes":
        if gap_after_empty and r.get("deficiency_type") not in ("None", ""):
            corrections.append(f"CORRECTED: deficiency {r['deficiency_type']}→None (PASS+Yes+no gap)")
            r["deficiency_type"] = "None"

    # ── Rule 7: PASS + Partially + real gap → Control Deficiency ─────────
    if r.get("result") == "PASS" and r.get("design_adequate") == "Partially":
        if not gap_after_empty and r.get("deficiency_type") in ("None", ""):
            corrections.append(
                f"CORRECTED: deficiency None→Control Deficiency "
                f"(design=Partially with gap: {gap_after[:60]})"
            )
            r["deficiency_type"] = "Control Deficiency"

    # ── Rule 8: FAIL + Partially → ensure at least Sig Deficiency ────────
    if r.get("result") == "FAIL" and r.get("design_adequate") == "Partially":
        if r.get("deficiency_type") not in ("Significant Deficiency", "Material Weakness"):
            corrections.append(
                f"CORRECTED: deficiency {r.get('deficiency_type')}→Significant Deficiency "
                f"(FAIL + Partially requires Significant Deficiency)"
            )
            r["deficiency_type"] = "Significant Deficiency"

    # Append corrections to remarks if any
    if corrections:
        correction_note = " [AUTO-CORRECTED: " + "; ".join(corrections) + "]"
        r["remarks"] = r.get("remarks", "") + correction_note

    return r


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  AZURE OPENAI CALLER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class AzureOpenAIEvaluator:
    def __init__(self, api_key, model="gpt-5.2-chat", max_tokens=4096,
                 retry_attempts=2, retry_delay=2.0,
                 azure_endpoint=None, azure_api_key=None, azure_deployment=None,
                 azure_api_version=None):
        if not azure_endpoint:
            raise ValueError("AZURE endpoint is required for AzureOpenAIEvaluator")

        self.api_key = azure_api_key or api_key
        self.model = azure_deployment or model
        self.api_url = (
            f"{azure_endpoint.rstrip('/')}/openai/deployments/{self.model}"
            f"/chat/completions?api-version={azure_api_version or '2024-12-01-preview'}"
        )
        self.auth_header = {"api-key": self.api_key}
        self.max_tokens = max_tokens
        self.retry_attempts = retry_attempts
        self.retry_delay = retry_delay

    def evaluate(self, rcm: RCMRow, sample: SampleEvidence) -> dict:
        system_prompt = (
            TOD_SYSTEM_PROMPT_AUTOMATED if rcm.control_type == "Automated"
            else TOD_SYSTEM_PROMPT_MANUAL
        )
        user_prompt = build_tod_prompt(rcm, sample)

        # Warn if prompt is very large (but still send it)
        prompt_tokens = _estimate_tokens(user_prompt)
        if prompt_tokens > PROMPT_WARNING_THRESHOLD:
            print(f"  ⚠️  {rcm.control_id}: prompt is ~{prompt_tokens:,} tokens "
                  f"(threshold: {PROMPT_WARNING_THRESHOLD:,}). "
                  f"Consider using a 128K model or reducing supporting docs.")

        headers = {**self.auth_header, "Content-Type": "application/json"}
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "max_completion_tokens": self.max_tokens,
            "response_format": {"type": "json_object"},
        }

        last_error = None
        for attempt in range(1, self.retry_attempts + 1):
            try:
                resp = requests.post(self.api_url, headers=headers, json=payload, timeout=60)
                resp.raise_for_status()
                text = resp.json()["choices"][0]["message"]["content"]
                clean = text.strip().removeprefix("```json").removesuffix("```").strip()
                return json.loads(clean)
            except requests.exceptions.HTTPError:
                last_error = f"HTTP {resp.status_code}: {resp.text[:300]}"
                if resp.status_code == 429:
                    time.sleep(self.retry_delay * attempt * 2)
                elif resp.status_code >= 500:
                    time.sleep(self.retry_delay * attempt)
                else:
                    break
            except (requests.RequestException, json.JSONDecodeError, KeyError, IndexError) as e:
                last_error = str(e)
                if attempt < self.retry_attempts:
                    time.sleep(self.retry_delay * attempt)

        return {
            "result": "ERROR", "design_adequate": "Unable to Assess",
            "control_exists": "Unable to Assess", "process_aligned": "Unable to Assess",
            "confidence": "Low", "remarks": f"LLM evaluation failed: {last_error}",
            "gap_identified": "Evaluation error — manual review required",
            "design_recommendation": "Re-run or perform manual review",
            "deficiency_type": "Unable to Assess",
        }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  RCM LOADER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

COLUMN_MAP = {
    "Process": "process", "process": "process",
    "SubProcess": "subprocess", "Sub Process": "subprocess", "subprocess": "subprocess",
    "Control Objective": "control_objective", "control objective": "control_objective",
    "Risk Id": "risk_id", "Risk ID": "risk_id", "risk id": "risk_id",
    "Risk Title": "risk_title", "Risk Description": "risk_description",
    "Control Id": "control_id", "Control ID": "control_id", "control id": "control_id",
    "Control Description": "control_description",
    "Control Owner": "control_owner", "Control Rating": "control_rating",
    "Nature of Control": "nature_of_control", "Control Type": "control_type",
    "Control Frequency": "control_frequency", "Application/System": "application_system",
    "risk_level": "risk_level", "Risk Level": "risk_level",
}

REQUIRED_RCM_COLUMNS = [
    "process", "subprocess", "control_objective", "risk_id", "risk_title",
    "risk_description", "control_id", "control_description", "control_owner",
    "control_rating", "nature_of_control", "control_type", "control_frequency",
    "application_system", "risk_level",
]


def load_rcm(filepath, sheet_name=None):
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"RCM file not found: {filepath}")

    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(path, sheet_name=sheet_name or 0, dtype=str)

    # Normalize header whitespace and map to dataclass names.
    df.columns = [str(c).strip() for c in df.columns]
    df.rename(columns=COLUMN_MAP, inplace=True)

    missing = [col for col in REQUIRED_RCM_COLUMNS if col not in df.columns]
    if missing:
        pretty = ", ".join(missing)
        raise ValueError(
            "RCM file is missing required columns after normalization: "
            f"{pretty}"
        )

    # Ensure all required fields are string-safe and present.
    for col in REQUIRED_RCM_COLUMNS:
        df[col] = df[col].fillna("").astype(str).str.strip()

    rows = []
    for _, row in df.iterrows():
        rows.append(RCMRow(**{col: str(row.get(col, "")) for col in RCMRow.__dataclass_fields__}))
    return rows



# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EVIDENCE LOADER — reads one subfolder per control
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#
#  Folder structure:
#
#    evidence/
#    ├── C-P2P-001/
#    │   ├── evidence.txt              ← walkthrough (has ---EVIDENCE--- marker)
#    │   ├── Procurement_SOP_v1.2.pdf  ← supporting doc (PDF)
#    │   ├── SAP_Config_Export.xlsx     ← supporting doc (Excel)
#    │   └── Approval_Matrix.docx      ← supporting doc (Word)
#    ├── C-P2P-002/
#    │   ├── evidence.txt
#    │   └── Budget_Policy.pptx        ← supporting doc (PowerPoint)
#    └── ...
#
#  Supported formats for text extraction:
#    .txt .md .csv .log    — plain text (UTF-8)
#    .eml                  — parsed email headers + body (plain/html)
#    .msg                  — Outlook message extraction (if extract_msg installed)
#    .pdf                  — text extracted via pdftotext CLI / PyPDF2
#    .docx                 — text extracted via python-docx (paragraphs + tables)
#    .xlsx .xls            — cell contents extracted via openpyxl (all sheets)
#    .pptx                 — slide text extracted via python-pptx
#    .png .jpg .jpeg .tiff — OCR via pytesseract (if installed)
#    other                 — file noted in prompt but content not extracted
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


# NOTE: All format-specific extractors (_extract_pdf, _extract_docx, _extract_xlsx,
# _extract_pptx, _extract_image_ocr, _extract_eml, _extract_msg, _azure_docint_extract,
# _azure_vision_describe_image) have been removed. All parsing is now handled by
# the central Document_Intelligence module via _di_extract_text().


def _llm_refine_extraction(content: str, filename: str, doc_type: str) -> str:
    """Optionally refine extracted document text with LLM for higher parsing accuracy."""
    text = (content or "").strip()
    if not text:
        return content

    enabled = os.getenv("LLM_EVIDENCE_PARSING", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return content

    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or "gpt-5.2-chat"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()

    if not endpoint or not api_key:
        return content

    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}

    try:
        chunk_chars = int(os.getenv("LLM_PARSER_CHUNK_CHARS", "12000"))
        overlap = int(os.getenv("LLM_PARSER_CHUNK_OVERLAP", "200"))
        chunks = []
        i = 0
        while i < len(text):
            j = min(len(text), i + chunk_chars)
            chunks.append(text[i:j])
            if j >= len(text):
                break
            i = max(j - overlap, i + 1)

        refined_chunks = []
        for idx, ch in enumerate(chunks, 1):
            payload = {
                "model": deployment,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a forensic document parser. Clean extraction noise while preserving all facts. "
                            "Never invent. Keep dates, amounts, IDs, approvers, headers, and key workflow language. "
                            "Return JSON only."
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"Document: {filename}\nType: {doc_type}\nChunk: {idx}/{len(chunks)}\n\n"
                            "Normalize formatting and OCR artifacts. Do not drop material evidence details.\n\n"
                            f"Extracted text:\n{ch}"
                        ),
                    },
                ],
                "max_completion_tokens": 2000,
                "response_format": {"type": "json_object"},
            }
            resp = requests.post(url, headers=headers, json=payload, timeout=90)
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"]
            data = json.loads(raw.strip().removeprefix("```json").removesuffix("```").strip())
            refined = str(data.get("clean_text") or data.get("parsed_text") or data.get("text") or "").strip()
            if refined:
                refined_chunks.append(refined)

        merged = "\n\n".join(refined_chunks).strip()
        if merged:
            return merged
        return content
    except Exception:
        return content


# ── File-type registry ───────────────────────────────────────────────────

# Plain text extensions — only these can contain the ---EVIDENCE--- marker
_TEXT_EXTS = {".txt", ".md", ".csv", ".log"}


def _extract_file(filepath: Path) -> tuple[str, str, bool]:
    """
    Extract text content from any supported file.

    Delegates to Document_Intelligence for all structured/binary formats.
    Plain text files are read directly.

    Returns: (content_text, doc_type_label, extraction_succeeded)
    """
    ext = filepath.suffix.lower()

    # Plain text files — read directly (fast, no external parser needed)
    if ext in _TEXT_EXTS:
        try:
            content = filepath.read_text(encoding="utf-8", errors="replace").strip()
            return content, _infer_doc_type(filepath.name), True
        except Exception:
            return "", "Text File", False

    # All other formats — delegate to Document_Intelligence
    text, doc_type, ok = _di_extract_text(str(filepath))

    return text, doc_type, ok


def _infer_doc_type(filename: str) -> str:
    """Infer document type from filename keywords."""
    name = filename.lower()
    if any(k in name for k in ("sop", "procedure", "policy", "manual", "guideline")):
        return "Policy/SOP"
    if any(k in name for k in ("config", "setting", "param", "spro")):
        return "System Configuration"
    if any(k in name for k in ("log", "trail", "audit")):
        return "Audit Log"
    if any(k in name for k in ("screenshot", "screen", "capture")):
        return "Screenshot/Capture"
    if any(k in name for k in ("report", "summary", "dashboard")):
        return "Report"
    if any(k in name for k in ("email", "approval", "sign", "ticket")):
        return "Approval/Communication"
    if any(k in name for k in ("checklist", "form", "template")):
        return "Checklist/Form"
    if any(k in name for k in ("matrix", "role", "access")):
        return "Access/Role Matrix"
    if any(k in name for k in ("invoice", "receipt", "payment", "credit")):
        return "Transaction Document"
    return "Supporting Document"


def _score_text_file_for_walkthrough(filepath: Path) -> float:
    """Heuristic score for selecting the main walkthrough file from mixed text artifacts."""
    name = filepath.stem.lower()
    score = 0.0

    # Filename cues
    if re.search(r"\b(sample|evidence|walkthrough|transaction)\b", name):
        score += 3.0
    if re.search(r"\b(email|mail|approval|policy|sop|report|matrix|config|log|audit|screenshot)\b", name):
        score -= 2.0

    # Content cues (lightweight read)
    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")
        head = "\n".join(text.splitlines()[:40]).lower()

        if re.search(r"^\s*(from|to|subject|cc|bcc|date)\s*:", head, flags=re.MULTILINE):
            score -= 4.0

        meta_hits = 0
        for k in ("sample id:", "source document:", "test date:", "tested by:", "control id:"):
            if k in head:
                meta_hits += 1
        if meta_hits:
            score += 2.0 + 0.5 * meta_hits

        if "---evidence---" in head:
            score += 2.0

        if any(k in head for k in ("walkthrough", "transaction", "control", "tested", "sample")):
            score += 1.0
    except Exception:
        pass

    return score


def _llm_classify_text_role(filepath: Path, mode: str = "tod") -> tuple[str, float]:
    """Classify text file role using LLM: returns (role, confidence).

    role is one of: "sample", "supporting", "unknown"
    """
    enabled = os.getenv("LLM_ROLE_CLASSIFIER", "1").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return "unknown", 0.0

    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
    api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
    deployment = (
        os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or "gpt-5.2-chat"
    )
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview").strip()
    if not endpoint or not api_key:
        return "unknown", 0.0

    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return "unknown", 0.0

    prompt_mode = "TOD walkthrough" if mode == "tod" else "TOE sample"
    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    headers = {"api-key": api_key, "Content-Type": "application/json"}
    payload = {
        "model": deployment,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Classify file role for audit evidence. "
                    "Return JSON {role, confidence}. role must be sample|supporting|unknown."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Target role type: {prompt_mode}.\n"
                    f"Filename: {filepath.name}\n\n"
                    "sample means primary walkthrough/transaction narrative to test control; "
                    "supporting means email/policy/report/log/screenshot/reference evidence.\n\n"
                    f"Content:\n{text}"
                ),
            },
        ],
        "max_completion_tokens": 300,
        "response_format": {"type": "json_object"},
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=60)
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"]
        data = json.loads(raw.strip().removeprefix("```json").removesuffix("```").strip())
        role = str(data.get("role", "unknown")).strip().lower()
        if role not in {"sample", "supporting", "unknown"}:
            role = "unknown"
        conf = float(data.get("confidence", 0.0) or 0.0)
        conf = max(0.0, min(1.0, conf))
        return role, conf
    except Exception:
        return "unknown", 0.0


# ── Evidence folder loader ───────────────────────────────────────────────

def load_evidence_folder(folder_path: str) -> dict[str, SampleEvidence]:
    """
    Read evidence from a folder of subfolders (one per control).

    Each subfolder should contain:
      - One text file with ---EVIDENCE--- marker (usually evidence.txt)
      - Zero or more supporting documents in any format

    Folder structure:
      evidence/
        C-P2P-001/
          evidence.txt          ← main walkthrough (plain text, no markers)
          SOP_v1.2.pdf          ← supporting doc
          SAP_Config.xlsx       ← supporting doc
        C-P2P-002/
          evidence.txt
        ...

    The evidence.txt file is PLAIN TEXT — just the walkthrough narrative.
    Optionally, metadata can be included at the top as Key: Value lines
    separated from the narrative by a blank line:

        Sample ID: S1
        Source Document: PR-20240315
        Test Date: 2024-06-10
        Tested By: Priya Nair

        Purchase Requisition PR-20240315 for $12,500 submitted by...

    If no metadata lines are present, defaults are used (Sample ID: S1, etc.).

    Supported extraction:
      .txt .md .csv .log — plain text
      .pdf               — pdftotext / PyPDF2
      .docx              — python-docx (paragraphs + tables)
      .xlsx .xls         — openpyxl (all sheets, all cells)
      .pptx              — python-pptx (all slides + tables)
      .png .jpg .tiff    — pytesseract OCR
      other              — noted but not extracted

    Returns dict[control_id -> SampleEvidence] ready for test_all().
    """
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Evidence folder not found: {folder_path}")

    sample_bank: dict[str, SampleEvidence] = {}
    errors: list[str] = []
    format_stats: dict[str, int] = {}

    subdirs = sorted(d for d in folder.iterdir() if d.is_dir())
    if not subdirs:
        raise FileNotFoundError(
            f"No subfolders found in {folder_path}. "
            f"Expected one folder per control (e.g. C-P2P-001/, C-P2P-002/)."
        )

    for control_dir in subdirs:
        control_id = control_dir.name
        all_files = sorted(f for f in control_dir.iterdir() if f.is_file())

        if not all_files:
            errors.append(f"  ⚠️  {control_id}/: empty folder, skipped")
            continue

        # ── Treat ALL files as evidence equally (no primary/supporting split) ──
        evidence_blocks = []
        for ef in all_files:
            ext = ef.suffix.lower()
            format_stats[ext] = format_stats.get(ext, 0) + 1

            if ext in _TEXT_EXTS:
                try:
                    file_text = ef.read_text(encoding="utf-8", errors="replace").strip()
                    ok = True
                except Exception as e:
                    file_text = f"[Read error for {ef.stem}: {e}]"
                    ok = False
            else:
                file_text, _, ok = _extract_file(ef)

            if not (file_text or "").strip():
                file_text = f"[No extractable content in {ef.stem}]"
                ok = False

            evidence_blocks.append(f"[EVIDENCE FILE: {ef.stem}]\n{file_text}")
            if not ok:
                errors.append(f"  ⚠️  {control_id}/{ef.stem}: extraction may be incomplete")

        if not evidence_blocks:
            errors.append(f"  ⚠️  {control_id}/: no parseable evidence files found, skipped")
            continue

        evidence_text = "\n\n".join(evidence_blocks)

        sample_bank[control_id] = SampleEvidence(
            sample_id="S1",
            description=evidence_text,
            source_document=", ".join(f.stem for f in all_files),
            test_date="",
            tested_by="",
            supporting_docs=[],
        )

    # ── Print summary
    total_files = sum(format_stats.values())
    _EXT_LABELS = {".txt": "text", ".md": "text", ".csv": "csv", ".log": "log",
                   ".pdf": "pdf", ".docx": "docx", ".xlsx": "xlsx", ".doc": "doc",
                   ".xls": "xls", ".pptx": "pptx", ".ppt": "ppt",
                   ".png": "image", ".jpg": "image", ".jpeg": "image",
                   ".tiff": "image", ".tif": "image", ".bmp": "image",
                   ".eml": "email", ".msg": "email"}

    print(f"[Evidence Loaded] {len(sample_bank)} controls from {folder_path}")
    print(f"[Evidence Files]  {total_files} files across {len(sample_bank)} controls")

    if format_stats:
        label_counts: dict[str, int] = {}
        for ext, n in format_stats.items():
            label = _EXT_LABELS.get(ext, ext.lstrip("."))
            label_counts[label] = label_counts.get(label, 0) + n
        fmt_summary = ", ".join(f"{lbl}: {n}" for lbl, n in sorted(label_counts.items()))
        print(f"[File Formats]    {fmt_summary}")

    for cid in sorted(sample_bank):
        n_files = sample_bank[cid].source_document.count(",") + 1 if sample_bank[cid].source_document else 0
        print(f"  📁 {cid}: {n_files} evidence file{'s' if n_files != 1 else ''}")

    if errors:
        print(f"[Evidence Warnings]")
        for err in errors:
            print(err)

    return sample_bank

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MAIN TESTER CLASS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class RCMControlTester:
    """
    Test of Design engine.

    Input:  dict[control_id, SampleEvidence] — one sample per control
    Output: list[DesignTestResult]

    Usage:
        tester = RCMControlTester(...)
        sample_bank = {
            "C-P2P-001": SampleEvidence(sample_id="S1", description="..."),
            "C-P2P-009": SampleEvidence(sample_id="S1", description="..."),
        }
        results = tester.test_all(sample_bank)
    """

    def __init__(self, rcm_path, openai_api_key=None, openai_model="gpt-5.2-chat",
                 azure_endpoint=None, azure_api_key=None, azure_deployment=None,
                 azure_api_version="2024-12-01-preview", sheet_name=None):
        self.rcm_rows = load_rcm(rcm_path, sheet_name)
        self.rcm_lookup = {r.control_id: r for r in self.rcm_rows}
        self.evaluator = AzureOpenAIEvaluator(
            api_key=openai_api_key or azure_api_key,
            model=openai_model or azure_deployment,
            azure_endpoint=azure_endpoint, azure_api_key=azure_api_key,
            azure_deployment=azure_deployment, azure_api_version=azure_api_version,
        )
        print(f"[RCM Loaded] {len(self.rcm_rows)} controls from {rcm_path}")
        provider = "Azure OpenAI" if azure_endpoint else "OpenAI"
        model_name = openai_model or azure_deployment
        print(f"[{provider}] {model_name}")
        print(f"[Mode] Test of Design (TOD) — full evidence + supporting docs per control\n")

    def list_controls(self):
        return pd.DataFrame([{
            "Control ID": r.control_id, "Risk ID": r.risk_id,
            "Risk Title": r.risk_title, "Risk Level": r.risk_level,
            "Nature": r.nature_of_control,
            "Type": r.control_type, "Frequency": r.control_frequency,
        } for r in self.rcm_rows])

    def _evaluate_single(self, control_id: str, sample: SampleEvidence) -> DesignTestResult:
        """Evaluate one control's design (thread-safe, no printing)."""
        rcm = self.rcm_lookup.get(control_id)
        if not rcm:
            raise ValueError(f"Control '{control_id}' not found.")

        eval_result = self.evaluator.evaluate(rcm, sample)
        eval_result = validate_and_correct(eval_result, rcm)

        return DesignTestResult(
            control_id=control_id,
            risk_id=rcm.risk_id,
            risk_title=rcm.risk_title,
            control_type=rcm.control_type,
            nature_of_control=rcm.nature_of_control,
            control_frequency=rcm.control_frequency,
            sample_id=sample.sample_id,
            result=eval_result.get("result", "ERROR"),
            design_adequate=eval_result.get("design_adequate", "Unable to Assess"),
            control_exists=eval_result.get("control_exists", "Unable to Assess"),
            process_aligned=eval_result.get("process_aligned", "Unable to Assess"),
            confidence=eval_result.get("confidence", "Low"),
            remarks=eval_result.get("remarks", ""),
            gap_identified=eval_result.get("gap_identified", ""),
            design_recommendation=eval_result.get("design_recommendation", ""),
            deficiency_type=eval_result.get("deficiency_type", "Unable to Assess"),
            raw_evidence=sample.description or "",
            evaluation_timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
        )

    def test_control(self, control_id: str, sample: SampleEvidence) -> DesignTestResult:
        """Evaluate one control (with console output). For single-control use."""
        rcm = self.rcm_lookup.get(control_id)
        if not rcm:
            raise ValueError(f"Control '{control_id}' not found.")

        n_docs = len(sample.supporting_docs)
        docs_note = f" + {n_docs} doc{'s' if n_docs != 1 else ''}" if n_docs else ""
        print(f"  Evaluating {control_id} | Sample {sample.sample_id}{docs_note}...", end=" ")
        result = self._evaluate_single(control_id, sample)
        icon = "✅" if result.result == "PASS" else "❌" if result.result == "FAIL" else "⚠️"
        print(f"{icon} {result.result} | Design: {result.design_adequate}")
        return result

    def test_all(self, sample_bank: dict[str, SampleEvidence],
                 delay: float = 0.5, max_workers: int = 5) -> list[DesignTestResult]:
        """
        Run TOD for all controls in sample_bank — parallel API calls.

        Args:
            sample_bank: dict mapping control_id → SampleEvidence
            delay: (ignored in parallel mode, kept for backward compat)
            max_workers: number of parallel API calls (default 5)
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed

        # Filter valid controls
        valid_items = []
        for cid, sample in sample_bank.items():
            if cid not in self.rcm_lookup:
                print(f"  ⚠️  Skipping unknown control: {cid}")
            else:
                valid_items.append((cid, sample))

        total_docs = sum(len(s.supporting_docs) for _, s in valid_items)

        print(f"{'='*70}")
        print(f"  TEST OF DESIGN (TOD)")
        print(f"  Controls to evaluate: {len(valid_items)}")
        print(f"  Supporting documents: {total_docs}")
        print(f"  Parallel workers: {max_workers}")
        print(f"{'='*70}\n")

        results_map = {}
        completed = 0
        total = len(valid_items)
        start_time = time.time()

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_cid = {
                executor.submit(self._evaluate_single, cid, sample): cid
                for cid, sample in valid_items
            }

            for future in as_completed(future_to_cid):
                cid = future_to_cid[future]
                completed += 1
                try:
                    result = future.result()
                    results_map[cid] = result
                    icon = "✅" if result.result == "PASS" else "❌" if result.result == "FAIL" else "⚠️"
                    n_docs = len(sample_bank[cid].supporting_docs)
                    docs_note = f" +{n_docs} docs" if n_docs else ""
                    print(f"  [{completed:2d}/{total}] {icon} {cid}{docs_note} | {result.result} | Design: {result.design_adequate}")
                except Exception as e:
                    print(f"  [{completed:2d}/{total}] ⚠️  {cid} | ERROR: {e}")

        elapsed = time.time() - start_time

        # Return results in original sample_bank order
        results = [results_map[cid] for cid, _ in valid_items if cid in results_map]

        adequate = sum(1 for r in results if r.result == "PASS")
        gaps = sum(1 for r in results if r.result == "FAIL")

        print(f"\n{'='*70}")
        print(f"  TOD COMPLETE in {elapsed:.1f}s")
        print(f"  Total: {len(results)} | ✅ Adequate: {adequate} | ❌ Design Gap: {gaps}")
        print(f"{'='*70}\n")

        return results

    # ── Export ────────────────────────────────────────────────────────────

    def results_to_dataframe(self, results: list[DesignTestResult]) -> pd.DataFrame:
        return pd.DataFrame([{
            "Control ID": r.control_id, "Risk ID": r.risk_id,
            "Risk Title": r.risk_title,
            "Risk Level": (self.rcm_lookup.get(r.control_id).risk_level if self.rcm_lookup.get(r.control_id) else ""),
            "Type": r.control_type,
            "Nature": r.nature_of_control, "Frequency": r.control_frequency,
            "Sample ID": r.sample_id,
            "TOD Result": r.result, "Design Adequate": r.design_adequate,
            "Control Exists": r.control_exists, "Process Aligned": r.process_aligned,
            "Confidence": r.confidence, "Remarks": r.remarks,
            "Gap Identified": r.gap_identified,
            "Design Recommendation": r.design_recommendation,
            "Deficiency Type": r.deficiency_type,
            "Evaluated At": r.evaluation_timestamp,
        } for r in results])

    def export_results(self, results: list[DesignTestResult], output_path: str,
                       sample_bank: dict[str, SampleEvidence] = None):
        """
        Export to Excel with columns flowing left to right:
          RCM Details → Sample Evidence → TOD Results → Gap & Reasoning

        Pass sample_bank so the full evidence text is included in the export.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "TOD Results"

        # ── Styles ────────────────────────────────────────────────────────
        bdr = Border(left=Side("thin"), right=Side("thin"),
                     top=Side("thin"), bottom=Side("thin"))
        wrap = Alignment(wrap_text=True, vertical="top")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header styles — 3 color groups
        rcm_hfill  = PatternFill("solid", fgColor="1F3864")   # dark blue  — RCM
        evi_hfill  = PatternFill("solid", fgColor="4472C4")   # mid blue   — Evidence
        res_hfill  = PatternFill("solid", fgColor="2E75B6")   # blue       — Results
        gap_hfill  = PatternFill("solid", fgColor="833C0B")   # brown      — Gap & Reasoning
        hfont      = Font(name="Arial", bold=True, color="FFFFFF", size=10)

        # Data cell colors
        pass_fill    = PatternFill("solid", fgColor="C6EFCE")
        fail_fill    = PatternFill("solid", fgColor="FFC7CE")
        partial_fill = PatternFill("solid", fgColor="FFF2CC")
        warn_fill    = PatternFill("solid", fgColor="FFEB9C")

        # ── Headers ───────────────────────────────────────────────────────
        # Group 1: RCM Details
        rcm_headers = [
            "Process", "SubProcess", "Control Objective",
            "Risk ID", "Risk Title", "Risk Level", "Risk Description",
            "Control ID", "Control Description", "Control Owner",
            "Control Type", "Nature of Control", "Control Frequency",
            "Application/System",
        ]
        # Group 2: Sample Evidence (cols 14-18)
        evi_headers = [
            "Sample ID", "Source Document", "Test Date", "Tested By",
            "Sample Evidence",
        ]
        # Group 3: TOD Results (cols 19-23)
        res_headers = [
            "TOD Result", "Design Adequate", "Control Exists",
            "Process Aligned", "Confidence", "Deficiency Type",
        ]
        # Group 4: Gap & Reasoning (cols 24-26)
        gap_headers = [
            "Remarks / Reasoning", "Gap Identified", "Design Recommendation",
        ]

        all_headers = rcm_headers + evi_headers + res_headers + gap_headers
        header_fills = (
            [rcm_hfill] * len(rcm_headers) +
            [evi_hfill] * len(evi_headers) +
            [res_hfill] * len(res_headers) +
            [gap_hfill] * len(gap_headers)
        )

        for col, (h, fill) in enumerate(zip(all_headers, header_fills), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border = hfont, fill, bdr
            c.alignment = center

        # ── Data Rows ─────────────────────────────────────────────────────
        for ri, r in enumerate(results, 2):
            rcm = self.rcm_lookup.get(r.control_id)

            # Get full evidence from sample_bank if provided
            evidence_text = ""
            source_doc = ""
            test_date = ""
            tested_by = ""
            if sample_bank and r.control_id in sample_bank:
                s = sample_bank[r.control_id]
                evidence_text = s.description.strip()
                if s.supporting_docs:
                    doc_names = ", ".join(d.filename for d in s.supporting_docs)
                    evidence_text += f"\n\n[Supporting Documents: {doc_names}]"
                source_doc = s.source_document
                test_date = s.test_date
                tested_by = s.tested_by
            else:
                evidence_text = r.raw_evidence

            # Build row: RCM → Evidence → Results → Gap
            vals = [
                # RCM Details
                rcm.process if rcm else "",
                rcm.subprocess if rcm else "",
                rcm.control_objective if rcm else "",
                r.risk_id,
                r.risk_title,
                rcm.risk_level if rcm else "",
                rcm.risk_description if rcm else "",
                r.control_id,
                rcm.control_description if rcm else "",
                rcm.control_owner if rcm else "",
                r.control_type,
                r.nature_of_control,
                r.control_frequency,
                rcm.application_system if rcm else "",
                # Sample Evidence
                r.sample_id,
                source_doc,
                test_date,
                tested_by,
                evidence_text,
                # TOD Results
                r.result,
                r.design_adequate,
                r.control_exists,
                r.process_aligned,
                r.confidence,
                r.deficiency_type,
                # Gap & Reasoning
                r.remarks,
                r.gap_identified,
                r.design_recommendation,
            ]

            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=col, value=v)
                cell.alignment, cell.border = wrap, bdr

            # ── Color coding ──────────────────────────────────────────────
            col_offset_res = len(rcm_headers) + len(evi_headers)

            # TOD Result
            rc = ws.cell(row=ri, column=col_offset_res + 1)
            if r.result == "PASS":
                rc.fill = pass_fill
                rc.font = Font(name="Arial", bold=True, color="006100")
            elif r.result == "FAIL":
                rc.fill = fail_fill
                rc.font = Font(name="Arial", bold=True, color="9C0006")
            else:
                rc.fill = warn_fill

            # Design Adequate
            da = ws.cell(row=ri, column=col_offset_res + 2)
            if r.design_adequate == "Yes": da.fill = pass_fill
            elif r.design_adequate == "No": da.fill = fail_fill
            elif r.design_adequate == "Partially": da.fill = partial_fill

            # Control Exists
            ce = ws.cell(row=ri, column=col_offset_res + 3)
            if r.control_exists == "Yes": ce.fill = pass_fill
            elif r.control_exists == "No": ce.fill = fail_fill

            # Process Aligned
            pa = ws.cell(row=ri, column=col_offset_res + 4)
            if r.process_aligned == "Yes": pa.fill = pass_fill
            elif r.process_aligned == "No": pa.fill = fail_fill

            # Deficiency Type
            dt = ws.cell(row=ri, column=col_offset_res + 6)
            if r.deficiency_type == "None":
                dt.fill = pass_fill
            elif r.deficiency_type == "Material Weakness":
                dt.fill = fail_fill
                dt.font = Font(name="Arial", bold=True, color="9C0006")
            elif r.deficiency_type == "Significant Deficiency":
                dt.fill = warn_fill
                dt.font = Font(name="Arial", bold=True, color="9C6500")
            elif r.deficiency_type == "Control Deficiency":
                dt.fill = partial_fill

        # ── Column Widths ─────────────────────────────────────────────────
        widths = [
            # RCM (14 cols)
            16, 16, 30, 12, 28, 14, 40, 14, 45, 22, 12, 14, 15, 18,
            # Evidence (5 cols)
            10, 20, 12, 18, 55,
            # Results (6 cols)
            12, 16, 14, 15, 12, 22,
            # Gap (3 cols)
            55, 45, 45,
        ]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        ws.sheet_properties.pageSetUpPr = None

        wb.save(output_path)
        print(f"[Export] TOD results saved to {output_path}")

    def generate_summary_report(self, results: list[DesignTestResult]) -> str:
        lines = []
        lines.append("=" * 75)
        lines.append("  TEST OF DESIGN (TOD) — SUMMARY REPORT")
        lines.append(f"  Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 75)

        for r in results:
            lines.append(f"\n{'─'*75}")
            lines.append(f"  {r.control_id} | {r.risk_title}")
            lines.append(f"  {r.control_type} | {r.nature_of_control} | {r.control_frequency}")
            icon = "✅" if r.result == "PASS" else "❌"
            lines.append(f"  {icon} TOD: {r.result} | Design: {r.design_adequate} | Exists: {r.control_exists} | Aligned: {r.process_aligned}")
            lines.append(f"     Deficiency: {r.deficiency_type}")
            lines.append(f"     {r.remarks}")
            if r.gap_identified and r.gap_identified not in ("None", "N/A", ""):
                lines.append(f"     🔍 GAP: {r.gap_identified}")
                lines.append(f"     💡 FIX: {r.design_recommendation}")

        adequate = sum(1 for r in results if r.result == "PASS")
        gaps = sum(1 for r in results if r.result == "FAIL")
        mw = sum(1 for r in results if r.deficiency_type == "Material Weakness")
        sd = sum(1 for r in results if r.deficiency_type == "Significant Deficiency")

        lines.append(f"\n{'='*75}")
        lines.append(f"  OVERALL")
        lines.append(f"  Controls Evaluated:     {len(results)}")
        lines.append(f"  ✅ Adequately Designed:  {adequate}")
        lines.append(f"  ❌ Design Gap:           {gaps}")
        if mw: lines.append(f"  🔴 Material Weaknesses: {mw}")
        if sd: lines.append(f"  🟡 Significant Deficiencies: {sd}")
        lines.append(f"{'='*75}")
        return "\n".join(lines)


def _build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run TOD using RCM file (CSV/XLSX) and evidence folder."
    )
    parser.add_argument("--rcm-path", required=True, help="Path to RCM CSV/XLSX file")
    parser.add_argument("--evidence-folder", required=True, help="Path to evidence folder")
    parser.add_argument("--output-xlsx", default="tod_results.xlsx", help="Output Excel filename")
    parser.add_argument("--output-report", default="tod_report.txt", help="Output text report filename")
    parser.add_argument("--max-workers", type=int, default=5, help="Parallel worker count")
    parser.add_argument("--sheet-name", default=None, help="Excel sheet name/index (ignored for CSV)")
    parser.add_argument("--openai-api-key", default=os.getenv("AZURE_OPENAI_API_KEY", os.getenv("OPENAI_API_KEY", "")))
    parser.add_argument("--openai-model", default=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", os.getenv("OPENAI_MODEL", "gpt-5.2-chat")))
    parser.add_argument("--azure-endpoint", default=os.getenv("AZURE_OPENAI_ENDPOINT", "https://entgptaiuat.openai.azure.com"))
    parser.add_argument("--azure-api-version", default=os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"))
    return parser


def cli_main(argv=None):
    parser = _build_cli_parser()
    args = parser.parse_args(argv)

    if not args.openai_api_key:
        parser.error("Missing required: OPENAI_API_KEY or --openai-api-key")

    sample_bank = load_evidence_folder(args.evidence_folder)
    tester = RCMControlTester(
        rcm_path=args.rcm_path,
        openai_api_key=args.openai_api_key,
        openai_model=args.openai_model,
        azure_endpoint=args.azure_endpoint,
        azure_api_key=args.openai_api_key,
        azure_deployment=args.openai_model,
        azure_api_version=args.azure_api_version,
        sheet_name=args.sheet_name,
    )

    results = tester.test_all(sample_bank, max_workers=args.max_workers)
    tester.export_results(results, args.output_xlsx, sample_bank=sample_bank)

    report = tester.generate_summary_report(results)
    print(report)
    with open(args.output_report, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"\n[Report] Saved to {args.output_report}")

    print("\n\nTOD Summary:")
    df = tester.results_to_dataframe(results)
    print(df[[
        "Control ID", "Risk Level", "TOD Result", "Design Adequate",
        "Control Exists", "Process Aligned", "Deficiency Type",
    ]].to_string(index=False))


if __name__ == "__main__":
    cli_main()
