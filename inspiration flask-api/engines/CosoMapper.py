"""
COSO Internal Control Framework Mapper
=======================================
Maps RCM controls to COSO 2013 components, principles, and points of focus
using direct LLM classification (no RAG/embedding needed — COSO is a finite,
well-structured framework that GPT-5.2 knows thoroughly).

Usage:
    from engines.CosoMapper import map_coso_for_controls

    # From TOD/TOE engines:
    coso_map = map_coso_for_controls(rcm_rows, evaluator)
    # coso_map[control_id] → CosoMapping(...)

    # From scoping engine:
    coso_map = map_coso_for_controls_standalone(controls_dicts)
    # Uses its own Azure OpenAI client
"""

from __future__ import annotations

import json
import logging
import os
import re
import requests
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Any
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger("engines.coso_mapper")

# ── Module-level cache for COSO mapping results ──────────────────────────────
# Key: frozenset of control_ids → value: Dict[str, CosoMapping]
# Avoids re-running expensive LLM classification when TOD + TOE (or re-calls)
# request the same mapping.  Cleared on module reload.
_coso_cache: Dict[frozenset, Dict[str, "CosoMapping"]] = {}


def _cache_key(control_ids: list[str]) -> frozenset:
    """Build a hashable cache key from a list of control IDs."""
    return frozenset(control_ids)


# ═══════════════════════════════════════════════════════════════════════════════
#  COSO 2013 REFERENCE TABLE — Baked into the LLM prompt
# ═══════════════════════════════════════════════════════════════════════════════

COSO_REFERENCE = """
COSO 2013 INTERNAL CONTROL — INTEGRATED FRAMEWORK
===================================================

COMPONENT 1: CONTROL ENVIRONMENT
  Principle 1: Commitment to integrity and ethical values
    1.1 Sets the tone at the top  1.2 Establishes standards of conduct
    1.3 Evaluates adherence  1.4 Addresses deviations timely
  Principle 2: Board exercises oversight responsibility
    2.1 Establishes oversight responsibilities  2.2 Applies relevant expertise
    2.3 Operates independently  2.4 Oversees internal control system
    2.5 Oversees entity-level and key controls
  Principle 3: Establishes structure, authority, and responsibility
    3.1 Considers all entity structures  3.2 Establishes reporting lines
    3.3 Defines authorities and responsibilities  3.4 Limits defined by management and board
  Principle 4: Demonstrates commitment to competence
    4.1 Establishes policies and practices  4.2 Evaluates competence and addresses shortcomings
  Principle 5: Enforces accountability
    5.1 Enforces accountability through structures and authorities
    5.2 Establishes performance measures and incentives
    5.3 Evaluates performance and rewards or disciplines

COMPONENT 2: RISK ASSESSMENT
  Principle 6: Specifies suitable objectives
    6.1 Operations objectives  6.2 External financial reporting objectives
    6.3 External non-financial reporting objectives  6.4 Internal reporting objectives
    6.5 Compliance objectives
  Principle 7: Identifies and analyzes risk
    7.1 Includes entity, subsidiary, division, operating unit, and functional levels
    7.2 Analyzes internal and external factors  7.3 Involves appropriate levels of management
    7.4 Estimates significance of risks  7.5 Determines how to respond to risks
  Principle 8: Assesses fraud risk
    8.1 Considers various types of fraud  8.2 Assesses incentive and pressures
    8.3 Assesses opportunity  8.4 Assesses attitudes and rationalizations
    8.5 Anti-fraud programs and controls
  Principle 9: Identifies and analyzes significant change
    9.1 Assesses changes in external environment  9.2 Assesses changes in business model
    9.3 Assesses changes in leadership  9.4 Assesses changes in foreign operations
    9.5 Assesses changes in economic conditions

COMPONENT 3: CONTROL ACTIVITIES
  Principle 10: Selects and develops control activities
    10.1 Integrates with risk assessment  10.2 Considers entity-specific factors
    10.3 Determines relevant business processes  10.4 Evaluates a mix of control activity types
    10.5 Considers at what level activities are applied  10.6 Addresses segregation of duties
  Principle 11: Selects and develops general controls over technology
    11.1 Determines dependency between use of technology and general controls
    11.2 Establishes relevant technology infrastructure controls
    11.3 Establishes relevant security management process controls
    11.4 Establishes relevant technology acquisition, development, and maintenance controls
  Principle 12: Deploys through policies and procedures
    12.1 Establishes policies and procedures to support deployment
    12.2 Establishes responsibility and accountability for executing policies
    12.3 Performs in a timely manner  12.4 Takes corrective action
    12.5 Performs using competent personnel  12.6 Reassesses policies and procedures

COMPONENT 4: INFORMATION & COMMUNICATION
  Principle 13: Uses relevant information
    13.1 Identifies information requirements  13.2 Captures internal and external sources
    13.3 Processes relevant data into information  13.4 Maintains quality throughout processing
    13.5 Considers costs and benefits
  Principle 14: Communicates internally
    14.1 Communicates internal control information  14.2 Communicates with the board
    14.3 Provides separate communication lines  14.4 Selects relevant method of communication
    14.5 Board communicates downward and across
  Principle 15: Communicates externally
    15.1 Communicates to external parties  15.2 Enables inbound communications
    15.3 Communicates with board on external matters

COMPONENT 5: MONITORING
  Principle 16: Conducts ongoing and/or separate evaluations
    16.1 Considers a mix of ongoing and separate evaluations
    16.2 Considers rate of change  16.3 Establishes baseline understanding
    16.4 Uses knowledgeable personnel  16.5 Integrates with business processes
  Principle 17: Evaluates and communicates deficiencies
    17.1 Assesses results  17.2 Communicates deficiencies
    17.3 Monitors corrective actions  17.4 Reports to board or management
    17.5 Timeliness of corrective action

ADDITIONAL OBJECTIVE: SAFEGUARDING OF ASSETS
  Relates to controls that protect assets from unauthorized acquisition, use, or disposition.
  Typically maps to Principles 10, 11, 12 (Control Activities) when the control's purpose
  is specifically about asset protection, physical security, or access restrictions.

OBJECTIVE CATEGORIES:
  - Operations: Controls over effectiveness and efficiency of operations
  - Financial Reporting (Finance): Controls over reliability of financial reporting
  - Compliance: Controls over compliance with applicable laws and regulations
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class CosoMapping:
    """Full classification result for a single control — Financial Assertions, IFC, and COSO."""
    control_id: str
    # ── Financial Assertions (6 Yes/No — blank if not classified) ──
    existence_or_occurrence: str = ""
    completeness: str = ""
    rights_and_obligations: str = ""
    valuation_and_allocation: str = ""
    presentation_and_disclosure: str = ""
    accuracy: str = ""
    # ── IFC Components (5 Yes/No — blank if not classified) ──
    ifc_accuracy: str = ""
    timely_preparation_of_financial_info: str = ""
    anti_fraud: str = ""
    policies_and_procedures: str = ""
    safeguarding_of_assets: str = ""
    # ── COSO Components (5 Yes/No — blank if not classified) ──
    control_environment: str = ""
    risk_assessment: str = ""
    control_activities: str = ""
    info_and_communication: str = ""
    monitoring: str = ""
    # ── COSO Details (text) ──
    coso_principles: str = ""
    coso_points_of_focus: str = ""
    reporting_category: str = ""
    # Legacy field for backward compatibility
    coso_objective: str = ""


# ── Column header constants ──────────────────────────────────────────────────

FINANCIAL_ASSERTION_HEADERS = [
    "Existence or Occurrence",
    "Completeness",
    "Rights and Obligations",
    "Valuation and Allocation",
    "Presentation and Disclosure",
    "Accuracy",
]

IFC_COMPONENT_HEADERS = [
    "Accuracy",
    "Timely Preparation of Reliable Financial Information",
    "Anti-Fraud",
    "Policies and Procedures",
    "Safeguarding of Assets",
]

COSO_COLUMN_HEADERS = [
    "Control Environment",
    "Risk Assessment",
    "Control Activities",
    "Info & Communication",
    "Monitoring",
    "COSO Principle",
    "COSO Point of Focus",
    "Reporting Category",
]

# Counts for formatting
N_ASSERTION_COLS = len(FINANCIAL_ASSERTION_HEADERS)     # 6
N_IFC_COLS = len(IFC_COMPONENT_HEADERS)                 # 5
N_COSO_COLS = len(COSO_COLUMN_HEADERS)                  # 8
N_COSO_COMPONENT_COLS = 5  # Yes/No tick columns within COSO


# ── Row converters ───────────────────────────────────────────────────────────

def assertion_mapping_to_row(m: CosoMapping) -> list:
    """Financial Assertions values in header order."""
    return [
        m.existence_or_occurrence,
        m.completeness,
        m.rights_and_obligations,
        m.valuation_and_allocation,
        m.presentation_and_disclosure,
        m.accuracy,
    ]


def ifc_mapping_to_row(m: CosoMapping) -> list:
    """IFC Components values in header order."""
    return [
        m.ifc_accuracy,
        m.timely_preparation_of_financial_info,
        m.anti_fraud,
        m.policies_and_procedures,
        m.safeguarding_of_assets,
    ]


def coso_mapping_to_row(m: CosoMapping) -> list:
    """COSO values in header order."""
    return [
        m.control_environment,
        m.risk_assessment,
        m.control_activities,
        m.info_and_communication,
        m.monitoring,
        m.coso_principles,
        m.coso_points_of_focus,
        m.reporting_category or m.coso_objective,
    ]


def empty_assertion_row() -> list:
    return [""] * N_ASSERTION_COLS


def empty_ifc_row() -> list:
    return [""] * N_IFC_COLS


def empty_coso_row() -> list:
    return [""] * N_COSO_COLS


# ═══════════════════════════════════════════════════════════════════════════════
#  LLM PROMPT
# ═══════════════════════════════════════════════════════════════════════════════

# ── COSO Prompt ──
COSO_SYSTEM_PROMPT = (
    "You are a senior audit expert specializing in the COSO 2013 Internal Control — "
    "Integrated Framework. Your task is to map each control to its applicable COSO components, "
    "principles, and points of focus.\n\n"
    + COSO_REFERENCE +
    "\n\nINSTRUCTIONS:\n"
    "For each control provided, determine:\n"
    "1. Which COSO component(s) the control primarily addresses — mark Yes or No for each:\n"
    "   - control_environment: Yes if the control relates to tone at top, ethics, governance, competence\n"
    "   - risk_assessment: Yes if the control identifies, analyzes, or responds to risks\n"
    "   - control_activities: Yes if the control is an operational/transaction control (most common)\n"
    "   - info_and_communication: Yes if the control ensures information quality or communication\n"
    "   - monitoring: Yes if the control reviews/evaluates other controls or monitors performance\n"
    "2. The specific COSO Principle number(s) — e.g. \"10\" or \"10 & 12\"\n"
    "3. The specific Point(s) of Focus — e.g. \"10.1\" or \"10.1 & 12.1\"\n"
    "4. The Reporting Category — EXACTLY one of:\n"
    "   - \"Operational\" — for operational effectiveness controls\n"
    "   - \"Finance\" — for financial reporting controls\n"
    "   - \"Operational and Finance\" — for controls covering both\n"
    "   - \"Reporting\" — for controls related to internal/external reporting\n"
    "   - \"Compliance\" — for regulatory compliance controls\n\n"
    "IMPORTANT RULES:\n"
    "- Most controls map to Principle 10 (Control Activities) and/or 12 (Policies & Procedures)\n"
    "- IT controls map to Principle 11 (Technology Controls)\n"
    "- Management review controls often also map to Principle 16 (Monitoring)\n"
    "- Reconciliation controls are typically Principle 10\n"
    "- Access controls map to Principle 11\n"
    "- Segregation of duties maps to Principle 10.6\n"
    "- Every control MUST have at least one component marked Yes and at least one principle\n"
    "- Most financial reporting controls have reporting_category = \"Finance\"\n\n"
    "Return JSON: {\"controls\": [{\"control_id\": \"...\", \"control_environment\": \"Yes/No\", "
    "\"risk_assessment\": \"Yes/No\", \"control_activities\": \"Yes/No\", "
    "\"info_and_communication\": \"Yes/No\", \"monitoring\": \"Yes/No\", "
    "\"coso_principles\": \"10 & 12\", \"coso_points_of_focus\": \"10.1 & 12.1\", "
    "\"reporting_category\": \"Finance\"}]}"
)

# ── Financial Assertions Prompt ──
ASSERTION_SYSTEM_PROMPT = (
    "You are a senior audit expert. For each control, determine which PCAOB/COSO "
    "financial statement assertions the control addresses. Mark Yes or No for each.\n\n"
    "THE SIX FINANCIAL ASSERTIONS:\n"
    "1. existence_or_occurrence — The control verifies that assets, liabilities, or transactions "
    "actually exist and that recorded transactions actually occurred during the period.\n"
    "2. completeness — The control ensures ALL transactions, accounts, and disclosures that should "
    "be recorded or disclosed are included. Nothing is omitted.\n"
    "3. rights_and_obligations — The control verifies that the entity holds or controls the rights "
    "to assets, and that liabilities are genuine obligations of the entity.\n"
    "4. valuation_and_allocation — The control ensures assets, liabilities, equity, revenues, and "
    "expenses are recorded at appropriate amounts, and any valuation adjustments are proper.\n"
    "5. presentation_and_disclosure — The control ensures transactions are properly classified, "
    "described, and disclosed in accordance with the applicable financial reporting framework.\n"
    "6. accuracy — The control ensures amounts and data relating to recorded transactions are "
    "recorded correctly (mathematically correct, proper period, right account).\n\n"
    "RULES:\n"
    "- A control can address MULTIPLE assertions simultaneously\n"
    "- Reconciliation controls typically address Completeness and Accuracy\n"
    "- Three-way match controls address Existence/Occurrence, Accuracy, and Valuation\n"
    "- Journal entry review controls address Completeness, Accuracy, and Presentation\n"
    "- Access controls address Existence/Occurrence and Rights and Obligations\n"
    "- Period-end close controls often address Completeness, Valuation, and Presentation\n"
    "- Every control MUST have at least one assertion marked Yes\n\n"
    "Return JSON: {\"controls\": [{\"control_id\": \"...\", "
    "\"existence_or_occurrence\": \"Yes/No\", \"completeness\": \"Yes/No\", "
    "\"rights_and_obligations\": \"Yes/No\", \"valuation_and_allocation\": \"Yes/No\", "
    "\"presentation_and_disclosure\": \"Yes/No\", \"accuracy\": \"Yes/No\"}]}"
)

# ── IFC Components Prompt ──
IFC_SYSTEM_PROMPT = (
    "You are a senior audit expert. For each control, determine which Internal Financial "
    "Control (IFC) components the control addresses. Mark Yes or No for each.\n\n"
    "THE FIVE IFC COMPONENTS (per Companies Act Section 134 / SOX framework):\n"
    "1. ifc_accuracy — The control ensures accuracy and completeness of accounting records, "
    "ledger entries, and financial data. Reconciliations, data validations, input checks, "
    "and mathematical verification controls map here.\n"
    "2. timely_preparation_of_financial_info — The control contributes to the timely preparation "
    "of reliable financial statements and information. Period-end closing controls, reporting "
    "deadline controls, and financial reporting review controls map here.\n"
    "3. anti_fraud — The control prevents or detects fraud, errors, or irregularities. "
    "Segregation of duties, authorization limits, exception monitoring, anomaly detection, "
    "and whistleblower-related controls map here.\n"
    "4. policies_and_procedures — The control enforces adherence to company policies and "
    "procedures, or represents a documented policy itself. Approval workflows, standard "
    "operating procedures, compliance checks, and mandate-enforcement controls map here.\n"
    "5. safeguarding_of_assets — The control protects physical or digital assets from "
    "unauthorized acquisition, use, or disposition. Access controls, physical security, "
    "backup controls, asset verification, and custody controls map here.\n\n"
    "RULES:\n"
    "- A control can address MULTIPLE IFC components simultaneously\n"
    "- Most transaction-level controls address both ifc_accuracy and policies_and_procedures\n"
    "- Access and authorization controls typically address anti_fraud and safeguarding_of_assets\n"
    "- Month-end/quarter-end controls address timely_preparation_of_financial_info\n"
    "- Every control MUST have at least one component marked Yes\n\n"
    "Return JSON: {\"controls\": [{\"control_id\": \"...\", "
    "\"ifc_accuracy\": \"Yes/No\", \"timely_preparation_of_financial_info\": \"Yes/No\", "
    "\"anti_fraud\": \"Yes/No\", \"policies_and_procedures\": \"Yes/No\", "
    "\"safeguarding_of_assets\": \"Yes/No\"}]}"
)


def _build_user_prompt(controls: list[dict], context: str = "COSO") -> str:
    """Build the user prompt with control details for classification.

    Args:
        controls: List of control dicts with control_id, control_description, etc.
        context: Classification context — "COSO", "Assertions", or "IFC".
    """
    _CONTEXT_INSTRUCTIONS = {
        "COSO": "Classify the following controls by COSO components, principles, and points of focus.",
        "Assertions": "Classify the following controls by financial statement assertions (existence/occurrence, completeness, rights & obligations, valuation & allocation, presentation & disclosure, accuracy).",
        "IFC": "Classify the following controls by IFC components (accuracy, timely preparation of financial info, anti-fraud, policies & procedures, safeguarding of assets).",
    }
    instruction = _CONTEXT_INSTRUCTIONS.get(context, _CONTEXT_INSTRUCTIONS["COSO"])
    lines = [f"{instruction}\n"]
    for ctrl in controls:
        lines.append(
            f"- control_id: {ctrl['control_id']}\n"
            f"  control_description: {ctrl.get('control_description', '')}\n"
            f"  control_objective: {ctrl.get('control_objective', '')}\n"
            f"  control_type: {ctrl.get('control_type', '')}\n"
            f"  nature_of_control: {ctrl.get('nature_of_control', '')}\n"
            f"  control_frequency: {ctrl.get('control_frequency', '')}\n"
            f"  risk_description: {ctrl.get('risk_description', '')}\n"
            f"  process: {ctrl.get('process', '')}\n"
        )
    lines.append("\nReturn a JSON object with key \"controls\" containing one object per control_id.")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
#  PARSING
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_mappings(data) -> list[dict]:
    """Extract list of mapping dicts from either a parsed dict/list or raw text.

    Handles:
      - Already-parsed dict (from _chat_json): {"controls": [...]}
      - Already-parsed list: [...]
      - Raw text string (from standalone calls)
    """
    # Already parsed by _chat_json
    if isinstance(data, dict):
        for key in ("controls", "mappings", "results", "data"):
            if key in data and isinstance(data[key], list):
                return data[key]
        # Single mapping object
        if "control_id" in data:
            return [data]
        return []

    if isinstance(data, list):
        return data

    # Raw text — parse it
    if not isinstance(data, str) or not data.strip():
        return []

    try:
        result = json.loads(data)
        return _extract_mappings(result)  # recurse on parsed result
    except json.JSONDecodeError:
        pass

    # Try extracting JSON from markdown code block
    match = re.search(r"```(?:json)?\s*(\{[\s\S]*?\}|\[[\s\S]*?\])\s*```", data)
    if match:
        try:
            return _extract_mappings(json.loads(match.group(1)))
        except json.JSONDecodeError:
            pass

    # Try finding raw JSON object or array
    for pattern in [r"\{[\s\S]*\}", r"\[[\s\S]*\]"]:
        match = re.search(pattern, data)
        if match:
            try:
                return _extract_mappings(json.loads(match.group(0)))
            except json.JSONDecodeError:
                pass

    logger.warning("Failed to parse COSO mapping response")
    return []


def _yn(val) -> str:
    """Normalize to Yes/No."""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    s = str(val).strip().lower()
    return "Yes" if s in ("yes", "true", "1", "y", "●", "✓") else "No"


_FIELD_KEYWORDS: Dict[str, list[str]] = {
    # Financial Assertions — each field mapped to distinctive substrings
    "existence_or_occurrence": ["existence", "occurrence"],
    "completeness": ["completeness"],
    "rights_and_obligations": ["obligation", "rights_and"],
    "valuation_and_allocation": ["valuation", "allocation"],
    "presentation_and_disclosure": ["presentation", "disclosure"],
    # IFC Components
    "ifc_accuracy": ["ifc_accuracy"],
    "timely_preparation_of_financial_info": ["timely"],
    "anti_fraud": ["fraud"],
    "policies_and_procedures": ["policies", "procedures"],
    "safeguarding_of_assets": ["safeguard"],
    # COSO Components
    "control_environment": ["control_environment"],
    "risk_assessment": ["risk_assessment"],
    "control_activities": ["control_activities"],
    "info_and_communication": ["communication"],
    "monitoring": ["monitoring"],
    # COSO text fields
    "coso_principles": ["principle"],
    "coso_points_of_focus": ["point", "focus"],
    "reporting_category": ["reporting_category", "coso_objective", "objective_category"],
}


def _dict_to_coso_mapping(d: dict) -> Optional[CosoMapping]:
    """Convert a raw dict from LLM response to a CosoMapping dataclass.

    Uses keyword matching on normalized keys. Each field has distinctive keywords
    that uniquely identify it regardless of how the LLM formats the key name.
    """
    # Normalize all keys
    nd: Dict[str, Any] = {}
    for k, v in d.items():
        nk = re.sub(r"[\s\-/&.]+", "_", k.strip().lower()).strip("_")
        nd[nk] = v

    cid = nd.get("control_id", "")
    if not cid:
        return None

    m = CosoMapping(control_id=str(cid).strip())

    # Detect stream type so "accuracy" resolves correctly
    is_ifc = any("timely" in k or "fraud" in k or "safeguard" in k for k in nd)
    is_assertion = any("existence" in k or "obligation" in k or "valuation" in k for k in nd)

    # Build reverse lookup: for each normalized response key, find which field it maps to
    matched: Dict[str, Any] = {}
    for resp_key, resp_val in nd.items():
        if resp_key == "control_id":
            continue
        for field_name, keywords in _FIELD_KEYWORDS.items():
            if any(kw in resp_key for kw in keywords):
                matched[field_name] = resp_val
                break

    # Handle "accuracy" ambiguity — could be assertion or IFC
    if "accuracy" in nd and "ifc_accuracy" not in matched:
        for resp_key, resp_val in nd.items():
            if resp_key == "accuracy" or (resp_key.endswith("accuracy") and "ifc" not in resp_key):
                if is_ifc:
                    matched["ifc_accuracy"] = resp_val
                elif is_assertion or not is_ifc:
                    matched["accuracy"] = resp_val
                break

    # Apply matched values to dataclass
    yn_fields = {
        "existence_or_occurrence", "completeness", "rights_and_obligations",
        "valuation_and_allocation", "presentation_and_disclosure", "accuracy",
        "ifc_accuracy", "timely_preparation_of_financial_info", "anti_fraud",
        "policies_and_procedures", "safeguarding_of_assets",
        "control_environment", "risk_assessment", "control_activities",
        "info_and_communication", "monitoring",
    }
    text_fields = {"coso_principles", "coso_points_of_focus", "reporting_category"}

    for field_name, val in matched.items():
        if field_name in yn_fields:
            setattr(m, field_name, _yn(val))
        elif field_name in text_fields:
            setattr(m, field_name, str(val).strip())

    # Debug: log if nothing matched
    set_count = sum(1 for f in yn_fields if getattr(m, f) != "")
    if set_count == 0 and len(nd) > 1:
        logger.warning("[Classification] %s: 0 fields matched from keys: %s", cid, list(nd.keys()))
    else:
        logger.debug("[Classification] %s: %d fields set from %d response keys", cid, set_count, len(nd) - 1)

    return m


def _merge_mappings(base: CosoMapping, overlay: CosoMapping) -> CosoMapping:
    """Merge fields from overlay into base — non-default values from overlay win."""
    for field_name in CosoMapping.__dataclass_fields__:
        if field_name == "control_id":
            continue
        overlay_val = getattr(overlay, field_name)
        default_val = CosoMapping.__dataclass_fields__[field_name].default
        if overlay_val != default_val:
            setattr(base, field_name, overlay_val)
    return base


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN API — for TOD/TOE engines (uses AzureOpenAIEvaluator)
# ═══════════════════════════════════════════════════════════════════════════════

def map_coso_for_controls(
    rcm_rows: list,
    evaluator,
    batch_size: int = 25,
    max_workers: int = 3,
) -> Dict[str, CosoMapping]:
    """Map COSO principles for a list of RCMRow objects.

    Uses the existing AzureOpenAIEvaluator from TOD/TOE engines.
    Runs all 3 classification streams (Assertions, IFC, COSO) in parallel.

    Args:
        rcm_rows:   List of RCMRow dataclass instances.
        evaluator:  AzureOpenAIEvaluator instance (has ._call_llm method).
        batch_size: Controls per LLM call (default 25).
        max_workers: Parallel streams (default 3).

    Returns:
        Dict mapping control_id → CosoMapping.
    """
    if not rcm_rows:
        return {}

    # Deduplicate by control_id (same control may appear multiple times in RCM)
    seen = set()
    unique_controls = []
    for r in rcm_rows:
        if r.control_id and r.control_id not in seen:
            seen.add(r.control_id)
            unique_controls.append({
                "control_id": r.control_id,
                "control_description": r.control_description,
                "control_objective": getattr(r, "control_objective", ""),
                "control_type": getattr(r, "control_type", ""),
                "nature_of_control": getattr(r, "nature_of_control", ""),
                "control_frequency": getattr(r, "control_frequency", ""),
                "risk_description": r.risk_description,
                "process": r.process,
            })

    if not unique_controls:
        return {}

    # Check cache — avoid re-running LLM calls for the same controls
    cids = [c["control_id"] for c in unique_controls]
    key = _cache_key(cids)
    if key in _coso_cache:
        logger.info("[Classification] Cache hit — reusing COSO mapping for %d controls", len(cids))
        return _coso_cache[key]

    logger.info(
        "[Classification] Mapping %d unique controls — 3 concurrent LLM streams "
        "(Financial Assertions, IFC Components, COSO) batch_size=%d",
        len(unique_controls), batch_size,
    )

    # Split into batches
    batches = [unique_controls[i:i + batch_size] for i in range(0, len(unique_controls), batch_size)]

    # ── Helper: run one classification stream across all batches ──
    def _run_stream(system_prompt: str, stream_name: str) -> Dict[str, CosoMapping]:
        stream_map: Dict[str, CosoMapping] = {}
        for bi, batch in enumerate(batches):
            user_prompt = _build_user_prompt(batch, context=stream_name)
            try:
                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ]
                parsed = evaluator._chat_json(messages, max_tokens=4096, timeout=120)
                raw = _extract_mappings(parsed)
                for d in raw:
                    m = _dict_to_coso_mapping(d)
                    if m:
                        stream_map[m.control_id] = m
                logger.info(
                    "[%s] Batch %d/%d: classified %d controls",
                    stream_name, bi + 1, len(batches), len(raw),
                )
            except Exception as e:
                logger.error("[%s] Batch %d failed: %s", stream_name, bi + 1, e)
        return stream_map

    # ── Run 3 streams in parallel — they are independent classifications ──
    coso_map: Dict[str, CosoMapping] = {}
    assertion_map: Dict[str, CosoMapping] = {}
    ifc_map: Dict[str, CosoMapping] = {}

    streams = [
        (ASSERTION_SYSTEM_PROMPT, "Assertions"),
        (IFC_SYSTEM_PROMPT, "IFC"),
        (COSO_SYSTEM_PROMPT, "COSO"),
    ]

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_run_stream, prompt, name): name
            for prompt, name in streams
        }
        for future in as_completed(futures):
            stream_name = futures[future]
            try:
                result = future.result()
                if stream_name == "Assertions":
                    assertion_map = result
                elif stream_name == "IFC":
                    ifc_map = result
                else:
                    coso_map = result
            except Exception as e:
                logger.error("[%s] Stream failed: %s", stream_name, e)

    # ── Merge all 3 streams into a single mapping per control ──
    result_map: Dict[str, CosoMapping] = {}
    all_cids = set(list(coso_map.keys()) + list(assertion_map.keys()) + list(ifc_map.keys()))

    for cid in all_cids:
        base = CosoMapping(control_id=cid)
        if cid in coso_map:
            base = _merge_mappings(base, coso_map[cid])
        if cid in assertion_map:
            base = _merge_mappings(base, assertion_map[cid])
        if cid in ifc_map:
            base = _merge_mappings(base, ifc_map[cid])
        result_map[cid] = base

    # Fill in controls not returned by any stream — leave blank (not "No")
    for ctrl in unique_controls:
        if ctrl["control_id"] not in result_map:
            logger.warning("[Classification] No mapping for %s — columns will be blank", ctrl["control_id"])
            result_map[ctrl["control_id"]] = CosoMapping(control_id=ctrl["control_id"])

    logger.info(
        "[Classification] Complete: %d controls — COSO=%d, Assertions=%d, IFC=%d",
        len(result_map), len(coso_map), len(assertion_map), len(ifc_map),
    )
    # Cache the result so TOD + TOE exports don't re-run LLM calls
    _coso_cache[key] = result_map
    return result_map


# ═══════════════════════════════════════════════════════════════════════════════
#  STANDALONE API — for Scoping Engine (uses its own LLM client)
# ═══════════════════════════════════════════════════════════════════════════════

def _call_azure_openai(system_prompt: str, user_prompt: str, max_tokens: int = 4096) -> str:
    """Direct Azure OpenAI call for the scoping engine (which uses a different LLM client)."""
    try:
        from config import (
            AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY,
            AZURE_OPENAI_DEPLOYMENT, AZURE_OPENAI_API_VERSION,
        )
    except ImportError:
        AZURE_OPENAI_ENDPOINT = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
        AZURE_OPENAI_API_KEY = os.environ.get("AZURE_OPENAI_API_KEY", "")
        AZURE_OPENAI_DEPLOYMENT = os.environ.get("AZURE_OPENAI_DEPLOYMENT_NAME", "")
        AZURE_OPENAI_API_VERSION = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

    if not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_API_KEY:
        raise RuntimeError("Azure OpenAI credentials not configured for COSO mapping")

    url = (
        f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/{AZURE_OPENAI_DEPLOYMENT}"
        f"/chat/completions?api-version={AZURE_OPENAI_API_VERSION}"
    )
    headers = {"api-key": AZURE_OPENAI_API_KEY, "Content-Type": "application/json"}
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_completion_tokens": max_tokens,
        "response_format": {"type": "json_object"},
    }

    for attempt in range(3):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            if attempt < 2:
                import time
                wait = 2 ** (attempt + 1)
                logger.warning(f"[COSO] LLM call attempt {attempt + 1}/3 failed: {e}. Retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise


def map_coso_for_controls_standalone(
    controls: list[dict],
    batch_size: int = 25,
    max_workers: int = 3,
) -> Dict[str, CosoMapping]:
    """Map COSO principles for a list of control dicts (used by the scoping engine).

    Each dict should have keys: control_id, control_description, control_objective,
    control_type, nature_of_control, control_frequency, risk_description, process.

    Returns:
        Dict mapping control_id → CosoMapping.
    """
    if not controls:
        return {}

    # Deduplicate
    seen = set()
    unique = []
    for c in controls:
        cid = c.get("control_id") or c.get("id", "")
        if cid and cid not in seen:
            seen.add(cid)
            unique.append({
                "control_id": cid,
                "control_description": c.get("control_description") or c.get("desc", ""),
                "control_objective": c.get("control_objective") or c.get("objective", ""),
                "control_type": c.get("control_type") or c.get("type", ""),
                "nature_of_control": c.get("nature_of_control") or c.get("nature", ""),
                "control_frequency": c.get("control_frequency") or c.get("freq", ""),
                "risk_description": c.get("risk_description", ""),
                "process": c.get("process", ""),
            })

    if not unique:
        return {}

    # Check cache — avoid re-running LLM calls for the same controls
    cids = [c["control_id"] for c in unique]
    key = _cache_key(cids)
    if key in _coso_cache:
        logger.info("[Classification] Cache hit — reusing standalone COSO mapping for %d controls", len(cids))
        return _coso_cache[key]

    logger.info(
        "[Classification] Standalone mapping for %d controls — 3 parallel streams",
        len(unique),
    )

    batches = [unique[i:i + batch_size] for i in range(0, len(unique), batch_size)]

    def _run_stream(system_prompt: str, stream_name: str) -> Dict[str, CosoMapping]:
        stream_map: Dict[str, CosoMapping] = {}
        for bi, batch in enumerate(batches):
            user_prompt = _build_user_prompt(batch, context=stream_name)
            try:
                response_text = _call_azure_openai(system_prompt, user_prompt)
                raw = _extract_mappings(response_text)
                for d in raw:
                    m = _dict_to_coso_mapping(d)
                    if m:
                        stream_map[m.control_id] = m
                logger.info("[%s] Standalone batch %d/%d: %d controls",
                            stream_name, bi + 1, len(batches), len(raw))
            except Exception as e:
                logger.error("[%s] Standalone batch %d failed: %s", stream_name, bi + 1, e)
        return stream_map

    # ── Run 3 streams in parallel — they are independent classifications ──
    coso_map: Dict[str, CosoMapping] = {}
    assertion_map: Dict[str, CosoMapping] = {}
    ifc_map: Dict[str, CosoMapping] = {}

    streams = [
        (ASSERTION_SYSTEM_PROMPT, "Assertions"),
        (IFC_SYSTEM_PROMPT, "IFC"),
        (COSO_SYSTEM_PROMPT, "COSO"),
    ]

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_run_stream, prompt, name): name
            for prompt, name in streams
        }
        for future in as_completed(futures):
            stream_name = futures[future]
            try:
                result = future.result()
                if stream_name == "Assertions":
                    assertion_map = result
                elif stream_name == "IFC":
                    ifc_map = result
                else:
                    coso_map = result
            except Exception as e:
                logger.error("[%s] Standalone stream failed: %s", stream_name, e)

    # Merge all 3 streams
    result_map: Dict[str, CosoMapping] = {}
    all_cids = set(list(coso_map.keys()) + list(assertion_map.keys()) + list(ifc_map.keys()))
    for cid in all_cids:
        base = CosoMapping(control_id=cid)
        if cid in coso_map:
            base = _merge_mappings(base, coso_map[cid])
        if cid in assertion_map:
            base = _merge_mappings(base, assertion_map[cid])
        if cid in ifc_map:
            base = _merge_mappings(base, ifc_map[cid])
        result_map[cid] = base

    # Fill in controls not returned by any stream — leave blank
    for c in unique:
        if c["control_id"] not in result_map:
            result_map[c["control_id"]] = CosoMapping(control_id=c["control_id"])

    logger.info(
        "[Classification] Standalone complete: %d controls — COSO=%d, Assertions=%d, IFC=%d",
        len(result_map), len(coso_map), len(assertion_map), len(ifc_map),
    )
    # Cache the result so subsequent calls don't re-run LLM classification
    _coso_cache[key] = result_map
    return result_map


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL HELPERS — write classification columns into openpyxl worksheets
# ═══════════════════════════════════════════════════════════════════════════════

def _write_headers(ws, headers, start_col, header_row, font, fill, border, alignment):
    """Generic helper: write a list of header strings into a worksheet row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment
    return start_col + len(headers)


def _write_data_row(ws, row_num, start_col, values, border, alignment, fill_yes):
    """Generic helper: write a list of values, highlighting 'Yes' cells."""
    for i, val in enumerate(values):
        cell = ws.cell(row=row_num, column=start_col + i, value=val)
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment
        if fill_yes and str(val).strip().lower() == "yes":
            cell.fill = fill_yes


# ── Financial Assertions ──

def write_assertion_headers(ws, start_col: int, header_row: int = 1,
                            font=None, fill=None, border=None, alignment=None):
    """Write Financial Assertion column headers. Returns next column index."""
    return _write_headers(ws, FINANCIAL_ASSERTION_HEADERS, start_col, header_row,
                          font, fill, border, alignment)


def write_assertion_data_row(ws, row_num: int, start_col: int,
                             mapping: Optional[CosoMapping] = None,
                             border=None, alignment=None, fill_yes=None):
    """Write one row of Financial Assertion data."""
    values = assertion_mapping_to_row(mapping) if mapping else empty_assertion_row()
    _write_data_row(ws, row_num, start_col, values, border, alignment, fill_yes)


def set_assertion_column_widths(ws, start_col: int):
    """Set column widths for Financial Assertion columns."""
    from openpyxl.utils import get_column_letter
    widths = [18, 14, 18, 18, 20, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


# ── IFC Components ──

def write_ifc_headers(ws, start_col: int, header_row: int = 1,
                      font=None, fill=None, border=None, alignment=None):
    """Write IFC Component column headers. Returns next column index."""
    return _write_headers(ws, IFC_COMPONENT_HEADERS, start_col, header_row,
                          font, fill, border, alignment)


def write_ifc_data_row(ws, row_num: int, start_col: int,
                       mapping: Optional[CosoMapping] = None,
                       border=None, alignment=None, fill_yes=None):
    """Write one row of IFC Component data."""
    values = ifc_mapping_to_row(mapping) if mapping else empty_ifc_row()
    _write_data_row(ws, row_num, start_col, values, border, alignment, fill_yes)


def set_ifc_column_widths(ws, start_col: int):
    """Set column widths for IFC Component columns."""
    from openpyxl.utils import get_column_letter
    widths = [12, 32, 12, 18, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


# ── COSO ──

def write_coso_headers(ws, start_col: int, header_row: int = 1,
                       font=None, fill=None, border=None, alignment=None):
    """Write COSO column headers. Returns next column index."""
    return _write_headers(ws, COSO_COLUMN_HEADERS, start_col, header_row,
                          font, fill, border, alignment)


def write_coso_data_row(ws, row_num: int, start_col: int,
                        mapping: Optional[CosoMapping] = None,
                        border=None, alignment=None, fill_yes=None):
    """Write one row of COSO data."""
    values = coso_mapping_to_row(mapping) if mapping else empty_coso_row()
    _write_data_row(ws, row_num, start_col, values, border, alignment, fill_yes)


def set_coso_column_widths(ws, start_col: int):
    """Set reasonable column widths for COSO columns."""
    from openpyxl.utils import get_column_letter
    widths = [16, 14, 16, 18, 12, 16, 18, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w
