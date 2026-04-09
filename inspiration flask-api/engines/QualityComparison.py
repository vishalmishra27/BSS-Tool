"""
Quality Comparison Engine
=========================
Runs the full control testing pipeline autonomously (TOD -> Sampling -> TOE),
then compares our tool's results against pre-existing human auditor
workpapers to identify gaps in the human's work.

Direction: Single-directional — only gaps in human work
(what our tool caught that the human missed/got wrong).

Comparison granularity:
  - Per-control (effectiveness verdict)
  - Per-attribute (attributes tested, PASS/FAIL differences)
  - Per-sample (samples evaluated, result differences)
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger("engines.QualityComparison")

# Regex to strip characters illegal in XML/Excel (openpyxl raises
# IllegalCharacterError for control chars 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F).
_ILLEGAL_XML_CHARS = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
)


def _sanitize_for_excel(obj):
    """Recursively strip illegal XML characters from all strings in obj."""
    if isinstance(obj, str):
        return _ILLEGAL_XML_CHARS.sub("", obj)
    if isinstance(obj, list):
        return [_sanitize_for_excel(item) for item in obj]
    if isinstance(obj, dict):
        return {k: _sanitize_for_excel(v) for k, v in obj.items()}
    if hasattr(obj, "__dataclass_fields__"):
        from dataclasses import fields as dc_fields
        for f in dc_fields(obj):
            val = getattr(obj, f.name)
            cleaned = _sanitize_for_excel(val)
            if cleaned is not val:
                try:
                    object.__setattr__(obj, f.name, cleaned)
                except (AttributeError, TypeError):
                    pass
        return obj
    return obj


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class HumanSampleResult:
    """One sample row parsed from a per-control sheet in the human workpaper."""
    control_id: str
    sample_id: str
    sample_details: Dict[str, str]    # column_header -> value
    attribute_results: Dict[str, str] # {"A": "Yes", "B": "No", ...}
    remarks: str = ""
    result: str = ""                  # PASS / FAIL (derived: any "No" = FAIL)


@dataclass
class HumanControlResult:
    """Parsed from both the summary sheet and per-control sheet."""
    control_id: str
    # From summary / conclusion
    operating_effectiveness: str = ""   # Effective / Effective with Exceptions / Not Effective
    deficiency_type: str = ""
    total_samples: int = 0
    passed_samples: int = 0
    failed_samples: int = 0
    deviation_rate: float = 0.0
    overall_remarks: str = ""
    # From per-control sheet
    worksteps: List[str] = field(default_factory=list)
    attributes_tested: List[Dict] = field(default_factory=list)  # [{"id", "name", "description"}]
    sample_results: List[HumanSampleResult] = field(default_factory=list)


@dataclass
class ControlComparison:
    """Comparison result for one control."""
    control_id: str
    phase: str = ""  # "TOD" or "TOE"
    # Verdict
    our_effectiveness: str = ""
    human_effectiveness: str = ""
    verdict_gap: bool = False
    verdict_detail: str = ""
    # Attributes
    our_attributes: List[str] = field(default_factory=list)
    human_attributes: List[str] = field(default_factory=list)
    attributes_human_missed: List[str] = field(default_factory=list)
    attributes_human_extra: List[Dict] = field(default_factory=list)  # [{id, name}] — human tested but tool didn't
    attribute_differences: List[Dict] = field(default_factory=list)
    # Matched attribute pairs with similarity details
    # [{"our_id", "our_name", "human_id", "human_name", "similarity", "match_type"}]
    attribute_match_details: List[Dict] = field(default_factory=list)
    # Samples
    our_sample_count: int = 0
    human_sample_count: int = 0
    samples_human_missed: List[str] = field(default_factory=list)
    sample_differences: List[Dict] = field(default_factory=list)
    # Classification
    gap_severity: str = "NONE"  # CRITICAL / HIGH / MEDIUM / LOW / NONE
    # LLM narrative (populated when LLM comparison is used)
    overall_narrative: str = ""
    # Full per-attribute (TOD) or per-sample-per-attribute (TOE) comparison detail
    attribute_comparison_detail: List[Dict] = field(default_factory=list)


@dataclass
class ComparisonReport:
    """Full comparison report across all controls."""
    timestamp: str = ""
    total_controls: int = 0
    controls_with_gaps: int = 0
    severity_counts: Dict[str, int] = field(default_factory=lambda: {
        "CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "NONE": 0,
    })
    control_comparisons: List[ControlComparison] = field(default_factory=list)
    # Pipeline outputs
    tod_output_path: str = ""
    toe_output_path: str = ""
    comparison_output_path: str = ""


# ═══════════════════════════════════════════════════════════════════════════════
#  HUMAN WORKPAPER PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class HumanWorkpaperParser:
    """
    Parses a human auditor's Excel workbook into structured data.

    Expected format:
      - Sheet 1 = Summary (original RCM cols + separator + test result cols)
      - Sheet 2+ = Per-control workpapers (header, attributes legend,
        testing table, conclusion banner)
    """

    # Keywords used to detect the summary header row
    SUMMARY_KEYWORDS = {
        "deviation rate", "deficiency type", "passed", "failed",
        "samples tested", "detailed testing",
    }

    # Keywords for detecting per-control testing table headers
    TESTING_TABLE_KEYWORDS = {"sr. no", "sr.no", "serial", "sample no", "sample id"}

    # Attribute column pattern
    ATTR_COL_PATTERN = re.compile(r"^attribute\s+([a-z0-9]+)", re.IGNORECASE)

    def parse(self, workbook_path: str) -> Dict[str, HumanControlResult]:
        """
        Parse human workpaper Excel into structured results.

        Returns: dict[control_id -> HumanControlResult]
        """
        import openpyxl

        logger.info("Parsing human workpaper: %s", workbook_path)
        wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
        sheet_names = wb.sheetnames

        if not sheet_names:
            logger.warning("Workbook has no sheets")
            return {}

        results: Dict[str, HumanControlResult] = {}

        # Parse summary sheet (first sheet)
        summary_ws = wb[sheet_names[0]]
        summary_data = self._parse_summary_sheet(summary_ws)

        # Seed results from summary
        for cid, data in summary_data.items():
            results[cid] = data

        # Parse per-control sheets (remaining sheets)
        for sheet_name in sheet_names[1:]:
            ws = wb[sheet_name]
            control_id = sheet_name.strip()

            control_data = self._parse_control_sheet(ws, control_id)
            if control_data is None:
                logger.warning("Could not parse control sheet: %s", sheet_name)
                continue

            # Merge with summary data if exists
            if control_id in results:
                existing = results[control_id]
                existing.worksteps = control_data.worksteps
                existing.attributes_tested = control_data.attributes_tested
                existing.sample_results = control_data.sample_results
                # Override effectiveness from conclusion banner if available
                if control_data.operating_effectiveness:
                    existing.operating_effectiveness = control_data.operating_effectiveness
                if control_data.deficiency_type:
                    existing.deficiency_type = control_data.deficiency_type
            else:
                results[control_id] = control_data

        wb.close()
        logger.info("Parsed %d controls from human workpaper", len(results))
        return results

    def _parse_summary_sheet(self, ws) -> Dict[str, HumanControlResult]:
        """Parse the summary/first sheet to extract per-control result summaries."""
        results = {}

        # Find the header row by scanning for keywords
        header_row_idx = None
        header_map = {}  # col_idx -> column_name

        for row_idx in range(1, min(ws.max_row + 1, 15)):  # Scan first 15 rows
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value or "").strip().lower()
                row_values.append(val)

            # Check if this row contains enough summary keywords
            row_text = " ".join(row_values)
            matches = sum(1 for kw in self.SUMMARY_KEYWORDS if kw in row_text)
            if matches >= 2:
                header_row_idx = row_idx
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value or "").strip()
                    if val:
                        header_map[col_idx] = val
                break

        if header_row_idx is None:
            logger.warning("Could not find summary header row")
            return results

        # Identify key columns by keyword matching
        col_indices = self._identify_summary_columns(header_map)

        # Parse data rows
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            control_id = self._get_cell_value(ws, row_idx, col_indices.get("control_id"))
            if not control_id:
                continue

            result = HumanControlResult(control_id=control_id)

            # Parse samples tested
            samples_str = self._get_cell_value(ws, row_idx, col_indices.get("samples_tested"))
            result.total_samples = self._parse_int(samples_str)

            # Parse passed/total
            passed_str = self._get_cell_value(ws, row_idx, col_indices.get("passed"))
            if passed_str and "/" in str(passed_str):
                parts = str(passed_str).split("/")
                result.passed_samples = self._parse_int(parts[0])
            else:
                result.passed_samples = self._parse_int(passed_str)

            # Parse failed/total
            failed_str = self._get_cell_value(ws, row_idx, col_indices.get("failed"))
            if failed_str and "/" in str(failed_str):
                parts = str(failed_str).split("/")
                result.failed_samples = self._parse_int(parts[0])
            else:
                result.failed_samples = self._parse_int(failed_str)

            # Parse deviation rate
            dev_str = self._get_cell_value(ws, row_idx, col_indices.get("deviation_rate"))
            result.deviation_rate = self._parse_percentage(dev_str)

            # Parse deficiency type
            result.deficiency_type = self._get_cell_value(
                ws, row_idx, col_indices.get("deficiency_type")) or ""

            # Parse operating effectiveness from summary
            eff_str = self._get_cell_value(
                ws, row_idx, col_indices.get("effectiveness")) or ""
            if eff_str:
                eff_lower = eff_str.lower().strip()
                if "not effective" in eff_lower:
                    result.operating_effectiveness = "Not Effective"
                elif "effective with exception" in eff_lower:
                    result.operating_effectiveness = "Effective with Exceptions"
                elif "effective" in eff_lower:
                    result.operating_effectiveness = "Effective"
                else:
                    result.operating_effectiveness = eff_str.strip()

            # Parse overall remarks
            result.overall_remarks = self._get_cell_value(
                ws, row_idx, col_indices.get("overall_remarks")) or ""

            results[control_id] = result

        return results

    def _identify_summary_columns(self, header_map: Dict[int, str]) -> Dict[str, int]:
        """Map semantic column names to their column indices."""
        col_indices = {}
        for col_idx, header in header_map.items():
            h = header.lower().strip()

            if any(kw in h for kw in ("control id", "control no", "control_id")):
                col_indices["control_id"] = col_idx
            elif "samples tested" in h:
                col_indices["samples_tested"] = col_idx
            elif "passed" in h and "fail" not in h:
                col_indices["passed"] = col_idx
            elif "failed" in h or ("fail" in h and "pass" not in h):
                col_indices["failed"] = col_idx
            elif "deviation" in h and "rate" in h:
                col_indices["deviation_rate"] = col_idx
            elif "deficiency" in h and "type" in h:
                col_indices["deficiency_type"] = col_idx
            elif any(kw in h for kw in ("operating effectiveness", "effectiveness", "conclusion")):
                if "effectiveness" not in col_indices:
                    col_indices["effectiveness"] = col_idx
            elif "overall" in h and "remark" in h:
                col_indices["overall_remarks"] = col_idx
            elif "remark" in h and "fail" not in h:
                if "overall_remarks" not in col_indices:
                    col_indices["overall_remarks"] = col_idx

        return col_indices

    def _parse_control_sheet(self, ws, control_id: str) -> Optional[HumanControlResult]:
        """Parse a per-control workpaper sheet."""
        result = HumanControlResult(control_id=control_id)

        max_row = ws.max_row
        if max_row is None or max_row < 3:
            return None

        # ── Parse attributes legend ────────────────────────────────────
        attr_legend_start = None
        for row_idx in range(1, min(max_row + 1, 30)):
            for col_idx in range(1, min(ws.max_column + 1, 12)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if "attribute" in val and ("to be tested" in val or "tested" in val):
                    attr_legend_start = row_idx + 1
                    break
                # Also detect "Attribute" header in first column
                if val == "attribute" and col_idx <= 2:
                    # Check if next column says something like "Attribute to be tested"
                    next_val = str(ws.cell(row=row_idx, column=col_idx + 1).value or "").strip().lower()
                    if "attribute" in next_val or "tested" in next_val or "description" in next_val:
                        attr_legend_start = row_idx + 1
                        break
            if attr_legend_start is not None:
                break

        if attr_legend_start:
            for row_idx in range(attr_legend_start, min(max_row + 1, attr_legend_start + 20)):
                attr_id_val = str(ws.cell(row=row_idx, column=1).value or "").strip()
                if not attr_id_val:
                    # Try column 2
                    attr_id_val = str(ws.cell(row=row_idx, column=2).value or "").strip()
                if not attr_id_val:
                    break  # End of attributes

                # Check if this looks like an attribute ID (single letter or short code)
                if len(attr_id_val) > 5 or attr_id_val.lower() in ("total", "testing", "conclusion"):
                    break

                # Get description from merged cells to the right
                desc_parts = []
                for col_idx in range(3, min(ws.max_column + 1, 12)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        desc_parts.append(str(val).strip())
                description = " ".join(desc_parts).strip()

                # Split "Name: Description" if present
                name = attr_id_val
                if ":" in description:
                    name_part, desc_part = description.split(":", 1)
                    name = name_part.strip()
                    description = desc_part.strip()

                result.attributes_tested.append({
                    "id": attr_id_val,
                    "name": name,
                    "description": description,
                })

        # ── Parse testing table ────────────────────────────────────────
        table_header_row = None
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 6)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if val in self.TESTING_TABLE_KEYWORDS or any(kw in val for kw in self.TESTING_TABLE_KEYWORDS):
                    table_header_row = row_idx
                    break
            if table_header_row is not None:
                break

        if table_header_row is None:
            # Try to find by "Testing Sample Details" header
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, min(ws.max_column + 1, 20)):
                    val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                    if "testing sample" in val or "sample details" in val:
                        table_header_row = row_idx
                        break
                if table_header_row is not None:
                    break

        if table_header_row is None:
            return result  # No testing table found, return what we have

        # Detect column layout from header row (may be 2-row header)
        # Check if the row below also has headers (sub-headers)
        sub_header_row = table_header_row + 1
        columns = self._detect_testing_table_columns(ws, table_header_row, sub_header_row)

        # Data starts after header(s)
        data_start = sub_header_row + 1 if columns.get("has_sub_header") else table_header_row + 1

        # Parse sample rows until TOTAL row
        for row_idx in range(data_start, max_row + 1):
            # Check for TOTAL row
            is_total = False
            for col_idx in range(1, min(ws.max_column + 1, 6)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if val == "total":
                    is_total = True
                    break
            if is_total:
                break

            # Check for Conclusion row
            is_conclusion = False
            for col_idx in range(1, min(ws.max_column + 1, 4)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if "conclusion" in val:
                    is_conclusion = True
                    break
            if is_conclusion:
                break

            # Parse sample row
            sample = self._parse_sample_row(ws, row_idx, columns, control_id)
            if sample and sample.sample_id:
                result.sample_results.append(sample)

        # ── Remap attribute_results keys ──────────────────────────────
        # The legend may use IDs "1","2","3" while the testing table
        # columns use "A","B","C" (from "Attribute A" headers).
        # Build a positional mapping: testing-table-col-ID → legend-ID
        # so attribute_results can be looked up by legend ID.
        table_attr_ids = [aid for _, aid in columns.get("attribute_cols", [])]
        legend_attr_ids = [a["id"] for a in result.attributes_tested]
        if (table_attr_ids and legend_attr_ids
                and set(table_attr_ids) != set(legend_attr_ids)
                and len(table_attr_ids) == len(legend_attr_ids)):
            col_to_legend = dict(zip(table_attr_ids, legend_attr_ids))
            for sample in result.sample_results:
                remapped = {}
                for col_id, val in sample.attribute_results.items():
                    legend_id = col_to_legend.get(col_id, col_id)
                    remapped[legend_id] = val
                sample.attribute_results = remapped

        # ── Parse conclusion banner ────────────────────────────────────
        for row_idx in range(max(1, max_row - 15), max_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 4)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if "conclusion" in val:
                    # The conclusion text is in the adjacent merged cells
                    conclusion_text = ""
                    for c in range(col_idx + 1, min(ws.max_column + 1, col_idx + 8)):
                        cell_val = ws.cell(row=row_idx, column=c).value
                        if cell_val:
                            conclusion_text += str(cell_val).strip() + " "
                    conclusion_text = conclusion_text.strip()
                    result.operating_effectiveness, result.deficiency_type = (
                        self._parse_conclusion(conclusion_text)
                    )
                    break

        # Derive sample counts
        result.total_samples = len(result.sample_results)
        result.passed_samples = sum(1 for s in result.sample_results if s.result == "PASS")
        result.failed_samples = sum(1 for s in result.sample_results if s.result == "FAIL")
        if result.total_samples > 0:
            result.deviation_rate = result.failed_samples / result.total_samples

        return result

    def _detect_testing_table_columns(
        self, ws, header_row: int, sub_header_row: int,
    ) -> Dict[str, Any]:
        """Detect column layout of the testing table."""
        columns: Dict[str, Any] = {
            "sr_no": None,
            "sample_id": None,
            "detail_cols": [],        # [(col_idx, header_name), ...]
            "attribute_cols": [],     # [(col_idx, attr_id), ...]
            "remarks": None,
            "analysis": None,
            "has_sub_header": False,
        }

        max_col = ws.max_column or 1

        # Read both header rows
        row1_vals = {}
        row2_vals = {}
        for col_idx in range(1, max_col + 1):
            v1 = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
            v2 = str(ws.cell(row=sub_header_row, column=col_idx).value or "").strip()
            if v1:
                row1_vals[col_idx] = v1
            if v2:
                row2_vals[col_idx] = v2

        # Check if row2 has sub-headers (e.g., "Sample No.", detail column names)
        sub_header_keywords = {"sample no", "sample id", "pr number", "invoice", "date", "amount"}
        has_sub = any(
            any(kw in v.lower() for kw in sub_header_keywords)
            for v in row2_vals.values()
        )
        columns["has_sub_header"] = has_sub

        # Use the appropriate header row
        effective_headers = row2_vals if has_sub else row1_vals
        all_headers = {**row1_vals, **row2_vals}  # Merged for keyword detection

        # Identify columns
        attr_start_col = None
        for col_idx in sorted(effective_headers.keys()):
            h = effective_headers[col_idx].lower().strip()
            h_r1 = row1_vals.get(col_idx, "").lower().strip()

            if h in ("sr. no.", "sr.no", "sr. no", "s.no", "serial", "sno", "#"):
                columns["sr_no"] = col_idx
            elif any(kw in h for kw in ("sample no", "sample id", "sample_id")):
                columns["sample_id"] = col_idx
            elif "remark" in h and "attribute" not in h:
                columns["remarks"] = col_idx
            elif "attribute analysis" in h or "analysis" in h:
                columns["analysis"] = col_idx
            elif "workpaper" in h and "ref" in h:
                pass  # Skip workpaper reference column
            elif self.ATTR_COL_PATTERN.match(effective_headers.get(col_idx, "")) or \
                 self.ATTR_COL_PATTERN.match(row1_vals.get(col_idx, "")):
                # This is an attribute column
                match = self.ATTR_COL_PATTERN.match(
                    effective_headers.get(col_idx, "") or row1_vals.get(col_idx, "")
                )
                if match:
                    attr_id = match.group(1).upper()
                    columns["attribute_cols"].append((col_idx, attr_id))
                    if attr_start_col is None:
                        attr_start_col = col_idx

        # Columns between sample_id and first attribute column are detail columns
        if columns["sample_id"] and attr_start_col:
            for col_idx in range(columns["sample_id"] + 1, attr_start_col):
                header = effective_headers.get(col_idx, row1_vals.get(col_idx, ""))
                if header:
                    columns["detail_cols"].append((col_idx, header))

        return columns

    def _parse_sample_row(
        self, ws, row_idx: int, columns: Dict, control_id: str,
    ) -> Optional[HumanSampleResult]:
        """Parse one sample data row from the testing table."""
        # Get sample ID
        sample_id_col = columns.get("sample_id")
        if not sample_id_col:
            return None

        sample_id = str(ws.cell(row=row_idx, column=sample_id_col).value or "").strip()
        if not sample_id or sample_id.lower() in ("", "nan", "none", "null"):
            return None

        # Get sample details
        sample_details = {}
        for col_idx, header_name in columns.get("detail_cols", []):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if val and val.lower() not in ("nan", "none"):
                sample_details[header_name] = val

        # Get attribute results
        attribute_results = {}
        for col_idx, attr_id in columns.get("attribute_cols", []):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            attribute_results[attr_id] = self._normalize_yes_no(val)

        # Get remarks
        remarks = ""
        if columns.get("remarks"):
            remarks = str(ws.cell(row=row_idx, column=columns["remarks"]).value or "").strip()

        # Derive result: any "No" attribute = FAIL
        has_no = any(v.lower() == "no" for v in attribute_results.values())
        result = "FAIL" if has_no else "PASS"

        return HumanSampleResult(
            control_id=control_id,
            sample_id=sample_id,
            sample_details=sample_details,
            attribute_results=attribute_results,
            remarks=remarks,
            result=result,
        )

    def _parse_conclusion(self, text: str) -> Tuple[str, str]:
        """Parse conclusion banner text into (effectiveness, deficiency_type)."""
        text_lower = text.lower()

        effectiveness = ""
        deficiency = ""

        if "not effective" in text_lower:
            effectiveness = "Not Effective"
        elif "effective with exception" in text_lower:
            effectiveness = "Effective with Exceptions"
        elif "effective" in text_lower:
            effectiveness = "Effective"

        if "material weakness" in text_lower:
            deficiency = "Material Weakness"
        elif "significant deficiency" in text_lower:
            deficiency = "Significant Deficiency"
        elif "control deficiency" in text_lower:
            deficiency = "Control Deficiency"
        elif re.search(r'\bnone\b', text_lower) and (
            "deficiency" in text_lower or text_lower.strip() == "none"
            or re.search(r'—\s*none|–\s*none|-\s*none|:\s*none', text_lower)
        ):
            deficiency = "None"

        # Try splitting by " — " or " - " for structured conclusions
        if not effectiveness:
            for sep in [" — ", " - ", " – "]:
                if sep in text:
                    parts = text.split(sep)
                    effectiveness = parts[0].strip()
                    if len(parts) > 1:
                        deficiency = parts[1].strip()
                    break

        return effectiveness, deficiency

    @staticmethod
    def _get_cell_value(ws, row: int, col: Optional[int]) -> Optional[str]:
        """Safely get a cell value as string."""
        if col is None:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        return str(val).strip()

    @staticmethod
    def _parse_int(val) -> int:
        """Parse a value to int, defaulting to 0."""
        if val is None:
            return 0
        try:
            return int(float(str(val).replace(",", "").strip()))
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def _parse_percentage(val) -> float:
        """Parse a percentage string like '15.0%' to float 0.15."""
        if val is None:
            return 0.0
        raw = str(val).strip()
        has_pct = "%" in raw
        s = raw.replace("%", "").strip()
        try:
            num = float(s)
        except (ValueError, TypeError):
            return 0.0
        # If the raw value had a '%' sign, divide by 100 (e.g. "15%" -> 0.15).
        # If it's already a decimal <= 1 with no '%', assume Excel stored it
        # as a fraction (e.g. 0.15 for 15%).  Only divide when > 1 and no '%'.
        if has_pct:
            return num / 100.0
        if num > 1.0:
            return num / 100.0
        return num

    @staticmethod
    def _normalize_yes_no(val: str) -> str:
        """Normalize Yes/No/N/A values."""
        v = val.strip().lower()
        if v in ("yes", "y", "true", "1", "pass"):
            return "Yes"
        elif v in ("no", "n", "false", "0", "fail"):
            return "No"
        elif v in ("n/a", "na", "not applicable", "-"):
            return "N/A"
        return val.strip()


# ═══════════════════════════════════════════════════════════════════════════════
#  RISK LEVEL INFERENCE (mirrors agent/tools/infer_risk_level.py)
# ═══════════════════════════════════════════════════════════════════════════════
#
# Weighted, non-linear scoring — same logic as the main interactive pipeline.
# In QC mode the inferred values are auto-applied (no user approval step).

DEFAULT_SCORE_MAP = {"low": 1, "medium": 3, "high": 6}
DEFAULT_BANDS = [(5, "Low"), (17, "Medium"), (35, "High")]

_RISK_INPUT_ALIASES = {
    "low": "Low", "l": "Low", "minor": "Low", "minimal": "Low",
    "medium": "Medium", "med": "Medium", "moderate": "Medium", "m": "Medium",
    "high": "High", "h": "High", "major": "High", "significant": "High",
    "very high": "High", "critical": "High",
}


def _score_to_level(score: int, bands=None) -> str:
    """Map a numeric risk score to a risk level string using band thresholds."""
    for threshold, label in (bands or DEFAULT_BANDS):
        if score <= threshold:
            return label
    return "Critical"

KEYWORD_RISK_MAP = [
    {
        "keywords": ["financial reporting", "financial statement", "material misstatement",
                      "regulatory compliance", "regulatory requirement", "legal compliance",
                      "fraud", "fraud risk", "revenue recognition", "asset misappropriation"],
        "probability": "High", "impact": "High", "risk_level": "Critical",
    },
    {
        "keywords": ["access control", "privileged access", "system access",
                      "segregation of duties", "sod violation", "unauthorized access",
                      "data breach", "cybersecurity", "information security"],
        "probability": "Medium", "impact": "High", "risk_level": "High",
    },
    {
        "keywords": ["reconciliation", "bank reconciliation", "account reconciliation",
                      "journal entry", "manual journal", "period-end close",
                      "month-end", "quarter-end"],
        "probability": "Medium", "impact": "Medium", "risk_level": "Medium",
    },
    {
        "keywords": ["change management", "change control", "system change",
                      "configuration change", "patch management"],
        "probability": "Medium", "impact": "Medium", "risk_level": "Medium",
    },
    {
        "keywords": ["backup", "disaster recovery", "business continuity",
                      "data retention", "archival"],
        "probability": "Low", "impact": "High", "risk_level": "Medium",
    },
    {
        "keywords": ["training", "awareness", "policy acknowledgment",
                      "documentation", "record keeping"],
        "probability": "Low", "impact": "Low", "risk_level": "Low",
    },
]


def _normalize_risk_input(raw) -> str:
    """Normalize a raw risk input value to Low/Medium/High."""
    if raw is None:
        return ""
    return _RISK_INPUT_ALIASES.get(str(raw).strip().lower(), "")


def _compute_risk_score(probability: str, impact: str,
                        score_map=None, bands=None) -> tuple:
    """Compute risk score and rating from probability and impact strings.

    The risk matrix is never a separate hardcoded artefact — it is the
    arithmetic product of the weights mapped through the band thresholds.

    Returns (score, risk_level).
    """
    p = probability.strip().lower()
    i = impact.strip().lower()
    sm = score_map or DEFAULT_SCORE_MAP
    p_score = sm.get(p, 0)
    i_score = sm.get(i, 0)
    if p_score == 0 or i_score == 0:
        return 0, ""
    score = p_score * i_score
    level = _score_to_level(score, bands)
    return score, level


def _match_risk_keyword(description: str) -> Optional[Dict[str, str]]:
    """Try to match a control/risk description against keyword table."""
    desc_lower = description.lower()
    for entry in KEYWORD_RISK_MAP:
        for keyword in entry["keywords"]:
            if keyword in desc_lower:
                return {
                    "probability": entry["probability"],
                    "impact": entry["impact"],
                    "risk_level": entry["risk_level"],
                    "matched_keyword": keyword,
                }
    return None


def _llm_infer_risk_qc(
    controls: List[Dict[str, str]],
    infer_mode: str = "prob_impact",
) -> Dict[str, Dict[str, str]]:
    """Use LLM to infer missing risk probability/impact or risk level.

    One API call per control for reliability. Same approach as the main
    pipeline but used autonomously in QC.
    """
    if not controls:
        return {}

    try:
        from engines.config import (
            AZURE_OPENAI_API_KEY, AZURE_OPENAI_ENDPOINT,
            AZURE_OPENAI_DEPLOYMENT, AZURE_OPENAI_API_VERSION,
        )
        from openai import AzureOpenAI
    except Exception as e:
        logger.warning("Cannot load LLM config for risk inference: %s", e)
        return {}

    client = AzureOpenAI(
        api_key=AZURE_OPENAI_API_KEY,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_version=AZURE_OPENAI_API_VERSION,
    )

    valid_outputs = ("Low", "Medium", "High", "Critical")

    if infer_mode == "risk_level":
        system_prompt = (
            "You are an audit expert. Given a control description, determine "
            "the Risk Level, Probability, and Impact.\n\n"
            "Risk Levels: Low, Medium, High, Critical (Critical ONLY when both "
            "probability and impact are clearly very high).\n"
            "Probability & Impact: Low, Medium, or High.\n\n"
            "Be conservative — when uncertain, rate higher.\n\n"
            "Return a JSON object (NOT an array):\n"
            '{"probability": "<Low|Medium|High>", "impact": "<Low|Medium|High>", '
            '"risk_level": "<Low|Medium|High|Critical>", '
            '"reasoning": "<one sentence>"}\n\n'
            "Return ONLY valid JSON. No markdown, no code fences."
        )
    else:
        system_prompt = (
            "You are an audit expert. Given a control description, determine "
            "the Risk Probability and Risk Impact.\n\n"
            "Probability: Low = unlikely, Medium = possible, High = likely.\n"
            "Impact: Low = minor, Medium = moderate, High = severe.\n\n"
            "If one value is already provided, use it as context. "
            "Be conservative — when uncertain, rate higher.\n\n"
            "Return a JSON object (NOT an array):\n"
            '{"probability": "<Low|Medium|High>", "impact": "<Low|Medium|High>", '
            '"reasoning": "<one sentence>"}\n\n'
            "Return ONLY valid JSON. No markdown, no code fences."
        )

    def _infer_single(ctrl: dict) -> tuple:
        """Call the LLM for a single control. Returns (control_id, entry_or_None)."""
        cid = ctrl["control_id"]
        user_prompt = f"Control: {ctrl['description']}"
        if ctrl.get("existing_prob"):
            user_prompt += f"\nExisting Probability: {ctrl['existing_prob']}"
        if ctrl.get("existing_impact"):
            user_prompt += f"\nExisting Impact: {ctrl['existing_impact']}"

        try:
            response = client.chat.completions.create(
                model=AZURE_OPENAI_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                max_completion_tokens=256,
                response_format={"type": "json_object"},
            )

            raw = response.choices[0].message.content or "{}"
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                parsed = parsed[0] if parsed else {}
            if not isinstance(parsed, dict):
                raise ValueError(f"Unexpected response type: {type(parsed)}")

            prob = _normalize_risk_input(parsed.get("probability"))
            impact = _normalize_risk_input(parsed.get("impact"))
            risk_level = str(parsed.get("risk_level") or "").strip()

            if risk_level and risk_level not in valid_outputs:
                for vo in valid_outputs:
                    if vo.lower() == risk_level.lower():
                        risk_level = vo
                        break
                else:
                    risk_level = ""

            entry = {"probability": prob or "Medium", "impact": impact or "Medium"}
            if risk_level:
                entry["risk_level"] = risk_level
            entry["reasoning"] = parsed.get("reasoning") or ""
            return cid, entry

        except Exception as e:
            logger.warning("LLM risk inference failed for %s in QC: %s", cid, e)
            return cid, None

    # Run all controls in parallel
    result = {}
    succeeded = 0
    failed = 0
    max_workers = min(5, len(controls))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_infer_single, ctrl): ctrl for ctrl in controls}
        for future in as_completed(futures):
            cid, entry = future.result()
            if entry is not None:
                result[cid] = entry
                succeeded += 1
            else:
                failed += 1

    logger.info("QC LLM risk inference: %d succeeded, %d failed out of %d",
                succeeded, failed, len(controls))
    return result


def _infer_risk_levels_for_qc(rcm_df: pd.DataFrame,
                              score_map=None, bands=None) -> pd.DataFrame:
    """Run the full risk level inference pipeline on the RCM (autonomous, no approval).

    Uses the same 3-stage strategy as the main interactive pipeline:
      1. Compute from P x I (weighted non-linear scoring) when both present
      2. Keyword matching on control/risk description
      3. LLM inference for remaining controls

    If custom *score_map* or *bands* are provided (user changed weights via
    chat), the matrix auto-recomputes — no separate hardcoded matrix is used.

    Writes risk_level, risk_probability, risk_impact back into the DataFrame.
    """
    prob_col = "risk_probability" if "risk_probability" in rcm_df.columns else None
    impact_col = "risk_impact" if "risk_impact" in rcm_df.columns else None
    level_col = "risk_level" if "risk_level" in rcm_df.columns else None

    desc_col = None
    risk_desc_col = None
    cid_col = None
    for col in rcm_df.columns:
        c = col.lower().replace("_", " ").strip()
        if c in ("control description", "controldescription", "control_description"):
            desc_col = col
        elif c in ("risk description", "riskdescription", "risk_description"):
            risk_desc_col = col
        elif c in ("control id", "controlid", "control_id"):
            cid_col = col

    if not cid_col:
        logger.warning("No Control Id column found — skipping risk inference")
        if "risk_level" not in rcm_df.columns:
            rcm_df["risk_level"] = "High"
        return rcm_df

    # Ensure columns exist
    if "risk_level" not in rcm_df.columns:
        rcm_df["risk_level"] = ""
    if "risk_probability" not in rcm_df.columns:
        rcm_df["risk_probability"] = ""
    if "risk_impact" not in rcm_df.columns:
        rcm_df["risk_impact"] = ""

    # Refresh column references after ensuring they exist
    prob_col = "risk_probability"
    impact_col = "risk_impact"
    level_col = "risk_level"

    needs_keyword = []      # All three missing — try keyword, then LLM direct
    needs_llm_pi = []       # Has one of P/I — need the other inferred
    needs_llm_direct = []   # All three missing, keyword didn't match
    seen = set()

    computed_count = 0
    keyword_count = 0

    for idx, row in rcm_df.iterrows():
        control_id = str(row.get(cid_col, "")).strip()
        if not control_id or control_id.lower() in ("nan", "none", "null", ""):
            continue
        if control_id in seen:
            continue
        seen.add(control_id)

        raw_prob = str(row.get(prob_col, "")).strip()
        raw_impact = str(row.get(impact_col, "")).strip()
        raw_level = str(row.get(level_col, "")).strip()

        norm_prob = _normalize_risk_input(raw_prob)
        norm_impact = _normalize_risk_input(raw_impact)

        _missing = ("nan", "none", "null", "n/a", "na", "", "-", "tbd")
        is_prob_missing = not norm_prob or raw_prob.lower() in _missing
        is_impact_missing = not norm_impact or raw_impact.lower() in _missing
        is_level_missing = not raw_level or raw_level.lower() in _missing

        description = str(row.get(desc_col, "")).strip() if desc_col else ""
        risk_desc = str(row.get(risk_desc_col, "")).strip() if risk_desc_col else ""
        combined_desc = f"{description} {risk_desc}".strip()

        # ── Stage 1: Both P and I present → compute directly ──
        if not is_prob_missing and not is_impact_missing:
            score, level = _compute_risk_score(norm_prob, norm_impact, score_map, bands)
            if score > 0:
                # Write to ALL rows with this control_id
                mask = rcm_df[cid_col].astype(str).str.strip() == control_id
                rcm_df.loc[mask, level_col] = level
                rcm_df.loc[mask, prob_col] = norm_prob
                rcm_df.loc[mask, impact_col] = norm_impact
                computed_count += 1
                logger.debug("Computed risk for %s: P=%s, I=%s, Score=%d → %s",
                             control_id, norm_prob, norm_impact, score, level)
                continue

        # Already has a valid risk_level — skip
        if not is_level_missing and is_prob_missing and is_impact_missing:
            continue

        # ── Stage 2: Keyword matching ──
        if combined_desc:
            match = _match_risk_keyword(combined_desc)
            if match:
                prob = norm_prob if not is_prob_missing else match["probability"]
                impact = norm_impact if not is_impact_missing else match["impact"]
                score, level = _compute_risk_score(prob, impact, score_map, bands)
                if score > 0:
                    mask = rcm_df[cid_col].astype(str).str.strip() == control_id
                    rcm_df.loc[mask, level_col] = level
                    rcm_df.loc[mask, prob_col] = prob
                    rcm_df.loc[mask, impact_col] = impact
                    keyword_count += 1
                logger.debug("Keyword-matched risk for %s: %s → %s (kw: %s)",
                             control_id, match["matched_keyword"], level, match["matched_keyword"])
                continue

        # ── Stage 3: Needs LLM ──
        ctrl_info = {
            "index": idx,
            "control_id": control_id,
            "description": combined_desc,
            "existing_prob": norm_prob if not is_prob_missing else "",
            "existing_impact": norm_impact if not is_impact_missing else "",
            "all_missing": is_prob_missing and is_impact_missing and is_level_missing,
        }
        if ctrl_info["all_missing"]:
            needs_llm_direct.append(ctrl_info)
        else:
            needs_llm_pi.append(ctrl_info)

    # ── LLM inference: infer P/I for controls that have one missing ──
    llm_pi_count = 0
    if needs_llm_pi:
        llm_results = _llm_infer_risk_qc(
            [{"control_id": c["control_id"], "description": c["description"],
              "existing_prob": c["existing_prob"], "existing_impact": c["existing_impact"]}
             for c in needs_llm_pi],
            infer_mode="prob_impact",
        )
        for ctrl in needs_llm_pi:
            cid = ctrl["control_id"]
            inf = llm_results.get(cid, {})
            prob = ctrl["existing_prob"] or inf.get("probability", "Medium")
            impact = ctrl["existing_impact"] or inf.get("impact", "Medium")
            score, level = _compute_risk_score(prob, impact, score_map, bands)
            mask = rcm_df[cid_col].astype(str).str.strip() == cid
            rcm_df.loc[mask, level_col] = level
            rcm_df.loc[mask, prob_col] = prob
            rcm_df.loc[mask, impact_col] = impact
            llm_pi_count += 1
            logger.debug("LLM-inferred risk for %s: P=%s, I=%s → %s", cid, prob, impact, level)

    # ── LLM inference: direct risk level for controls with all three missing ──
    llm_direct_count = 0
    if needs_llm_direct:
        llm_results = _llm_infer_risk_qc(
            [{"control_id": c["control_id"], "description": c["description"],
              "existing_prob": "", "existing_impact": ""}
             for c in needs_llm_direct],
            infer_mode="risk_level",
        )
        for ctrl in needs_llm_direct:
            cid = ctrl["control_id"]
            inf = llm_results.get(cid, {})
            prob = inf.get("probability", "Medium")
            impact = inf.get("impact", "Medium")
            risk_level = inf.get("risk_level", "")
            score, computed_level = _compute_risk_score(prob, impact, score_map, bands)
            final_level = risk_level if risk_level else computed_level
            mask = rcm_df[cid_col].astype(str).str.strip() == cid
            rcm_df.loc[mask, level_col] = final_level
            rcm_df.loc[mask, prob_col] = prob
            rcm_df.loc[mask, impact_col] = impact
            llm_direct_count += 1
            logger.debug("LLM-direct risk for %s: P=%s, I=%s → %s", cid, prob, impact, final_level)

    # ── Final fallback: any row still without risk_level → default "High" ──
    empty_mask = (
        rcm_df[level_col].isna()
        | rcm_df[level_col].astype(str).str.strip().isin(["", "nan", "none", "null"])
    )
    fallback_count = empty_mask.sum()
    if fallback_count > 0:
        rcm_df.loc[empty_mask, level_col] = "High"
        logger.info("Defaulted %d remaining controls to risk_level='High'", fallback_count)

    logger.info(
        "QC risk inference complete: %d computed, %d keyword, %d LLM (P/I), "
        "%d LLM (direct), %d fallback-defaulted",
        computed_count, keyword_count, llm_pi_count, llm_direct_count, fallback_count,
    )
    return rcm_df


# ═══════════════════════════════════════════════════════════════════════════════
#  SAMPLING LOGIC (extracted from agent/tools/sampling_engine.py)
# ═══════════════════════════════════════════════════════════════════════════════

KPMG_TABLE = [
    {"frequency": "Annual",                                    "Low": 1,  "High": 1},
    {"frequency": "Quarterly (including period end, i.e. +1)", "Low": 2,  "High": 2},
    {"frequency": "Monthly",                                   "Low": 2,  "High": 3},
    {"frequency": "Weekly",                                    "Low": 5,  "High": 8},
    {"frequency": "Daily",                                     "Low": 15, "High": 25},
    {"frequency": "Recurring (multiple times per day)",        "Low": 25, "High": 40},
]

FREQUENCY_ALIASES = {
    "annual": "Annual", "annually": "Annual", "yearly": "Annual",
    "quarterly": "Quarterly (including period end, i.e. +1)",
    "quarter": "Quarterly (including period end, i.e. +1)",
    "monthly": "Monthly", "every month": "Monthly",
    "weekly": "Weekly", "every week": "Weekly",
    "daily": "Daily", "every day": "Daily",
    "recurring": "Recurring (multiple times per day)",
    "multiple times per day": "Recurring (multiple times per day)",
    "per transaction": "Recurring (multiple times per day)",
    "each transaction": "Recurring (multiple times per day)",
    "transactional": "Recurring (multiple times per day)",
    "continuous": "Recurring (multiple times per day)",
    "as needed": "Recurring (multiple times per day)",
    "ad hoc": "Recurring (multiple times per day)",
    "on demand": "Recurring (multiple times per day)",
    "event driven": "Recurring (multiple times per day)",
}

RISK_ALIASES = {
    "high": "High", "h": "High", "higher": "High", "critical": "High",
    "low": "Low", "l": "Low", "lower": "Low", "minimal": "Low",
    "medium": "High", "med": "High", "moderate": "High", "m": "High",
}


def _infer_frequency_from_description(description: str) -> Optional[str]:
    """Infer KPMG frequency from control description keywords (same as agent tool)."""
    desc_lower = description.lower()
    _KW_MAP = [
        (["policy approval", "policy review", "annual review", "annual assessment",
          "annual certification", "annual audit", "board approval", "charter review"], "Annual"),
        (["user access review", "user recertification", "access recertification",
          "quarterly review", "quarterly assessment", "quarterly reconciliation",
          "quarter-end", "quarter end", "period-end review", "period end review"],
         "Quarterly (including period end, i.e. +1)"),
        (["monthly review", "monthly reconciliation", "month-end", "month end",
          "monthly report", "monthly assessment", "monthly closing", "bank reconciliation"], "Monthly"),
        (["weekly review", "weekly report", "weekly check", "weekly reconciliation"], "Weekly"),
        (["daily backup", "nightly job", "daily reconciliation", "daily review",
          "daily check", "daily report", "daily monitoring", "end of day",
          "daily batch", "daily log review"], "Daily"),
        (["change approval", "cab approval", "change request", "change management",
          "incident response", "emergency change", "password expiry", "account lockout",
          "password policy", "go-live approval", "uat sign-off", "deployment approval",
          "three-way match", "3-way match", "invoice matching", "payment approval",
          "journal entry", "journal approval", "segregation of duties", "sod",
          "access provisioning", "access request", "per transaction",
          "real-time", "continuous monitoring", "automated alert"],
         "Recurring (multiple times per day)"),
    ]
    for keywords, kpmg_freq in _KW_MAP:
        for kw in keywords:
            if kw in desc_lower:
                return kpmg_freq
    return None


def _apply_sampling(rcm_df: pd.DataFrame, custom_table=None) -> pd.DataFrame:
    """Apply sampling to populate count_of_samples column.

    If *custom_table* is provided (list of dicts with frequency/Low/High keys),
    it is used instead of the default KPMG_TABLE.
    """
    table = custom_table or KPMG_TABLE
    freq_col = None
    risk_col = None
    control_type_col = None
    desc_col = None

    for col in rcm_df.columns:
        c = col.lower().replace("_", " ").strip()
        if c in ("control frequency", "controlfrequency", "frequency", "control_frequency"):
            freq_col = col
        elif c in ("risk level", "risk_level", "risklevel"):
            risk_col = col
        elif c in ("control type", "controltype", "control_type"):
            control_type_col = col
        elif c in ("control description", "controldescription", "control_description"):
            desc_col = col

    sample_counts = []
    for _, row in rcm_df.iterrows():
        raw_freq = str(row.get(freq_col, "")) if freq_col else ""
        raw_risk = str(row.get(risk_col, "High")) if risk_col else "High"

        # Normalize frequency
        cleaned_freq = raw_freq.strip().lower()
        norm_freq = FREQUENCY_ALIASES.get(cleaned_freq)
        if not norm_freq:
            # Try startswith match
            for alias, canonical in sorted(FREQUENCY_ALIASES.items(), key=lambda x: -len(x[0])):
                if cleaned_freq.startswith(alias):
                    norm_freq = canonical
                    break
        if not norm_freq and cleaned_freq in ("", "nan", "none", "null", "n/a", "na", "-",
                                               "tbd", "to be determined"):
            # Frequency missing — try to infer from control description
            if desc_col:
                description = str(row.get(desc_col, "")).strip()
                norm_freq = _infer_frequency_from_description(description)
                if norm_freq:
                    logger.info("Inferred frequency '%s' from description for %s",
                                norm_freq, row.get("Control Id", "?"))
        if not norm_freq:
            norm_freq = "Recurring (multiple times per day)"  # Conservative default

        # Normalize risk
        cleaned_risk = raw_risk.strip().lower()
        norm_risk = RISK_ALIASES.get(cleaned_risk, "High")

        # Lookup
        count = 1
        for entry in table:
            if entry["frequency"] == norm_freq:
                count = entry.get(norm_risk, 1)
                break

        # Automated controls always get 1
        if control_type_col:
            raw_type = str(row.get(control_type_col, "")).strip().lower()
            if raw_type in ("automated", "automatic", "auto", "it", "it automated"):
                count = 1

        sample_counts.append(count)

    rcm_df["count_of_samples"] = sample_counts
    logger.info("Sampling applied: %d controls, %d total samples",
                len(rcm_df), sum(sample_counts))
    return rcm_df


# ═══════════════════════════════════════════════════════════════════════════════
#  COMPARISON ALGORITHM
# ═══════════════════════════════════════════════════════════════════════════════

EFFECTIVENESS_RANK = {
    "effective": 0,
    "effective with exceptions": 1,
    "not effective": 2,
}


def _normalize_sample_id(sid: str) -> str:
    """Normalize sample IDs for matching: strip prefixes, lowercase."""
    s = sid.strip().lower()
    # Strip common prefixes
    for prefix in ("sample-", "sample_", "sample ", "s-", "s_", "s"):
        if s.startswith(prefix) and len(s) > len(prefix):
            rest = s[len(prefix):]
            if rest[0:1].isdigit():
                s = rest
                break
    return s


_LLM_COMPARE_SYSTEM = """You are a senior audit quality reviewer with deep expertise in PCAOB standards, COSO framework, and internal control testing. You are comparing the results of an automated control testing tool against a human auditor's workpaper for the SAME control.

Your job is to produce a rigorous, structured gap analysis identifying what the human auditor missed or got wrong, compared to the automated tool's findings.

IMPORTANT RULES:
- This is SINGLE-DIRECTIONAL: only flag gaps in the HUMAN's work. If the human found something the tool missed, do NOT report it.
- Match attributes by MEANING and INTENT, not by ID or position. Attribute "1" from the tool and attribute "A" from the human may test the same thing. Look at the name AND description to determine semantic equivalence.
- For verdict comparison: "Not Effective" is the strictest, "Effective with Exceptions" is middle, "Effective" is the most lenient.
- A verdict gap exists only when our tool is STRICTER than the human (e.g. we say Not Effective, human says Effective).
- When assessing attribute gaps, consider whether the human tested the underlying concept under a different name or merged it into another attribute before flagging as "Missing".
- For sample-level differences, examine the remarks and deviation details carefully — a human may have noted an issue in narrative form rather than marking FAIL explicitly.
- When the human has fewer samples than our tool, check if the human's sample size meets the PCAOB/KPMG sampling guidance for the population — if not, flag it.
- Be precise: do not flag cosmetic differences (e.g. wording variations) as gaps. Only flag substantive audit quality gaps.

SEVERITY GUIDELINES:
- CRITICAL: Our tool says "Not Effective" but human says "Effective" — human missed material control failure.
- HIGH: Verdict mismatch (e.g. Exceptions vs Effective) OR 3+ substantive attribute/sample gaps.
- MEDIUM: Human missed 1-2 attributes entirely OR has 1-2 sample-level result differences.
- LOW: Minor gaps — sample count differences, cosmetic attribute coverage gaps with same verdict.
- NONE: No substantive gaps found.

Return a JSON object with this exact structure:
{
  "attribute_matches": [
    {
      "our_id": "<our attr id>",
      "our_name": "<our attr name + description>",
      "human_id": "<matched human attr id, or null if human didn't test this>",
      "human_name": "<human attr name + description, or null>",
      "confidence": "high" | "medium" | "low",
      "reasoning": "<one sentence: why they match or why no match>"
    }
  ],
  "verdict_assessment": {
    "verdict_gap": true | false,
    "detail": "<thorough explanation of the verdict comparison, referencing specific evidence>",
    "severity_recommendation": "CRITICAL" | "HIGH" | "MEDIUM" | "LOW" | "NONE",
    "severity_reasoning": "<concrete reasoning citing specific gaps found>"
  },
  "sample_gaps": [
    {
      "sample_id": "<sample id>",
      "gap_type": "Different" | "Missing",
      "detail": "<what the human missed on this sample, with specific evidence>"
    }
  ],
  "attribute_gaps": [
    {
      "attr_id": "<our attr id>",
      "attr_name": "<name>",
      "gap_type": "Missing" | "Different" | "Human Skipped",
      "sample_id": "<if Different, which sample>",
      "detail": "<explanation with specific evidence>"
    }
  ],
  "overall_narrative": "<3-5 sentence summary of the most important findings for this control, including the audit risk implications of any gaps found>"
}
"""


def _llm_compare_one_control(
    cid: str,
    our_result,
    human: Optional[HumanControlResult],
    schema,
    phase: str,
    evaluator,
) -> Optional[Dict]:
    """Use the LLM to produce a full comparison for one control.

    Returns the parsed JSON dict, or None if the LLM call fails.
    """
    # Build our side
    our_effectiveness = getattr(our_result, "operating_effectiveness", "")
    our_samples_raw = getattr(our_result, "sample_results", [])

    our_attrs = []
    if schema:
        for a in getattr(schema, "attributes", []):
            our_attrs.append({
                "id": a.get("id", ""),
                "name": a.get("name", ""),
                "description": a.get("description", ""),
            })

    our_samples = []
    for s in our_samples_raw:
        sr = {
            "sample_id": getattr(s, "sample_id", ""),
            "result": getattr(s, "result", ""),
            "remarks": getattr(s, "remarks", ""),
            "deviation_details": getattr(s, "deviation_details", ""),
        }
        attr_res = getattr(s, "attribute_results", None)
        if attr_res:
            sr["attribute_results"] = dict(attr_res)
        our_samples.append(sr)

    # Build human side
    if human:
        human_effectiveness = _normalize_verdict(human.operating_effectiveness or "")
        human_attrs = [
            {"id": a["id"], "name": a.get("name", ""), "description": a.get("description", "")}
            for a in human.attributes_tested
        ]
        human_samples = []
        for s in human.sample_results:
            hs = {
                "sample_id": s.sample_id,
                "result": s.result,
                "attribute_results": dict(s.attribute_results),
                "remarks": s.remarks or "",
            }
            human_samples.append(hs)
    else:
        human_effectiveness = "NOT FOUND"
        human_attrs = []
        human_samples = []

    # Build control context from schema if available
    control_context = ""
    if schema:
        ctrl_desc = getattr(schema, "control_description", "") or ""
        worksteps = getattr(schema, "worksteps", []) or []
        if ctrl_desc or worksteps:
            control_context = f"═══ CONTROL CONTEXT ═══\n"
            if ctrl_desc:
                control_context += f"Description: {ctrl_desc}\n"
            if worksteps:
                control_context += f"Worksteps: {json.dumps(worksteps)}\n"
            control_context += "\n"

    user_prompt = (
        f"CONTROL: {cid}  |  PHASE: {phase}\n\n"
        f"{control_context}"
        f"═══ OUR TOOL'S RESULTS ═══\n"
        f"Verdict: {our_effectiveness}\n"
        f"Attributes tested ({len(our_attrs)}):\n{json.dumps(our_attrs, indent=2)}\n"
        f"Samples ({len(our_samples)}):\n{json.dumps(our_samples, indent=2)}\n\n"
        f"═══ HUMAN AUDITOR'S RESULTS ═══\n"
        f"Verdict: {human_effectiveness}\n"
        f"Attributes tested ({len(human_attrs)}):\n{json.dumps(human_attrs, indent=2)}\n"
        f"Samples ({len(human_samples)}):\n{json.dumps(human_samples, indent=2)}\n\n"
        f"Produce the comparison JSON."
    )

    try:
        return evaluator._chat_json(
            messages=[
                {"role": "system", "content": _LLM_COMPARE_SYSTEM},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=8000,
            timeout=180,
        )
    except Exception as e:
        logger.warning("LLM comparison failed for %s: %s", cid, e)
        return None


def _build_comparison_from_llm(
    cid: str,
    llm_result: Dict,
    our_result,
    human: Optional[HumanControlResult],
    schema,
    phase: str,
) -> ControlComparison:
    """Convert the LLM's JSON response into a ControlComparison dataclass."""
    comp = ControlComparison(control_id=cid, phase=phase)

    # Our side
    comp.our_effectiveness = getattr(our_result, "operating_effectiveness", "")
    comp.our_sample_count = getattr(our_result, "total_samples", 0)
    if schema:
        comp.our_attributes = [a["id"] for a in getattr(schema, "attributes", [])]

    # Human side
    if human:
        comp.human_effectiveness = _normalize_verdict(human.operating_effectiveness or "")
        comp.human_sample_count = human.total_samples
        comp.human_attributes = [a["id"] for a in human.attributes_tested]
    else:
        comp.human_effectiveness = "NOT FOUND"

    # Verdict — compute deterministically from actual data, NEVER trust LLM
    verdict = llm_result.get("verdict_assessment", {})
    verdicts_match = _verdicts_same(comp.our_effectiveness, comp.human_effectiveness)

    if verdicts_match:
        comp.verdict_gap = False
        comp.verdict_detail = f"Both agree: '{_normalize_verdict(comp.our_effectiveness)}'."
    else:
        # Check directionality — gap only when our tool is stricter
        our_rank = EFFECTIVENESS_RANK.get(
            _normalize_verdict(comp.our_effectiveness).lower(), -1)
        human_rank = EFFECTIVENESS_RANK.get(
            _normalize_verdict(comp.human_effectiveness).lower(), -1)
        if our_rank > human_rank and our_rank >= 0 and human_rank >= 0:
            comp.verdict_gap = True
            comp.verdict_detail = (
                f"Our tool assessed '{comp.our_effectiveness}' but human assessed "
                f"'{comp.human_effectiveness}' — human may have missed issues"
            )
        else:
            comp.verdict_gap = False
            comp.verdict_detail = (
                f"Ours: '{comp.our_effectiveness}', Human: '{comp.human_effectiveness}'. "
                f"Verdicts differ but human was not more lenient."
            )

    # Attribute matches from LLM
    confidence_to_sim = {"high": 0.90, "medium": 0.65, "low": 0.35}
    llm_matched_our_ids = set()  # track which of our attrs the LLM matched
    for am in llm_result.get("attribute_matches", []):
        human_id = am.get("human_id")
        our_id = am.get("our_id", "")
        conf = str(am.get("confidence", "low")).lower()
        comp.attribute_match_details.append({
            "our_id": our_id,
            "our_name": am.get("our_name", ""),
            "human_id": human_id or "",
            "human_name": am.get("human_name", ""),
            "similarity": confidence_to_sim.get(conf, 0.5),
            "match_type": f"LLM ({conf})",
            "reasoning": am.get("reasoning", ""),
        })
        if human_id is None:
            comp.attributes_human_missed.append(our_id)
        if our_id:
            llm_matched_our_ids.add(our_id)

    # Cross-check: any of our attributes NOT covered by LLM matching = missed
    # This catches the case where LLM simply didn't return an attribute at all
    if schema:
        all_our_attr_ids = {str(a["id"]) for a in getattr(schema, "attributes", [])}
        human_attr_ids = {str(a["id"]) for a in (human.attributes_tested if human else [])}
        matched_human_ids = {
            d["human_id"] for d in comp.attribute_match_details if d.get("human_id")
        }
        # Our attrs the LLM forgot to mention
        for our_id in all_our_attr_ids - llm_matched_our_ids:
            if our_id not in comp.attributes_human_missed:
                comp.attributes_human_missed.append(our_id)

    # Also: if we have more attributes than human AND some of ours weren't
    # matched to any human attr, they should be in attributes_human_missed
    # (deduplicate)
    comp.attributes_human_missed = list(dict.fromkeys(comp.attributes_human_missed))

    # Attribute-level gaps from LLM
    for ag in llm_result.get("attribute_gaps", []):
        gap_type = ag.get("gap_type", "Different")
        if gap_type in ("Different", "Missing", "Human Skipped"):
            comp.attribute_differences.append({
                "attr_id": ag.get("attr_id", ""),
                "attr_name": ag.get("attr_name", ""),
                "sample_id": ag.get("sample_id", ""),
                "our_result": ag.get("our_result", "No (FAIL)" if gap_type == "Different" else "Tested"),
                "human_result": ag.get("human_result", "Yes (PASS)" if gap_type == "Different" else "Not tested"),
                "gap_type": gap_type,
                "detail": ag.get("detail", ""),
            })

    # Sample-level gaps from LLM
    for sg in llm_result.get("sample_gaps", []):
        if sg.get("gap_type") == "Missing":
            comp.samples_human_missed.append(sg.get("sample_id", ""))
        elif sg.get("gap_type") == "Different":
            comp.sample_differences.append({
                "sample_id": sg.get("sample_id", ""),
                "our_result": "FAIL",
                "human_result": "PASS",
                "our_deviation": sg.get("detail", ""),
            })

    # LLM narrative — prefix with factual summary so it's grounded
    llm_narrative = llm_result.get("overall_narrative", "")
    factual_prefix = (
        f"Verdict: Both={_normalize_verdict(comp.our_effectiveness)}" if verdicts_match
        else f"Verdict difference: Ours={comp.our_effectiveness}, Human={comp.human_effectiveness}"
    )
    if comp.attributes_human_missed:
        factual_prefix += f" | {len(comp.attributes_human_missed)} attribute(s) missed by human"
    if comp.attribute_differences:
        factual_prefix += f" | {len(comp.attribute_differences)} attribute result difference(s)"
    comp.overall_narrative = f"{factual_prefix}. {llm_narrative}" if llm_narrative else factual_prefix

    # ── Build per-sample per-attribute comparison detail ───────────
    # Reconstruct attribute mapping from the LLM match details
    attr_mapping: Dict[str, Optional[str]] = {}
    for md in comp.attribute_match_details:
        our_id = md.get("our_id", "")
        human_id = md.get("human_id") or None
        if our_id:
            attr_mapping[our_id] = human_id
    # Include attributes the LLM didn't mention
    if schema:
        for a in getattr(schema, "attributes", []):
            if a["id"] not in attr_mapping:
                attr_mapping[a["id"]] = None

    our_samples = getattr(our_result, "sample_results", [])
    human_samples = human.sample_results if human else []
    comp.attribute_comparison_detail = _build_attribute_comparison_detail(
        attr_mapping, comp.attribute_match_details,
        our_samples, human_samples, schema,
    )

    # Severity — ALWAYS compute deterministically from actual data, never trust LLM
    comp.gap_severity = _classify_severity(comp)

    return comp


def _llm_match_attributes(
    our_attrs: List[Dict],
    human_attrs: List[Dict],
    control_id: str,
    evaluator,
) -> Tuple[Dict[str, Optional[str]], List[Dict]]:
    """
    Use the LLM to semantically match our attributes to the human's attributes.
    Returns the same (mapping, match_details) as _fuzzy_match_attributes.
    """
    # Build a compact representation for the prompt
    our_list = []
    for a in our_attrs:
        our_list.append({
            "id": a["id"],
            "name": a.get("name", ""),
            "description": a.get("description", ""),
        })
    human_list = []
    for a in human_attrs:
        human_list.append({
            "id": a["id"],
            "name": a.get("name", ""),
            "description": a.get("description", ""),
        })

    system_prompt = (
        "You are an audit attribute matching expert. Your job is to match "
        "testing attributes from two different audit workpapers for the same control. "
        "One set is from an automated tool (labelled 'ours'), the other is from a "
        "human auditor (labelled 'human'). Match attributes that test the SAME aspect "
        "of the control, even if they use different wording.\n\n"
        "Return a JSON object with:\n"
        "  \"matches\": [\n"
        "    {\n"
        "      \"our_id\": \"<our attribute id>\",\n"
        "      \"human_id\": \"<human attribute id or null if no match>\",\n"
        "      \"confidence\": \"high\" | \"medium\" | \"low\",\n"
        "      \"reasoning\": \"<one sentence explaining why they match or don't>\"\n"
        "    }\n"
        "  ]\n\n"
        "Rules:\n"
        "- Each 'our' attribute must appear exactly once in the output.\n"
        "- A 'human' attribute can be matched to at most one 'our' attribute.\n"
        "- If an 'our' attribute has no semantic match in the human list, set human_id to null.\n"
        "- 'high' confidence = clearly the same test. 'medium' = related but not identical. "
        "'low' = weak/uncertain match.\n"
        "- Prefer matching by meaning over matching by position.\n"
    )

    user_prompt = (
        f"Control ID: {control_id}\n\n"
        f"OUR ATTRIBUTES (automated tool):\n{json.dumps(our_list, indent=2)}\n\n"
        f"HUMAN ATTRIBUTES (auditor workpaper):\n{json.dumps(human_list, indent=2)}\n\n"
        "Match each of our attributes to the human's. Return JSON."
    )

    try:
        result = evaluator._chat_json(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=4096,
            timeout=120,
        )
    except Exception as e:
        logger.warning(
            "LLM attribute matching failed for %s, falling back to fuzzy: %s",
            control_id, e,
        )
        return _fuzzy_match_attributes(our_attrs, human_attrs)

    # Parse LLM response into mapping + match_details
    mapping: Dict[str, Optional[str]] = {}
    match_details: List[Dict] = []
    human_used: set = set()

    # Build name lookups
    our_by_id = {str(a["id"]): a for a in our_attrs}
    human_by_id = {str(a["id"]): a for a in human_attrs}

    def _attr_name(a):
        return (a.get("name", "") + " " + a.get("description", "")).strip()

    confidence_to_sim = {"high": 0.90, "medium": 0.65, "low": 0.35}

    matches = result.get("matches", [])
    for m in matches:
        our_id = str(m.get("our_id", ""))
        human_id = m.get("human_id")
        conf = str(m.get("confidence", "low")).lower()
        reasoning = str(m.get("reasoning", ""))

        if our_id not in our_by_id:
            continue

        if human_id is not None:
            human_id = str(human_id)
            if human_id in human_by_id and human_id.upper() not in human_used:
                mapping[our_id] = human_id
                human_used.add(human_id.upper())
                match_details.append({
                    "our_id": our_id,
                    "our_name": _attr_name(our_by_id[our_id]),
                    "human_id": human_id,
                    "human_name": _attr_name(human_by_id[human_id]),
                    "similarity": confidence_to_sim.get(conf, 0.5),
                    "match_type": f"LLM ({conf})",
                    "reasoning": reasoning,
                })
            else:
                mapping[our_id] = None
        else:
            mapping[our_id] = None

    # Ensure all our attrs are in the mapping (LLM may have missed some)
    for a in our_attrs:
        if a["id"] not in mapping:
            mapping[a["id"]] = None

    # Identify human attributes not matched to any of ours
    matched_human_ids = {v.upper() for v in mapping.values() if v}
    human_extra = [
        {"id": a["id"], "name": _attr_name(a)}
        for a in human_attrs if a["id"].upper() not in matched_human_ids
    ]

    return mapping, match_details, human_extra


def _fuzzy_match_attributes(
    our_attrs: List[Dict], human_attrs: List[Dict],
) -> Tuple[Dict[str, Optional[str]], List[Dict]]:
    """
    Match our attribute IDs to human attribute IDs.

    Returns:
        mapping:       {our_attr_id -> human_attr_id or None}
        match_details: [{our_id, our_name, human_id, human_name, similarity, match_type}]

    Uses four matching passes:
      1. Exact ID match (e.g. "A" == "A")
      2. Ordinal position match (our "1"→human "A", "2"→"B", etc.)
      3. Fuzzy name/description match (SequenceMatcher >= 0.55)
      4. Remaining unmatched → None (human missed)
    """
    mapping: Dict[str, Optional[str]] = {}
    match_details: List[Dict] = []
    human_used: set = set()

    # Lookup helpers
    our_by_id = {str(a["id"]): a for a in our_attrs}
    human_by_id_full = {str(a["id"]).upper(): a for a in human_attrs}

    def _our_name(attr):
        return (attr.get("name", "") + " " + attr.get("description", "")).strip()

    def _similarity(a, b):
        return SequenceMatcher(
            None, _our_name(a).lower(), _our_name(b).lower()
        ).ratio()

    def _record(our_attr, h_attr, match_type):
        sim = _similarity(our_attr, h_attr)
        match_details.append({
            "our_id": our_attr["id"],
            "our_name": _our_name(our_attr),
            "human_id": h_attr["id"],
            "human_name": _our_name(h_attr),
            "similarity": round(sim, 2),
            "match_type": match_type,
        })

    # Pass 1: exact ID match (case-insensitive)
    human_by_id = {a["id"].upper(): a["id"] for a in human_attrs}
    for a in our_attrs:
        our_id = a["id"].upper()
        if our_id in human_by_id and our_id not in human_used:
            mapping[a["id"]] = human_by_id[our_id]
            human_used.add(our_id)
            h = human_by_id_full.get(our_id, {})
            _record(a, h, "Exact ID")

    # Pass 2: ordinal / positional match
    unmatched_ours = [a for a in our_attrs if a["id"] not in mapping]
    unmatched_humans = [a for a in human_attrs if a["id"].upper() not in human_used]

    if unmatched_ours and unmatched_humans:
        our_ids_numeric = all(str(a["id"]).strip().isdigit() for a in unmatched_ours)
        human_ids_alpha = all(
            str(a["id"]).strip().isalpha() for a in unmatched_humans
        )
        human_ids_numeric = all(
            str(a["id"]).strip().isdigit() for a in unmatched_humans
        )
        our_ids_alpha = all(str(a["id"]).strip().isalpha() for a in unmatched_ours)

        schemes_differ = (our_ids_numeric and human_ids_alpha) or \
                         (our_ids_alpha and human_ids_numeric)

        if schemes_differ:
            for our_attr, h_attr in zip(unmatched_ours, unmatched_humans):
                mapping[our_attr["id"]] = h_attr["id"]
                human_used.add(h_attr["id"].upper())
                _record(our_attr, h_attr, "Positional")

    # Pass 3: fuzzy name match for still-unmatched (threshold: 0.65)
    unmatched_ours = [a for a in our_attrs if a["id"] not in mapping]
    unmatched_humans = [a for a in human_attrs if a["id"].upper() not in human_used]

    for our_attr in unmatched_ours:
        our_name = _our_name(our_attr).lower()
        best_score = 0.0
        best_match_attr = None
        for h_attr in unmatched_humans:
            h_name = _our_name(h_attr).lower()
            score = SequenceMatcher(None, our_name, h_name).ratio()
            if score > best_score and score >= 0.65:
                best_score = score
                best_match_attr = h_attr
        if best_match_attr:
            mapping[our_attr["id"]] = best_match_attr["id"]
            human_used.add(best_match_attr["id"].upper())
            _record(our_attr, best_match_attr, "Fuzzy Name")
            unmatched_humans = [a for a in unmatched_humans
                                if a["id"].upper() != best_match_attr["id"].upper()]
        else:
            mapping[our_attr["id"]] = None  # Human missed this attribute

    # Identify human attributes that were NOT matched to any of ours
    unmatched_human_final = [
        a for a in human_attrs if a["id"].upper() not in human_used
    ]
    human_extra = [
        {"id": a["id"], "name": _our_name(a)}
        for a in unmatched_human_final
    ]

    return mapping, match_details, human_extra


def _adapt_tod_results(tod_results: List) -> List:
    """
    Adapt TODControlResult objects to match the interface _compare_controls expects.

    TODControlResult has:
      result (PASS/FAIL), design_adequate, deficiency_type, sample_result (singular)
    _compare_controls expects:
      operating_effectiveness, total_samples, sample_results (list)

    Returns a list of lightweight wrapper objects with the expected attributes.
    """
    adapted = []
    for r in tod_results:
        # Detect no-evidence / unable-to-assess cases from the sample result
        sample_result = getattr(r, "sample_result", None)
        is_no_evidence = False
        if sample_result is not None:
            ev_suff = getattr(sample_result, "evidence_sufficient", "").lower().strip()
            ctrl_perf = getattr(sample_result, "control_performed", "").lower().strip()
            remarks = getattr(sample_result, "remarks", "").lower()
            if ev_suff == "no" or ctrl_perf == "unable to assess" or \
               "no evidence" in remarks or "insufficient" in remarks or \
               "no extractable" in remarks:
                is_no_evidence = True
        elif sample_result is None:
            # No sample result at all — no evidence was available
            is_no_evidence = True

        # Also check TOD-level remarks for no-evidence indicators
        tod_remarks = (getattr(r, "overall_remarks", "") or "").lower()
        tod_gap = (getattr(r, "gap_identified", "") or "").lower()
        if any(kw in tod_remarks or kw in tod_gap for kw in
               ("no evidence", "no extractable", "insufficient", "unable to assess")):
            is_no_evidence = True

        # Map TOD result to effectiveness terminology
        # Conservative: no evidence = Not Effective (cannot confirm control operated)
        if getattr(r, "result", "") == "PASS":
            effectiveness = "Effective"
        else:
            effectiveness = "Not Effective"

        # Wrap sample_result (singular) into sample_results (list)
        sample_results = [sample_result] if sample_result else []

        # Create a lightweight namespace with the expected fields
        class _Adapted:
            pass
        a = _Adapted()
        a.control_id = r.control_id
        a.operating_effectiveness = effectiveness
        a.deficiency_type = getattr(r, "deficiency_type", "")
        a.total_samples = 1 if sample_result else 0
        a.sample_results = sample_results
        adapted.append(a)
    return adapted


# ═══════════════════════════════════════════════════════════════════════════════
#  LLM-FIRST: Raw Excel rows → extract + compare in one call
#  Eliminates all parser brittleness (regex patterns, column-name detection,
#  merged-cell artefacts). The LLM reads the workpaper exactly as a human
#  would and extracts the results directly.
# ═══════════════════════════════════════════════════════════════════════════════

def _read_workbook_sheets_raw(workbook_path: str) -> Dict[str, List[List[str]]]:
    """Read all sheets from an Excel workbook as raw string rows.

    Returns {sheet_name: [[cell_str, ...], ...]} — empty/fully-blank rows omitted.
    """
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    result: Dict[str, List[List[str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for row in ws.iter_rows():
            cells = [str(c.value).strip() if c.value is not None else "" for c in row]
            if any(c for c in cells):
                rows.append(cells)
        result[sheet_name] = rows
    wb.close()
    return result


def _find_control_raw_rows(
    control_id: str,
    raw_sheets: Dict[str, List[List[str]]],
) -> Tuple[Optional[List[List[str]]], Optional[str]]:
    """Find the raw rows for a control ID across workbook sheets.

    Returns (rows, sheet_name) or (None, None).
    The first sheet is treated as the summary and skipped for per-control matching.
    """
    sheet_names = list(raw_sheets.keys())
    cid_upper = control_id.strip().upper()
    control_sheets = sheet_names[1:] if len(sheet_names) > 1 else sheet_names

    # Pass 1: exact sheet name match
    for name in control_sheets:
        if name.strip().upper() == cid_upper:
            return raw_sheets[name], name

    # Pass 2: sheet name contains the control ID
    for name in control_sheets:
        if cid_upper in name.strip().upper():
            return raw_sheets[name], name

    # Pass 3: control ID appears in first 5 rows of sheet
    for name in control_sheets:
        for row in raw_sheets[name][:5]:
            if any(cid_upper in str(cell).upper() for cell in row):
                return raw_sheets[name], name

    return None, None


def _get_summary_context_for_control(
    control_id: str,
    raw_sheets: Dict[str, List[List[str]]],
) -> Optional[str]:
    """Extract the summary sheet row for a control as a 'Header: Value' string."""
    sheet_names = list(raw_sheets.keys())
    if not sheet_names:
        return None
    summary_rows = raw_sheets[sheet_names[0]]
    cid_upper = control_id.strip().upper()

    header_row: Optional[List[str]] = None
    header_idx = 0
    for i, row in enumerate(summary_rows[:15]):
        row_text = " ".join(str(c) for c in row).lower()
        if "control" in row_text and ("id" in row_text or "effectiveness" in row_text):
            header_row = row
            header_idx = i
            break

    if header_row is None:
        return None

    for row in summary_rows[header_idx + 1:]:
        if any(cid_upper in str(cell).upper() for cell in row):
            pairs = [
                f"{h}: {v}"
                for h, v in zip(header_row, row)
                if str(h).strip() and str(v).strip()
            ]
            return " | ".join(pairs)
    return None


def _format_rows_as_text_table(rows: List[List[str]], max_rows: int = 300) -> str:
    """Format raw rows as a pipe-separated text table for the LLM."""
    if not rows:
        return "(empty)"
    display_rows = rows[:max_rows]
    lines = [
        " | ".join(str(c).replace("\n", " ").replace("|", "/")[:80] for c in row)
        for row in display_rows
    ]
    result = "\n".join(lines)
    if len(rows) > max_rows:
        result += f"\n... ({len(rows) - max_rows} more rows not shown)"
    return result


_LLM_EXTRACT_AND_COMPARE_SYSTEM = """You are a senior audit quality reviewer with deep expertise in PCAOB standards, COSO framework, and internal control testing.

You will receive:
1. RAW ROWS from a human auditor's Excel workpaper sheet for a specific control (pipe-separated text)
2. Our automated tool's STRUCTURED RESULTS for the same control (clean JSON)

YOUR TWO TASKS:
A. EXTRACT what the human auditor tested from the raw Excel rows
B. COMPARE against our tool's results, identifying only gaps in the HUMAN's work

════════════════════════════════════════════════════════════════
EXACT WORKPAPER FORMAT (both human and tool produce identical layout):
════════════════════════════════════════════════════════════════

SHEET 1 — "Control Summary" sheet:
  Row 1: Merged group header cells: "RCM" | "Financial Assertions" | "IFC Components" | "COSO" | (separator) | "Results"
  Row 2: Actual column headers. The Results group columns are:
    "Detailed Testing Steps" | "Samples Tested" | "Passed / Total" | "Failed / Total" |
    "Deviation Rate" | "Validator Overrides" | "Deficiency Type" | "Failed Sample Details" |
    "Overall Remarks" | "Remedial Actions"
  Row 3+: One data row per control ID.
  → To find a control's summary: match the control ID in the RCM columns, then read "Deficiency Type",
    "Passed / Total", "Failed / Total", "Deviation Rate" from the Results group.
  → Operating Effectiveness is NOT a direct column in the summary — derive it from Deficiency Type:
    "None" = Effective, "Control Deficiency" = Effective with Exceptions,
    "Significant Deficiency" / "Material Weakness" = Not Effective.

PER-CONTROL SHEETS (Sheet 2+, one sheet per control):
  Sheet name = Control ID (e.g. "C-001", truncated to 31 chars).

  TOP SECTION — label/value pairs in columns A (label, blue fill) and B–H (value, merged):
    "Process:" | "Sub Process:" | "Control No:" | "Control Description:" |
    "Control Type:" | "Control Frequency:" | "Control Owner:" | "RCM Sample Size:" | "Sample Size:"

  WORKSTEPS SECTION (if present):
    Label: "Worksteps Performed:" | value: newline-separated workstep text

  ATTRIBUTES LEGEND SECTION (light green fill, comes before the testing table):
    Header row: "Attribute" (merged A–B) | "Attribute to be tested" (merged C–J)
    Data rows: col A = attribute ID (numeric: 1, 2, 3 …) | col C = "Name: Description" text
    → This section defines what each attribute number means.

  TESTING TABLE (2-row merged header, then data rows):
    Col A: "Testing:" label merged vertically across all data rows
    Col B: "Sr. No." (row sequence number, merged across both header rows)
    Row 1 of header:
      Col C: "Testing Sample Details" (merged across C through detail columns)
      Then: "Attribute {id}" columns (e.g. "Attribute 1", "Attribute 2") with light green fill
      Then: "Attribute Analysis" column
      Then: "Remarks" column
      Then: "Workpaper Reference" column
    Row 2 of header (sub-headers):
      Col C: "Sample No." (the actual sample identifier, e.g. S1, Q1-2024)
      Col D+: individual detail column names (e.g. "Source Document", "Date", etc.)
      Under each "Attribute {id}": short attribute name
    DATA ROWS (one per sample):
      Col B: Sr. No. (1, 2, 3 …)
      Col C: Sample ID (e.g. "S1", "S2", "Q1-2024")
      Col D+: sample detail values
      Attribute columns: "Yes" (green fill) / "No" (red fill) / "N/A"
      "Attribute Analysis" column: multi-line text with "id (name): Yes/No — reasoning" per attribute
      "Remarks" column: FAIL samples have red text starting with "[FAIL]"; PASS may have remarks
      "Workpaper Reference" column: "{control_id}.T.{sample_id}"

  SUMMARY TOTALS ROW (after last data row):
    Col B: "TOTAL"
    Col C+: "Samples: N | Passed: N | Failed: N | Deviation Rate: X%"

  CONCLUSION BANNER (after summary totals, ~2 rows below):
    Col A: "Conclusion:" (bold)
    Col B–F: conclusion text, e.g.:
      "Effective" (green fill, white bold text)
      "Effective with Exceptions -- Control Deficiency" (amber fill)
      "Not Effective -- Material Weakness" (red fill)
    → THIS IS THE PRIMARY SOURCE FOR operating_effectiveness. Always read this cell.
    → If the conclusion banner is missing, fall back to the Summary Sheet Row.

  SIGN-OFF SECTION (below conclusion):
    "Prepared by" | name | "Reviewed by" | name
    "Date" | date | "Date" | (blank)

════════════════════════════════════════════════════════════════
HOW TO EXTRACT (step by step):
════════════════════════════════════════════════════════════════
1. Find the CONCLUSION BANNER — look for a row where col A contains "Conclusion:" and col B has a colored verdict text. This is operating_effectiveness.
   Normalize: "Effective -- None" → "Effective"; "Effective with Exceptions -- Control Deficiency" → "Effective with Exceptions"; "Not Effective -- ..." → "Not Effective"
   If no banner found, derive from the Summary Sheet Row using Deficiency Type mapping above.

2. Find the ATTRIBUTES LEGEND — rows between the worksteps section and testing table where col A is a short number/letter and col C has "Name: Description" format. These define the attribute IDs and names.

3. Find the TESTING TABLE — look for the 2-row header where:
   - Row 1 has "Testing Sample Details" and "Attribute {id}" columns
   - Row 2 has "Sample No." as sub-header
   The actual sample data rows start on the row after row 2 of this header.

4. For each data row: read col C as sample_id, read each "Attribute {id}" column as Yes/No/N/A.
   A sample FAILS if ANY attribute is "No". A sample PASSES if all attributes are "Yes" or "N/A".

5. total_samples = count of actual data rows (not header, not totals row, not legend).

6. Ignore: page-header rows, blank/decorative rows, the "TOTAL" summary row, legend rows, sign-off rows.

════════════════════════════════════════════════════════════════
COMPARISON RULES:
════════════════════════════════════════════════════════════════
- SINGLE-DIRECTIONAL: flag only gaps in the HUMAN's work. If the human found something extra, ignore it.
- Match attributes by MEANING and INTENT, not by ID. Our "Attribute 1" may match human "Attribute A" if they test the same aspect — use the name/description from the legend to determine this.
- A verdict gap exists ONLY when our tool is STRICTER (e.g. tool = Not Effective, human = Effective).
- Do NOT flag cosmetic wording differences. Only flag substantive audit quality gaps.

SEVERITY:
- CRITICAL: Tool = Not Effective, Human = Effective
- HIGH: Verdict mismatch (e.g. Exceptions vs Effective) OR 3+ attribute/sample gaps
- MEDIUM: 1-2 attributes missed entirely OR 1-2 sample result differences
- LOW: Sample count difference only, or cosmetic gaps with matching verdict
- NONE: No substantive gaps

════════════════════════════════════════════════════════════════
Return ONLY a valid JSON object (no markdown fences, no explanation outside the JSON):
════════════════════════════════════════════════════════════════
{
  "human_extracted": {
    "operating_effectiveness": "Effective" | "Effective with Exceptions" | "Not Effective",
    "total_samples": <integer>,
    "attributes_tested": [
      {"id": "<column label e.g. Attribute 1>", "name": "<attribute name from legend>", "description": "<description from legend>"}
    ],
    "sample_results": [
      {
        "sample_id": "<Sample No. value from col C>",
        "result": "PASS" | "FAIL",
        "attribute_results": {"<attr_id>": "Yes" | "No" | "N/A"}
      }
    ],
    "deficiency_type": "Material Weakness" | "Significant Deficiency" | "Control Deficiency" | "None" | "",
    "extraction_notes": "<brief note on anything ambiguous, or empty string>"
  },
  "attribute_matches": [
    {
      "our_id": "<our attr id>",
      "our_name": "<our attr name>",
      "human_id": "<matched human attr id, or null if human didn't test this>",
      "human_name": "<human attr name from legend, or null>",
      "confidence": "high" | "medium" | "low",
      "reasoning": "<one sentence: why they match or why no match>"
    }
  ],
  "verdict_assessment": {
    "verdict_gap": true | false,
    "detail": "<explanation of the verdict comparison, citing specific evidence from the workpaper>",
    "severity_recommendation": "CRITICAL" | "HIGH" | "MEDIUM" | "LOW" | "NONE",
    "severity_reasoning": "<concrete reasoning citing specific gaps>"
  },
  "sample_gaps": [
    {"sample_id": "<id>", "gap_type": "Different" | "Missing", "detail": "<what human missed>"}
  ],
  "attribute_gaps": [
    {
      "attr_id": "<our attr id>", "attr_name": "<name>",
      "gap_type": "Missing" | "Different" | "Human Skipped",
      "sample_id": "<sample id if Different, else empty>",
      "detail": "<explanation with specific evidence>"
    }
  ],
  "overall_narrative": "<3-5 sentence summary of the most important findings for this control>"
}
"""


def _llm_extract_and_compare_one_control(
    cid: str,
    our_result,
    human_sheet_rows: List[List[str]],
    human_summary_context: Optional[str],
    schema,
    phase: str,
    evaluator,
) -> Optional[Dict]:
    """Single LLM call: extract human data from raw workpaper rows AND compare to our results.

    Returns combined JSON dict, or None on failure.
    """
    our_effectiveness = getattr(our_result, "operating_effectiveness", "")
    our_samples_raw = getattr(our_result, "sample_results", [])

    our_attrs = []
    if schema:
        for a in getattr(schema, "attributes", []):
            our_attrs.append({
                "id": a.get("id", ""),
                "name": a.get("name", ""),
                "description": a.get("description", ""),
            })

    our_samples = []
    for s in our_samples_raw:
        sr = {
            "sample_id": getattr(s, "sample_id", ""),
            "result": getattr(s, "result", ""),
            "remarks": getattr(s, "remarks", ""),
            "deviation_details": getattr(s, "deviation_details", ""),
        }
        attr_res = getattr(s, "attribute_results", None)
        if attr_res:
            sr["attribute_results"] = dict(attr_res)
        our_samples.append(sr)

    control_context = ""
    if schema:
        ctrl_desc = getattr(schema, "control_description", "") or ""
        worksteps = getattr(schema, "worksteps", []) or []
        if ctrl_desc or worksteps:
            control_context = "═══ CONTROL CONTEXT ═══\n"
            if ctrl_desc:
                control_context += f"Description: {ctrl_desc}\n"
            if worksteps:
                control_context += f"Worksteps: {json.dumps(worksteps)}\n"
            control_context += "\n"

    raw_table = _format_rows_as_text_table(human_sheet_rows)
    summary_section = (
        f"═══ HUMAN SUMMARY SHEET ROW ═══\n{human_summary_context}\n\n"
        if human_summary_context else ""
    )

    user_prompt = (
        f"CONTROL: {cid}  |  PHASE: {phase}\n\n"
        f"{control_context}"
        f"═══ OUR TOOL'S RESULTS ═══\n"
        f"Verdict: {our_effectiveness}\n"
        f"Attributes tested ({len(our_attrs)}):\n{json.dumps(our_attrs, indent=2)}\n"
        f"Samples ({len(our_samples)}):\n{json.dumps(our_samples, indent=2)}\n\n"
        f"{summary_section}"
        f"═══ HUMAN AUDITOR'S RAW WORKPAPER ROWS ═══\n"
        f"{raw_table}\n\n"
        f"Extract the human's results from the raw rows above, then compare against our tool's results. Return the combined JSON."
    )

    try:
        return evaluator._chat_json(
            messages=[
                {"role": "system", "content": _LLM_EXTRACT_AND_COMPARE_SYSTEM},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=16000,
            timeout=300,
        )
    except Exception as e:
        logger.warning("LLM extract+compare failed for %s: %s", cid, e)
        return None


def _build_comparison_from_llm_full(
    cid: str,
    llm_result: Dict,
    our_result,
    schema,
    phase: str,
) -> ControlComparison:
    """Build ControlComparison from the combined LLM extraction+comparison JSON.

    Human data comes from llm_result['human_extracted'] rather than a
    pre-parsed HumanControlResult, so no parser brittleness affects the output.
    """
    comp = ControlComparison(control_id=cid, phase=phase)

    # Our side
    comp.our_effectiveness = getattr(our_result, "operating_effectiveness", "")
    comp.our_sample_count = getattr(our_result, "total_samples", 0)
    if schema:
        comp.our_attributes = [a["id"] for a in getattr(schema, "attributes", [])]

    # Human side — entirely from LLM extraction
    extracted = llm_result.get("human_extracted", {})
    raw_human_eff = str(extracted.get("operating_effectiveness", "") or "").strip()
    if not raw_human_eff or raw_human_eff.lower() in ("", "none", "null", "not found", "unknown"):
        comp.human_effectiveness = "NOT FOUND"
    else:
        comp.human_effectiveness = _normalize_verdict(raw_human_eff)
    comp.human_sample_count = int(extracted.get("total_samples", 0) or 0)
    extracted_attrs = extracted.get("attributes_tested", [])
    comp.human_attributes = [a.get("id", "") for a in extracted_attrs]

    # Verdict — deterministic from the extracted human effectiveness
    verdicts_match = _verdicts_same(comp.our_effectiveness, comp.human_effectiveness)
    if verdicts_match:
        comp.verdict_gap = False
        comp.verdict_detail = f"Both agree: '{_normalize_verdict(comp.our_effectiveness)}'."
    else:
        our_rank = EFFECTIVENESS_RANK.get(_normalize_verdict(comp.our_effectiveness).lower(), -1)
        human_rank = EFFECTIVENESS_RANK.get(_normalize_verdict(comp.human_effectiveness).lower(), -1)
        if our_rank > human_rank and our_rank >= 0 and human_rank >= 0:
            comp.verdict_gap = True
            comp.verdict_detail = (
                f"Our tool assessed '{comp.our_effectiveness}' but human assessed "
                f"'{comp.human_effectiveness}' — human may have missed issues"
            )
        else:
            comp.verdict_gap = False
            comp.verdict_detail = (
                f"Ours: '{comp.our_effectiveness}', Human: '{comp.human_effectiveness}'. "
                f"Verdicts differ but human was not more lenient."
            )

    # Attribute matches
    confidence_to_sim = {"high": 0.90, "medium": 0.65, "low": 0.35}
    llm_matched_our_ids: set = set()
    for am in llm_result.get("attribute_matches", []):
        human_id = am.get("human_id")
        our_id = am.get("our_id", "")
        conf = str(am.get("confidence", "low")).lower()
        comp.attribute_match_details.append({
            "our_id": our_id,
            "our_name": am.get("our_name", ""),
            "human_id": human_id or "",
            "human_name": am.get("human_name", ""),
            "similarity": confidence_to_sim.get(conf, 0.5),
            "match_type": f"LLM ({conf})",
            "reasoning": am.get("reasoning", ""),
        })
        if human_id is None:
            comp.attributes_human_missed.append(our_id)
        if our_id:
            llm_matched_our_ids.add(our_id)

    # Cross-check — catch any of our attrs the LLM forgot to mention
    if schema:
        all_our_attr_ids = {str(a["id"]) for a in getattr(schema, "attributes", [])}
        for our_id in all_our_attr_ids - llm_matched_our_ids:
            if our_id not in comp.attributes_human_missed:
                comp.attributes_human_missed.append(our_id)
    comp.attributes_human_missed = list(dict.fromkeys(comp.attributes_human_missed))

    # Attribute gaps
    for ag in llm_result.get("attribute_gaps", []):
        gap_type = ag.get("gap_type", "Different")
        if gap_type in ("Different", "Missing", "Human Skipped"):
            comp.attribute_differences.append({
                "attr_id": ag.get("attr_id", ""),
                "attr_name": ag.get("attr_name", ""),
                "sample_id": ag.get("sample_id", ""),
                "our_result": ag.get("our_result", "No (FAIL)" if gap_type == "Different" else "Tested"),
                "human_result": ag.get("human_result", "Yes (PASS)" if gap_type == "Different" else "Not tested"),
                "gap_type": gap_type,
                "detail": ag.get("detail", ""),
            })

    # Sample gaps
    for sg in llm_result.get("sample_gaps", []):
        if sg.get("gap_type") == "Missing":
            comp.samples_human_missed.append(sg.get("sample_id", ""))
        elif sg.get("gap_type") == "Different":
            comp.sample_differences.append({
                "sample_id": sg.get("sample_id", ""),
                "our_result": "FAIL",
                "human_result": "PASS",
                "our_deviation": sg.get("detail", ""),
            })

    # Narrative
    llm_narrative = llm_result.get("overall_narrative", "")
    factual_prefix = (
        f"Verdict: Both={_normalize_verdict(comp.our_effectiveness)}" if verdicts_match
        else f"Verdict difference: Ours={comp.our_effectiveness}, Human={comp.human_effectiveness}"
    )
    if comp.attributes_human_missed:
        factual_prefix += f" | {len(comp.attributes_human_missed)} attribute(s) missed by human"
    if comp.attribute_differences:
        factual_prefix += f" | {len(comp.attribute_differences)} attribute result difference(s)"
    comp.overall_narrative = f"{factual_prefix}. {llm_narrative}" if llm_narrative else factual_prefix

    # Per-sample per-attribute comparison detail — using LLM-extracted human samples
    class _ExtractedSample:
        """Lightweight shim so extracted sample dicts work with _build_attribute_comparison_detail."""
        def __init__(self, d: Dict) -> None:
            self.sample_id = str(d.get("sample_id", ""))
            self.result = str(d.get("result", ""))
            self.attribute_results = {
                str(k): str(v) for k, v in d.get("attribute_results", {}).items()
            }
            self.remarks = str(d.get("remarks", ""))

    human_samples_list = [_ExtractedSample(s) for s in extracted.get("sample_results", [])]
    our_samples_list = getattr(our_result, "sample_results", [])

    attr_mapping: Dict[str, Optional[str]] = {}
    for md in comp.attribute_match_details:
        our_id = md.get("our_id", "")
        human_id_val = md.get("human_id") or None
        if our_id:
            attr_mapping[our_id] = human_id_val
    if schema:
        for a in getattr(schema, "attributes", []):
            if a["id"] not in attr_mapping:
                attr_mapping[a["id"]] = None

    comp.attribute_comparison_detail = _build_attribute_comparison_detail(
        attr_mapping, comp.attribute_match_details,
        our_samples_list, human_samples_list, schema,
    )

    comp.gap_severity = _classify_severity(comp)
    return comp


def _compare_controls(
    our_results: List,
    human_results: Dict[str, HumanControlResult],
    schemas: Dict,
    total_rcm_controls: int = 0,
    phase: str = "TOE",
    evaluator=None,
    max_workers: int = 5,
    human_raw_sheets: Optional[Dict[str, List[List[str]]]] = None,
) -> ComparisonReport:
    """
    Compare our results against human workpaper results.
    Single-directional: only gaps in human work.
    Works for both TOD and TOE phases.

    If *evaluator* is provided, uses the LLM for semantic attribute matching,
    verdict assessment, and gap analysis (one call per control, threaded).
    Falls back to deterministic string comparison if the LLM call fails.
    """
    report = ComparisonReport(
        timestamp=datetime.now().isoformat(),
        total_controls=total_rcm_controls or len(our_results),
    )

    # Index human results by uppercase control ID for matching
    human_by_id = {cid.strip().upper(): hr for cid, hr in human_results.items()}

    # ── Primary: LLM extract+compare from raw Excel rows ─────────
    # Sends raw sheet rows to the LLM so it extracts AND compares in one call,
    # eliminating all parser brittleness (column detection, regex, etc.).
    raw_llm_results_map: Dict[str, Dict] = {}
    if evaluator and human_raw_sheets:
        _raw_eligible = []
        for our_result in our_results:
            cid = our_result.control_id
            ctrl_rows, _ = _find_control_raw_rows(cid, human_raw_sheets)
            if ctrl_rows is not None:
                summary_ctx = _get_summary_context_for_control(cid, human_raw_sheets)
                _raw_eligible.append((cid, our_result, ctrl_rows, summary_ctx, schemas.get(cid)))

        if _raw_eligible:
            logger.info(
                "Running LLM extract+compare for %d controls (%s phase)",
                len(_raw_eligible), phase,
            )
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {}
                for cid, our_r, rows, summary_ctx, sch in _raw_eligible:
                    fut = pool.submit(
                        _llm_extract_and_compare_one_control,
                        cid, our_r, rows, summary_ctx, sch, phase, evaluator,
                    )
                    futures[fut] = cid
                for fut in as_completed(futures):
                    cid = futures[fut]
                    try:
                        result = fut.result()
                        if result:
                            raw_llm_results_map[cid] = result
                            logger.info("LLM extract+compare OK for %s", cid)
                    except Exception as e:
                        logger.warning("LLM extract+compare failed for %s: %s", cid, e)

    # ── Fallback: LLM comparison using pre-parsed human data ──────
    # Used for controls where extract+compare failed or raw sheets not available.
    llm_results_map: Dict[str, Dict] = {}
    if evaluator:
        _eligible = []
        for our_result in our_results:
            cid = our_result.control_id
            if cid in raw_llm_results_map:
                continue  # already handled by extract+compare
            human = human_by_id.get(cid.strip().upper())
            if human:
                _eligible.append((cid, our_result, human, schemas.get(cid)))

        if _eligible:
            logger.info(
                "Running LLM comparison (parsed fallback) for %d controls (%s phase)",
                len(_eligible), phase,
            )
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {}
                for cid, our_r, h, sch in _eligible:
                    fut = pool.submit(
                        _llm_compare_one_control,
                        cid, our_r, h, sch, phase, evaluator,
                    )
                    futures[fut] = cid
                for fut in as_completed(futures):
                    cid = futures[fut]
                    try:
                        llm_json = fut.result()
                        if llm_json:
                            llm_results_map[cid] = llm_json
                            logger.info("LLM comparison (fallback) OK for %s", cid)
                    except Exception as e:
                        logger.warning("LLM comparison (fallback) failed for %s: %s", cid, e)

    # ── Build comparisons ─────────────────────────────────────────
    for our_result in our_results:
        cid = our_result.control_id
        cid_upper = cid.strip().upper()
        human = human_by_id.get(cid_upper)
        schema = schemas.get(cid)

        # Primary: LLM extract+compare (raw rows, no parser dependency)
        if cid in raw_llm_results_map:
            comp = _build_comparison_from_llm_full(
                cid, raw_llm_results_map[cid], our_result, schema, phase,
            )
            report.control_comparisons.append(comp)
            continue

        # Fallback: LLM comparison from pre-parsed human data
        if cid in llm_results_map:
            comp = _build_comparison_from_llm(
                cid, llm_results_map[cid], our_result, human, schema, phase,
            )
            report.control_comparisons.append(comp)
            continue

        # ── Deterministic fallback ────────────────────────────────
        comp = ControlComparison(control_id=cid, phase=phase)

        # Our results
        comp.our_effectiveness = getattr(our_result, "operating_effectiveness", "")
        comp.our_sample_count = getattr(our_result, "total_samples", 0)

        if schema:
            comp.our_attributes = [a["id"] for a in getattr(schema, "attributes", [])]

        if human is None:
            comp.human_effectiveness = "NOT FOUND"
            comp.verdict_gap = True
            comp.verdict_detail = "Control not present in human workpaper"
            comp.gap_severity = _classify_severity(comp)
            report.control_comparisons.append(comp)
            continue

        # Human results — normalize unknown/unable to "Not Effective"
        _raw_human_eff = human.operating_effectiveness or "Unknown"
        comp.human_effectiveness = _normalize_verdict(_raw_human_eff)
        comp.human_sample_count = human.total_samples
        comp.human_attributes = [a["id"] for a in human.attributes_tested]

        # Normalize our side too — unknown/unable becomes "Not Effective"
        comp.our_effectiveness = _normalize_verdict(comp.our_effectiveness)

        # ── 1. Verdict comparison (deterministic, normalised) ─────
        verdicts_match = _verdicts_same(comp.our_effectiveness, comp.human_effectiveness)
        if verdicts_match:
            comp.verdict_gap = False
            comp.verdict_detail = f"Both agree: '{_normalize_verdict(comp.our_effectiveness)}'."
        else:
            our_rank = EFFECTIVENESS_RANK.get(
                _normalize_verdict(comp.our_effectiveness).lower(), -1)
            human_rank = EFFECTIVENESS_RANK.get(
                _normalize_verdict(comp.human_effectiveness).lower(), -1)
            if our_rank > human_rank and our_rank >= 0 and human_rank >= 0:
                comp.verdict_gap = True
                comp.verdict_detail = (
                    f"Our tool assessed '{comp.our_effectiveness}' but human assessed "
                    f"'{comp.human_effectiveness}' — human may have missed issues"
                )
            else:
                comp.verdict_gap = False
                comp.verdict_detail = (
                    f"Ours: '{comp.our_effectiveness}', Human: '{comp.human_effectiveness}'. "
                    f"Verdicts differ but human was not more lenient."
                )

        # ── 2. Attribute comparison ────────────────────────────────
        attr_mapping: Dict[str, Optional[str]] = {}
        if schema and human.attributes_tested:
            our_attrs = getattr(schema, "attributes", [])
            attr_mapping, match_details, human_extra = _fuzzy_match_attributes(
                our_attrs, human.attributes_tested,
            )
            comp.attribute_match_details = match_details
            comp.attributes_human_extra = human_extra

            # Attributes human missed entirely
            comp.attributes_human_missed = [
                aid for aid, hid in attr_mapping.items() if hid is None
            ]

            # Attribute result differences (per-sample)
            our_samples = getattr(our_result, "sample_results", [])
            human_samples_by_id = {
                _normalize_sample_id(s.sample_id): s for s in human.sample_results
            }

            for our_sample in our_samples:
                our_sid_norm = _normalize_sample_id(our_sample.sample_id)
                human_sample = human_samples_by_id.get(our_sid_norm)
                if not human_sample:
                    continue  # Sample-level gap handled below

                our_attr_results = getattr(our_sample, "attribute_results", {})
                for our_aid, human_aid in attr_mapping.items():
                    if human_aid is None:
                        continue  # Already counted as missed

                    our_val = _normalize_yes_no(our_attr_results.get(our_aid, ""))
                    human_val = _normalize_yes_no(
                        human_sample.attribute_results.get(human_aid, "")
                    )

                    # Gap: our tool said No but human said Yes
                    if our_val == "No" and human_val == "Yes":
                        comp.attribute_differences.append({
                            "attr_id": our_aid,
                            "attr_name": next(
                                (a.get("name", "") for a in our_attrs if a["id"] == our_aid),
                                our_aid,
                            ),
                            "sample_id": our_sample.sample_id,
                            "our_result": "No (FAIL)",
                            "human_result": "Yes (PASS)",
                            "gap_type": "Different",
                        })
                    # Gap: our tool tested (Yes/No) but human marked N/A or skipped
                    elif our_val in ("Yes", "No") and human_val not in ("Yes", "No"):
                        comp.attribute_differences.append({
                            "attr_id": our_aid,
                            "attr_name": next(
                                (a.get("name", "") for a in our_attrs if a["id"] == our_aid),
                                our_aid,
                            ),
                            "sample_id": our_sample.sample_id,
                            "our_result": f"{our_val} (tested)",
                            "human_result": human_val or "N/A (not tested)",
                            "gap_type": "Human Skipped",
                        })

        elif schema and not human.attributes_tested:
            # Human tested NO attributes at all
            comp.attributes_human_missed = [a["id"] for a in getattr(schema, "attributes", [])]

        # ── 2b. Build full per-sample attribute comparison detail ──
        if schema:
            _our_attrs_list = getattr(schema, "attributes", [])
            _attr_mapping_full: Dict[str, Optional[str]] = {}
            if attr_mapping:
                _attr_mapping_full = attr_mapping
            else:
                # No fuzzy match ran (e.g. human had no attributes)
                for _a in _our_attrs_list:
                    _attr_mapping_full[_a["id"]] = None

            _our_samps = getattr(our_result, "sample_results", [])
            comp.attribute_comparison_detail = _build_attribute_comparison_detail(
                _attr_mapping_full,
                comp.attribute_match_details,
                _our_samps,
                human.sample_results,
                schema,
            )

        # ── 3. Sample comparison ───────────────────────────────────
        our_samples = getattr(our_result, "sample_results", [])
        human_sid_set = {
            _normalize_sample_id(s.sample_id) for s in human.sample_results
        }

        for our_sample in our_samples:
            our_sid_norm = _normalize_sample_id(our_sample.sample_id)
            if our_sid_norm not in human_sid_set:
                comp.samples_human_missed.append(our_sample.sample_id)
            else:
                # Compare results for matching samples
                human_sample = next(
                    (s for s in human.sample_results
                     if _normalize_sample_id(s.sample_id) == our_sid_norm),
                    None,
                )
                if human_sample:
                    our_res = getattr(our_sample, "result", "")
                    human_res = human_sample.result
                    # Gap: our tool found FAIL but human said PASS
                    if our_res == "FAIL" and human_res == "PASS":
                        comp.sample_differences.append({
                            "sample_id": our_sample.sample_id,
                            "our_result": "FAIL",
                            "human_result": "PASS",
                            "our_deviation": getattr(our_sample, "deviation_details", ""),
                        })

        # ── 4. Gap severity ────────────────────────────────────────
        comp.gap_severity = _classify_severity(comp)

        report.control_comparisons.append(comp)

    # Aggregate severity counts
    for comp in report.control_comparisons:
        report.severity_counts[comp.gap_severity] = (
            report.severity_counts.get(comp.gap_severity, 0) + 1
        )
    report.controls_with_gaps = sum(
        1 for c in report.control_comparisons if c.gap_severity != "NONE"
    )

    return report


def _normalize_yes_no(val: str) -> str:
    """Normalize Yes/No values for comparison."""
    v = val.strip().lower()
    if v in ("yes", "y", "true", "1", "pass"):
        return "Yes"
    elif v in ("no", "n", "false", "0", "fail"):
        return "No"
    return val.strip()


def _normalize_verdict(verdict: str) -> str:
    """Normalize an effectiveness verdict to a canonical form for comparison.

    'Unknown', 'Unable to Assess', 'No Evidence', and blank verdicts are all
    treated as 'Not Effective' — if the control could not be assessed, it
    cannot be considered effective.
    """
    v = verdict.strip().lower()
    if not v or v in ("unknown", "not found", "unable to assess", "no evidence"):
        return "Not Effective"
    if "not effective" in v:
        return "Not Effective"
    if "effective with exception" in v:
        return "Effective with Exceptions"
    if "effective" in v:
        return "Effective"
    return verdict.strip()


def _verdicts_same(our: str, human: str) -> bool:
    """Return True if two verdicts are semantically the same after normalisation."""
    return _normalize_verdict(our) == _normalize_verdict(human)


def _build_attribute_comparison_detail(
    attr_mapping: Dict[str, Optional[str]],
    match_details: List[Dict],
    our_samples: List,
    human_samples: List,
    schema,
) -> List[Dict]:
    """Build per-sample per-attribute comparison detail for the export sheets.

    Returns a list of dicts, each representing one attribute comparison for one
    sample.  For TOD (1 sample) the sample_id may be empty.
    """
    details: List[Dict] = []

    # Build match-info lookup keyed by our_attr_id
    match_info: Dict[str, Dict] = {}
    for md in match_details:
        match_info[md.get("our_id", "")] = {
            "match_method": md.get("match_type", ""),
            "match_confidence": md.get("similarity", 0),
            "our_name": md.get("our_name", ""),
            "human_name": md.get("human_name", ""),
        }

    # Attribute name lookup from schema
    attr_names: Dict[str, str] = {}
    if schema:
        for a in getattr(schema, "attributes", []):
            aid = str(a.get("id", ""))
            attr_names[aid] = (a.get("name", "") + " " + a.get("description", "")).strip()

    # Human samples indexed by normalised sample-id
    human_samples_by_id: Dict[str, Any] = {}
    for s in (human_samples or []):
        sid = _normalize_sample_id(s.sample_id)
        human_samples_by_id[sid] = s

    if not our_samples:
        # No sample data (typical for TOD) — list attribute matches only.
        # Try to pull human attribute results from the first human sample
        # (TOD human workpapers often have one implicit "sample" row).
        _human_first_sample = (human_samples or [None])[0] if human_samples else None
        _human_first_attr_res = (
            getattr(_human_first_sample, "attribute_results", {}) or {}
        ) if _human_first_sample else {}

        for our_id, human_id in attr_mapping.items():
            mi = match_info.get(our_id, {})

            # Look up human result from first human sample's attribute_results
            if not human_id:
                human_val = "Not Tested"
            elif _human_first_attr_res:
                human_val = _normalize_yes_no(str(_human_first_attr_res.get(human_id, "")))
                if not human_val:
                    human_val = "—"
            else:
                human_val = "—"

            details.append({
                "sample_id": "",
                "our_attr_id": our_id,
                "our_attr_name": mi.get("our_name", attr_names.get(our_id, "")),
                "human_attr_id": human_id or "",
                "human_attr_name": mi.get("human_name", ""),
                "match_method": mi.get("match_method", "No Match") if human_id else "No Match",
                "match_confidence": mi.get("match_confidence", 0),
                "our_result": "",
                "human_result": human_val,
                "same": human_id is not None,
                "direction": "" if human_id else "Human Did Not Test",
            })
        return details

    for our_sample in our_samples:
        sample_id = getattr(our_sample, "sample_id", "")
        our_sid_norm = _normalize_sample_id(sample_id)
        human_sample = human_samples_by_id.get(our_sid_norm)

        our_attr_results = getattr(our_sample, "attribute_results", {}) or {}
        if not isinstance(our_attr_results, dict):
            our_attr_results = dict(our_attr_results)

        for our_id, human_id in attr_mapping.items():
            mi = match_info.get(our_id, {})
            our_val = _normalize_yes_no(str(our_attr_results.get(our_id, "")))

            if human_id and human_sample:
                h_attr_results = getattr(human_sample, "attribute_results", {}) or {}
                human_val = _normalize_yes_no(str(h_attr_results.get(human_id, "")))
            elif human_id:
                human_val = "Sample Not Found"
            else:
                human_val = "Not Tested"

            # Determine if same
            non_comparable = ("Not Tested", "Sample Not Found", "N/A", "")
            if human_val in non_comparable or our_val in ("N/A", ""):
                same = False
            else:
                same = (our_val == human_val)

            direction = ""
            if not same:
                if our_val == "No" and human_val == "Yes":
                    direction = "We Failed, Human Passed"
                elif our_val == "Yes" and human_val == "No":
                    direction = "We Passed, Human Failed"
                elif human_val == "Not Tested":
                    direction = "Human Did Not Test"
                elif human_val == "Sample Not Found":
                    direction = "Human Did Not Test This Sample"
                elif our_val in ("N/A", ""):
                    direction = f"Ours: N/A, Human: {human_val}"
                else:
                    direction = f"Ours: {our_val}, Human: {human_val}"

            details.append({
                "sample_id": sample_id,
                "our_attr_id": our_id,
                "our_attr_name": mi.get("our_name", attr_names.get(our_id, "")),
                "human_attr_id": human_id or "",
                "human_attr_name": mi.get("human_name", ""),
                "match_method": mi.get("match_method", "No Match") if human_id else "No Match",
                "match_confidence": mi.get("match_confidence", 0),
                "our_result": our_val,
                "human_result": human_val,
                "same": same,
                "direction": direction,
            })

    return details


def _classify_severity(comp: ControlComparison) -> str:
    """Classify gap severity based on verdict, attribute, and sample gaps.

    Note: unknown/unable-to-assess verdicts are already normalised to
    'Not Effective' before this function is called, so no special-casing
    for unassessable verdicts is needed here.
    """
    our_eff = comp.our_effectiveness.lower().strip()
    human_eff = comp.human_effectiveness.lower().strip()

    # CRITICAL: Our tool says Not Effective, human says Effective
    our_rank = EFFECTIVENESS_RANK.get(our_eff, -1)
    human_rank = EFFECTIVENESS_RANK.get(human_eff, -1)

    if our_rank == 2 and human_rank == 0:  # Not Effective vs Effective
        return "CRITICAL"

    # HIGH: verdict mismatch (Exceptions vs Effective) OR 3+ missed differences
    total_missed = (
        len(comp.attributes_human_missed)
        + len(comp.attribute_differences)
        + len(comp.sample_differences)
    )
    if comp.verdict_gap and our_rank > human_rank:
        return "HIGH"
    if total_missed >= 3:
        return "HIGH"

    # MEDIUM: some attributes missed or 1-2 sample/attribute differences
    if comp.attributes_human_missed or (0 < total_missed < 3):
        return "MEDIUM"

    # LOW: sample count difference only
    if comp.our_sample_count > comp.human_sample_count and comp.human_sample_count > 0:
        return "LOW"

    # Human didn't have the control at all
    if comp.human_effectiveness == "NOT FOUND":
        if comp.our_effectiveness == "Effective":
            return "LOW"
        return "HIGH"

    if comp.verdict_gap or comp.samples_human_missed:
        return "LOW"

    return "NONE"


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

# Color scheme for severity
SEVERITY_COLORS = {
    "CRITICAL": "FF4444",
    "HIGH":     "F4B084",
    "MEDIUM":   "FFEB9C",
    "LOW":      "D6E4F0",
    "NONE":     "C6EFCE",
}


def _export_comparison_report(report: ComparisonReport, output_path: str,
                              schemas: Dict = None, rcm_df=None):
    """Export the comparison report as a 5-sheet Excel workbook.

    Sheets:
      1. Executive Summary  (Gold tab)
      2. TOD Comparison     (Blue tab)
      3. TOD Attribute Detail (Light-Blue tab)
      4. TOE Comparison     (Green tab)
      5. TOE Attribute Detail (Light-Green tab)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    TAB_GOLD = "FFD700"
    TAB_BLUE = "4472C4"
    TAB_LIGHT_BLUE = "9DC3E6"
    TAB_GREEN = "548235"
    TAB_LIGHT_GREEN = "A9D18E"

    def _write_header(ws, row_num, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = wrap
            c.border = thin

    def _cell(ws, r, c, val, *, align=None, fill=None):
        """Write a bordered cell with optional fill / alignment."""
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = thin
        if align:
            cell.alignment = align
        if fill:
            cell.fill = fill
        return cell

    # ── Lookup helpers ────────────────────────────────────────────
    _EMPTY = ("nan", "none", "null", "", "n/a", "na")

    ctrl_desc: Dict[str, str] = {}
    ctrl_freq: Dict[str, str] = {}
    ctrl_risk: Dict[str, str] = {}
    ctrl_process: Dict[str, str] = {}
    if schemas:
        for cid, sch in schemas.items():
            ctrl_desc[cid] = getattr(sch, "control_description", "") or ""
    if rcm_df is not None and "Control Id" in rcm_df.columns:
        for _, rw in rcm_df.iterrows():
            cid = str(rw.get("Control Id", "")).strip()
            for col in rcm_df.columns:
                cl = col.lower().replace("_", " ").strip()
                if cl in ("control description", "controldescription", "control_description",
                          "control narrative", "control activity", "control desc"):
                    val = str(rw.get(col, "")).strip()
                    if val.lower() not in _EMPTY and not ctrl_desc.get(cid):
                        ctrl_desc[cid] = val
                elif cl in ("control frequency", "controlfrequency", "frequency",
                            "control_frequency", "frequence", "periodicity"):
                    val = str(rw.get(col, "")).strip()
                    if val.lower() not in _EMPTY:
                        ctrl_freq[cid] = val
                elif cl in ("risk level", "risk_level", "risklevel",
                            "risk rating", "inherent risk", "risk severity"):
                    val = str(rw.get(col, "")).strip()
                    if val.lower() not in _EMPTY:
                        ctrl_risk[cid] = val
                elif cl in ("process", "mega process", "business process", "cycle"):
                    val = str(rw.get(col, "")).strip()
                    if val.lower() not in _EMPTY:
                        ctrl_process.setdefault(cid, val)

    def _build_fallback_narrative(comp) -> str:
        """Generate a plain-English gap summary when LLM narrative is unavailable."""
        parts = []
        our_v = comp.our_effectiveness or "Unknown"
        hum_v = comp.human_effectiveness or "Unknown"
        if _verdicts_same(our_v, hum_v):
            parts.append(f"Both tool and human concluded: {our_v}.")
        else:
            parts.append(f"Verdict mismatch — Tool: {our_v}, Human: {hum_v}.")
        n_missed = len(comp.attributes_human_missed)
        n_extra = len(comp.attributes_human_extra)
        n_diff = len(comp.attribute_differences)
        if n_missed:
            parts.append(f"Human did not test {n_missed} attribute(s) that the tool tested.")
        if n_extra:
            parts.append(f"Human tested {n_extra} additional attribute(s) not covered by the tool.")
        if n_diff:
            parts.append(f"{n_diff} attribute result(s) differ between tool and human.")
        s_missed = len(comp.samples_human_missed)
        if s_missed:
            parts.append(f"Human did not test {s_missed} sample(s) that the tool tested.")
        if not n_missed and not n_diff and not s_missed and not n_extra:
            parts.append("No substantive gaps identified.")
        return " ".join(parts)

    # Split comparisons by phase
    tod_comps = [c for c in report.control_comparisons if c.phase == "TOD"]
    toe_comps = [c for c in report.control_comparisons if c.phase == "TOE"]
    tod_by_cid = {c.control_id: c for c in tod_comps}
    toe_by_cid = {c.control_id: c for c in toe_comps}
    all_cids = list(dict.fromkeys(
        [c.control_id for c in tod_comps] + [c.control_id for c in toe_comps]
    ))

    # ═══════════════════════════════════════════════════════════════
    #  Sheet 1 — Executive Summary
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = TAB_GOLD

    ws1.merge_cells("A1:I1")
    t = ws1["A1"]
    t.value = "Quality Comparison Report"
    t.font = Font(bold=True, size=18, color="2F5496")
    t.alignment = Alignment(horizontal="center")

    ws1.cell(row=3, column=1, value="Generated:").font = Font(bold=True)
    ws1.cell(row=3, column=2, value=report.timestamp)
    ws1.cell(row=4, column=1, value="Total Controls in RCM:").font = Font(bold=True)
    ws1.cell(row=4, column=2, value=report.total_controls)
    ws1.cell(row=5, column=1, value="Controls Compared (TOD):").font = Font(bold=True)
    ws1.cell(row=5, column=2, value=len(tod_comps))
    ws1.cell(row=6, column=1, value="Controls Compared (TOE):").font = Font(bold=True)
    ws1.cell(row=6, column=2, value=len(toe_comps))

    tod_match = sum(1 for c in tod_comps if _verdicts_same(c.our_effectiveness, c.human_effectiveness))
    toe_match = sum(1 for c in toe_comps if _verdicts_same(c.our_effectiveness, c.human_effectiveness))
    ws1.cell(row=8, column=1, value="TOD Verdict Agreement:").font = Font(bold=True)
    tod_pct = f" ({tod_match/len(tod_comps):.0%})" if tod_comps else ""
    ws1.cell(row=8, column=2, value=f"{tod_match} / {len(tod_comps)}{tod_pct}" if tod_comps else "N/A")
    ws1.cell(row=9, column=1, value="TOE Verdict Agreement:").font = Font(bold=True)
    toe_pct = f" ({toe_match/len(toe_comps):.0%})" if toe_comps else ""
    ws1.cell(row=9, column=2, value=f"{toe_match} / {len(toe_comps)}{toe_pct}" if toe_comps else "N/A")

    # Severity breakdown
    sev_row = 11
    ws1.cell(row=sev_row, column=1, value="Gap Severity Breakdown:").font = Font(bold=True)
    sev_col = 2
    for sev_label in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        cnt = report.severity_counts.get(sev_label, 0)
        c = ws1.cell(row=sev_row, column=sev_col, value=f"{sev_label}: {cnt}")
        c.font = Font(bold=True, color="FFFFFF" if sev_label == "CRITICAL" else "000000")
        c.fill = PatternFill(start_color=SEVERITY_COLORS[sev_label],
                             end_color=SEVERITY_COLORS[sev_label], fill_type="solid")
        sev_col += 1

    # Control overview table
    ws1.cell(row=13, column=1, value="Control Overview").font = Font(bold=True, size=14, color="2F5496")
    ov_headers = [
        "Control ID", "Control Description",
        "TOD: Our Verdict", "TOD: Human Verdict", "TOD Match?",
        "TOE: Our Verdict", "TOE: Human Verdict", "TOE Match?",
        "Severity",
    ]
    _write_header(ws1, 14, ov_headers)

    rn = 15
    for cid in all_cids:
        tc = tod_by_cid.get(cid)
        ec = toe_by_cid.get(cid)
        tod_our = tc.our_effectiveness if tc else "—"
        tod_hum = tc.human_effectiveness if tc else "—"
        tod_same = ("Yes" if tc and _verdicts_same(tc.our_effectiveness, tc.human_effectiveness)
                    else ("No" if tc else "—"))
        toe_our = ec.our_effectiveness if ec else "—"
        toe_hum = ec.human_effectiveness if ec else "—"
        toe_same = ("Yes" if ec and _verdicts_same(ec.our_effectiveness, ec.human_effectiveness)
                    else ("No" if ec else "—"))

        # Determine worst severity across TOD + TOE for this control
        worst_sev = "NONE"
        sev_rank = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1, "NONE": 0}
        for comp in (tc, ec):
            if comp and sev_rank.get(comp.gap_severity, 0) > sev_rank.get(worst_sev, 0):
                worst_sev = comp.gap_severity

        _cell(ws1, rn, 1, cid)
        _cell(ws1, rn, 2, (ctrl_desc.get(cid, "") or "—")[:120], align=wrap)
        _cell(ws1, rn, 3, tod_our)
        _cell(ws1, rn, 4, tod_hum)
        _cell(ws1, rn, 5, tod_same, fill=green_fill if tod_same == "Yes" else (red_fill if tod_same == "No" else None))
        _cell(ws1, rn, 6, toe_our)
        _cell(ws1, rn, 7, toe_hum)
        _cell(ws1, rn, 8, toe_same, fill=green_fill if toe_same == "Yes" else (red_fill if toe_same == "No" else None))
        sev_color = SEVERITY_COLORS.get(worst_sev, "C6EFCE")
        sev_font_color = "FFFFFF" if worst_sev == "CRITICAL" else "000000"
        c = _cell(ws1, rn, 9, worst_sev,
                  fill=PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid"))
        c.font = Font(bold=True, color=sev_font_color)
        rn += 1

    for i, w in enumerate([14, 40, 20, 20, 12, 20, 20, 12, 14], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    if all_cids:
        ws1.auto_filter.ref = f"A14:{get_column_letter(len(ov_headers))}{rn - 1}"
    ws1.freeze_panes = "A15"

    # ═══════════════════════════════════════════════════════════════
    #  Helper — write a comparison sheet (reused for TOD & TOE)
    # ═══════════════════════════════════════════════════════════════

    # Verdict cell coloring helper
    _verdict_fill = {
        "effective": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "effective with exceptions": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "not effective": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    def _write_comparison_sheet(ws, comparisons, tab_color):
        ws.sheet_properties.tabColor = tab_color
        headers = [
            "Control ID", "Process", "Control Description",
            "Frequency", "Risk Level",
            "Our Verdict", "Human Verdict", "Verdict Match",
            "Severity",
            "Our Attributes", "Human Attributes",
            "Attribute Gaps",
            "Gap Analysis",
        ]
        _write_header(ws, 1, headers)

        r = 2
        for comp in comparisons:
            same = "Yes" if _verdicts_same(comp.our_effectiveness, comp.human_effectiveness) else "No"

            # Build human-readable attribute gap descriptions
            diff_parts = []
            seen_attrs = set()
            for d in comp.attribute_differences:
                aid = d.get("attr_id", "")
                aname = d.get("attr_name", aid)
                label = f"{aid} ({aname})" if aname and aname != aid else aid
                if label not in seen_attrs:
                    seen_attrs.add(label)
                    diff_parts.append(label)
            for missed in comp.attributes_human_missed:
                label = f"{missed} (not tested by human)"
                if label not in seen_attrs:
                    seen_attrs.add(label)
                    diff_parts.append(label)
            for extra in comp.attributes_human_extra:
                eid = extra.get("id", "")
                ename = extra.get("name", eid)
                label = f"{eid} ({ename}) — extra by human, not in tool"
                diff_parts.append(label)
            diff_str = "; ".join(diff_parts) if diff_parts else "None"

            _cell(ws, r, 1, comp.control_id)
            _cell(ws, r, 2, ctrl_process.get(comp.control_id, "") or "—", align=wrap)
            _cell(ws, r, 3, (ctrl_desc.get(comp.control_id, "") or "—")[:200], align=wrap)
            _cell(ws, r, 4, ctrl_freq.get(comp.control_id, "") or "—")
            _cell(ws, r, 5, ctrl_risk.get(comp.control_id, "") or "—")

            # Verdict cells with conditional coloring
            our_v = comp.our_effectiveness or "—"
            hum_v = comp.human_effectiveness or "—"
            _cell(ws, r, 6, our_v, fill=_verdict_fill.get(our_v.lower(), None))
            _cell(ws, r, 7, hum_v, fill=_verdict_fill.get(hum_v.lower(), None))
            _cell(ws, r, 8, same, fill=green_fill if same == "Yes" else red_fill)

            # Severity with color
            sev = comp.gap_severity
            sev_color = SEVERITY_COLORS.get(sev, "C6EFCE")
            sev_font_color = "FFFFFF" if sev == "CRITICAL" else "000000"
            c = _cell(ws, r, 9, sev,
                      fill=PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid"))
            c.font = Font(bold=True, color=sev_font_color)

            _cell(ws, r, 10, len(comp.our_attributes))
            _cell(ws, r, 11, len(comp.human_attributes))
            _cell(ws, r, 12, diff_str, align=wrap)
            narrative = comp.overall_narrative or _build_fallback_narrative(comp)
            _cell(ws, r, 13, narrative, align=wrap)
            r += 1

        for i, w in enumerate([14, 18, 40, 14, 12, 22, 22, 12, 12, 12, 12, 35, 55], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        if comparisons:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r - 1}"
        ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    #  Helper — write an attribute detail sheet (reused for TOD & TOE)
    # ═══════════════════════════════════════════════════════════════
    # Result cell coloring helper (Yes/No)
    _yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _not_tested_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    def _result_fill(val):
        v = str(val).strip().lower()
        if v == "yes":
            return _yes_fill
        if v == "no":
            return _no_fill
        if "not tested" in v:
            return _not_tested_fill
        return None

    def _write_attr_detail_sheet(ws, comparisons, tab_color, include_sample_col=False):
        ws.sheet_properties.tabColor = tab_color
        headers = ["Control ID", "Process"]
        if include_sample_col:
            headers.append("Sample #")
        headers += [
            "Our Attr ID", "Our Attribute Name",
            "Matched Human Attr ID", "Matched Human Attr Name",
            "Match Method", "Match Confidence",
            "Our Result", "Human Result",
            "Same?", "Direction",
        ]
        _write_header(ws, 1, headers)

        r = 2
        for comp in comparisons:
            proc = ctrl_process.get(comp.control_id, "") or "—"
            details = comp.attribute_comparison_detail
            if not details and comp.attribute_match_details:
                # Fallback: show match info without per-sample results
                for md in comp.attribute_match_details:
                    ci = 1
                    _cell(ws, r, ci, comp.control_id); ci += 1
                    _cell(ws, r, ci, proc); ci += 1
                    if include_sample_col:
                        _cell(ws, r, ci, "—"); ci += 1
                    _cell(ws, r, ci, md.get("our_id", "") or "—"); ci += 1
                    _cell(ws, r, ci, md.get("our_name", "") or "—", align=wrap); ci += 1
                    human_id = md.get("human_id", "")
                    _cell(ws, r, ci, human_id or "—"); ci += 1
                    _cell(ws, r, ci, md.get("human_name", "") or "—", align=wrap); ci += 1
                    mt = md.get("match_type", "") or "No Match"
                    _cell(ws, r, ci, mt); ci += 1
                    sim = md.get("similarity", 0)
                    _cell(ws, r, ci, f"{sim:.0%}" if sim else "—"); ci += 1
                    _cell(ws, r, ci, "—"); ci += 1  # our result (no per-sample data)
                    hr = "Not Tested" if not human_id else "—"
                    _cell(ws, r, ci, hr, fill=_not_tested_fill if not human_id else None); ci += 1
                    same_val = "No" if not human_id else "—"
                    _cell(ws, r, ci, same_val,
                          fill=red_fill if not human_id else None); ci += 1
                    direction = "Human Did Not Test" if not human_id else "—"
                    _cell(ws, r, ci, direction); ci += 1
                    r += 1

                # Also append human-extra rows in fallback path
                extra_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                for extra in comp.attributes_human_extra:
                    ci = 1
                    _cell(ws, r, ci, comp.control_id); ci += 1
                    _cell(ws, r, ci, proc); ci += 1
                    if include_sample_col:
                        _cell(ws, r, ci, "—"); ci += 1
                    _cell(ws, r, ci, "—"); ci += 1
                    _cell(ws, r, ci, "— (not tested by tool)", align=wrap, fill=extra_fill); ci += 1
                    _cell(ws, r, ci, extra.get("id", "—")); ci += 1
                    _cell(ws, r, ci, extra.get("name", "—"), align=wrap); ci += 1
                    _cell(ws, r, ci, "Human Extra"); ci += 1
                    _cell(ws, r, ci, "—"); ci += 1
                    _cell(ws, r, ci, "—", fill=extra_fill); ci += 1
                    _cell(ws, r, ci, "Tested", fill=_yes_fill); ci += 1
                    _cell(ws, r, ci, "No", fill=yellow_fill); ci += 1
                    _cell(ws, r, ci, "Tool Did Not Test This Attribute"); ci += 1
                    r += 1

                continue

            for d in details:
                ci = 1
                _cell(ws, r, ci, comp.control_id); ci += 1
                _cell(ws, r, ci, proc); ci += 1
                if include_sample_col:
                    _cell(ws, r, ci, d.get("sample_id", "") or "—"); ci += 1
                _cell(ws, r, ci, d.get("our_attr_id", "") or "—"); ci += 1
                _cell(ws, r, ci, d.get("our_attr_name", "") or "—", align=wrap); ci += 1
                _cell(ws, r, ci, d.get("human_attr_id", "") or "—"); ci += 1
                _cell(ws, r, ci, d.get("human_attr_name", "") or "—", align=wrap); ci += 1
                _cell(ws, r, ci, d.get("match_method", "") or "—"); ci += 1
                conf = d.get("match_confidence", 0)
                _cell(ws, r, ci, f"{conf:.0%}" if conf else "—"); ci += 1
                our_res = d.get("our_result", "") or "—"
                hum_res = d.get("human_result", "") or "—"
                _cell(ws, r, ci, our_res, fill=_result_fill(our_res)); ci += 1
                _cell(ws, r, ci, hum_res, fill=_result_fill(hum_res)); ci += 1
                same = d.get("same", True)
                _cell(ws, r, ci, "Yes" if same else "No",
                      fill=green_fill if same else red_fill); ci += 1
                direction = d.get("direction", "")
                _cell(ws, r, ci, direction if direction else ("—" if same else "—"), align=wrap); ci += 1
                r += 1

            # Append rows for human-extra attributes (human tested, tool did not)
            extra_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            for extra in comp.attributes_human_extra:
                ci = 1
                _cell(ws, r, ci, comp.control_id); ci += 1
                _cell(ws, r, ci, proc); ci += 1
                if include_sample_col:
                    _cell(ws, r, ci, "—"); ci += 1
                _cell(ws, r, ci, "—"); ci += 1                                          # Our Attr ID
                _cell(ws, r, ci, "— (not tested by tool)", align=wrap, fill=extra_fill); ci += 1  # Our Attr Name
                _cell(ws, r, ci, extra.get("id", "—")); ci += 1                         # Human Attr ID
                _cell(ws, r, ci, extra.get("name", "—"), align=wrap); ci += 1            # Human Attr Name
                _cell(ws, r, ci, "Human Extra"); ci += 1                                 # Match Method
                _cell(ws, r, ci, "—"); ci += 1                                           # Match Confidence
                _cell(ws, r, ci, "—", fill=extra_fill); ci += 1                          # Our Result
                _cell(ws, r, ci, "Tested", fill=_yes_fill); ci += 1                      # Human Result
                _cell(ws, r, ci, "No", fill=yellow_fill); ci += 1                        # Same?
                _cell(ws, r, ci, "Tool Did Not Test This Attribute"); ci += 1             # Direction
                r += 1

        # Column widths
        if include_sample_col:
            widths = [14, 18, 10, 10, 35, 10, 35, 16, 12, 12, 12, 10, 30]
        else:
            widths = [14, 18, 10, 35, 10, 35, 16, 12, 12, 12, 10, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        if comparisons:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r - 1}"
        ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    #  Sheet 2 — TOD Comparison
    # ═══════════════════════════════════════════════════════════════
    ws_tod = wb.create_sheet("TOD Comparison")
    _write_comparison_sheet(ws_tod, tod_comps, TAB_BLUE)

    # ═══════════════════════════════════════════════════════════════
    #  Sheet 3 — TOD Attribute Detail
    # ═══════════════════════════════════════════════════════════════
    ws_tod_attr = wb.create_sheet("TOD Attribute Detail")
    _write_attr_detail_sheet(ws_tod_attr, tod_comps, TAB_LIGHT_BLUE, include_sample_col=False)

    # ═══════════════════════════════════════════════════════════════
    #  Sheet 4 — TOE Comparison
    # ═══════════════════════════════════════════════════════════════
    ws_toe = wb.create_sheet("TOE Comparison")
    _write_comparison_sheet(ws_toe, toe_comps, TAB_GREEN)

    # ═══════════════════════════════════════════════════════════════
    #  Sheet 5 — TOE Attribute Detail  (per-sample breakdown)
    # ═══════════════════════════════════════════════════════════════
    ws_toe_attr = wb.create_sheet("TOE Attribute Detail")
    _write_attr_detail_sheet(ws_toe_attr, toe_comps, TAB_LIGHT_GREEN, include_sample_col=True)

    # ── Save ──────────────────────────────────────────────────────
    wb.save(output_path)
    logger.info("Comparison report saved to: %s", output_path)


# ═══════════════════════════════════════════════════════════════════════════════
#  QUALITY COMPARISON ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class QualityComparisonEngine:
    """
    Runs the full control testing pipeline autonomously, then compares
    our results against human auditor workpapers.
    """

    def __init__(
        self,
        rcm_path: str,
        tod_evidence_folder: str,
        toe_evidence_folder: str,
        human_tod_workpaper_path: str,
        human_toe_workpaper_path: str,
        output_dir: str,
        *,
        openai_api_key: str = "",
        openai_model: str = "gpt-4o-mini",
        azure_endpoint: str = "",
        azure_api_key: str = "",
        azure_deployment: str = "gpt-4o-mini",
        azure_api_version: str = "2024-12-01-preview",
        max_workers: int = 5,
        company_name: str = "",
        prepared_by: str = "",
        reviewed_by: str = "",
        risk_score_map: dict = None,
        risk_bands: list = None,
        sampling_methodology: str = "kpmg",
        custom_sampling_file_path: str = None,
    ):
        self.rcm_path = rcm_path
        self.tod_evidence_folder = tod_evidence_folder
        self.toe_evidence_folder = toe_evidence_folder
        self.human_tod_workpaper_path = human_tod_workpaper_path
        self.human_toe_workpaper_path = human_toe_workpaper_path
        self.output_dir = output_dir
        self.openai_api_key = openai_api_key or azure_api_key
        self.openai_model = openai_model
        self.azure_endpoint = azure_endpoint
        self.azure_api_key = azure_api_key or openai_api_key
        self.azure_deployment = azure_deployment
        self.azure_api_version = azure_api_version
        self.max_workers = max_workers
        self.company_name = company_name
        self.prepared_by = prepared_by
        self.reviewed_by = reviewed_by
        self.risk_score_map = risk_score_map   # custom weights (None = defaults)
        self.risk_bands = risk_bands           # custom bands (None = defaults)
        self.sampling_methodology = sampling_methodology or "kpmg"
        self.custom_sampling_file_path = custom_sampling_file_path
        self.progress_callback = None  # optional (current, total, message) callback

    def _emit_progress(self, current: int, total: int, message: str):
        """Emit progress if a callback is registered."""
        if self.progress_callback:
            try:
                self.progress_callback(current, total, message)
            except Exception:
                pass

    def run(self) -> ComparisonReport:
        """Execute the full autonomous pipeline and comparison."""
        os.makedirs(self.output_dir, exist_ok=True)

        # Ensure engines directory is on path
        engines_dir = str(Path(__file__).resolve().parent)
        if engines_dir not in sys.path:
            sys.path.insert(0, engines_dir)

        import TOD_Engine
        from rcm_reader import smart_read_file

        # ── 1. Load RCM ───────────────────────────────────────────
        self._emit_progress(0, 12, "Loading RCM")
        logger.info("Step 1: Loading RCM from %s", self.rcm_path)

        # Read original (unmapped column names) for workpaper export
        original_rcm_df = smart_read_file(self.rcm_path, normalize_columns=False)

        # Read normalised (canonical column names) for engine processing
        rcm_df, read_info = smart_read_file(self.rcm_path, normalize_columns=True)

        logger.info("RCM loaded: %d rows, %d columns", len(rcm_df), len(rcm_df.columns))

        # Drop rows with empty/null Control Id (trailing merged-cell artefacts)
        if "Control Id" in rcm_df.columns:
            _empty_cid = rcm_df["Control Id"].isna() | (
                rcm_df["Control Id"].astype(str).str.strip().isin(
                    ["", "nan", "none", "null"]
                )
            )
            n_dropped = _empty_cid.sum()
            if n_dropped > 0:
                logger.info("Dropping %d rows with no Control Id", n_dropped)
                keep_mask = ~_empty_cid
                rcm_df = rcm_df[keep_mask].reset_index(drop=True)
                # Keep original_rcm_df in sync (same row indices)
                if len(original_rcm_df) == len(keep_mask):
                    original_rcm_df = original_rcm_df[keep_mask.values].reset_index(drop=True)

        # ── Risk Level Inference (same logic as main pipeline) ─────
        # Uses weighted non-linear scoring (P x I), keyword matching,
        # and LLM inference — auto-applied (no user approval in QC).
        self._emit_progress(0.5, 12, "Inferring risk levels")
        logger.info("Step 1b: Running risk level inference (weighted scoring + keyword + LLM)")
        rcm_df = _infer_risk_levels_for_qc(rcm_df, self.risk_score_map, self.risk_bands)

        # ── 2. Apply sampling ──────────────────────────────────────
        self._emit_progress(1, 12, "Applying sampling")
        custom_table = None
        if self.sampling_methodology == "custom" and self.custom_sampling_file_path:
            logger.info("Step 2: Loading custom sampling table from %s",
                        self.custom_sampling_file_path)
            try:
                from agent.tools.sampling_engine import _parse_custom_table
                custom_table = _parse_custom_table(self.custom_sampling_file_path)
                logger.info("Custom sampling table loaded: %d rows", len(custom_table))
            except Exception as e:
                logger.warning("Failed to load custom sampling table: %s — falling back to KPMG", e)
        else:
            logger.info("Step 2: Applying KPMG sampling")
        rcm_df = _apply_sampling(rcm_df, custom_table)

        # ── 3. Initialize tester ───────────────────────────────────
        logger.info("Step 3: Initializing RCMControlTester")
        tester = TOD_Engine.RCMControlTester(
            normalized_rcm_df=rcm_df,
            openai_api_key=self.openai_api_key,
            openai_model=self.openai_model,
            azure_endpoint=self.azure_endpoint,
            azure_api_key=self.azure_api_key,
            azure_deployment=self.azure_deployment,
            azure_api_version=self.azure_api_version,
            original_rcm_df=original_rcm_df,
        )

        # ── 4. Load TOD evidence ───────────────────────────────────
        self._emit_progress(3, 12, "Loading TOD evidence")
        logger.info("Step 4: Loading TOD evidence from %s", self.tod_evidence_folder)
        tod_bank = TOD_Engine.load_tod_evidence_folder(self.tod_evidence_folder)
        logger.info("TOD evidence: %d controls, %d samples",
                     len(tod_bank), sum(len(v) for v in tod_bank.values()))

        # Normalize evidence keys to match RCM
        if "Control Id" in rcm_df.columns:
            rcm_ids = {cid.strip().upper(): cid.strip()
                       for cid in rcm_df["Control Id"].astype(str)}
            normalized_bank = {}
            for folder_key, evidence in tod_bank.items():
                upper_key = folder_key.strip().upper()
                matched_id = rcm_ids.get(upper_key, folder_key.strip())
                normalized_bank[matched_id] = evidence
            tod_bank = normalized_bank

        if not tod_bank:
            raise ValueError(
                f"No evidence found in {self.tod_evidence_folder}. "
                "Expected subfolders named by Control ID."
            )

        # ── 5. Generate schemas autonomously ───────────────────────
        self._emit_progress(4, 12, "Generating TOD schemas")
        logger.info("Step 5: Generating TOD schemas (autonomous, no approval)")
        schemas = {}
        valid_cids = list(tod_bank.keys())

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_cid = {}
            for cid in valid_cids:
                if cid in tester.rcm_lookup:
                    rcm_row = tester.rcm_lookup[cid]
                    future = executor.submit(tester.evaluator.generate_schema, rcm_row)
                    future_to_cid[future] = cid

            for future in as_completed(future_to_cid):
                cid = future_to_cid[future]
                try:
                    schema = future.result()
                    schemas[cid] = schema
                    logger.info("Schema generated for %s", cid)
                except Exception as e:
                    logger.error("Schema generation failed for %s: %s", cid, e)

        logger.info("Schemas generated for %d / %d controls", len(schemas), len(valid_cids))

        # ── 6. Run TOD ─────────────────────────────────────────────
        self._emit_progress(5, 12, "Running Test of Design")
        logger.info("Step 6: Running Test of Design")
        tod_results, tod_schemas = tester.test_all_tod(
            tod_bank, max_workers=self.max_workers, pre_schemas=schemas,
        )

        tod_output = os.path.join(self.output_dir, "QC_TOD_Results.xlsx")
        # Sanitize LLM-generated text to strip illegal XML/Excel characters
        for r in tod_results:
            _sanitize_for_excel(r)
        tester.export_tod_workpaper(
            tod_results, tod_output, tod_bank=tod_bank,
            company_name=self.company_name,
            prepared_by=self.prepared_by,
            reviewed_by=self.reviewed_by,
        )
        logger.info("TOD complete: %d results, saved to %s", len(tod_results), tod_output)

        # ── 7. Filter TOD-PASS controls for TOE ───────────────────
        tod_pass_cids = {
            r.control_id for r in tod_results
            if getattr(r, "result", None) == "PASS"
        }
        logger.info("TOD PASS: %d controls proceed to TOE", len(tod_pass_cids))

        # ── 8. Load TOE evidence ───────────────────────────────────
        self._emit_progress(7, 12, "Loading TOE evidence")
        logger.info("Step 8: Loading TOE evidence from %s", self.toe_evidence_folder)
        toe_bank = TOD_Engine.load_toe_evidence_folder(
            self.toe_evidence_folder,
            include_control_ids=tod_pass_cids,
        )
        logger.info("TOE evidence: %d controls, %d samples",
                     len(toe_bank), sum(len(v) for v in toe_bank.values()))

        # ── 9. Run TOE ─────────────────────────────────────────────
        self._emit_progress(8, 12, "Running Test of Effectiveness")
        logger.info("Step 9: Running Test of Effectiveness")
        toe_results = tester.test_all_toe(
            toe_bank, max_workers=self.max_workers,
            pre_schemas=tod_schemas,
            tod_results=tod_results,
        )

        toe_output = os.path.join(self.output_dir, "QC_TOE_Results.xlsx")
        for r in toe_results:
            _sanitize_for_excel(r)
        tester.export_toe_workpaper(
            toe_results, toe_output, toe_bank=toe_bank,
            company_name=self.company_name,
            prepared_by=self.prepared_by,
            reviewed_by=self.reviewed_by,
        )
        logger.info("TOE complete: %d results, saved to %s", len(toe_results), toe_output)

        # ── 10. Parse human workpapers ─────────────────────────────
        self._emit_progress(9, 12, "Parsing human workpapers")
        parser = HumanWorkpaperParser()

        logger.info("Step 10a: Parsing human TOD workpaper from %s", self.human_tod_workpaper_path)
        human_tod_results = parser.parse(self.human_tod_workpaper_path)
        logger.info("Human TOD workpaper: %d controls parsed", len(human_tod_results))

        logger.info("Step 10b: Parsing human TOE workpaper from %s", self.human_toe_workpaper_path)
        human_toe_results = parser.parse(self.human_toe_workpaper_path)
        logger.info("Human TOE workpaper: %d controls parsed", len(human_toe_results))

        # Read raw sheets for LLM-first extraction (format-agnostic, no regex dependency)
        tod_raw_sheets: Optional[Dict[str, List[List[str]]]] = None
        toe_raw_sheets: Optional[Dict[str, List[List[str]]]] = None
        try:
            logger.info("Step 10c: Reading raw TOD workpaper sheets for LLM extraction")
            tod_raw_sheets = _read_workbook_sheets_raw(self.human_tod_workpaper_path)
            logger.info("Raw TOD sheets: %d sheets loaded", len(tod_raw_sheets))
        except Exception as e:
            logger.warning("Failed to read raw TOD sheets (will fall back to parsed data): %s", e)
        try:
            logger.info("Step 10d: Reading raw TOE workpaper sheets for LLM extraction")
            toe_raw_sheets = _read_workbook_sheets_raw(self.human_toe_workpaper_path)
            logger.info("Raw TOE sheets: %d sheets loaded", len(toe_raw_sheets))
        except Exception as e:
            logger.warning("Failed to read raw TOE sheets (will fall back to parsed data): %s", e)

        # ── 11. Compare ────────────────────────────────────────────
        self._emit_progress(10, 12, "Comparing results")
        logger.info("Step 11a: Comparing TOD results")
        adapted_tod = _adapt_tod_results(tod_results)
        tod_report = _compare_controls(
            adapted_tod, human_tod_results, tod_schemas,
            total_rcm_controls=len(rcm_df), phase="TOD",
            evaluator=tester.evaluator, max_workers=self.max_workers,
            human_raw_sheets=tod_raw_sheets,
        )

        # Filter out "TOD Failed" stubs — test_all_toe includes them for
        # the workpaper export, but they shouldn't be compared against the
        # human's TOE workpaper (our tool never actually ran TOE on them).
        actual_toe_results = [
            r for r in toe_results
            if "TOD Failed" not in (getattr(r, "operating_effectiveness", "") or "")
            and "No Evidence" not in (getattr(r, "operating_effectiveness", "") or "")
        ]
        logger.info(
            "Step 11b: Comparing TOE results (%d actual, %d TOD-failed stubs excluded)",
            len(actual_toe_results), len(toe_results) - len(actual_toe_results),
        )
        toe_report = _compare_controls(
            actual_toe_results, human_toe_results, tod_schemas,
            total_rcm_controls=len(rcm_df), phase="TOE",
            evaluator=tester.evaluator, max_workers=self.max_workers,
            human_raw_sheets=toe_raw_sheets,
        )

        # Merge both reports into one
        report = ComparisonReport(
            timestamp=datetime.now().isoformat(),
            total_controls=len(rcm_df),
            control_comparisons=(
                tod_report.control_comparisons + toe_report.control_comparisons
            ),
        )
        for comp in report.control_comparisons:
            report.severity_counts[comp.gap_severity] = (
                report.severity_counts.get(comp.gap_severity, 0) + 1
            )
        report.controls_with_gaps = len({
            c.control_id for c in report.control_comparisons if c.gap_severity != "NONE"
        })
        report.tod_output_path = tod_output
        report.toe_output_path = toe_output

        # ── 12. Export comparison report ───────────────────────────
        self._emit_progress(11, 12, "Exporting comparison report")
        comparison_output = os.path.join(self.output_dir, "QC_Comparison_Report.xlsx")
        _export_comparison_report(report, comparison_output, schemas=tod_schemas, rcm_df=rcm_df)
        report.comparison_output_path = comparison_output

        logger.info(
            "Quality Comparison complete: %d controls, %d with gaps "
            "(CRITICAL=%d, HIGH=%d, MEDIUM=%d, LOW=%d)",
            report.total_controls,
            report.controls_with_gaps,
            report.severity_counts.get("CRITICAL", 0),
            report.severity_counts.get("HIGH", 0),
            report.severity_counts.get("MEDIUM", 0),
            report.severity_counts.get("LOW", 0),
        )

        return report
